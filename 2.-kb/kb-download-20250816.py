from selenium import webdriver
import sys, os, json
from selenium.webdriver.chrome.service import Service
from PyQt6.QtGui import QPalette, QColor, QIcon
from PyQt6.QtCore import Qt
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import openpyxl  # 이 줄 추가!
from openpyxl.utils import get_column_letter  # 이 줄 추가!
from datetime import datetime
from selenium.webdriver.common.keys import Keys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QLabel, QLineEdit, QPushButton, QMessageBox,
                           QHBoxLayout, QButtonGroup, QGridLayout, QFileDialog, QDialog,
                           QTextEdit, QTabWidget, QCheckBox, QProgressBar)  # QProgressBar 추가
from PyQt6.QtWidgets import QRadioButton

# 설정 파일 경로
CONFIG_FILE = 'apartment_config.json'

def load_config():
    """설정 파일 로드"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {'save_path': ''}
    return {'save_path': ''}

def save_config(config):
    """설정 파일 저장"""
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)


class SettingsDialog(QDialog):
    def __init__(self, parent=None, save_path=""):
        super().__init__(parent)
        self.setWindowTitle("설정")
        self.setFixedSize(500, 200)
        
        layout = QVBoxLayout(self)
        
        # 저장 경로 선택 부분
        path_container = QWidget()
        path_layout = QHBoxLayout(path_container)
        
        path_label = QLabel("저장 경로:")
        self.path_input = QLineEdit()
        self.path_input.setText(save_path)
        browse_button = QPushButton("찾아보기")
        browse_button.clicked.connect(self.browse_folder)
        
        path_layout.addWidget(path_label)
        path_layout.addWidget(self.path_input)
        path_layout.addWidget(browse_button)
        
        # 확인 버튼
        confirm_button = QPushButton("확인")
        confirm_button.clicked.connect(self.accept)
        
        layout.addWidget(path_container)
        layout.addWidget(confirm_button)
    
    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "저장 경로 선택")
        if folder:
            self.path_input.setText(folder)
    
    def get_save_path(self):
        return self.path_input.text()


def setup_driver():
    import os
    import sys
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    
    print("=== ChromeDriver 디버깅 시작 ===")
    
    # Chrome 설치 여부 확인
    chrome_paths = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    ]
    
    chrome_found = False
    for chrome_path in chrome_paths:
        if os.path.exists(chrome_path):
            print(f"Chrome 브라우저 발견: {chrome_path}")
            chrome_found = True
            break
    
    if not chrome_found:
        print("❌ Chrome 브라우저가 설치되지 않았습니다!")
        print("Chrome 브라우저를 설치한 후 다시 시도하세요.")
        return None
    
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-plugins')
    chrome_options.add_argument('--disable-images')
    
    # Chrome 경로 명시적 지정
    if chrome_found:
        for chrome_path in chrome_paths:
            if os.path.exists(chrome_path):
                chrome_options.binary_location = chrome_path
                print(f"Chrome 바이너리 경로 설정: {chrome_path}")
                break
    
    try:
        # webdriver-manager를 사용하여 ChromeDriver 자동 관리
        print("webdriver-manager로 ChromeDriver를 설정합니다...")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        print("✅ WebDriver 생성 성공!")
        return driver
        
    except Exception as e:
        print(f"❌ ChromeDriver 실행 실패: {e}")
        print(f"오류 타입: {type(e).__name__}")
        
        # Selenium 4의 자동 드라이버 관리 기능 사용 (fallback)
        try:
            print("Selenium 4 자동 드라이버 관리 기능으로 재시도...")
            driver = webdriver.Chrome(options=chrome_options)
            print("✅ Selenium 4 자동 관리로 성공!")
            return driver
        except Exception as e2:
            print(f"❌ Selenium 4 자동 관리도 실패: {e2}")
            raise Exception(f"모든 방법이 실패했습니다. 마지막 오류: {e2}")
# click_elements 함수 추가 - 조건별 검색에 필요한 함수
# 지역 선택 기능을 추가한 코드

def click_elements(driver, min_area, max_area, max_year, period='6개월', max_price=None, rate_change=None, complex_count=3, save_path="", selected_regions=None, gui=None):
    wait = WebDriverWait(driver, 30)
    long_wait = WebDriverWait(driver, 60)  # 더 긴 대기 시간
    
    # 첫 페이지 여부를 추적하는 변수
    is_first_page = True
    
    try:
        # 1. 지역 선택 박스 클릭
        try:
            # 첫 번째 방법: 직접 선택자로 시도
            region_select = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 
                "#sectionWrapper > div.topselarea > div.selectbox.type2 > span")))
            driver.execute_script("arguments[0].click();", region_select)
            print("지역 선택 박스 클릭 완료")
            time.sleep(3)
        except Exception as e:
            print(f"지역 선택 박스 클릭 첫 번째 방법 실패: {str(e)}")
            try:
                # 두 번째 방법: 텍스트 검색으로 시도
                region_selects = driver.find_elements(By.XPATH, "//span[contains(text(), '지역') or contains(text(), '서울')]")
                if region_selects:
                    driver.execute_script("arguments[0].click();", region_selects[0])
                    print("지역 선택 박스 클릭 완료 (대체 방법)")
                    time.sleep(3)
                else:
                    print("지역 선택 요소를 찾을 수 없음")
            except Exception as e2:
                print(f"지역 선택 박스 클릭 두 번째 방법도 실패: {str(e2)}")
        
        # 2. 지역 '+' 버튼 클릭 (selected_regions에 따라)
        # 기본값 설정: selected_regions가 없거나 비어있으면 '서울'만 선택
        if selected_regions is None or len(selected_regions) == 0:
            selected_regions = ['서울']
        
        print(f"선택할 지역: {selected_regions}")
        
        # 모든 '+' 버튼 찾기
        add_button_xpath = "//span[@role='button' and @title='지역추가' and @class='add']"
        add_buttons = driver.find_elements(By.XPATH, add_button_xpath)
        
        if len(add_buttons) > 0:
            print(f"총 {len(add_buttons)}개의 '+' 버튼 발견")
            
            # 각 버튼의 부모 요소 텍스트 가져오기
            button_regions = {}
            for i, btn in enumerate(add_buttons):
                try:
                    parent = driver.execute_script("return arguments[0].parentNode;", btn)
                    parent_text = driver.execute_script("return arguments[0].textContent;", parent).strip()
                    button_regions[parent_text] = btn
                    print(f"버튼 {i+1}의 부모 요소 텍스트: '{parent_text}'")
                except Exception as e:
                    print(f"버튼 {i+1}의 부모 요소 텍스트 가져오기 실패: {str(e)}")
            
            # 선택한 지역 버튼 클릭
            clicked_regions = []
            for region in selected_regions:
                # 정확히 일치하는 지역명 찾기
                if region in button_regions:
                    btn = button_regions[region]
                    try:
                        driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", btn)
                        print(f"{region} 지역 추가 버튼 클릭 완료")
                        clicked_regions.append(region)
                        time.sleep(1)  # 클릭 후 잠시 대기
                    except Exception as e:
                        print(f"{region} 지역 버튼 클릭 실패: {str(e)}")
                else:
                    # 부분 일치하는 지역명 찾기
                    found = False
                    for region_name, btn in button_regions.items():
                        if region.lower() in region_name.lower() or region_name.lower() in region.lower():
                            try:
                                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                                time.sleep(0.5)
                                driver.execute_script("arguments[0].click();", btn)
                                print(f"{region} 지역 추가 버튼 클릭 완료 (일치: '{region_name}')")
                                clicked_regions.append(region_name)
                                found = True
                                time.sleep(1)  # 클릭 후 잠시 대기
                                break
                            except Exception as e:
                                print(f"{region} 지역 버튼 클릭 실패 (일치: '{region_name}'): {str(e)}")
                    
                    if not found:
                        print(f"{region} 지역을 찾을 수 없어 선택하지 않았습니다.")
            
            # 클릭한 지역이 없으면 서울 선택
            if len(clicked_regions) == 0:
                print("선택한 지역이 없어 기본값 '서울' 선택")
                for i, btn in enumerate(add_buttons):
                    try:
                        parent = driver.execute_script("return arguments[0].parentNode;", btn)
                        parent_text = driver.execute_script("return arguments[0].textContent;", parent).strip()
                        if parent_text == '서울':
                            driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                            time.sleep(0.5)
                            driver.execute_script("arguments[0].click();", btn)
                            print("서울 지역 추가 버튼 클릭 완료 (기본값)")
                            clicked_regions.append('서울')
                            time.sleep(1)
                            break
                    except:
                        continue
                
                # 여전히 클릭한 지역이 없으면 첫 번째 버튼 클릭
                if len(clicked_regions) == 0:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView(true);", add_buttons[0])
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", add_buttons[0])
                        print("첫 번째 지역 추가 버튼 클릭 완료 (최후의 수단)")
                        time.sleep(1)
                    except Exception as e:
                        print(f"첫 번째 지역 추가 버튼 클릭 실패: {str(e)}")
                        
            print(f"선택된 지역: {clicked_regions}")
        else:
            print("'+' 버튼을 찾을 수 없음, JavaScript로 대체 시도")
            # JavaScript로 버튼 찾기 및 클릭 시도
            js_result = driver.execute_script("""
                var regions = arguments[0];
                var clickedRegions = [];
                var allButtons = document.querySelectorAll('span.add[role="button"]');
                
                if (allButtons.length === 0) {
                    console.log('JavaScript: 지역추가 버튼을 찾을 수 없음');
                    return clickedRegions;
                }
                
                // 모든 버튼의 부모 텍스트 가져오기
                var buttonRegions = {};
                for (var i = 0; i < allButtons.length; i++) {
                    var btn = allButtons[i];
                    var parent = btn.parentNode;
                    var regionName = parent ? parent.textContent.trim() : '';
                    if (regionName) {
                        buttonRegions[regionName] = btn;
                    }
                }
                
                // 선택한 지역 버튼 클릭
                for (var i = 0; i < regions.length; i++) {
                    var region = regions[i];
                    
                    // 정확히 일치하는 지역명 찾기
                    if (buttonRegions[region]) {
                        try {
                            buttonRegions[region].scrollIntoView(true);
                            setTimeout(function() {
                                buttonRegions[region].click();
                            }, 500);
                            clickedRegions.push(region);
                            console.log('JavaScript: ' + region + ' 지역 추가 버튼 클릭 완료');
                        } catch (e) {
                            console.log('JavaScript: ' + region + ' 지역 버튼 클릭 실패: ' + e);
                        }
                    } else {
                        // 부분 일치하는 지역명 찾기
                        var found = false;
                        for (var regionName in buttonRegions) {
                            if (regionName.toLowerCase().includes(region.toLowerCase()) || 
                                region.toLowerCase().includes(regionName.toLowerCase())) {
                                try {
                                    buttonRegions[regionName].scrollIntoView(true);
                                    setTimeout(function() {
                                        buttonRegions[regionName].click();
                                    }, 500);
                                    clickedRegions.push(regionName);
                                    found = true;
                                    console.log('JavaScript: ' + region + ' 지역 추가 버튼 클릭 완료 (일치: ' + regionName + ')');
                                    break;
                                } catch (e) {
                                    console.log('JavaScript: ' + region + ' 지역 버튼 클릭 실패 (일치: ' + regionName + '): ' + e);
                                }
                            }
                        }
                        
                        if (!found) {
                            console.log('JavaScript: ' + region + ' 지역을 찾을 수 없어 선택하지 않았습니다.');
                        }
                    }
                }
                
                // 클릭한 지역이 없으면 서울 선택
                if (clickedRegions.length === 0) {
                    console.log('JavaScript: 선택한 지역이 없어 기본값 서울 선택');
                    
                    if (buttonRegions['서울']) {
                        try {
                            buttonRegions['서울'].scrollIntoView(true);
                            setTimeout(function() {
                                buttonRegions['서울'].click();
                            }, 500);
                            clickedRegions.push('서울');
                            console.log('JavaScript: 서울 지역 추가 버튼 클릭 완료 (기본값)');
                        } catch (e) {
                            console.log('JavaScript: 서울 지역 버튼 클릭 실패: ' + e);
                        }
                    } else if (allButtons.length > 0) {
                        try {
                            allButtons[0].scrollIntoView(true);
                            setTimeout(function() {
                                allButtons[0].click();
                            }, 500);
                            console.log('JavaScript: 첫 번째 지역 추가 버튼 클릭 완료 (최후의 수단)');
                        } catch (e) {
                            console.log('JavaScript: 첫 번째 지역 추가 버튼 클릭 실패: ' + e);
                        }
                    }
                }
                
                return clickedRegions;
            """, selected_regions)
            
            if js_result and len(js_result) > 0:
                print(f"JavaScript로 선택된 지역: {js_result}")
            else:
                print("JavaScript로도 지역 선택 실패")

        # 지역 선택 후 충분한 대기 시간 추가
        print("지역 선택 완료 후 추가 대기 중...")
        time.sleep(5)  # 지역 선택 후 5초 대기
        if gui:
            gui.update_progress(40, "지역 선택이 완료되었습니다...")
        
        # 3. 확인 버튼 클릭 - 더 안정적인 방법으로 변경
        try:
            print("확인 버튼 찾는 중...")
            # 확인 버튼이 로드될 때까지 대기
            confirm_button = long_wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "button.btn.round.r30.blue")))
            
            # 버튼이 보이도록 스크롤
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_button)
            time.sleep(2)  # 스크롤 후 잠시 대기
            
            # 버튼 클릭
            driver.execute_script("arguments[0].click();", confirm_button)
            print("확인 버튼 클릭 완료")
            
            # 클릭 후 추가 대기
            print("확인 버튼 클릭 후 데이터 로딩 중...")
            time.sleep(10)  # 확인 버튼 클릭 후 10초 대기
            if gui:
                gui.update_progress(50, "데이터를 로딩하는 중...")
            
        except Exception as e:
            print(f"확인 버튼 클릭 첫 번째 방법 실패: {str(e)}")
            try:
                # 두 번째 방법: 텍스트로 확인 버튼 찾기
                confirm_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '확인')]")
                if confirm_buttons:
                    # 버튼이 사용 가능한 상태가 될 때까지 대기 로직 추가
                    for i in range(10):  # 최대 10회 시도
                        try:
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_buttons[0])
                            time.sleep(1)
                            driver.execute_script("arguments[0].click();", confirm_buttons[0])
                            print(f"확인 버튼 클릭 완료 (대체 방법, {i+1}번째 시도)")
                            # 클릭 후 데이터 로딩 대기
                            print("확인 버튼 클릭 후 데이터 로딩 중...")
                            time.sleep(10)  # 확인 버튼 클릭 후 10초 대기
                            break
                        except Exception as click_error:
                            print(f"클릭 시도 {i+1} 실패: {str(click_error)}")
                            time.sleep(2)  # 재시도 전 대기
                else:
                    # 세 번째 시도: 강제 JavaScript 실행
                    print("확인 버튼을 찾을 수 없어 JavaScript로 시도")
                    driver.execute_script("""
                        var buttons = document.querySelectorAll('button');
                        for (var i = 0; i < buttons.length; i++) {
                            if (buttons[i].textContent.includes('확인')) {
                                buttons[i].scrollIntoView({block: 'center'});
                                setTimeout(function() {
                                    buttons[i].click();
                                    console.log('JavaScript로 확인 버튼 클릭');
                                }, 1000);
                                return true;
                            }
                        }
                        return false;
                    """)
                    print("JavaScript로 확인 버튼 클릭 시도")
                    # 클릭 후 데이터 로딩 대기
                    print("JavaScript 확인 버튼 클릭 후 데이터 로딩 중...")
                    time.sleep(10)  # 확인 버튼 클릭 후 10초 대기
            except Exception as e2:
                print(f"확인 버튼 클릭 대체 방법도 실패: {str(e2)}")
                print("확인 버튼 클릭 실패 - 계속 진행")
        
        # 4. 기간 선택 (인덱스 기반 접근)
        try:
            print(f"선택하려는 기간: {period}")
            
            # 기간에 따른 인덱스 매핑
            period_index = {
                '1년': 0,
                '6개월': 1,
                '3개월': 2,
                '1개월': 3
            }.get(period, 1)  # 기본값 6개월(인덱스 1)
            
            # JavaScript로 인덱스에 해당하는 버튼 클릭
            clicked = driver.execute_script("""
                var buttons = document.querySelectorAll('button.btn.textline');
                var index = arguments[0];
                
                if (buttons.length > index) {
                    console.log("클릭할 버튼 텍스트: " + buttons[index].textContent);
                    buttons[index].click();
                    return true;
                }
                return false;
            """, period_index)
            
            if clicked:
                print(f"{period} 기간 선택 완료 (인덱스 {period_index} 사용)")
            else:
                print(f"{period} 기간 선택 실패")
                
                # 대체 방법 시도
                period_elements = driver.find_elements(By.XPATH, f"//button[text()='{period}']")
                if period_elements:
                    driver.execute_script("arguments[0].click();", period_elements[0])
                    print(f"{period} 기간 선택 완료 (대체 방법)")
            
            # 기간 선택 후 데이터 로딩 대기
            print("기간 선택 후 데이터 로딩 중...")
            time.sleep(5)  # 기간 선택 후 5초 대기
            if gui:
                gui.update_progress(70, "검색 조건을 적용하는 중...")
            
        except Exception as e:
            print(f"기간 선택 중 오류 발생: {str(e)}")
            # 기간 선택을 건너뛰고 계속 진행
            print("기간 선택을 건너뛰고 계속 진행합니다.")
        
        # 5. 전용면적, 연차, 증감률, 평균평단가 입력
        try:
            print("조건 입력 시작...")
            
            # 전용면적 최소값 입력 - 값이 있는 경우에만
            if min_area is not None:
                first_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                    "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2020 > div.fonmhead > div:nth-child(2) > div > input[type=text]:nth-child(1)")))
                first_input.clear()
                first_input.send_keys(str(min_area))
                print(f"최소 면적 {min_area} 입력 완료")
                time.sleep(1)  # 입력 후 잠시 대기
            
            # 전용면적 최대값 입력 - 값이 있는 경우에만
            if max_area is not None:
                second_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                    "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2020 > div.fonmhead > div:nth-child(2) > div > input[type=text]:nth-child(3)")))
                second_input.clear()
                second_input.send_keys(str(max_area))
                print(f"최대 면적 {max_area} 입력 완료")
                time.sleep(1)  # 입력 후 잠시 대기
            
            # 연차 입력 - 값이 있는 경우에만
            if max_year is not None:
                third_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                    "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2040.type3 > div.fonmhead > div > div > input[type=text]:nth-child(3)")))
                third_input.clear()
                third_input.send_keys(str(max_year))
                print(f"연차 최대값 {max_year} 입력 완료")
                time.sleep(1)  # 입력 후 잠시 대기
            
            # 증감률 입력 (입력값이 있는 경우에만)
            if rate_change is not None:
                rate_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                    "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2050 > div.fonmhead > div:nth-child(2) > div > input[type=text]:nth-child(3)")))
                rate_input.clear()
                rate_input.send_keys(str(rate_change))
                print(f"증감률 {rate_change}% 입력 완료")
                time.sleep(1)  # 입력 후 잠시 대기
            
            # 평균평단가 최대값 입력 (입력값이 있는 경우에만)
            # 평균평단가 최대값 입력 (입력값이 있는 경우에만)
            if max_price is not None:
                # GUI에서 입력된 평당 가격을 제곱미터당 가격으로 변환 (3.3으로 나누기)
                converted_price = round(max_price / 3.305785, 2)
                print(f"GUI 입력값(평당): {max_price}, 웹페이지 입력값(㎡당): {converted_price}")
                
                price_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                    "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2050 > div.fonmhead > div:nth-child(1) > div > input[type=text]:nth-child(3)")))
                price_input.clear()
                price_input.send_keys(str(converted_price))
                print(f"평균평단가 최대값 {converted_price}(㎡당) 입력 완료 (원래 입력값: {max_price}평당)")
                
                # 모든 입력 완료 후 엔터키 입력 (검색 실행)
                price_input.send_keys(Keys.ENTER)
                print("엔터키 입력하여 검색 시작")
                
                # 조건 입력 후 충분한 시간 대기하며 데이터 로딩 확인
                print("조건 입력 후 데이터 로딩 중... (충분한 시간 대기)")
                time.sleep(10)  # 조건 입력 후 10초 대기
                
                # 데이터 로딩 상태 확인 (최대 30초 동안 확인)
                start_time = time.time()
                data_loaded = False
                
                while time.time() - start_time < 30:
                    try:
                        # 데이터가 로드되었는지 확인 (bodyitem 요소가 있는지)
                        elements = driver.find_elements(By.CSS_SELECTOR, 
                            "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body > div.bodyitem")
                        if len(elements) > 0:
                            print(f"데이터 로딩 확인: {len(elements)}개 요소 발견")
                            data_loaded = True
                            break
                        # 로딩중 메시지 또는 요소 확인
                        loading_elements = driver.find_elements(By.XPATH, "//div[contains(text(), '로딩') or contains(@class, 'loading')]")
                        if loading_elements:
                            print("로딩 중인 것으로 확인됨...")
                        
                        time.sleep(2)  # 2초마다 확인
                    except Exception as e:
                        print(f"데이터 로딩 확인 중 오류: {str(e)}")
                        time.sleep(2)
                
                if data_loaded:
                    print("데이터가 성공적으로 로드되었습니다.")
                else:
                    print("데이터 로딩 대기 시간이 초과되었습니다. 계속 진행합니다.")
                
                # 추가 대기 시간
                time.sleep(5)
                
            elif min_area is not None or max_area is not None or max_year is not None or rate_change is not None:
                # max_price가 없지만 다른 조건이 있는 경우, 마지막으로 입력한 필드에서 엔터키 입력
                last_input = None
                if rate_change is not None:
                    last_input = rate_input
                elif max_year is not None:
                    last_input = third_input
                elif max_area is not None:
                    last_input = second_input
                elif min_area is not None:
                    last_input = first_input
                
                if last_input:
                    last_input.send_keys(Keys.ENTER)
                    print("엔터키 입력하여 검색 시작")
                    
                    # 조건 입력 후 충분한 시간 대기
                    print("조건 입력 후 데이터 로딩 중...")
                    time.sleep(10)  # 조건 입력 후 10초 대기
            
        except Exception as e:
            print(f"입력 필드 설정 중 오류 발생: {str(e)}")
            print("입력 필드 설정을 건너뛰고 계속 진행합니다.")

        # 매매시세(억원) 버튼 클릭 및 정렬
        try:
            print("시세(만원) 버튼 클릭 시도...")
            price_button = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[@data-v-9bca83a2 and contains(text(), '시세(만원)') and @style='cursor: pointer;']")))
            driver.execute_script("arguments[0].click();", price_button)
            print("시세(만원) 버튼 클릭 완료")
            
            # 버튼 클릭 후 데이터 정렬 대기
            print("데이터 정렬 중...")
            time.sleep(5)  # 정렬 후 5초 대기
            
        except Exception as e:
            print(f"시세(만원) 버튼 클릭 실패: {str(e)}")
            try:
                # JavaScript로 대체 시도
                driver.execute_script("""
                    var headers = document.querySelectorAll('div[data-v-9bca83a2]');
                    for (var i = 0; i < headers.length; i++) {
                        if (headers[i].textContent.trim().includes('시세(만원)') && 
                            headers[i].style.cursor === 'pointer') {
                            headers[i].click();
                            console.log("JavaScript로 시세(만원) 버튼 클릭");
                            break;
                        }
                    }
                """)
                print("JavaScript로 시세(만원) 버튼 클릭 시도")
                time.sleep(5)  # 대기
            except:
                print("JavaScript로 시세(만원) 버튼 클릭도 실패")
        
        # 최종 데이터 로딩 확인
        print("최종 데이터 로딩 확인 중...")
        try:
            # 데이터가 로드될 때까지 대기 (최대 30초)
            elements = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 
                "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body > div.bodyitem")))
            print(f"최종 데이터 로딩 확인 완료: {len(elements)}개 요소 발견")
            
            # 추가 로딩 대기
            time.sleep(5)
        except:
            print("데이터 로딩 대기 시간 초과, 추가 5초 대기 후 계속 진행")
            time.sleep(5)
        
        # 이하 기존 코드와 동일...
        
        # 최대 추출 개수, 한 번에 추출할 개수
        total_count = complex_count
        batch_size = 50
        total_extracted = 0
        all_data = []
        
        # 현재 최대 시세값 (다음 검색에 사용될 필터)
        current_max_price = max_price
        
        # 데이터 스크롤 및 파싱
        while total_extracted < total_count:
            # 현재 배치에서 추출할 개수 계산
            current_batch_size = min(batch_size, total_count - total_extracted)
            print(f"현재 배치 목표 크기: {current_batch_size}")
            
            if gui:
                progress_percentage = 70 + int((total_extracted / total_count) * 20)  # 70-90% 범위
                gui.update_progress(progress_percentage, f"데이터 수집 중... ({total_extracted}/{total_count})")

            
            # 첫 페이지에서만 스크롤 처리 수행
            if is_first_page:
                # 스크롤 처리 함수가 정의되어 있다면 호출 (ensure_all_elements_loaded)
                try:
                    loaded_count = ensure_all_elements_loaded(driver, current_batch_size)
                    print(f"로드된 요소 수: {loaded_count}, 현재 배치 목표: {current_batch_size}")
                except NameError:
                    print("ensure_all_elements_loaded 함수가 정의되지 않았습니다. 스크롤 처리를 건너뜁니다.")
                    # 대체 스크롤 처리
                    try:
                        complex_container = driver.find_element(By.CSS_SELECTOR, 
                            "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body")
                        
                        # 스크롤을 위한 JavaScript 실행
                        for i in range(5):
                            driver.execute_script(
                                f"arguments[0].scrollTop = arguments[0].scrollHeight * {i/4};", 
                                complex_container
                            )
                            time.sleep(0.5)
                        
                        # 다시 맨 위로 스크롤
                        driver.execute_script("arguments[0].scrollTop = 0;", complex_container)
                        time.sleep(1)
                    except Exception as e:
                        print(f"대체 스크롤 처리 중 오류 발생: {str(e)}")
                
                is_first_page = False  # 첫 페이지 처리 완료
            else:
                # 두 번째 페이지부터는 스크롤 없이 바로 현재 요소 수 확인만 수행
                loaded_count = len(driver.find_elements(By.CSS_SELECTOR, 
                    "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body > div.bodyitem"))
                print(f"두 번째 이후 페이지 - 스크롤 없이 확인된 요소 수: {loaded_count}")
            
            # 데이터 파싱
            try:
                data, prices = parse_complex_data(driver, current_batch_size)
            except Exception as e:
                print(f"데이터 파싱 중 오류 발생: {str(e)}")
                data, prices = [], []
            
            # 추출한 데이터가 없으면 종료
            if not data:
                print("추출된 데이터가 없습니다. 데이터 수집을 종료합니다.")
                break
            
            # 추출한 데이터를 전체 리스트에 추가
            all_data.extend(data)
            total_extracted += len(data)
            
            print(f"현재까지 {total_extracted}/{total_count} 개 데이터 추출 완료")
            
            # 다음 배치를 위한 설정 부분 수정
            if len(prices) > 0 and total_extracted < total_count:
                # 현재 배치의 가장 작은 시세 값을 다음 검색의 최대값으로 설정
                min_price_in_batch = min(prices)
                print(f"다음 검색을 위한 최대 시세 값: {min_price_in_batch}")
                
                # 시세 최대값 입력 필드를 찾아 새 값 입력
                try:
                    # 기존 시세 최대값 입력 필드 클리어 및 새 값 입력
                    price_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                        "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2050 > div.fonmhead > div:nth-child(1) > div > input[type=text]:nth-child(3)")))
                    price_input.clear()
                    time.sleep(1)  # 클리어 후 잠시 대기
                    
                    # 약간의 여유를 두기 위해 0.01 뺌
                    # 약간의 여유를 두기 위해 0.01 뺌 (이미 ㎡당 가격이므로 그대로 사용)
                    new_max_price = min_price_in_batch - 0.01
                    price_input.send_keys(str(new_max_price))
                    print(f"최대 시세값 {new_max_price}(㎡당) 입력 완료"
                         )
                    time.sleep(1)
                    
                    # 엔터 키만 입력
                    price_input.send_keys(Keys.ENTER)
                    print("엔터 키 입력 완료")
                    
                    # 다음 페이지 로딩 대기 (최적화된 버전)
                    print("다음 배치 데이터 로딩 중...")
                    try:
                        # 즉시 이전 요소 수 확인
                        old_elements = driver.find_elements(By.CSS_SELECTOR, 
                            "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body > div.bodyitem")
                        old_count = len(old_elements)
                        
                        # 최대 4초만 대기하며 요소 수 변화 확인 (0.5초 간격으로 폴링)
                        start_time = time.time()
                        changed = False
                        
                        while time.time() - start_time < 4:
                            current_elements = driver.find_elements(By.CSS_SELECTOR, 
                                "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body > div.bodyitem")
                            current_count = len(current_elements)
                            
                            # 요소 수가 변하면 페이지 전환이 시작된 것
                            if current_count != old_count:
                                print(f"페이지 전환 감지: {old_count} -> {current_count}")
                                changed = True
                                break
                            
                            time.sleep(0.5)
                        
                        # 짧은 대기 후 진행
                        time.sleep(1)
                        print("다음 페이지 데이터 확인 완료, 계속 진행합니다.")
                        
                    except Exception as e:
                        print(f"다음 페이지 로딩 대기 중 오류: {str(e)}")
                        # 최소 대기 후 진행 (1초로 단축)
                        time.sleep(1)
                        print("오류 발생 후 계속 진행")
                        
                except Exception as e:
                    print(f"다음 검색 설정 중 오류: {str(e)}")
                    break
            else:
                # 더 이상 처리할 데이터가 없거나 요청 개수에 도달
                break
        
        # 모든 데이터 수집 완료 후 엑셀 저장
        # click_elements 함수와 search_by_complex_names 함수 두 곳 모두 수정 필요
        
        # click_elements 함수 내에서 데이터프레임 생성 부분:
        # 모든 데이터 수집 완료 후 엑셀 저장
        # 모든 데이터 수집 완료 후 엑셀 저장
        if all_data:
            if gui:
                gui.update_progress(90, "엑셀 파일을 생성하는 중...")
            
            # 매매시세(억원)와 전세시세 값을 만원 단위로 변환
            # 매매시세(억원)와 전세시세 값을 만원 단위로 변환
            for item in all_data:
                try:
                    # 매매시세(억원) 변환 (억 단위 -> 만원 단위)
                    if '매매시세(억원)' in item and item['매매시세(억원)']:
                        sale_price_str = item['매매시세(억원)'].replace(',', '')
                        if sale_price_str.replace('.', '', 1).isdigit():
                            sale_price_float = float(sale_price_str)
                            # 억 단위를 만원 단위로 변환 (1억 = 10000만원)
                            sale_price_man = int(sale_price_float * 10000)
                            item['매매시세(만원)'] = format(sale_price_man, ',')
                            del item['매매시세(억원)']
                    
                    # 전세시세 변환 (억 단위 -> 만원 단위)
                    if '전세시세' in item and item['전세시세']:
                        rent_price_str = item['전세시세'].replace(',', '')
                        if rent_price_str.replace('.', '', 1).isdigit():
                            rent_price_float = float(rent_price_str)
                            # 억 단위를 만원 단위로 변환
                            rent_price_man = int(rent_price_float * 10000)
                            item['전세시세(만원)'] = format(rent_price_man, ',')
                            del item['전세시세']
                    
                    # 매매전세차 변환 (억 단위 -> 만원 단위)
                    if '매매전세차' in item and item['매매전세차']:
                        diff_price_str = item['매매전세차'].replace(',', '')
                        if diff_price_str.replace('.', '', 1).replace('-', '', 1).isdigit():
                            diff_price_float = float(diff_price_str)
                            # 억 단위를 만원 단위로 변환 (1억 = 10000만원)
                            diff_price_man = int(diff_price_float * 10000)
                            item['매매전세차(만원)'] = format(diff_price_man, ',')
                            del item['매매전세차']
                    
                except Exception as e:
                    print(f"가격 변환 중 오류 발생: {str(e)}")
                    # 오류 발생 시 원래 값 유지
                    continue
            
            # 현재 시간을 파일명에 포함
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

            # 지역명을 파일명에 포함
            region_name = ""
            if selected_regions and len(selected_regions) > 0:
                region_name = "_".join(selected_regions) + "_"

            file_name = f'apartment_data_{region_name}{current_time}.xlsx'
            
            # 저장 경로가 있으면 경로 추가
            if save_path:
                file_path = os.path.join(save_path, file_name)
            else:
                file_path = file_name
            
            # DataFrame 생성 및 엑셀 저장
            # DataFrame 생성 및 엑셀 저장
            # DataFrame 생성 및 엑셀 저장
            df = pd.DataFrame(all_data)
            
            print("DataFrame의 실제 컬럼들:", df.columns.tolist())
            
            # 숫자 형식으로 변환 (조건별 검색)
            def convert_to_numeric_safe(series, is_price=True):
                """안전한 숫자 변환 함수"""
                if series.empty:
                    return series
                
                # 문자열로 변환하고 콤마 제거
                series_str = series.astype(str).str.replace(',', '').str.strip()
                
                if is_price:
                    # 가격 컬럼: 빈 문자열이나 'nan'은 0으로
                    series_str = series_str.replace(['', 'nan', 'None'], '0')
                else:
                    # 일반 숫자 컬럼: 빈 문자열이나 'nan'은 0으로
                    series_str = series_str.replace(['', 'nan', 'None'], '0')
                
                # 퍼센트 기호 제거
                series_str = series_str.str.replace('%', '')
                
                # 숫자로 변환 (오류 시 0으로)
                return pd.to_numeric(series_str, errors='coerce').fillna(0)
            
            # 가격 관련 컬럼들 숫자 변환
            price_columns = ['시세(만원)', '평당시세(만원)', '매매시세(만원)', '전세시세(만원)', '매매전세차(만원)']
            for col in price_columns:
                if col in df.columns:
                    print(f"{col} 컬럼 숫자 변환 중...")
                    df[col] = convert_to_numeric_safe(df[col], is_price=True)
            
            # 증감률 컬럼들 숫자 변환 (퍼센트 제거)
            rate_columns = ['매매증감률', '전세증감률', '전세가율']
            for col in rate_columns:
                if col in df.columns:
                    print(f"{col} 컬럼 숫자 변환 중...")
                    df[col] = convert_to_numeric_safe(df[col], is_price=False)
            
            # 기타 숫자 컬럼들 변환
            # 기타 숫자 컬럼들 변환 (세대수 추가)
            # 기타 숫자 컬럼들 변환 (세대수 포함)
            numeric_columns = ['전용면적', '세대수', '연차']  # 세대수 위치 변경
            for col in numeric_columns:
                if col in df.columns:
                    print(f"{col} 컬럼 숫자 변환 중...")
                    df[col] = convert_to_numeric_safe(df[col], is_price=False)
            
            print("모든 숫자 변환 완료")


            # 시세(만원)을 평당시세(만원)로 변환 (3.3 곱하기)
            if '시세(만원)' in df.columns:
                print("시세(만원)을 평당시세(만원)로 변환 중...")
                # 이미 숫자로 변환된 상태이므로 바로 계산
                df['평당시세(만원)'] = df['시세(만원)'] * 3.305785
                df['평당시세(만원)'] = df['평당시세(만원)'].round(2)  # 소수점 둘째 자리까지
                # 기존 시세(만원) 컬럼 삭제
                df = df.drop('시세(만원)', axis=1)
                print("평당시세 변환 완료")
 


            # 컬럼 순서 정의 (매매시세를 매매증감률 앞으로, 전세시세를 전세증감률 앞으로)
            desired_column_order = [
                '지역명',
                '단지명', 
                '전용면적',
                '세대수',              # 전용면적 바로 옆으로 이동
                '연차', 
                '평당시세(만원)',
                '매매시세(만원)',      # 매매시세를 매매증감률 앞으로
                '매매증감률',          # 매매증감률
                '전세시세(만원)',      # 전세시세를 전세증감률 앞으로  
                '전세증감률',          # 전세증감률
                '전세가율', 
                '매매전세차(만원)'
            ]
            
            # 실제 존재하는 컬럼만 선택하고 순서 적용
            final_columns = []
            for col in desired_column_order:
                if col in df.columns:
                    final_columns.append(col)
            
            # 원하는 순서에 없는 컬럼들도 추가 (누락 방지)
            for col in df.columns:
                if col not in final_columns:
                    final_columns.append(col)
            
            # 컬럼 순서대로 정렬
            df = df[final_columns]
            
            print("최종 컬럼 순서:", final_columns)
            

            
            # ExcelWriter를 사용하여 열 너비 자동 조정 (한 번만!)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='검색결과')
                
                # 워크시트 가져오기
                worksheet = writer.sheets['검색결과']
                
                # 열 너비 자동 조정
                for idx, col in enumerate(df.columns):
                    # 열 문자 (A, B, C, ...) 가져오기
                    column_letter = openpyxl.utils.get_column_letter(idx + 1)
                    
                    # 각 열의 최대 문자 길이 찾기
                    max_length = max([len(str(col))] + [len(str(x)) for x in df[col] if pd.notna(x)])
                    
                    # 너비 설정 (문자 길이 * 1.2 + 4 여백 추가)
                    adjusted_width = max_length * 1.2 + 4
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # 숫자 컬럼에 숫자 형식 적용 (세대수 추가)
                    if col in ['평당시세(만원)', '매매시세(만원)', '전세시세(만원)', '매매전세차(만원)', '매매증감률', '전세증감률', '전세가율', '연차', '전용면적', '세대수']:
                        # 첫 번째 행은 헤더이므로 건너뛰고 2행부터 시작
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=idx + 1)
                            # 가격 컬럼에 숫자 형식 적용
                            if col in ['평당시세(만원)', '매매시세(만원)', '전세시세(만원)', '매매전세차(만원)']:
                                cell.number_format = '#,##0'  # 천 단위 구분 기호 사용
                            # 증감률과 전세가율에 일반 숫자 형식 적용
                            elif col in ['매매증감률', '전세증감률', '전세가율']:
                                cell.number_format = '0.00'  # 일반 숫자 형식 (소수점 2자리)
                            # 다른 숫자 컬럼에 일반 숫자 형식 적용
                            else:
                                cell.number_format = '0.00'  # 소수점 2자리
            
            # ExcelWriter를 사용하여 열 너비 자동 조정
            # ExcelWriter를 사용하여 열 너비 자동 조정
            # ExcelWriter를 사용하여 열 너비 자동 조정
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='검색결과')
                
                # 워크시트 가져오기
                worksheet = writer.sheets['검색결과']
                
                # 열 너비 자동 조정 - 더 직접적인 방법
                for idx, col in enumerate(df.columns):
                    # 열 문자 (A, B, C, ...) 가져오기
                    column_letter = openpyxl.utils.get_column_letter(idx + 1)
                    
                    # 각 열의 최대 문자 길이 찾기
                    max_length = max([len(str(col))] + [len(str(x)) for x in df[col] if pd.notna(x)])
                    
                    # 너비 설정 (문자 길이 * 1.2 + 4 여백 추가)
                    adjusted_width = max_length * 1.2 + 4
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"\n데이터가 {file_path}에 저장되었습니다.")
            print(f"검색된 총 결과 수: {len(all_data)}")
            
            if gui:
                gui.update_progress(100, "저장이 완료되었습니다!")  # 수정된 부분
            
            # 드라이버 종료
            driver.quit()
            print("웹 브라우저를 종료했습니다.")
            
            # 저장 완료 메시지 표시 (드라이버 종료 후) - 추가할 부분
            QMessageBox.information(None, "조건별 검색 완료", 
                                   f"🔍 조건별 검색이 완료되었습니다!\n\n"
                                   f"💾 저장 위치: {file_path}\n"
                                   f"🏢 검색된 결과 수: {len(all_data)}개\n\n"
                                   f"✅ 데이터를 확인해보세요!")
        else:
            # 추출된 데이터가 없는 경우
            # 드라이버 종료
            driver.quit()
            print("웹 브라우저를 종료했습니다.")
            
            QMessageBox.warning(None, "조건별 검색 결과 없음", 
                               f"❌ 검색 결과가 없습니다.\n\n"
                               f"💡 검색 조건을 확인하시거나 다른 조건으로 시도해보세요.")
        # 크롬 창 닫기
        try:
            driver.quit()
            print("웹 브라우저를 종료했습니다.")
        except:
            pass
        
        return True
        
    except Exception as e:
        print(f"에러 발생: {str(e)}")
        # 예외 발생 시에도 크롬 창 닫기
        try:
            driver.quit()
            print("예외 발생으로 웹 브라우저를 종료했습니다.")
        except:
            print("웹 브라우저 종료 실패")
        return False




class ApartmentSearchGUI(QMainWindow):
    
    def __init__(self):
        super().__init__()
        # 설정 로드
        config = load_config()
        self.save_path = config.get('save_path', '')
        # 진행바 관련 변수 초기화
        self.progress_dialog = None
        self.progress_bar = None
        self.initUI()

    def open_blog_link(self, event):
            """크레딧 클릭 시 블로그 링크로 이동"""
            import webbrowser
            try:
                webbrowser.open('https://blog.naver.com/landlover333')
                print("블로그 링크를 열었습니다.")
            except Exception as e:
                print(f"링크 열기 실패: {str(e)}")
                QMessageBox.information(self, "링크", "https://blog.naver.com/landlover333")
    
    def show_progress_dialog(self, title="진행 중...", total_steps=100):
        """진행바 다이얼로그 표시"""
        self.progress_dialog = QDialog(self)
        self.progress_dialog.setWindowTitle(title)
        self.progress_dialog.setFixedSize(400, 120)
        self.progress_dialog.setModal(True)
        
        layout = QVBoxLayout(self.progress_dialog)
        
        # 상태 레이블
        self.status_label = QLabel("작업을 시작합니다...")
        layout.addWidget(self.status_label)
        
        # 진행바
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, total_steps)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        self.progress_dialog.show()
        QApplication.processEvents()  # UI 업데이트 강제
    
    def update_progress(self, value, status_text=""):
        """진행바 업데이트"""
        if self.progress_bar:
            self.progress_bar.setValue(value)
        if self.status_label and status_text:
            self.status_label.setText(status_text)
        QApplication.processEvents()  # UI 업데이트 강제
    
    def close_progress_dialog(self):
        """진행바 다이얼로그 닫기"""
        if self.progress_dialog:
            self.progress_dialog.close()
            self.progress_dialog = None
            self.progress_bar = None
            self.status_label = None


    def open_manual_link(self, event):
            """매뉴얼 클릭 시 매뉴얼 링크로 이동"""
            import webbrowser
            try:
                webbrowser.open('https://blog.naver.com/landlover333/223904190804')
                print("매뉴얼 링크를 열었습니다.")
            except Exception as e:
                print(f"링크 열기 실패: {str(e)}")
                QMessageBox.information(self, "링크", "https://blog.naver.com/landlover333/223904190804")
    
        
    def initUI(self):
        self.setWindowTitle('부태리의 KB시세 프로그램')
        self.setFixedSize(950, 1100)  # 가로 950px, 세로 1100px로 대폭 확장
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 탭 위젯 생성
        self.tabs = QTabWidget()
        
        # 조건별 검색 탭
        self.condition_search_tab = QWidget()
        self.initConditionSearchTab()
        
        # 단지명 검색 탭
        self.complex_search_tab = QWidget()
        self.initComplexSearchTab()
        
        # 탭 추가
        self.tabs.addTab(self.condition_search_tab, "조건별 검색")
        self.tabs.addTab(self.complex_search_tab, "단지명 검색")
        
        # 메인 레이아웃에 탭 추가
        main_layout = QVBoxLayout(central_widget)
        main_layout.addWidget(self.tabs)
        
        # 하단 버튼 영역 (설정 버튼과 매뉴얼 버튼)
        bottom_buttons_layout = QHBoxLayout()
        
        # 설정 버튼 (왼쪽)
        settings_button = QPushButton("설정")
        settings_button.setFixedWidth(100)
        settings_button.clicked.connect(self.open_settings)
        bottom_buttons_layout.addWidget(settings_button)
        
        # 가운데 공간
        bottom_buttons_layout.addStretch()
        
        # 매뉴얼 버튼 (오른쪽)
        manual_button = QLabel('매뉴얼')
        manual_button.setAlignment(Qt.AlignmentFlag.AlignCenter)
        manual_button.setFixedSize(80, 35)
        manual_button.setStyleSheet("""
            QLabel {
                font-size: 12pt;
                font-weight: bold;
                color: #4a90e2;
                border: 2px solid #4a90e2;
                border-radius: 5px;
                padding: 5px;
                cursor: pointer;
                background-color: white;
            }
            QLabel:hover {
                color: white;
                background-color: #4a90e2;
                border: 2px solid #357abd;
            }
        """)
        manual_button.mousePressEvent = self.open_manual_link
        bottom_buttons_layout.addWidget(manual_button)
        
        main_layout.addLayout(bottom_buttons_layout)
        
        # 하단 크레딧 추가
        credit_label = QLabel('만든이 부태리')
        credit_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        credit_label.setStyleSheet("""
            QLabel {
                font-size: 16pt;
                font-weight: bold;
                color: #9370DB;
                margin: 15px;
                padding: 10px;
                cursor: pointer;
            }
            QLabel:hover {
                color: #8A2BE2;
                background-color: #F8F0FF;
                border-radius: 5px;
            }
        """)
        credit_label.mousePressEvent = self.open_blog_link
        
        main_layout.addWidget(credit_label)
        
        # 전체 위젯에 기본 스타일시트 적용
        self.setStyleSheet("""
            QWidget {
                color: black;
                background-color: white;
            }
            QLabel {
                color: black;
            }
            QPushButton {
                color: black;
            }
            QLineEdit {
                color: black;
            }
            QTextEdit {
                color: black;
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            QTabWidget::pane {
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            QTabBar::tab {
                padding: 8px 20px;
                border: 1px solid #ddd;
                border-bottom: none;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background-color: #4a90e2;
                color: white;
            }
            QTabBar::tab:!selected {
                background-color: #f0f0f0;
            }
        """)



    
    def initConditionSearchTab(self):
        """조건별 검색 탭 초기화"""
        layout = QVBoxLayout(self.condition_search_tab)

        layout.setSpacing(25)  # 요소 간격 늘림
        layout.setContentsMargins(40, 40, 40, 40)  # 탭 내부 여백 늘림
        
        title_label = QLabel('아파트 검색 설정')
        title_label.setStyleSheet("""
            font-size: 22pt;
            font-weight: bold;
            color: #333;
            margin-bottom: 10px;
        """)
        layout.addWidget(title_label)
        
        # 지역 선택 섹션 추가
        region_section = QWidget()
        region_layout = QVBoxLayout(region_section)
        region_layout.setContentsMargins(0, 0, 0, 0)
        
        region_label = QLabel('지역 선택:')
        region_label.setStyleSheet("font-size: 12pt; font-weight: bold; color: black; margin-bottom: 5px;")
        region_layout.addWidget(region_label)
        
        # 지역 선택 라디오 버튼 그룹 (변경된 부분)
        region_buttons_layout = QGridLayout()
        region_buttons_layout.setSpacing(10)
        
        # 주요 지역 목록
        regions = ['서울', '경기', '인천', '부산', '대구', '광주', '대전', '울산', '세종', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주']
        
        # 지역 라디오 버튼 생성 (변경된 부분)
        self.region_group = QButtonGroup()
        self.region_buttons = {}
        for i, region in enumerate(regions):
            radio_button = QRadioButton(region)  # 체크박스 대신 라디오 버튼 사용
            radio_button.setStyleSheet("""
                QRadioButton {
                    font-size: 11pt;
                    min-height: 30px;
                    padding: 5px 10px;
                    border-radius: 5px;
                    background-color: #f0f0f0;
                }
                QRadioButton:hover {
                    background-color: #e0e0e0;
                }
                QRadioButton:checked {
                    background-color: #4a90e2;
                    color: white;
                    font-weight: bold;
                }
                QRadioButton::indicator {
                    width: 20px;
                    height: 20px;
                    border-radius: 10px;
                    border: 2px solid #666;
                }
                QRadioButton::indicator:checked {
                    background-color: white;
                    border: 2px solid white;
                }
                QRadioButton::indicator:checked:hover {
                    background-color: #f0f0f0;
                }
            """)
            # 서울 기본 선택
            if region == '서울':
                radio_button.setChecked(True)

            self.region_buttons[region] = radio_button
            self.region_group.addButton(radio_button)  # 라디오 버튼 그룹에 추가
            
            # 3열 그리드 레이아웃으로 배치
            row = i // 3
            col = i % 3
            region_buttons_layout.addWidget(radio_button, row, col)
        
        region_layout.addLayout(region_buttons_layout)
        layout.addWidget(region_section)
        
        
        # 기간 선택 섹션 - 기존 코드 그대로 사용
        period_section = QWidget()
        period_layout = QVBoxLayout(period_section)
        period_layout.setContentsMargins(0, 0, 0, 0)  # 여백 제거
        
        period_label = QLabel('기간 선택:')
        period_label.setStyleSheet("font-size: 12pt; font-weight: bold; color: black; margin-bottom: 5px;")
        period_layout.addWidget(period_label)
        
        # 기간 선택 버튼들
        period_buttons_layout = QHBoxLayout()
        period_buttons_layout.setSpacing(10)  # 버튼 간격 줄임
        self.period_buttons = QButtonGroup()
        
        periods = ['1년', '6개월', '3개월', '1개월']
        webTexts = ['1년', '6개월', '3개월', '1개월']  # 웹사이트의 실제 버튼 텍스트
        
        for i, (period, webText) in enumerate(zip(periods, webTexts)):
            btn = QPushButton(period)
            btn.setCheckable(True)
            btn.setMinimumWidth(100)
            btn.setMinimumHeight(35)
            # 웹사이트의 텍스트를 속성으로 저장
            btn.setProperty('webText', webText)
            if period == '6개월':  # 기본값
                btn.setChecked(True)
            self.period_buttons.addButton(btn, i)
            period_buttons_layout.addWidget(btn)
            
            btn.setStyleSheet("""
                QPushButton {
                    border: 1px solid #ddd;
                    border-radius: 5px;
                    padding: 5px 10px;
                    background-color: white;
                    color: black;
                    font-size: 11pt;  /* 글씨 크기 줄임 */
                }
                QPushButton:checked {
                    background-color: #4a90e2;
                    color: white;
                    border: 1px solid #357abd;
                }
                QPushButton:hover {
                    background-color: #f0f0f0;
                    color: black;
                }
                QPushButton:checked:hover {
                    background-color: #357abd;
                    color: white;
                }
            """)
        
        self.period_buttons.setExclusive(True)
        period_layout.addLayout(period_buttons_layout)
        layout.addWidget(period_section)
        
        # 이하 기존 코드와 동일
        # 입력 필드 초기화
        self.min_area_input = QLineEdit()
        self.max_area_input = QLineEdit()
        self.max_year_input = QLineEdit()
        self.max_price_input = QLineEdit()
        self.complex_count_input = QLineEdit()
        self.rate_change_input = QLineEdit()
        
        # 입력 필드와 레이블 배치
        # 입력 필드와 레이블 배치
        inputs_container = QWidget()
        inputs_layout = QGridLayout(inputs_container)
        inputs_layout.setSpacing(5)  # 간격 늘림
        inputs_layout.setVerticalSpacing(10)  # 세로 간격 늘림
        inputs_layout.setColumnMinimumWidth(0, 180)  # 첫 번째 열(레이블) 너비 늘림
        inputs_layout.setColumnMinimumWidth(1, 100)  # 두 번째 열(입력필드) 너비 늘림
        
        input_fields = [
            ('전용면적 최소값:', self.min_area_input),
            ('전용면적 최대값:', self.max_area_input),
            ('연차 최대값:', self.max_year_input),
            ('검색 최대평균평단가:', self.max_price_input),
            ('저장할 단지 개수:', self.complex_count_input),
            ('증감률(%):', self.rate_change_input)
        ]
        
        placeholder_texts = ['58', '60', '25', '3000', '50', '5']
        
        input_style = """
            QLineEdit {
                padding: 12px;  /* 패딩 늘림 */
                border: 1px solid #ddd;
                border-radius: 5px;
                font-size: 12pt;  /* 글씨 크기 늘림 */
                min-height: 45px;  /* 높이 늘림 */
                min-width: 400px;  /* 최소 너비 추가 */
                background-color: white;
                color: black !important;
            }
            QLineEdit:focus {
                border: 1px solid #4a90e2;
            }
            QLineEdit::placeholder {
                color: #888;
            }
        """
        
        for i, (label_text, input_field) in enumerate(input_fields):
            label = QLabel(label_text)
            label.setMinimumWidth(170)  # 레이블 너비 늘림
            label.setStyleSheet("font-size: 12pt; font-weight: bold; color: black;")
            label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)  # 오른쪽 정렬 추가
            input_field.setStyleSheet(input_style)
            input_field.setPlaceholderText(f'예: {placeholder_texts[i]}')

            inputs_layout.addWidget(label, i, 0)
            inputs_layout.addWidget(input_field, i, 1)
        
        layout.addWidget(inputs_container)
        
        # 검색 버튼
        search_button = QPushButton('검색 시작')
        search_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-size: 12pt;  /* 글씨 크기 줄임 */
                font-weight: bold;
                min-height: 40px;  /* 높이 줄임 */
                min-width: 150px;  /* 너비 줄임 */
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        search_button.clicked.connect(self.start_search)
        
        # 버튼을 가운데 정렬
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.addStretch()
        button_layout.addWidget(search_button)
        button_layout.addStretch()
        
        layout.addWidget(button_container)
        layout.addStretch()
        
    # initComplexSearchTab 메서드 수정 - UI에 적절한 레이블 사용
    # initComplexSearchTab 메서드 수정 - 레이아웃 개선
    def initComplexSearchTab(self):
        """단지명 검색 탭 초기화"""
        layout = QVBoxLayout(self.complex_search_tab)
        layout.setSpacing(25)  # 요소 간격 늘림
        layout.setContentsMargins(40, 40, 40, 40)  # 탭 내부 여백 늘림
        
        title_label = QLabel('단지명 검색')
        title_label.setStyleSheet("""
            font-size: 22pt;
            font-weight: bold;
            color: #333;
            margin-bottom: 10px;
        """)
        layout.addWidget(title_label)
        
        # 검색 설정 컨테이너
        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(20)
        
        # 왼쪽 컨테이너 - 단지명 입력
        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(10)
        
        # 설명 레이블 (25개 -> 50개로 수정)
        description_label = QLabel('검색할 단지명을 입력하세요 (각 줄에 하나씩, 최대 50개):')
        description_label.setStyleSheet("font-size: 11pt; font-weight: bold; color: black;")
        left_layout.addWidget(description_label)
        
        # 단지명 입력 영역
        self.complex_names_input = QTextEdit()
        self.complex_names_input.setStyleSheet("""
            font-size: 11pt;
            padding: 10px;
            min-height: 400px;
        """)
        self.complex_names_input.setPlaceholderText("예시:\n래미안 서초\n자이 김포\n힐스테이트 광교\n...")
        left_layout.addWidget(self.complex_names_input)
        
        # 오른쪽 컨테이너 - 검색 옵션
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(20)
        
        # 지역 선택 섹션 추가 (변경된 부분)
        region_section = QWidget()
        region_layout = QVBoxLayout(region_section)
        region_layout.setContentsMargins(0, 0, 0, 0)
        
        region_label = QLabel('지역 선택:')
        region_label.setStyleSheet("font-size: 12pt; font-weight: bold; color: black; margin-bottom: 5px;")
        region_layout.addWidget(region_label)
        
        # 지역 선택 라디오 버튼 그룹 (변경된 부분)
        region_buttons_layout = QGridLayout()
        region_buttons_layout.setSpacing(10)
        
        # 주요 지역 목록
        regions = ['서울', '경기', '인천', '부산', '대구', '광주', '대전', '울산', '세종', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주']
        
        # 지역 라디오 버튼 생성 (변경된 부분)
        self.complex_region_group = QButtonGroup()
        self.complex_region_buttons = {}
        for i, region in enumerate(regions):
            radio_button = QRadioButton(region)
            radio_button.setStyleSheet("""
                QRadioButton {
                    font-size: 11pt;
                    min-height: 30px;
                    padding: 5px 10px;
                    border-radius: 5px;
                    background-color: #f0f0f0;
                }
                QRadioButton:hover {
                    background-color: #e0e0e0;
                }
                QRadioButton:checked {
                    background-color: #4a90e2;
                    color: white;
                    font-weight: bold;
                }
                QRadioButton::indicator {
                    width: 20px;
                    height: 20px;
                    border-radius: 10px;
                    border: 2px solid #666;
                }
                QRadioButton::indicator:checked {
                    background-color: white;
                    border: 2px solid white;
                }
                QRadioButton::indicator:checked:hover {
                    background-color: #f0f0f0;
                }
            """)
            # 서울 기본 선택
            if region == '서울':
                radio_button.setChecked(True)

            self.complex_region_buttons[region] = radio_button
            self.complex_region_group.addButton(radio_button)
            
            # 3열 그리드 레이아웃으로 배치
            row = i // 3
            col = i % 3
            region_buttons_layout.addWidget(radio_button, row, col)
        
        region_layout.addLayout(region_buttons_layout)
        right_layout.addWidget(region_section)
        
        # 전용면적 입력 섹션 추가
        area_section = QWidget()
        area_layout = QVBoxLayout(area_section)
        area_layout.setContentsMargins(0, 0, 0, 0)
        area_layout.setSpacing(10)
        
        area_label = QLabel('전용면적 범위:')
        area_label.setStyleSheet("font-size: 12pt; font-weight: bold; color: black; margin-bottom: 5px;")
        area_layout.addWidget(area_label)
        
        # 전용면적 최소값/최대값 입력 필드
        area_inputs_layout = QGridLayout()
        area_inputs_layout.setSpacing(10)
        
        self.complex_min_area_input = QLineEdit()
        self.complex_max_area_input = QLineEdit()
        
        input_style = """
            padding: 8px;
            border: 1px solid #ddd;
            border-radius: 5px;
            font-size: 11pt;
            min-height: 35px;
            background-color: white;
            color: black !important;
        """
        
        self.complex_min_area_input.setStyleSheet(input_style)
        self.complex_max_area_input.setStyleSheet(input_style)
        
        self.complex_min_area_input.setPlaceholderText("최소값 (예: 58)")
        self.complex_max_area_input.setPlaceholderText("최대값 (예: 85)")
        
        min_area_label = QLabel("최소값:")
        min_area_label.setStyleSheet("font-size: 11pt;")
        max_area_label = QLabel("최대값:")
        max_area_label.setStyleSheet("font-size: 11pt;")
        
        area_inputs_layout.addWidget(min_area_label, 0, 0)
        area_inputs_layout.addWidget(self.complex_min_area_input, 0, 1)
        area_inputs_layout.addWidget(max_area_label, 1, 0)
        area_inputs_layout.addWidget(self.complex_max_area_input, 1, 1)
        
        area_layout.addLayout(area_inputs_layout)
        right_layout.addWidget(area_section)
        
        # 기간 선택 섹션
        period_section = QWidget()
        period_layout = QVBoxLayout(period_section)
        period_layout.setContentsMargins(0, 0, 0, 0)
        period_layout.setSpacing(10)
        
        period_label = QLabel('기간 선택:')
        period_label.setStyleSheet("font-size: 12pt; font-weight: bold; color: black; margin-bottom: 5px;")
        period_layout.addWidget(period_label)
        
        # 기간 선택 버튼들
        period_buttons_layout = QGridLayout()
        period_buttons_layout.setSpacing(10)
        self.complex_period_buttons = QButtonGroup()
        
        periods = ['1년', '6개월', '3개월', '1개월']
        webTexts = ['1년', '6개월', '3개월', '1개월']
        
        button_style = """
            QPushButton {
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 5px 10px;
                background-color: white;
                color: black;
                font-size: 11pt;
                min-height: 35px;
                min-width: 100px;
            }
            QPushButton:checked {
                background-color: #4a90e2;
                color: white;
                border: 1px solid #357abd;
            }
            QPushButton:hover {
                background-color: #f0f0f0;
                color: black;
            }
            QPushButton:checked:hover {
                background-color: #357abd;
                color: white;
            }
        """
        
        # 2x2 그리드 레이아웃으로 기간 버튼 배치
        for i, (period, webText) in enumerate(zip(periods, webTexts)):
            btn = QPushButton(period)
            btn.setCheckable(True)
            btn.setProperty('webText', webText)
            if period == '6개월':  # 기본값
                btn.setChecked(True)
            self.complex_period_buttons.addButton(btn, i)
            btn.setStyleSheet(button_style)
            
            # 2x2 그리드에 버튼 배치 (0,0), (0,1), (1,0), (1,1)
            row = i // 2
            col = i % 2
            period_buttons_layout.addWidget(btn, row, col)
        
        self.complex_period_buttons.setExclusive(True)
        period_layout.addLayout(period_buttons_layout)
        right_layout.addWidget(period_section)
        
        # 여백 추가
        right_layout.addStretch()
        
        # 검색 버튼
        search_button = QPushButton('단지명으로 검색')
        search_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-size: 12pt;
                font-weight: bold;
                min-height: 50px;
                min-width: 200px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        search_button.clicked.connect(self.start_complex_search)
        right_layout.addWidget(search_button)
        
        # 왼쪽 컨테이너와 오른쪽 컨테이너를 검색 컨테이너에 추가
        # 왼쪽은 더 넓게 (2/3), 오른쪽은 더 좁게 (1/3)
        search_layout.addWidget(left_container, 2)
        search_layout.addWidget(right_container, 1)
        
        layout.addWidget(search_container)

    def get_selected_period(self, buttons=None):
        """선택된 기간 반환 - 웹사이트에서 사용할 텍스트 반환"""
        if buttons is None:
            buttons = self.period_buttons
            
        for button in buttons.buttons():
            if button.isChecked():
                # 웹사이트에서 사용할 텍스트 반환
                return button.property('webText')
        return '6개월'  # 기본값

    def open_settings(self):
        dialog = SettingsDialog(self, self.save_path)
        if dialog.exec():
            self.save_path = dialog.get_save_path()
            # 설정 파일에 저장
            config = load_config()
            config['save_path'] = self.save_path
            save_config(config)
            print(f"저장 경로가 '{self.save_path}'로 설정되었습니다.")
    
    # validate_inputs 메서드 수정 - 대부분의 필드를 선택적으로 변경
    def validate_inputs(self):
        try:
            # 저장할 단지 개수만 필수, 나머지는 선택적
            complex_count_text = self.complex_count_input.text().strip()
            
            # 저장할 단지 개수는 필수 입력
            if not complex_count_text:
                QMessageBox.warning(self, '입력 오류', "저장할 단지 개수를 입력해주세요.")
                return False, None
            
            # 저장할 단지 개수 변환 시도
            try:
                complex_count = int(complex_count_text)
                if complex_count <= 0:
                    QMessageBox.warning(self, '입력 오류', "저장할 단지 개수는 양수여야 합니다.")
                    return False, None
            except ValueError:
                QMessageBox.warning(self, '입력 오류', "저장할 단지 개수는 정수여야 합니다.")
                return False, None
            
            # 나머지 필드는 입력된 경우에만 처리
            min_area = None
            max_area = None
            max_year = None
            max_price = None
            rate_change = None
            
            # 전용면적 최소값 (선택적)
            min_area_text = self.min_area_input.text().strip()
            if min_area_text:
                try:
                    min_area = float(min_area_text)
                    if min_area <= 0:
                        QMessageBox.warning(self, '입력 오류', "전용면적 최소값은 양수여야 합니다.")
                        return False, None
                except ValueError:
                    QMessageBox.warning(self, '입력 오류', "전용면적 최소값은 숫자여야 합니다.")
                    return False, None
            
            # 전용면적 최대값 (선택적)
            max_area_text = self.max_area_input.text().strip()
            if max_area_text:
                try:
                    max_area = float(max_area_text)
                    if max_area <= 0:
                        QMessageBox.warning(self, '입력 오류', "전용면적 최대값은 양수여야 합니다.")
                        return False, None
                except ValueError:
                    QMessageBox.warning(self, '입력 오류', "전용면적 최대값은 숫자여야 합니다.")
                    return False, None
            
            # 최소값과 최대값 비교 (둘 다 입력된 경우)
            if min_area is not None and max_area is not None:
                if min_area > max_area:
                    QMessageBox.warning(self, '입력 오류', "전용면적 최소값이 최대값보다 클 수 없습니다.")
                    return False, None
            
            # 연차 최대값 (선택적)
            max_year_text = self.max_year_input.text().strip()
            if max_year_text:
                try:
                    max_year = float(max_year_text)
                    if max_year <= 0:
                        QMessageBox.warning(self, '입력 오류', "연차 최대값은 양수여야 합니다.")
                        return False, None
                except ValueError:
                    QMessageBox.warning(self, '입력 오류', "연차 최대값은 숫자여야 합니다.")
                    return False, None
            
            # 검색 최대평균평단가 (선택적)
            max_price_text = self.max_price_input.text().strip()
            if max_price_text:
                try:
                    max_price = float(max_price_text)
                    if max_price <= 0:
                        QMessageBox.warning(self, '입력 오류', "검색 최대평균평단가는 양수여야 합니다.")
                        return False, None
                except ValueError:
                    QMessageBox.warning(self, '입력 오류', "검색 최대평균평단가는 숫자여야 합니다.")
                    return False, None
            
            # 증감률(%) (선택적)
            rate_change_text = self.rate_change_input.text().strip()
            if rate_change_text:
                try:
                    rate_change = float(rate_change_text)
                except ValueError:
                    QMessageBox.warning(self, '입력 오류', "증감률은 숫자여야 합니다.")
                    return False, None
            
            # 선택된 기간 가져오기
            period = self.get_selected_period()
            
            return True, (min_area, max_area, max_year, period, max_price, complex_count, rate_change)
                
        except Exception as e:
            QMessageBox.warning(self, '입력 오류', f"예기치 않은 오류: {str(e)}")
            return False, None
            
    # validate_complex_names 메서드 수정 - 전용면적 유효성 검사 추가
    def validate_complex_names(self):
        """단지명 입력 유효성 검사"""
        text = self.complex_names_input.toPlainText().strip()
        if not text:
            QMessageBox.warning(self, '입력 오류', "검색할 단지명을 입력해주세요.")
            return False, []
            
        # 각 줄을 단지명으로 처리
        complex_names = [line.strip() for line in text.split('\n') if line.strip()]
        
        # 중복 제거
        complex_names = list(dict.fromkeys(complex_names))
        
        # 최대 50개로 제한 (기존 25개에서 변경)
        if len(complex_names) > 50:
            complex_names = complex_names[:50]
            QMessageBox.warning(self, '입력 제한', "최대 50개의 단지명만 처리됩니다. 처음 50개만 검색합니다.")
        
        # 전용면적 유효성 검사
        min_area = None
        max_area = None
        
        min_area_text = self.complex_min_area_input.text().strip()
        max_area_text = self.complex_max_area_input.text().strip()
        
        # 최소값 검사
        if min_area_text:
            try:
                min_area = float(min_area_text)
                if min_area <= 0:
                    QMessageBox.warning(self, '입력 오류', "전용면적 최소값은 양수여야 합니다.")
                    return False, []
            except ValueError:
                QMessageBox.warning(self, '입력 오류', "전용면적 최소값은 숫자여야 합니다.")
                return False, []
        
        # 최대값 검사
        if max_area_text:
            try:
                max_area = float(max_area_text)
                if max_area <= 0:
                    QMessageBox.warning(self, '입력 오류', "전용면적 최대값은 양수여야 합니다.")
                    return False, []
            except ValueError:
                QMessageBox.warning(self, '입력 오류', "전용면적 최대값은 숫자여야 합니다.")
                return False, []
        
        # 최소값과 최대값 비교
        if min_area is not None and max_area is not None:
            if min_area > max_area:
                QMessageBox.warning(self, '입력 오류', "전용면적 최소값이 최대값보다 클 수 없습니다.")
                return False, []
        
        # 선택된 기간 가져오기
        period = self.get_selected_period(self.complex_period_buttons)
        
        return True, (complex_names, period, min_area, max_area)



    
    def start_search(self):
        valid, values = self.validate_inputs()
        if not valid:
            return
            
        min_area, max_area, max_year, period, max_price, complex_count, rate_change = values
        
        # 선택된 지역 가져오기 (하나만)
        selected_region = None
        for region, button in self.region_buttons.items():
            if button.isChecked():
                selected_region = region
                break
        
        if not selected_region:
            QMessageBox.warning(self, '입력 오류', "지역을 선택해주세요.")
            return
        
        # 단일 지역을 리스트로 만들어 전달
        selected_regions = [selected_region]
        
        # 진행바 표시
        self.show_progress_dialog("조건별 검색 진행 중...", 100)
        
        driver = None
        
        try:
            self.update_progress(10, "웹 브라우저를 시작하는 중...")
            driver = setup_driver()
            
            self.update_progress(20, "웹사이트에 접속하는 중...")
            url = "https://data.kbland.kr/kbstats/investment-table"
            driver.get(url)
            
            self.update_progress(30, "검색 조건을 설정하는 중...")
            # 선택된 지역 정보 전달
            click_elements(driver, min_area, max_area, max_year, period, max_price, rate_change, complex_count, self.save_path, selected_regions, self)
            
        except Exception as e:
            self.close_progress_dialog()
            QMessageBox.critical(self, '오류', f'검색 중 오류가 발생했습니다: {str(e)}')
            if driver:
                try:
                    driver.quit()
                    print("예외 발생으로 웹 브라우저를 종료했습니다.")
                except:
                    print("웹 브라우저 종료 실패")
        finally:
            self.close_progress_dialog()
 

        
    # start_complex_search 메서드 수정 - 전용면적 추가
    def start_complex_search(self):
        """단지명으로 검색 시작"""
        valid, values = self.validate_complex_names()
        if not valid:
            return
            
        complex_names, period, min_area, max_area = values
        
        # 선택된 지역 가져오기 (하나만)
        selected_region = None
        for region, button in self.complex_region_buttons.items():
            if button.isChecked():
                selected_region = region
                break
        
        if not selected_region:
            QMessageBox.warning(self, '입력 오류', "지역을 선택해주세요.")
            return
        
        # 단일 지역을 리스트로 만들어 전달
        selected_regions = [selected_region]
        
        # 진행바 표시 (단지명 개수에 따라 total_steps 설정)
        total_steps = len(complex_names) + 5  # 단지명 개수 + 초기 설정 단계
        self.show_progress_dialog("단지명 검색 진행 중...", total_steps)
        
        driver = None
        
        try:
            self.update_progress(1, "웹 브라우저를 시작하는 중...")
            driver = setup_driver()
            
            self.update_progress(2, "웹사이트에 접속하는 중...")
            url = "https://data.kbland.kr/kbstats/investment-table"
            driver.get(url)
            
            self.update_progress(3, "검색 조건을 설정하는 중...")
            # 선택된 지역 정보 전달
            search_by_complex_names(driver, complex_names, period, self.save_path, min_area, max_area, selected_regions, self)
            
        except Exception as e:
            self.close_progress_dialog()
            QMessageBox.critical(self, '오류', f'검색 중 오류가 발생했습니다: {str(e)}')
            if driver:
                try:
                    driver.quit()
                    print("예외 발생으로 웹 브라우저를 종료했습니다.")
                except:
                    print("웹 브라우저 종료 실패")
        finally:
            self.close_progress_dialog()



def parse_complex_data(driver, target_count):
    """페이지의 모든 데이터를 파싱하는 함수"""
    print(f"페이지 데이터 파싱 시작 (목표: {target_count}개)...")
    
    # 전용면적 디버깅을 위한 코드 추가
    try:
        area_debug_elements = driver.find_elements(By.CSS_SELECTOR, "div.bodygroup.I2020 div.bodyitem.I2020")
        print(f"전용면적 디버깅: {len(area_debug_elements)}개 요소 발견")
        if len(area_debug_elements) > 0:
            for i in range(min(3, len(area_debug_elements))):  # 처음 3개만 확인
                elem = area_debug_elements[i]
                print(f"전용면적 요소 {i}: 전체텍스트='{elem.text}', innerHTML길이={len(elem.get_attribute('innerHTML'))}")
                divs = elem.find_elements(By.TAG_NAME, "div")
                print(f"  하위 div 개수: {len(divs)}")
                for j, div in enumerate(divs):
                    print(f"  div[{j}]: '{div.text}'")
    except Exception as e:
        print(f"전용면적 디버깅 중 오류: {str(e)}")
    
    # 모든 칼럼 데이터 컨테이너 찾기
    try:
        # 지역명 추가
        try:
            region_container = driver.find_element(By.XPATH, "//div[contains(text(), '지역명')]/parent::div/following-sibling::div[@class='body']")
            region_elements = region_container.find_elements(By.CSS_SELECTOR, "div.bodyitem")
        except:
            # 대체 방법
            try:
                region_elements = driver.find_elements(By.XPATH, "//div[@data-v-9bca83a2 and contains(@class, 'bodyitem')][parent::div[@class='body' and preceding-sibling::div[@class='head']/div[contains(text(), '지역명')]]]")
            except:
                print("지역명 컬럼을 찾을 수 없습니다.")
                region_elements = []
        
        # 단지명
        complex_container = driver.find_element(By.CSS_SELECTOR, 
            "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body")
        complex_elements = complex_container.find_elements(By.CSS_SELECTOR, "div.bodyitem")
        
        # 각 컬럼 요소들 가져오기
        # 세대수 추가 (I2030 클래스) - 이 부분을 추가
        try:
            household_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2030")
            household_elements = household_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2030")
            print(f"세대수 요소 {len(household_elements)}개 발견")
        except Exception as e:
            print(f"세대수 컬럼을 찾을 수 없습니다: {str(e)}")
            household_elements = []
        
        # 각 컬럼 요소들 가져오기
        try:
            area_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2020")
            area_elements = area_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2020")
            print(f"전용면적 컨테이너 방법1: {len(area_elements)}개 발견")
        except Exception as e:
            print(f"전용면적 컨테이너 방법1 실패: {str(e)}")
            try:
                # 대체 방법
                area_elements = driver.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2020")
                print(f"전용면적 컨테이너 방법2: {len(area_elements)}개 발견")
            except Exception as e2:
                print(f"전용면적 컨테이너 방법2도 실패: {str(e2)}")
                area_elements = []
        
        year_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2040")
        year_elements = year_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2040")
        
        price_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2050")
        price_elements = price_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2050")
        
        sale_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2060")
        sale_elements = sale_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2060")
        
        rent_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2070")
        rent_elements = rent_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2070")
        
        # 전세가율 컨테이너 (I2080)
        try:
            rent_rate_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2080")
            rent_rate_elements = rent_rate_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2080")
        except:
            print("전세가율 컬럼을 찾을 수 없습니다.")
            rent_rate_elements = []
        
        # 매매전세차 컨테이너 (I2090)
        try:
            price_diff_container = driver.find_element(By.CSS_SELECTOR, "div.bodygroup.I2090")
            price_diff_elements = price_diff_container.find_elements(By.CSS_SELECTOR, "div.bodyitem.I2090")
        except:
            print("매매전세차 컬럼을 찾을 수 없습니다.")
            price_diff_elements = []
        
        # 실제 처리할 개수 계산 (모든 칼럼의 최소 개수와 목표 개수 중 작은 값)
        # 지역명 컬럼도 포함하여 계산
        columns_with_data = [complex_elements, area_elements, year_elements, price_elements, sale_elements, rent_elements]
        if region_elements:
            columns_with_data.append(region_elements)
        if household_elements:  # 세대수 추가 - 이 줄 추가
            columns_with_data.append(household_elements)
        if rent_rate_elements:
            columns_with_data.append(rent_rate_elements)
        if price_diff_elements:
            columns_with_data.append(price_diff_elements)
        
        min_elements = min(len(column) for column in columns_with_data)
        
        process_count = min(min_elements, target_count)
        print(f"처리할 데이터 수: {process_count} (목표: {target_count}, 가용: {min_elements})")
        
        # 각 칼럼별 개수 출력
        if region_elements:
            print(f"지역명 요소: {len(region_elements)}개")
        print(f"단지명 요소: {len(complex_elements)}개")
        if household_elements:  # 세대수 출력 추가 - 이 부분 추가
            print(f"세대수 요소: {len(household_elements)}개")
        print(f"전용면적 요소: {len(area_elements)}개")
        print(f"연차 요소: {len(year_elements)}개")
        print(f"시세 요소: {len(price_elements)}개")
        print(f"매매시세 요소: {len(sale_elements)}개")
        print(f"전세시세 요소: {len(rent_elements)}개")
        if rent_rate_elements:
            print(f"전세가율 요소: {len(rent_rate_elements)}개")
        if price_diff_elements:
            print(f"매매전세차 요소: {len(price_diff_elements)}개")
        
        # 각 행별 데이터 추출
        data = []
        prices = []  # 시세(만원) 값을 저장할 리스트
        
        # JavaScript를 사용하여 보다 효율적으로 데이터 추출

        # JavaScript를 사용하여 보다 효율적으로 데이터 추출
        js_data = driver.execute_script(r"""
            function extractText(elem) {
                return elem ? elem.textContent.trim() : "";
            }
            
            function findPercentageInText(text) {
                var matches = text.match(/([-+]?\d+\.?\d*)%/);
                return matches ? matches[0] : "";
            }
            
            var results = [];
            
            // 지역명 요소들 찾기
            var regionElems = [];
            var regionContainers = document.querySelectorAll('div[data-v-9bca83a2]');
            for (var i = 0; i < regionContainers.length; i++) {
                var container = regionContainers[i];
                var header = container.querySelector('div.head');
                if (header && header.textContent.includes('지역명')) {
                    var body = container.querySelector('div.body');
                    if (body) {
                        regionElems = body.querySelectorAll('div.bodyitem');
                        break;
                    }
                }
            }
            
            var complexElems = document.querySelectorAll("#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body > div.bodyitem");
            
            // 세대수 요소들 찾기 (I2030 클래스) - 이 줄 추가
            var householdElems = document.querySelectorAll("div.bodygroup.I2030 div.bodyitem.I2030");
            
            var areaElems = document.querySelectorAll("div.bodygroup.I2020 div.bodyitem.I2020");
            var yearElems = document.querySelectorAll("div.bodygroup.I2040 div.bodyitem.I2040");
            var priceElems = document.querySelectorAll("div.bodygroup.I2050 div.bodyitem.I2050");
            var saleElems = document.querySelectorAll("div.bodygroup.I2060 div.bodyitem.I2060");
            var rentElems = document.querySelectorAll("div.bodygroup.I2070 div.bodyitem.I2070");
            var rentRateElems = document.querySelectorAll("div.bodygroup.I2080 div.bodyitem.I2080");
            var priceDiffElems = document.querySelectorAll("div.bodygroup.I2090 div.bodyitem.I2090");
            
            var count = Math.min(
                complexElems.length,
                areaElems.length,
                yearElems.length,
                priceElems.length,
                saleElems.length,
                rentElems.length,
                arguments[0]
            );
            
            if (regionElems.length > 0) {
                count = Math.min(count, regionElems.length);
            }
            if (householdElems.length > 0) {  // 세대수 개수 체크 추가 - 이 부분 추가
                count = Math.min(count, householdElems.length);
            }
            if (rentRateElems.length > 0) {
                count = Math.min(count, rentRateElems.length);
            }
            
            console.log("Processing count:", count);
            
            for (var i = 0; i < count; i++) {
                var regionName = "";
                if (i < regionElems.length) {
                    regionName = extractText(regionElems[i]);
                }
                
            
                var complexName = extractText(complexElems[i]);
                
                // 세대수 데이터 추출 - 이 부분 추가
                var household = "";
                if (i < householdElems.length) {
                    var householdDivs = householdElems[i].querySelectorAll("div");
                    if (householdDivs.length >= 1) {
                        household = extractText(householdDivs[0]);
                    }
                }
                
                var area = "";
                if (i < areaElems.length) {
                    var areaElement = areaElems[i];
                    var areaDivs = areaElement.querySelectorAll("div");
                    
                    console.log("전용면적 요소 " + i + ": div개수=" + areaDivs.length);
                    
                    if (areaDivs.length >= 2) {
                        area = extractText(areaDivs[1]);
                        console.log("  방법1(div[1]): '" + area + "'");
                    } else if (areaDivs.length >= 1) {
                        area = extractText(areaDivs[0]);
                        console.log("  방법2(div[0]): '" + area + "'");
                    } else {
                        area = extractText(areaElement);
                        console.log("  방법3(전체): '" + area + "'");
                    }
                    
                    // 추가 시도: 숫자만 추출
                    if (!area || area.trim() === '') {
                        var fullText = areaElement.textContent || areaElement.innerText || '';
                        var numberMatch = fullText.match(/(\d+\.?\d*)/);
                        if (numberMatch) {
                            area = numberMatch[1];
                            console.log("  방법4(숫자추출): '" + area + "'");
                        }
                    }
                }
                
                // 연차 데이터 추출 - 누락된 부분 추가
                var year = "";
                if (i < yearElems.length) {
                    var yearDivs = yearElems[i].querySelectorAll("div");
                    if (yearDivs.length >= 1) {
                        year = extractText(yearDivs[0]);
                    }
                }
                
                var price = "";
                var changeRate = "";
                if (i < priceElems.length) {
                    var priceElement = priceElems[i];
                    var priceDivs = priceElement.querySelectorAll("div");
                    
                    if (priceDivs.length >= 2) {
                        price = extractText(priceDivs[0]);
                        changeRate = extractText(priceDivs[1]);
                    } else if (priceDivs.length >= 1) {
                        var fullText = extractText(priceDivs[0]);
                        var parts = fullText.split(/[\\s\\n]+/).filter(function(part) { 
                            return part.trim() !== ''; 
                        });
                        if (parts.length >= 2) {
                            price = parts[0];
                            changeRate = parts[1];
                        } else {
                            price = parts[0] || fullText;
                        }
                    }
                    
                    if (!changeRate || !changeRate.includes('%')) {
                        var fullElementText = priceElement.textContent || priceElement.innerText || "";
                        var foundPercentage = findPercentageInText(fullElementText);
                        if (foundPercentage) {
                            changeRate = foundPercentage;
                        }
                    }
                    
                    if (!changeRate || !changeRate.includes('%')) {
                        var siblings = priceElement.parentNode ? priceElement.parentNode.children : [];
                        for (var s = 0; s < siblings.length; s++) {
                            var siblingText = siblings[s].textContent || siblings[s].innerText || "";
                            var foundPercentage = findPercentageInText(siblingText);
                            if (foundPercentage) {
                                changeRate = foundPercentage;
                                break;
                            }
                        }
                    }
                }
                
                var salePrice = "";
                if (i < saleElems.length) {
                    var saleDivs = saleElems[i].querySelectorAll("div");
                    if (saleDivs.length >= 1) {
                        salePrice = extractText(saleDivs[0]);
                    }
                }
                
                var rentPrice = "";
                var rentChangeRate = "";
                if (i < rentElems.length) {
                    var rentDivs = rentElems[i].querySelectorAll("div");
                    if (rentDivs.length >= 2) {
                        rentPrice = extractText(rentDivs[0]);
                        rentChangeRate = extractText(rentDivs[1]);
                    } else if (rentDivs.length >= 1) {
                        rentPrice = extractText(rentDivs[0]);
                    }
                }
                
                var rentRate = "";
                if (i < rentRateElems.length) {
                    var rentRateDivs = rentRateElems[i].querySelectorAll("div");
                    if (rentRateDivs.length >= 1) {
                        rentRate = extractText(rentRateDivs[0]);
                    }
                }
                
                var priceDiff = "";
                if (i < priceDiffElems.length) {
                    var priceDiffDivs = priceDiffElems[i].querySelectorAll("div");
                    if (priceDiffDivs.length >= 1) {
                        priceDiff = extractText(priceDiffDivs[0]);
                    }
                }
                
                results.push({
                    regionName: regionName,
                    complexName: complexName,
                    household: household,  // 세대수 추가 - 이 줄 추가
                    area: area,
                    year: year,
                    price: price,
                    changeRate: changeRate,
                    salePrice: salePrice,
                    rentPrice: rentPrice,
                    rentChangeRate: rentChangeRate,
                    rentRate: rentRate,
                    priceDiff: priceDiff
                });
            }
            
            return results;
        """, target_count)
        
        # JavaScript 결과 처리
        if js_data:
            print(f"JavaScript로 {len(js_data)}개 데이터 추출 성공")
            
            # 데이터 변환
            for item in js_data:
                region_name = item.get('regionName', '')
                complex_name = item.get('complexName', '')
                household = item.get('household', '')  # 세대수 추가 - 이 줄 추가
                area = item.get('area', '')
                year = item.get('year', '')
                price = item.get('price', '')
                change_rate = item.get('changeRate', '')
                sale_price = item.get('salePrice', '')
                rent_price = item.get('rentPrice', '')
                rent_change_rate = item.get('rentChangeRate', '')  # 전세 증감률
                rent_rate = item.get('rentRate', '')              # 전세가율
                price_diff = item.get('priceDiff', '')            # 매매전세차
                
                # 시세값 변환 시도
                try:
                    if price:
                        price_value = float(price.replace(",", ""))
                        prices.append(price_value)
                except:
                    print(f"시세값 변환 실패: {price}")
                
                row_data = {
                    '지역명': region_name,
                    '단지명': complex_name,
                    '세대수': household,        # 세대수 추가 (전용면적 앞에 위치) - 이 줄 추가
                    '전용면적': area,
                    '연차': year,
                    '시세(만원)': price,
                    '매매증감률': change_rate,  # 매매 증감률
                    '매매시세(억원)': sale_price,
                    '전세시세': rent_price,
                    '전세증감률': rent_change_rate,  # 전세 증감률
                    '전세가율': rent_rate,          # 전세가율
                    '매매전세차': price_diff        # 매매전세차 - 원래대로
                }
                data.append(row_data)
            
            return data, prices
        
        # JavaScript 실패 시 기존 방식으로 진행
        print("JavaScript 추출 실패, 기존 방식으로 진행...")
        
        for i in range(process_count):
            try:
                # 지역명
                region_name = ""
                if i < len(region_elements):
                    region_name = region_elements[i].text.strip()
                
                # 단지명
                complex_name = ""
                if i < len(complex_elements):
                    complex_name = complex_elements[i].text.strip()
                
                # 세대수 추가 - 이 부분 추가
                household = ""
                if i < len(household_elements):
                    divs = household_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 1:
                        household = divs[0].text.strip()
                
                # 전용면적
                area = ""
                
                # 전용면적
                area = ""
                if i < len(area_elements):
                    divs = area_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 2:
                        area = divs[1].text.strip()
                    elif len(divs) >= 1:
                        # div가 1개만 있는 경우 첫 번째 div 사용
                        area = divs[0].text.strip()
                    else:
                        # div가 없는 경우 전체 텍스트 사용
                        area = area_elements[i].text.strip()
                
                # 연차
                year = ""
                if i < len(year_elements):
                    divs = year_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 1:
                        year = divs[0].text.strip()
                
                # 시세(만원)과 매매 증감률
                price = ""
                change_rate = ""
                if i < len(price_elements):
                    divs = price_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 2:
                        price = divs[0].text.strip()
                        change_rate = divs[1].text.strip()
                        try:
                            # 시세값을 숫자로 변환하여 저장 (콤마 제거)
                            price_value = float(price.replace(",", ""))
                            prices.append(price_value)
                        except:
                            print(f"시세값 변환 실패: {price}")
                
                # 매매시세(억원)
                sale_price = ""
                if i < len(sale_elements):
                    divs = sale_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 1:
                        sale_price = divs[0].text.strip()
                
                # 전세시세와 전세 증감률
                rent_price = ""
                rent_change_rate = ""
                if i < len(rent_elements):
                    divs = rent_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 2:
                        rent_price = divs[0].text.strip()
                        rent_change_rate = divs[1].text.strip()  # 전세 증감률 추가
                    elif len(divs) >= 1:
                        rent_price = divs[0].text.strip()
                
                # 전세가율
                rent_rate = ""
                if i < len(rent_rate_elements):
                    divs = rent_rate_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 1:
                        rent_rate = divs[0].text.strip()
                
                # 매매전세차
                price_diff = ""
                if i < len(price_diff_elements):
                    divs = price_diff_elements[i].find_elements(By.TAG_NAME, "div")
                    if len(divs) >= 1:
                        price_diff = divs[0].text.strip()
                
                row_data = {
                    '지역명': region_name,
                    '단지명': complex_name,
                    '세대수': household,        # 세대수 추가 (전용면적 앞에 위치) - 이 줄 추가
                    '전용면적': area,
                    '연차': year,
                    '시세(만원)': price,
                    '매매증감률': change_rate,  # 매매 증감률
                    '매매시세(억원)': sale_price,
                    '전세시세': rent_price,
                    '전세증감률': rent_change_rate,  # 전세 증감률
                    '전세가율': rent_rate,          # 전세가율
                    '매매전세차': price_diff        # 매매전세차 - 원래대로
                }
                data.append(row_data)
                
            except Exception as e:
                print(f"행 {i} 데이터 추출 중 오류: {str(e)}")
                continue
        
        print(f"기존 방식으로 {len(data)}개 데이터 추출 완료")
        return data, prices
        
    except Exception as e:
        print(f"데이터 파싱 중 오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        return [], []


# 새로운 함수 추가 - 여기에 넣어주세요
def find_most_expensive_areas(data_list):
    """
    각 단지에서 가장 비싼 전용면적을 찾아 반환합니다.
    """
    # 데이터가 없으면 빈 리스트 반환
    if not data_list:
        return []
    
    # pandas DataFrame으로 변환
    df = pd.DataFrame(data_list)
    
    # 단지명이 없는 경우 빈 리스트 반환
    if df.empty or '단지명' not in df.columns:
        return []
    
    # 디버깅: DataFrame의 컬럼 확인
    print(f"find_most_expensive_areas - 입력 데이터 컬럼: {df.columns.tolist()}")
    
    # 모든 컬럼이 숫자형으로 변환되도록 처리
    # 가격 컬럼들
    price_columns = ['시세(만원)', '매매시세(만원)', '전세시세(만원)']
    
    # 숫자형 변환 함수
    def convert_to_numeric(series, is_price=True):
        if is_price:
            # 가격 컬럼: 콤마 제거, 빈 문자열은 0으로
            return pd.to_numeric(series.astype(str).str.replace(',', '').replace('', '0').fillna('0'), errors='coerce').fillna(0)
        else:
            # 일반 컬럼: 빈 문자열은 0으로
            return pd.to_numeric(series.astype(str).replace('', '0').fillna('0'), errors='coerce').fillna(0)
    
    # 가격 컬럼 변환
    for col in price_columns:
        if col in df.columns:
            df[col] = convert_to_numeric(df[col])
    
    # 전용면적 및 연차 변환
    if '전용면적' in df.columns:
        df['전용면적'] = convert_to_numeric(df['전용면적'], is_price=False)
    
    if '연차' in df.columns:
        df['연차'] = convert_to_numeric(df['연차'], is_price=False)
    
    # 가격 컬럼 중 최대값을 가진 컬럼 찾기
    df['최대가격'] = 0
    
    for col in price_columns:
        if col in df.columns:
            # 각 행에서 현재 최대가격과 컬럼 값 중 큰 값을 선택
            df['최대가격'] = df.apply(lambda row: max(row['최대가격'], row[col]), axis=1)
    
    # 디버깅: 가격 변환 및 계산 결과 출력
    print(f"단지 수: {len(df['단지명'].unique())}")
    print(f"데이터 행 수: {len(df)}")
    
    # 단지명으로 그룹화하고, 각 그룹에서 최대가격 행만 선택
    result_data = []
    
    # 각 단지명별로 처리
    # 각 단지명별로 처리할 때 모든 컬럼 데이터 보존
    for name in df['단지명'].unique():
        # 단지명으로 필터링
        complex_df = df[df['단지명'] == name].copy()
        
        if not complex_df.empty:
            # 최대가격이 제대로 계산되었는지 확인
            max_price = complex_df['최대가격'].max()
            
            # 디버깅 출력
            print(f"단지명: {name}, 최대가격: {max_price}")
            print(f"해당 단지 데이터 수: {len(complex_df)}")
            
            # 가격 기준 내림차순, 전용면적 기준 내림차순 정렬
            sorted_df = complex_df.sort_values(['최대가격', '전용면적'], ascending=[False, False])
            
            # 가장 첫 번째 행 선택 (가장 비싼 전용면적)
            max_price_row = sorted_df.iloc[0].to_dict()
            
            # 임시 컬럼 제거 (모든 원본 데이터는 보존)
            if '최대가격' in max_price_row:
                del max_price_row['최대가격']
            
            # 증감률 데이터 보존 확인
            if '매매증감률' in max_price_row:
                print(f"단지 {name} - 매매증감률: {max_price_row['매매증감률']}")
            if '전세증감률' in max_price_row:
                print(f"단지 {name} - 전세증감률: {max_price_row['전세증감률']}")
            
            result_data.append(max_price_row)
    
    print(f"최종 선택된 단지 수: {len(result_data)}")
    return result_data




# search_by_complex_names 함수 수정 - 전용면적 입력 필드 식별 및 사용
# search_by_complex_names 함수 수정 - 서울 선택 및 전용면적 입력 프로세스 강화
def search_by_complex_names(driver, complex_names, period, save_path="", min_area=None, max_area=None, selected_regions=None, gui=None):
    """단지명 목록으로 검색하고 결과를 저장하는 함수"""
    wait = WebDriverWait(driver, 30)
    long_wait = WebDriverWait(driver, 60)  # 더 긴 대기 시간
    
    try:
        print(f"단지명 {len(complex_names)}개로 검색 시작")
        if min_area is not None:
            print(f"전용면적 최소값: {min_area}")
        if max_area is not None:
            print(f"전용면적 최대값: {max_area}")
        
        # 1. 지역 선택 박스 클릭 (click_elements와 동일하게)
        try:
            # 첫 번째 방법: 직접 선택자로 시도
            region_select = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 
                "#sectionWrapper > div.topselarea > div.selectbox.type2 > span")))
            driver.execute_script("arguments[0].click();", region_select)
            print("지역 선택 박스 클릭 완료")
            time.sleep(3)
        except Exception as e:
            print(f"지역 선택 박스 클릭 첫 번째 방법 실패: {str(e)}")
            try:
                # 두 번째 방법: 텍스트 검색으로 시도
                region_selects = driver.find_elements(By.XPATH, "//span[contains(text(), '지역') or contains(text(), '서울')]")
                if region_selects:
                    driver.execute_script("arguments[0].click();", region_selects[0])
                    print("지역 선택 박스 클릭 완료 (대체 방법)")
                    time.sleep(3)
                else:
                    print("지역 선택 요소를 찾을 수 없음")
            except Exception as e2:
                print(f"지역 선택 박스 클릭 두 번째 방법도 실패: {str(e2)}")
        
        # 2. 지역 '+' 버튼 클릭 (click_elements와 동일하게)
        # 기본값 설정: selected_regions가 없거나 비어있으면 '서울'만 선택
        if selected_regions is None or len(selected_regions) == 0:
            selected_regions = ['서울']
        
        print(f"선택할 지역: {selected_regions}")
        
        # 모든 '+' 버튼 찾기 (click_elements와 동일하게)
        add_button_xpath = "//span[@role='button' and @title='지역추가' and @class='add']"
        add_buttons = driver.find_elements(By.XPATH, add_button_xpath)
        
        if len(add_buttons) > 0:
            print(f"총 {len(add_buttons)}개의 '+' 버튼 발견")
            
            # 각 버튼의 부모 요소 텍스트 가져오기
            button_regions = {}
            for i, btn in enumerate(add_buttons):
                try:
                    parent = driver.execute_script("return arguments[0].parentNode;", btn)
                    parent_text = driver.execute_script("return arguments[0].textContent;", parent).strip()
                    button_regions[parent_text] = btn
                    print(f"버튼 {i+1}의 부모 요소 텍스트: '{parent_text}'")
                except Exception as e:
                    print(f"버튼 {i+1}의 부모 요소 텍스트 가져오기 실패: {str(e)}")
            
            # 선택한 지역 버튼 클릭
            clicked_regions = []
            for region in selected_regions:
                # 정확히 일치하는 지역명 찾기
                if region in button_regions:
                    btn = button_regions[region]
                    try:
                        driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", btn)
                        print(f"{region} 지역 추가 버튼 클릭 완료")
                        clicked_regions.append(region)
                        time.sleep(1)  # 클릭 후 잠시 대기
                    except Exception as e:
                        print(f"{region} 지역 버튼 클릭 실패: {str(e)}")
                else:
                    # 부분 일치하는 지역명 찾기
                    found = False
                    for region_name, btn in button_regions.items():
                        if region.lower() in region_name.lower() or region_name.lower() in region.lower():
                            try:
                                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                                time.sleep(0.5)
                                driver.execute_script("arguments[0].click();", btn)
                                print(f"{region} 지역 추가 버튼 클릭 완료 (일치: '{region_name}')")
                                clicked_regions.append(region_name)
                                found = True
                                time.sleep(1)  # 클릭 후 잠시 대기
                                break
                            except Exception as e:
                                print(f"{region} 지역 버튼 클릭 실패 (일치: '{region_name}'): {str(e)}")
                    
                    if not found:
                        print(f"{region} 지역을 찾을 수 없어 선택하지 않았습니다.")
            
            # 클릭한 지역이 없으면 서울 선택
            if len(clicked_regions) == 0:
                print("선택한 지역이 없어 기본값 '서울' 선택")
                for i, btn in enumerate(add_buttons):
                    try:
                        parent = driver.execute_script("return arguments[0].parentNode;", btn)
                        parent_text = driver.execute_script("return arguments[0].textContent;", parent).strip()
                        if parent_text == '서울':
                            driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                            time.sleep(0.5)
                            driver.execute_script("arguments[0].click();", btn)
                            print("서울 지역 추가 버튼 클릭 완료 (기본값)")
                            clicked_regions.append('서울')
                            time.sleep(1)
                            break
                    except:
                        continue
                
                # 여전히 클릭한 지역이 없으면 첫 번째 버튼 클릭
                if len(clicked_regions) == 0:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView(true);", add_buttons[0])
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", add_buttons[0])
                        print("첫 번째 지역 추가 버튼 클릭 완료 (최후의 수단)")
                        time.sleep(1)
                    except Exception as e:
                        print(f"첫 번째 지역 추가 버튼 클릭 실패: {str(e)}")
                        
            print(f"선택된 지역: {clicked_regions}")
        else:
            print("'+' 버튼을 찾을 수 없음, JavaScript로 대체 시도")
            # JavaScript로 버튼 찾기 및 클릭 시도 (click_elements와 동일하게)
            js_result = driver.execute_script("""
                var regions = arguments[0];
                var clickedRegions = [];
                var allButtons = document.querySelectorAll('span.add[role="button"]');
                
                if (allButtons.length === 0) {
                    console.log('JavaScript: 지역추가 버튼을 찾을 수 없음');
                    return clickedRegions;
                }
                
                // 모든 버튼의 부모 텍스트 가져오기
                var buttonRegions = {};
                for (var i = 0; i < allButtons.length; i++) {
                    var btn = allButtons[i];
                    var parent = btn.parentNode;
                    var regionName = parent ? parent.textContent.trim() : '';
                    if (regionName) {
                        buttonRegions[regionName] = btn;
                    }
                }
                
                // 선택한 지역 버튼 클릭
                for (var i = 0; i < regions.length; i++) {
                    var region = regions[i];
                    
                    // 정확히 일치하는 지역명 찾기
                    if (buttonRegions[region]) {
                        try {
                            buttonRegions[region].scrollIntoView(true);
                            setTimeout(function() {
                                buttonRegions[region].click();
                            }, 500);
                            clickedRegions.push(region);
                            console.log('JavaScript: ' + region + ' 지역 추가 버튼 클릭 완료');
                        } catch (e) {
                            console.log('JavaScript: ' + region + ' 지역 버튼 클릭 실패: ' + e);
                        }
                    } else {
                        // 부분 일치하는 지역명 찾기
                        var found = false;
                        for (var regionName in buttonRegions) {
                            if (regionName.toLowerCase().includes(region.toLowerCase()) || 
                                region.toLowerCase().includes(regionName.toLowerCase())) {
                                try {
                                    buttonRegions[regionName].scrollIntoView(true);
                                    setTimeout(function() {
                                        buttonRegions[regionName].click();
                                    }, 500);
                                    clickedRegions.push(regionName);
                                    found = true;
                                    console.log('JavaScript: ' + region + ' 지역 추가 버튼 클릭 완료 (일치: ' + regionName + ')');
                                    break;
                                } catch (e) {
                                    console.log('JavaScript: ' + region + ' 지역 버튼 클릭 실패 (일치: ' + regionName + '): ' + e);
                                }
                            }
                        }
                        
                        if (!found) {
                            console.log('JavaScript: ' + region + ' 지역을 찾을 수 없어 선택하지 않았습니다.');
                        }
                    }
                }
                
                // 클릭한 지역이 없으면 서울 선택
                if (clickedRegions.length === 0) {
                    console.log('JavaScript: 선택한 지역이 없어 기본값 서울 선택');
                    
                    if (buttonRegions['서울']) {
                        try {
                            buttonRegions['서울'].scrollIntoView(true);
                            setTimeout(function() {
                                buttonRegions['서울'].click();
                            }, 500);
                            clickedRegions.push('서울');
                            console.log('JavaScript: 서울 지역 추가 버튼 클릭 완료 (기본값)');
                        } catch (e) {
                            console.log('JavaScript: 서울 지역 버튼 클릭 실패: ' + e);
                        }
                    } else if (allButtons.length > 0) {
                        try {
                            allButtons[0].scrollIntoView(true);
                            setTimeout(function() {
                                allButtons[0].click();
                            }, 500);
                            console.log('JavaScript: 첫 번째 지역 추가 버튼 클릭 완료 (최후의 수단)');
                        } catch (e) {
                            console.log('JavaScript: 첫 번째 지역 추가 버튼 클릭 실패: ' + e);
                        }
                    }
                }
                
                return clickedRegions;
            """, selected_regions)
            
            if js_result and len(js_result) > 0:
                print(f"JavaScript로 선택된 지역: {js_result}")
            else:
                print("JavaScript로도 지역 선택 실패")

        # 지역 선택 후 충분한 대기 시간 추가
        print("지역 선택 완료 후 추가 대기 중...")
        time.sleep(5)  # 지역 선택 후 5초 대기
        
        # 3. 확인 버튼 클릭 - click_elements와 동일하게
        try:
            print("확인 버튼 찾는 중...")
            # 확인 버튼이 로드될 때까지 대기
            confirm_button = long_wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "button.btn.round.r30.blue")))
            
            # 버튼이 보이도록 스크롤
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_button)
            time.sleep(2)  # 스크롤 후 잠시 대기
            
            # 버튼 클릭
            driver.execute_script("arguments[0].click();", confirm_button)
            print("확인 버튼 클릭 완료")
            
            # 클릭 후 추가 대기
            print("확인 버튼 클릭 후 데이터 로딩 중...")
            time.sleep(10)  # 확인 버튼 클릭 후 10초 대기
            
        except Exception as e:
            print(f"확인 버튼 클릭 첫 번째 방법 실패: {str(e)}")
            try:
                # 두 번째 방법: 텍스트로 확인 버튼 찾기
                confirm_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '확인')]")
                if confirm_buttons:
                    # 버튼이 사용 가능한 상태가 될 때까지 대기 로직 추가
                    for i in range(10):  # 최대 10회 시도
                        try:
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", confirm_buttons[0])
                            time.sleep(1)
                            driver.execute_script("arguments[0].click();", confirm_buttons[0])
                            print(f"확인 버튼 클릭 완료 (대체 방법, {i+1}번째 시도)")
                            # 클릭 후 데이터 로딩 대기
                            print("확인 버튼 클릭 후 데이터 로딩 중...")
                            time.sleep(10)  # 확인 버튼 클릭 후 10초 대기
                            break
                        except Exception as click_error:
                            print(f"클릭 시도 {i+1} 실패: {str(click_error)}")
                            time.sleep(2)  # 재시도 전 대기
                else:
                    # 세 번째 시도: 강제 JavaScript 실행
                    print("확인 버튼을 찾을 수 없어 JavaScript로 시도")
                    driver.execute_script("""
                        var buttons = document.querySelectorAll('button');
                        for (var i = 0; i < buttons.length; i++) {
                            if (buttons[i].textContent.includes('확인')) {
                                buttons[i].scrollIntoView({block: 'center'});
                                setTimeout(function() {
                                    buttons[i].click();
                                    console.log('JavaScript로 확인 버튼 클릭');
                                }, 1000);
                                return true;
                            }
                        }
                        return false;
                    """)
                    print("JavaScript로 확인 버튼 클릭 시도")
                    # 클릭 후 데이터 로딩 대기
                    print("JavaScript 확인 버튼 클릭 후 데이터 로딩 중...")
                    time.sleep(10)  # 확인 버튼 클릭 후 10초 대기
            except Exception as e2:
                print(f"확인 버튼 클릭 대체 방법도 실패: {str(e2)}")
                print("확인 버튼 클릭 실패 - 계속 진행")
        
        # 4. 기간 선택 - click_elements와 동일하게
        try:
            print(f"선택하려는 기간: {period}")
            
            # 기간에 따른 인덱스 매핑
            period_index = {
                '1년': 0,
                '6개월': 1,
                '3개월': 2,
                '1개월': 3
            }.get(period, 1)  # 기본값 6개월(인덱스 1)
            
            # JavaScript로 인덱스에 해당하는 버튼 클릭
            clicked = driver.execute_script("""
                var buttons = document.querySelectorAll('button.btn.textline');
                var index = arguments[0];
                
                if (buttons.length > index) {
                    console.log("클릭할 버튼 텍스트: " + buttons[index].textContent);
                    buttons[index].click();
                    return true;
                }
                return false;
            """, period_index)
            
            if clicked:
                print(f"{period} 기간 선택 완료 (인덱스 {period_index} 사용)")
            else:
                print(f"{period} 기간 선택 실패")
                
                # 대체 방법 시도
                period_elements = driver.find_elements(By.XPATH, f"//button[text()='{period}']")
                if period_elements:
                    driver.execute_script("arguments[0].click();", period_elements[0])
                    print(f"{period} 기간 선택 완료 (대체 방법)")
            
            # 기간 선택 후 데이터 로딩 대기
            print("기간 선택 후 데이터 로딩 중...")
            time.sleep(5)  # 기간 선택 후 5초 대기
            
        except Exception as e:
            print(f"기간 선택 중 오류 발생: {str(e)}")
            # 기간 선택을 건너뛰고 계속 진행
            print("기간 선택을 건너뛰고 계속 진행합니다.")
        
        # 5. 전용면적 입력 (값이 있는 경우에만)
        if min_area is not None or max_area is not None:
            try:
                # 전용면적 최소값 입력 - 정확한 선택자 사용
                min_area_selector = "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2020 > div.fonmhead > div:nth-child(2) > div > input[type=text]:nth-child(1)"
                if min_area is not None:
                    min_area_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, min_area_selector)))
                    min_area_input.clear()
                    min_area_input.send_keys(str(min_area))
                    print(f"전용면적 최소값 {min_area} 입력 완료")
                
                # 전용면적 최대값 입력 - 정확한 선택자 사용
                max_area_selector = "#gridTableWrapper > div > div > div.gTable > div.gHead > div > div > div > div.headitem.I2020 > div.fonmhead > div:nth-child(2) > div > input[type=text]:nth-child(3)"
                if max_area is not None:
                    max_area_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, max_area_selector)))
                    max_area_input.clear()
                    max_area_input.send_keys(str(max_area))
                    print(f"전용면적 최대값 {max_area} 입력 완료")
                
                # 엔터 키 입력으로 필터 적용
                if max_area is not None:
                    max_area_input.send_keys(Keys.ENTER)
                elif min_area is not None:
                    min_area_input.send_keys(Keys.ENTER)
                
                print("전용면적 필터 적용 완료")
                time.sleep(2)  # 필터 적용 대기
            except Exception as e:
                print(f"전용면적 입력 실패: {str(e)}")
                print("전용면적 필터 적용을 건너뛰고 계속 진행합니다.")
        
        # 수집된 데이터를 저장할 리스트
        all_data = []
        
        # 각 단지명에 대해 검색 수행
        # 각 단지명에 대해 검색 수행
        for idx, complex_name in enumerate(complex_names):
            print(f"{idx+1}/{len(complex_names)} - 단지명 '{complex_name}' 검색 시작")
            
            if gui:
                progress_value = 5 + idx  # 5부터 시작해서 단지명 개수만큼 증가
                gui.update_progress(progress_value, f"'{complex_name}' 검색 중... ({idx+1}/{len(complex_names)})")
            
            try:
                # 검색 입력 필드 찾기
                try:
                    search_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                        "input[data-v-9bca83a2][type='text']:not([name])")))
                except:
                    # 대체 방법: 더 일반적인 선택자 시도
                    search_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 
                        "input[type='text']:not([name])")))
                
                # 이전 검색어 지우기
                search_input.clear()
                time.sleep(0.5)
                
                # 단지명 입력
                search_input.send_keys(complex_name)
                time.sleep(1)
                
                # 엔터 키 입력으로 검색
                search_input.send_keys(Keys.ENTER)
                print(f"'{complex_name}' 검색어 입력 및 엔터 완료")
                
                # 검색 결과 로딩 대기 - 더 길게 설정
                time.sleep(5)
                
                # 데이터가 나타날 때까지 명시적 대기 추가
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 
                        "#gridTableWrapper > div > div > div.gLeft.flexgroup > div:nth-child(2) > div.body > div.bodyitem"))
                    )
                    
                    # 금액별 내림차순 정렬 시도 (시세 헤더 클릭)
              
                    # 매매시세 오름차순 정렬 시도
                    try:
                        # 매매시세(억원) 헤더 요소 찾기 및 클릭
                        price_headers = driver.find_elements(By.CSS_SELECTOR, "div[data-v-9bca83a2][style='cursor: pointer;']")
                        price_header = None
                        
                        for header in price_headers:
                            if "시세(억원)" in header.text:
                                price_header = header
                                break
                        
                        if price_header:
                            driver.execute_script("arguments[0].click();", price_header)
                            print("매매시세(억원) 헤더 클릭 - 오름차순 정렬")
                            time.sleep(2)
                        else:
                            print("매매시세(억원) 헤더를 찾을 수 없음")
                    except Exception as e:
                        print(f"매매시세(억원) 헤더 클릭 실패: {str(e)}")
                        
                        # 대체 방법: JavaScript로 매매시세 헤더 찾기 및 클릭
                        try:
                            driver.execute_script("""
                                var headers = document.querySelectorAll('div[data-v-9bca83a2][style="cursor: pointer;"]');
                                for (var i = 0; i < headers.length; i++) {
                                    if (headers[i].textContent.trim().includes('시세(억원)')) {
                                        headers[i].click();
                                        console.log("매매시세(억원) 헤더 찾아 클릭 (JS)");
                                        break;
                                    }
                                }
                            """)
                            time.sleep(2)
                            print("JavaScript로 매매시세(억원) 헤더 클릭 시도")
                        except:
                            print("JavaScript로 매매시세(억원) 헤더 클릭도 실패")
                except:
                    print(f"'{complex_name}' 검색 결과가 로드되지 않았습니다.")
                
                # 검색 결과 파싱
                # 검색 결과 파싱
                data, _ = parse_complex_data(driver, 50)  # 최대 50개까지 조회
                
                if data:
                    print(f"'{complex_name}'에 대한 검색 결과 {len(data)}개 발견")
                    # 첫 번째 데이터의 컬럼 확인 (디버깅용)
                    if len(data) > 0:
                        print(f"첫 번째 데이터 컬럼: {list(data[0].keys())}")
                        if '매매증감률' in data[0]:
                            print(f"매매증감률 값: {data[0]['매매증감률']}")
                        if '증감률' in data[0]:
                            print(f"증감률 값: {data[0]['증감률']}")
                    
                    # 필터링 로직 개선: 부분 일치도 허용하고 더 유연하게 검색
                    filtered_data = []
                    for item in data:
                        db_complex_name = item.get('단지명', '')
                        
                        # 단지명 매칭 로직 (기존과 동일)
                        search_no_space = complex_name.replace(' ', '').lower()
                        db_no_space = db_complex_name.replace(' ', '').lower()
                        
                        name_match = (complex_name.lower() in db_complex_name.lower() or
                                    db_complex_name.lower() in complex_name.lower() or
                                    search_no_space in db_no_space)
                        
                        if name_match:
                            item['검색단지명'] = complex_name  # 원래 검색한 단지명 추가
                            filtered_data.append(item)
                    
                   
                    # search_by_complex_names 함수 내 수정 부분
                    # search_by_complex_names 함수 내 수정 부분
                    if filtered_data:
                        print(f"'{complex_name}'에 대한 필터링된 결과 {len(filtered_data)}개 발견")
                        
                        # 전용면적 범위가 설정되었는지 확인
                        area_filter_applied = min_area is not None or max_area is not None
                        
                        if area_filter_applied:
                            # 전용면적 범위가 설정된 경우: 가장 비싼 전용면적만 선택
                            print(f"전용면적 범위가 설정되어 최적 평형만 선택합니다.")
                            best_data = find_most_expensive_areas(filtered_data)
                            
                            if best_data:
                                print(f"'{complex_name}'에서 선택된 최적 전용면적 옵션: {len(best_data)}개")
                                all_data.extend(best_data)
                            else:
                                print(f"'{complex_name}'에서 적합한 전용면적 옵션을 찾지 못했습니다.")
                        else:
                            # 전용면적 범위가 설정되지 않은 경우: 모든 평형 포함
                            print(f"전용면적 범위가 설정되지 않아 모든 평형을 포함합니다.")
                            all_data.extend(filtered_data)
                            print(f"'{complex_name}'의 모든 평형 {len(filtered_data)}개가 추가되었습니다.")
                    else:
                        print(f"'{complex_name}'에 대한 필터링된 결과가 없습니다.")
                
            except Exception as e:
                print(f"'{complex_name}' 검색 중 오류 발생: {str(e)}")
                continue
      
           # 모든 검색 완료 후 결과 저장
        # 모든 검색 완료 후 결과 저장
        if all_data:
            if gui:
                gui.update_progress(len(complex_names) + 4, "검색 결과를 정리하는 중...")            
            # 매매시세(억원)와 전세시세 값을 만원 단위로 변환하고 매매전세차 계산
            for item in all_data:
                try:
                    sale_price_man = 0  # 매매시세(만원)
                    rent_price_man = 0  # 전세시세(만원)
                    
                    # 매매시세(억원) 변환 (억 단위 -> 만원 단위)
                    if '매매시세(억원)' in item and item['매매시세(억원)']:
                        # 억 단위 문자열 (예: '5.2')을 만원 단위 숫자로 변환 (예: 52000)
                        sale_price_str = item['매매시세(억원)'].replace(',', '')
                        if sale_price_str.replace('.', '', 1).isdigit():
                            sale_price_float = float(sale_price_str)
                            # 억 단위를 만원 단위로 변환 (1억 = 10000만원)
                            sale_price_man = int(sale_price_float * 10000)
                            # 컬럼명 변경 및 값 업데이트
                            item['매매시세(만원)'] = format(sale_price_man, ',')
                            del item['매매시세(억원)']
                    
                    # 전세시세 변환 (억 단위 -> 만원 단위)
                    if '전세시세' in item and item['전세시세']:
                        # 전세시세 형식이 '억' 단위인 경우 (예: '3.5')
                        rent_price_str = item['전세시세'].replace(',', '')
                        # 숫자만 있는지 확인
                        if rent_price_str.replace('.', '', 1).isdigit():
                            rent_price_float = float(rent_price_str)
                            # 억 단위를 만원 단위로 변환
                            rent_price_man = int(rent_price_float * 10000)
                            # 값 업데이트
                            item['전세시세(만원)'] = format(rent_price_man, ',')
                            del item['전세시세']
                    
                    # 매매전세차 계산 (매매시세 - 전세시세, 만원 단위)
                    # 매매전세차 계산 (매매시세 - 전세시세, 만원 단위)
                    if sale_price_man > 0 and rent_price_man > 0:
                        price_diff_man = sale_price_man - rent_price_man
                        item['매매전세차(만원)'] = format(price_diff_man, ',')
                        print(f"매매전세차 계산: {sale_price_man:,} - {rent_price_man:,} = {price_diff_man:,}")
                    elif sale_price_man > 0:
                        # 전세시세가 없는 경우 매매시세만 표시
                        item['매매전세차(만원)'] = format(sale_price_man, ',')
                    else:
                        # 매매시세나 전세시세가 없는 경우 빈 값
                        item['매매전세차(만원)'] = ''
                    
                    # 기존 매매전세차 컬럼이 있으면 항상 삭제 (새로 계산한 값으로 대체)
                    if '매매전세차' in item:
                        del item['매매전세차']

                    # 시세(만원)을 평당시세(만원)로 변환 (3.3 곱하기)
                    if '시세(만원)' in item and item['시세(만원)']:
                        try:
                            # 콤마 제거하고 숫자로 변환
                            price_str = str(item['시세(만원)']).replace(',', '')
                            if price_str.replace('.', '', 1).isdigit():
                                price_float = float(price_str)
                                # 3.3을 곱해서 평당시세 계산
                                pyeong_price = price_float * 3.305785
                                item['평당시세(만원)'] = format(round(pyeong_price, 2), ',')
                                print(f"평당시세 계산: {price_float:,} × 3.305785 = {pyeong_price:,.2f}")
                            else:
                                # 숫자가 아닌 경우 원래 값 유지
                                item['평당시세(만원)'] = item['시세(만원)']
                            
                            # 기존 시세(만원) 컬럼 삭제
                            del item['시세(만원)']
                        except Exception as e:
                            print(f"평당시세 계산 중 오류: {str(e)}")
                            # 오류 발생 시 컬럼명만 변경
                            item['평당시세(만원)'] = item['시세(만원)']
                            del item['시세(만원)']                        
                    
                except Exception as e:
                    print(f"가격 변환 및 매매전세차 계산 중 오류 발생: {str(e)}")
                    # 오류 발생 시 원래 값 유지
                    continue
                    
            # 현재 시간을 파일명에 포함
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

            # 지역명을 파일명에 포함
            region_name = ""
            if selected_regions and len(selected_regions) > 0:
                region_name = "_".join(selected_regions) + "_"

            file_name = f'complex_search_results_{region_name}{current_time}.xlsx'
            
            # 저장 경로가 있으면 경로 추가
            if save_path:
                file_path = os.path.join(save_path, file_name)
            else:
                file_path = file_name

            
            # DataFrame 생성 및 엑셀 저장
            df = pd.DataFrame(all_data)
            
            # 필요한 컬럼 순서 지정 (변경된 컬럼명 적용)
            # 필요한 컬럼 순서 지정 (매매시세를 매매증감률 앞으로, 전세시세를 전세증감률 앞으로)
            column_order = [
                '검색단지명',          
                '지역명',
                '단지명', 
                '전용면적',
                '세대수',              # 전용면적 바로 옆으로 이동
                '연차', 
                '평당시세(만원)',      # 시세(만원) → 평당시세(만원)로 변경
                '매매시세(만원)',      
                '매매증감률',          
                '전세시세(만원)',      
                '전세증감률',          
                '전세가율', 
                '매매전세차(만원)'
            ]
            # 실제 존재하는 컬럼만 선택하고 순서 적용
            existing_columns = []
            for col in column_order:
                if col in df.columns:
                    existing_columns.append(col)
            
            # 원하는 순서에 없는 컬럼들도 추가 (누락 방지)
            for col in df.columns:
                if col not in existing_columns:
                    existing_columns.append(col)
            
            # 컬럼 순서대로 정렬
            df = df[existing_columns]
            
            print("최종 컬럼 순서:", existing_columns)
            
            # 숫자 형식으로 변환 
            # 숫자 형식으로 변환 
            price_columns = ['평당시세(만원)', '매매시세(만원)', '전세시세(만원)', '매매전세차(만원)']  # '시세(만원)' → '평당시세(만원)'
            for column in price_columns:
                if column in df.columns:
                    # 콤마 제거하고 숫자로 변환
                    df[column] = df[column].astype(str).str.replace(',', '').replace('', '0')
                    df[column] = pd.to_numeric(df[column], errors='coerce')
            
            # 증감률과 전세가율 변환 - 백분율을 소수로 변환하지 않음
            rate_columns = ['매매증감률', '전세증감률', '전세가율']
            for column in rate_columns:
                if column in df.columns:
                    # % 기호만 제거하고 숫자로 변환 (100으로 나누지 않음)
                    df[column] = df[column].astype(str).str.replace('%', '').replace('', '0')
                    df[column] = pd.to_numeric(df[column], errors='coerce')
            
            # '증감률' 컬럼이 있는 경우 '매매증감률'로 매핑
            if '증감률' in df.columns:
                if '매매증감률' not in df.columns:
                    df['매매증감률'] = df['증감률']
                df = df.drop('증감률', axis=1)
                print("'증감률' 컬럼을 '매매증감률'로 변경했습니다.")
            
            # 연차와 전용면적도 숫자로 변환
            # 연차, 전용면적, 세대수도 숫자로 변환
            numeric_columns = ['연차', '전용면적', '세대수']  # 세대수 추가
            for column in numeric_columns:
                if column in df.columns:
                    df[column] = df[column].astype(str).replace('', '0')
                    df[column] = pd.to_numeric(df[column], errors='coerce')
            
            print("데이터 변환 후 컬럼들:", df.columns.tolist())
            
            # 연차와 전용면적도 숫자로 변환
            if '연차' in df.columns:
                df['연차'] = df['연차'].astype(str).replace('', '0')
                df['연차'] = pd.to_numeric(df['연차'], errors='coerce')
            
            if '전용면적' in df.columns:
                df['전용면적'] = df['전용면적'].astype(str).replace('', '0')
                df['전용면적'] = pd.to_numeric(df['전용면적'], errors='coerce')
            
            # 이 부분을 삭제하고 아래 메시지로 대체
            print(f"검색한 단지명 수: {len(complex_names)}, 검색 결과 수: {len(all_data)}")
            
            # ExcelWriter를 사용하여 열 너비 자동 조정 (중복된 두 번째 블록 제거)
            # with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            #     df.to_excel(writer, index=False, sheet_name='검색결과')
                
            #     # 워크시트 가져오기
            #     worksheet = writer.sheets['검색결과']
                
            #     # 열 너비 자동 조정 및 숫자 형식 적용
            #     for idx, col in enumerate(df.columns):
            #         # 열 문자 (A, B, C, ...) 가져오기
            #         column_letter = openpyxl.utils.get_column_letter(idx + 1)
                    
            #         # 각 열의 최대 문자 길이 찾기
            #         max_length = max([len(str(col))] + [len(str(x)) for x in df[col] if pd.notna(x)])
                    
            #         # 너비 설정 (문자 길이 * 1.2 + 4 여백 추가)
            #         adjusted_width = max_length * 1.2 + 4
            #         worksheet.column_dimensions[column_letter].width = adjusted_width
                    
            #         # 숫자 컬럼에 숫자 형식 적용
            #         # 숫자 컬럼에 숫자 형식 적용
            #         # 숫자 컬럼에 숫자 형식 적용
            #         if col in ['시세(만원)', '매매시세(만원)', '전세시세(만원)', '매매전세차(만원)', '매매증감률', '전세증감률', '전세가율', '연차', '전용면적']:
            #             # 첫 번째 행은 헤더이므로 건너뛰고 2행부터 시작
            #             for row_idx in range(2, len(df) + 2):
            #                 cell = worksheet.cell(row=row_idx, column=idx + 1)
            #                 # 가격 컬럼에 숫자 형식 적용
            #                 if col in ['시세(만원)', '매매시세(만원)', '전세시세(만원)', '매매전세차(만원)']:
            #                     cell.number_format = '#,##0'  # 천 단위 구분 기호 사용
            #                 # 증감률과 전세가율에 백분율 형식 적용 - 이미 백분율 값이므로 일반 숫자로 처리
            #                 elif col in ['매매증감률', '전세증감률', '전세가율']:
            #                     cell.number_format = '0.00'  # 일반 숫자 형식 (소수점 2자리)
            #                 # 다른 숫자 컬럼에 일반 숫자 형식 적용
            #                 else:
            #                     cell.number_format = '0.00'  # 소수점 2자리
            
            # # 컬럼 순서대로 정렬
            # df = df[column_order]
            
            # ExcelWriter를 사용하여 열 너비 자동 조정
            # ExcelWriter를 사용하여 열 너비 자동 조정
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='검색결과')
                
                # 워크시트 가져오기
                worksheet = writer.sheets['검색결과']
                
                # 열 너비 자동 조정 - 더 직접적인 방법
                for idx, col in enumerate(df.columns):
                    # 열 문자 (A, B, C, ...) 가져오기
                    column_letter = openpyxl.utils.get_column_letter(idx + 1)
                    
                    # 각 열의 최대 문자 길이 찾기
                    max_length = max([len(str(col))] + [len(str(x)) for x in df[col] if pd.notna(x)])
                    
                    # 너비 설정 (문자 길이 * 1.2 + 4 여백 추가)
                    adjusted_width = max_length * 1.2 + 4
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 드라이버 종료
            driver.quit()
            print("웹 브라우저를 종료했습니다.")
            
            # 저장 완료 메시지 표시 (드라이버 종료 후) - 추가할 부분
            QMessageBox.information(None, "단지명 검색 완료", 
                                   f"🔍 단지명 검색이 완료되었습니다!\n\n"
                                   f"💾 저장 위치: {file_path}\n"
                                   f"🏢 검색된 결과 수: {len(all_data)}개\n"
                                   f"📋 검색한 단지명: {len(complex_names)}개\n\n"
                                   f"✅ 데이터를 확인해보세요!")
                                   
        else:
            # 추출된 데이터가 없는 경우
            # 드라이버 종료
            driver.quit()
            print("웹 브라우저를 종료했습니다.")
            
            QMessageBox.warning(None, "단지명 검색 결과 없음", 
                               f"❌ 검색 결과가 없습니다.\n\n"
                               f"📋 검색한 단지명: {len(complex_names)}개\n"
                               f"💡 단지명을 확인하시거나 다른 지역을 선택해보세요.")
        
        return True
   
    except Exception as e:
        print(f"단지명 검색 중 오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # 예외 발생 시에도 크롬 창 닫기
        try:
            driver.quit()
            print("예외 발생으로 웹 브라우저를 종료했습니다.")
        except:
            print("웹 브라우저 종료 실패")
        
        # 오류 메시지 표시
        QMessageBox.critical(None, "검색 오류", f"단지명 검색 중 오류가 발생했습니다:\n{str(e)}")
        return False


# 기존 코드의 나머지 부분은 그대로 유지 (parse_complex_data 함수 등)

def main():
    app = QApplication(sys.argv)
    window = ApartmentSearchGUI()
    window.show()
    app.exec()  # sys.exit() 제거

if __name__ == '__main__':
    main()
