import tkinter as tk 
from tkinter import ttk, messagebox, filedialog 
from selenium import webdriver
import logging 
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from PIL import Image, ImageTk
import time
import os
import sys
import shutil
import json  # 맨 위에 import 추가
import warnings
import win32com.client
import numpy as np
from openpyxl import load_workbook
import requests
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import glob  # 파일 상단에 추가
import matplotlib.pyplot as plt
from openpyxl import Workbook  # 상단 임포트 구문에 추가
import fnmatch 


class KBLandApp:
    def __init__(self):
        self.downloaded_data = {}
        self.root = tk.Tk()
        self.root.title("KB부동산 시세 분석기")
        self.root.geometry("1400x750")
        self.root.withdraw()  # 초기 설정 동안 메인 창 숨기기
        
        # 초기 설정 파일 확인
        settings_file = os.path.join(os.getcwd(), 'settings.json')
        
        # 최초 실행 확인
        if not os.path.exists(settings_file) or not self.validate_settings(settings_file):
            # 최초 실행 시 초기 설정 대화상자 표시
            if not self.show_initial_setup():
                # 설정 취소 시 프로그램 종료
                self.root.destroy()
                sys.exit()
        
        # 설정 로드
        self.load_settings()
        
        # 설정 완료 후 메인 창 표시
        self.root.deiconify()
        
        # 나머지 초기화 진행
        self.initialize_after_settings()
    
        # 심리차트 표시 여부 상태 변수 수정 (두 개에서 하나로)
        self.show_una_sentiment = tk.BooleanVar(value=True)
        # PIR 차트 표시 여부 상태 변수 추가
        self.show_pir = tk.BooleanVar(value=True)

        # 데이터 표시 옵션 추가
        self.show_kb_price = tk.BooleanVar(value=True)      # KB시세 표시
        self.show_naver_deal = tk.BooleanVar(value=True)    # 네이버매물 표시
        self.show_real_trade = tk.BooleanVar(value=True)    # 실거래가 표시
        
        
        # KB부동산 가격 데이터 선택을 위한 변수 추가
        self.sale_price_type = tk.StringVar(value="normal")  # 기본값은 일반 평균
        self.lease_price_type = tk.StringVar(value="normal")  # 기본값은 일반 평균
        
        # 설정 파일 로드 (lawdong_path를 여기서 먼저 정의)
        settings_file = os.path.join(os.getcwd(), 'settings.json')
        
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings_data = json.load(f)
                    self.download_path = settings_data.get('download_path', "C:\\Download")
                    self.history_path = settings_data.get('history_path', os.path.join(self.download_path, "history"))
                    self.lawdong_path = settings_data.get('lawdong_path', "C:/law-dong/law-dong.txt")
                    self.sentiment_excel_path = settings_data.get('sentiment_excel_path', "")
            except:
                self.download_path = "C:\\Download"
                self.history_path = os.path.join(self.download_path, "history")
                self.lawdong_path = "C:/law-dong/law-dong.txt"
                self.sentiment_excel_path = ""
        else:
            self.download_path = "C:\\Download"
            self.history_path = os.path.join(self.download_path, "history")
            self.lawdong_path = "C:/law-dong/law-dong.txt"
            self.sentiment_excel_path = ""
        
        # lawdong_path가 정의된 후에 폰트 경로 설정
        font_path = os.path.join(os.path.dirname(self.lawdong_path), "KoPubWorld Dotum Medium.ttf")
        if os.path.exists(font_path):
            self.font_normal = ('KoPubWorld Dotum Medium', 9)
        else:
            self.font_normal = ('Helvetica', 9)  # 폴백

        
        # KB부동산 가격 데이터 선택을 위한 변수 추가 - 이 두 줄만 추가
        self.sale_price_type = tk.StringVar(value="normal")  # 기본값은 일반 평균
        self.lease_price_type = tk.StringVar(value="normal")  # 기본값은 일반 평균
        # GUI 생성 전에 폰트 설정
        self.font_normal = ('Helvetica', 9)  # 기본 폰트 먼저 설정
        self.downloaded_files = {}  # 다운로드된 파일 저장용 딕셔너리 추가
        self.completion_year = None  # 준공년 저장 변수 추가
        self.selected_apt_info = None  # 선택된 단지 정보 저장 변수 추가
        # 신축단지 정보를 위한 변수 추가
        self.new_apt_name = None
        self.new_apt_price = None

        # KB부동산 가격 데이터 선택을 위한 변수 추가 - 리스트와 딕셔너리로 변경
        self.sale_price_types = {}  # 매매가 유형 선택 상태 (키: 유형명, 값: BooleanVar)
        self.lease_price_types = {}  # 전세가 유형 선택 상태 (키: 유형명, 값: BooleanVar)
        
        # 초기 세팅
        for type_name in ["low", "normal", "high"]:
            self.sale_price_types[type_name] = tk.BooleanVar(value=(type_name == "normal"))
            self.lease_price_types[type_name] = tk.BooleanVar(value=(type_name == "normal"))
        

        # 로그 파일 초기화
        debug_file = 'debug.log'
        
        # 기존 로그 핸들러 제거
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
            
        # 안전하게 파일 제거 시도
        try:
            if os.path.exists(debug_file):
                os.remove(debug_file)
        except PermissionError:
            # 파일이 사용 중이면 새로운 이름으로 생성
            base_name = 'debug'
            ext = '.log'
            counter = 1
            while True:
                debug_file = f'{base_name}_{counter}{ext}'
                if not os.path.exists(debug_file):
                    break
                counter += 1
        
        # 로깅 설정
        logging.basicConfig(
            filename=debug_file,
            level=logging.INFO,
            format='%(asctime)s [%(levelname)s] - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        logging.info("=== 애플리케이션 시작 ===")

        # 설정 파일 로드
        settings_file = os.path.join(os.getcwd(), 'settings.json')
    
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings_data = json.load(f)
                    self.download_path = settings_data.get('download_path', "C:\\Download")
                    self.history_path = settings_data.get('history_path', os.path.join(self.download_path, "history"))
                    self.lawdong_path = settings_data.get('lawdong_path', "C:/law-dong/law-dong.txt")
                    # self.naver_excel_path = settings_data.get('naver_excel_path', "")
                    # 심리 데이터 엑셀 경로 추가
                    self.sentiment_excel_path = settings_data.get('sentiment_excel_path', "")
            except:
                self.download_path = "C:\\Download"
                self.history_path = os.path.join(self.download_path, "history")
                self.lawdong_path = "C:/law-dong/law-dong.txt"
                self.naver_excel_path = ""
                self.sentiment_excel_path = ""  # 기본값은 빈 문자열
        else:
            self.download_path = "C:\\Download"
            self.history_path = os.path.join(self.download_path, "history")
            self.lawdong_path = "C:/law-dong/law-dong.txt"
            self.naver_excel_path = ""
            self.sentiment_excel_path = ""  # 기본값은 빈 문자열



            # 폰트 설정 추가
        # 폰트 경로 디버깅을 위한 코드 추가
        font_path = os.path.join(os.path.dirname(self.lawdong_path), "KoPubWorld Dotum Medium.ttf")
        print("\n=== 폰트 설정 디버깅 ===")
        print(f"폰트 파일 경로: {font_path}")
        print(f"폴더 존재 여부: {os.path.exists(os.path.dirname(self.lawdong_path))}")
        print(f"폰트 파일 존재 여부: {os.path.exists(font_path)}")
        print(f"해당 폴더의 파일 목록:")
        try:
            for file in os.listdir(os.path.dirname(self.lawdong_path)):
                print(f"- {file}")
        except Exception as e:
            print(f"폴더 읽기 실패: {str(e)}")
        if os.path.exists(font_path):
            try:
                # GUI 폰트 설정
                from tkinter import font
                custom_font = font.Font(family="custom_font")
                self.root.option_add("*Font", custom_font)
                
                # matplotlib 폰트 설정
                from matplotlib import font_manager
                font_manager.fontManager.addfont(font_path)
                plt.rcParams['font.family'] = font_manager.FontProperties(fname=font_path).get_name()
                plt.rcParams['axes.unicode_minus'] = False  # 마이너스 기호 깨짐 방지
                
                logging.info(f"사용자 지정 폰트 적용 완료: {font_path}")
            except Exception as e:
                logging.error(f"폰트 설정 중 오류 발생: {str(e)}")
                # 기본 폰트 설정으로 폴백
                plt.rcParams['font.family'] = 'Malgun Gothic'
        else:
            logging.warning("폰트 파일을 찾을 수 없어 기본 폰트를 사용합니다.")
            plt.rcParams['font.family'] = 'Malgun Gothic'







        
        self.image_path = os.path.join(self.download_path, "graph.jpg")
############################
        
        # 폴더 생성
        for path in [self.download_path, self.history_path]:
            if not os.path.exists(path):
                os.makedirs(path)
                
        self.history_list = self.load_history()  # 히스토리 목록 로드



        
        # 법정동 코드 관련 변수 초기화
        self.region_codes = {}
        self.sido_list = []
        self.sigungu_dict = {}
        self.dong_dict = {}     

        
        # 나머지 초기화
        self.load_lawdong_file()

        self.setup_fonts()
        self.setup_gui()

        # GUI 설정 후 폰트 설정
        self.setup_fonts()


        
        self.check_chrome_installation()
        
        if not os.path.exists(self.download_path):
            os.makedirs(self.download_path)
            
        # API 키 설정
        self.service_key = "Vs5lXsSo6iEI8no3pP%2FT0udWF9s7Cc8oP1SIWnEI5F4h6dKq92fLvnKmxkoWGJxSeW2%2FSOLQECGxOJzWcjJEXQ%3D%3D"


    def validate_settings(self, settings_file):
        """설정 파일 유효성 검사"""
        try:
            with open(settings_file, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                # 필수 항목 확인
                if 'lawdong_path' not in settings:
                    return False
                if not os.path.exists(settings.get('lawdong_path', '')):
                    return False
                return True
        except:
            return False
    
    def show_initial_setup(self):
        """최초 실행 시 초기 설정 대화상자"""
        setup_dialog = tk.Toplevel(self.root)
        setup_dialog.title("초기 설정")
        setup_dialog.attributes('-topmost', True)
        
        # 창을 화면 중앙에 배치
        width = 700
        height = 450
        screen_width = setup_dialog.winfo_screenwidth()
        screen_height = setup_dialog.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        setup_dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # 결과 저장 변수
        self.setup_completed = False
        
        # 메인 프레임
        main_frame = ttk.Frame(setup_dialog, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # 제목
        title_label = ttk.Label(main_frame, 
                               text="KB부동산 시세 분석기 초기 설정",
                               font=('Arial', 14, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # 설명
        desc_label = ttk.Label(main_frame,
                              text="프로그램을 처음 실행하셨습니다.\n아래 필수 항목을 설정해주세요.",
                              font=('Arial', 10))
        desc_label.pack(pady=(0, 20))
        
        # 설정 프레임
        settings_frame = ttk.LabelFrame(main_frame, text="필수 설정", padding=15)
        settings_frame.pack(fill='both', expand=True, pady=(0, 20))
        
        # 1. 법정동 코드 파일 설정
        lawdong_frame = ttk.Frame(settings_frame)
        lawdong_frame.pack(fill='x', pady=10)
        
        ttk.Label(lawdong_frame, text="법정동 코드 파일:", width=15, anchor='w').pack(side='left')
        lawdong_path_var = tk.StringVar(value="")
        lawdong_entry = ttk.Entry(lawdong_frame, textvariable=lawdong_path_var, width=40)
        lawdong_entry.pack(side='left', padx=5)
        
        def select_lawdong():
            file_path = filedialog.askopenfilename(
                title="법정동 코드 파일(law-dong.txt) 선택",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                parent=setup_dialog
            )
            if file_path:
                lawdong_path_var.set(file_path)
                # 같은 폴더에서 폰트 파일 자동 검색
                font_dir = os.path.dirname(file_path)
                font_file = os.path.join(font_dir, "KoPubWorld Dotum Medium.ttf")
                if os.path.exists(font_file):
                    font_path_var.set(font_file)
                    font_status_label.config(text="✓ 폰트 파일이 자동으로 감지되었습니다.", 
                                           foreground="green")
        
        ttk.Button(lawdong_frame, text="찾아보기", command=select_lawdong).pack(side='left')
        
        # 2. 폰트 파일 설정 (선택사항)
        font_frame = ttk.Frame(settings_frame)
        font_frame.pack(fill='x', pady=10)
        
        ttk.Label(font_frame, text="폰트 파일(선택):", width=15, anchor='w').pack(side='left')
        font_path_var = tk.StringVar(value="")
        font_entry = ttk.Entry(font_frame, textvariable=font_path_var, width=40)
        font_entry.pack(side='left', padx=5)
        
        def select_font():
            file_path = filedialog.askopenfilename(
                title="폰트 파일 선택 (KoPubWorld Dotum Medium.ttf)",
                filetypes=[("TrueType Font", "*.ttf"), ("All files", "*.*")],
                parent=setup_dialog
            )
            if file_path:
                font_path_var.set(file_path)
                font_status_label.config(text="✓ 폰트 파일이 선택되었습니다.", 
                                       foreground="green")
        
        ttk.Button(font_frame, text="찾아보기", command=select_font).pack(side='left')
        
        # 폰트 상태 표시
        font_status_label = ttk.Label(settings_frame, text="※ 폰트를 선택하지 않으면 기본 폰트를 사용합니다.", 
                                     font=('Arial', 9), foreground="gray")
        font_status_label.pack(pady=5)
        
        # 3. 다운로드 경로 설정
        download_frame = ttk.Frame(settings_frame)
        download_frame.pack(fill='x', pady=10)
        
        ttk.Label(download_frame, text="다운로드 경로:", width=15, anchor='w').pack(side='left')
        download_path_var = tk.StringVar(value="C:\\Download")
        download_entry = ttk.Entry(download_frame, textvariable=download_path_var, width=40)
        download_entry.pack(side='left', padx=5)
        
        def select_download():
            folder_path = filedialog.askdirectory(
                title="다운로드 폴더 선택",
                parent=setup_dialog
            )
            if folder_path:
                download_path_var.set(folder_path)
        
        ttk.Button(download_frame, text="찾아보기", command=select_download).pack(side='left')
        
        # 버튼 프레임
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=10)
        
        def save_initial_settings():
            # 필수 파일 확인
            if not lawdong_path_var.get():
                messagebox.showerror("오류", "법정동 코드 파일을 선택해주세요.", parent=setup_dialog)
                return
            
            if not os.path.exists(lawdong_path_var.get()):
                messagebox.showerror("오류", "선택한 법정동 코드 파일이 존재하지 않습니다.", parent=setup_dialog)
                return
            
            # 폰트 파일 확인 (선택사항)
            font_path = font_path_var.get()
            if font_path and not os.path.exists(font_path):
                if messagebox.askyesno("확인", 
                                      "폰트 파일을 찾을 수 없습니다.\n기본 폰트를 사용하시겠습니까?",
                                      parent=setup_dialog):
                    font_path = ""
                else:
                    return
            
            # 폴더 생성
            download_path = download_path_var.get()
            history_path = os.path.join(download_path, "history")
            
            try:
                os.makedirs(download_path, exist_ok=True)
                os.makedirs(history_path, exist_ok=True)
            except Exception as e:
                messagebox.showerror("오류", f"폴더 생성 실패: {str(e)}", parent=setup_dialog)
                return
            
            # settings.json에 저장
            settings_data = {
                'download_path': download_path,
                'history_path': history_path,
                'lawdong_path': lawdong_path_var.get(),
                'font_path': font_path,
                'sentiment_excel_path': "",
                'first_run_completed': True
            }
            
            try:
                with open('settings.json', 'w', encoding='utf-8') as f:
                    json.dump(settings_data, f, ensure_ascii=False, indent=2)
                
                messagebox.showinfo("성공", "초기 설정이 완료되었습니다.\n프로그램을 시작합니다.", 
                                  parent=setup_dialog)
                self.setup_completed = True
                setup_dialog.destroy()
            except Exception as e:
                messagebox.showerror("오류", f"설정 저장 실패: {str(e)}", parent=setup_dialog)
        
        def cancel_setup():
            if messagebox.askyesno("확인", "설정을 취소하시겠습니까?\n프로그램이 종료됩니다.", 
                                  parent=setup_dialog):
                setup_dialog.destroy()
        
        ttk.Button(button_frame, text="설정 완료", command=save_initial_settings,
                  width=15).pack(side='right', padx=5)
        ttk.Button(button_frame, text="취소", command=cancel_setup,
                  width=15).pack(side='right')
        
        # 대화상자가 닫힐 때까지 대기
        setup_dialog.protocol("WM_DELETE_WINDOW", cancel_setup)
        setup_dialog.grab_set()
        self.root.wait_window(setup_dialog)
        
        return self.setup_completed
    
    def load_settings(self):
        """설정 파일 로드"""
        settings_file = os.path.join(os.getcwd(), 'settings.json')
        
        try:
            with open(settings_file, 'r', encoding='utf-8') as f:
                settings_data = json.load(f)
                self.download_path = settings_data.get('download_path', "C:\\Download")
                self.history_path = settings_data.get('history_path', 
                                                       os.path.join(self.download_path, "history"))
                self.lawdong_path = settings_data.get('lawdong_path', "C:/law-dong/law-dong.txt")
                self.custom_font_path = settings_data.get('font_path', "")
                self.sentiment_excel_path = settings_data.get('sentiment_excel_path', "")
                
                logging.info(f"설정 로드 완료: {settings_file}")
        except Exception as e:
            logging.error(f"설정 로드 실패: {str(e)}")
            # 기본값 설정
            self.download_path = "C:\\Download"
            self.history_path = os.path.join(self.download_path, "history")
            self.lawdong_path = "C:/law-dong/law-dong.txt"
            self.custom_font_path = ""
            self.sentiment_excel_path = ""
    
    def initialize_after_settings(self):
        """설정 로드 후 나머지 초기화 작업"""
        # 기존 __init__의 나머지 부분을 여기로 이동
        
        # 상태 변수들 초기화
        self.show_una_sentiment = tk.BooleanVar(value=True)
        self.show_pir = tk.BooleanVar(value=True)
        self.show_kb_price = tk.BooleanVar(value=True)
        self.show_naver_deal = tk.BooleanVar(value=True)
        self.show_real_trade = tk.BooleanVar(value=True)
        
        # 나머지 변수들 초기화...
        # (기존 __init__ 메서드의 나머지 코드를 여기로 이동)

        
    def setup_fonts(self):
        font_path = r"C:\law-dong\KoPubWorld Dotum Medium.ttf"
        
        if os.path.exists(font_path):
            from tkinter import font as tkfont
            
            # 원하는 크기로 지정
            self.font_normal = tkfont.Font(family="KoPubWorld Dotum Medium", size=9, weight="normal")
            self.font_large = tkfont.Font(family="KoPubWorld Dotum Medium", size=11, weight="normal")
            self.font_title = tkfont.Font(family="KoPubWorld Dotum Medium", size=14, weight="normal")  # bold를 normal로 변경
            self.font_button = tkfont.Font(family="KoPubWorld Dotum Medium", size=9, weight="normal")
            # 모든 기본 라벨/위젯의 폰트를 지정
            self.root.option_add("*Font", self.font_normal)
            
            # matplotlib에도 폰트 적용
            from matplotlib import font_manager
            font_manager.fontManager.addfont(font_path)
            font_name = font_manager.FontProperties(fname=font_path).get_name()
            plt.rcParams['font.family'] = font_name
            plt.rcParams['axes.unicode_minus'] = False
            
            logging.info(f"사용자 지정 폰트 적용 완료: {font_path}")
        else:
            messagebox.showwarning("폰트 파일 없음", f"폰트 파일을 찾을 수 없습니다:\n{font_path}\n기본 폰트를 사용합니다.")
            plt.rcParams['font.family'] = 'Malgun Gothic'


    def match_region_from_address(self, address, sentiment_regions):
        """주소에서 지역을 추출하여 심리 데이터의 지역과 매칭"""
        try:
            print(f"\n=== 주소 매칭 시작 ===")
            print(f"입력 주소: {address}")
            print(f"사용 가능한 지역: {sentiment_regions}")
            
            if not address:
                print("주소가 비어있습니다.")
                return None
                
            # 주소에서 도/시 추출
            # 경기도, 강원도, 충청북도, 충청남도, 전라북도, 전라남도, 경상북도, 경상남도, 제주도 등
            do_keywords = ['경기도', '강원도', '강원특별자치도', '충청북도', '충청남도', '전라북도', '전라남도', 
                           '전북특별자치도', '경상북도', '경상남도', '제주도', '제주특별자치도']
            
            # 주요 광역시
            si_keywords = ['서울', '서울특별시', '부산', '부산광역시', '대구', '대구광역시', 
                           '인천', '인천광역시', '광주', '광주광역시', '대전', '대전광역시', 
                           '울산', '울산광역시', '세종', '세종특별자치시']
            
            # 주소에서 '도' 직접 추출
            matched_do = None
            for do in do_keywords:
                if do in address:
                    matched_do = do
                    print(f"'도' 기준 매칭: {matched_do}")
                    break
                    
            # 주소에서 '시' 직접 추출
            matched_si = None
            for si in si_keywords:
                if si in address:
                    matched_si = si
                    print(f"'시' 기준 매칭: {matched_si}")
                    break
                    
            # 지역 매칭 (도 우선, 없으면 시)
            matched_region = matched_do if matched_do else matched_si
            
            if not matched_region:
                print("주소에서 도/시를 찾을 수 없습니다.")
                return None
                
            # 심리 데이터의 지역과 매칭
            for region in sentiment_regions:
                # 정확한 매칭 - 영어명 포함 가능성 처리
                region_name = region.split()[0]  # "강원특별자치도 Gangwon-do"에서 "강원특별자치도"만 추출
                
                # 주소 정보를 정확히 매칭하기 위한 가공 처리
                if matched_region == "강원도" and "강원특별자치도" in region:
                    print(f"강원도 -> 강원특별자치도로 매칭")
                    return region
                elif matched_region == "전라북도" and "전북특별자치도" in region:
                    print(f"전라북도 -> 전북특별자치도로 매칭")
                    return region
                elif matched_region == "제주도" and "제주특별자치도" in region:
                    print(f"제주도 -> 제주특별자치도로 매칭")
                    return region
                
                # 직접 매칭
                if matched_region in region:
                    print(f"최종 매칭된 지역: {region}")
                    return region
                    
            # 부분 매칭 (위에서 매칭되지 않은 경우)
            for region in sentiment_regions:
                # 특수한 경우 처리
                if matched_region == '서울' and ('서울' in region or '서울특별시' in region):
                    print(f"부분 매칭된 지역: {region}")
                    return region
                # 강원특별자치도 대응
                elif (matched_region == '강원도' or matched_region == '강원특별자치도') and ('강원' in region):
                    print(f"부분 매칭된 지역: {region}")
                    return region
                # 전북특별자치도 대응
                elif (matched_region == '전라북도' or matched_region == '전북특별자치도') and ('전북' in region or '전라북도' in region):
                    print(f"부분 매칭된 지역: {region}")
                    return region
                # 제주특별자치도 대응
                elif (matched_region == '제주도' or matched_region == '제주특별자치도') and ('제주' in region):
                    print(f"부분 매칭된 지역: {region}")
                    return region
                # 기타 부분 매칭
                elif matched_region in region or region in matched_region:
                    print(f"부분 매칭된 지역: {region}")
                    return region
                    
            print(f"일치하는 지역을 찾을 수 없습니다: {matched_region}")
            return None
            
        except Exception as e:
            print(f"지역 매칭 중 오류: {str(e)}")
            return None


    def load_sentiment_data(self, file_path=None):
        """심리 데이터 엑셀 파일에서 데이터 로드"""
        try:
            print("\n=== 심리 데이터 로드 시작 ===")
            if not file_path and hasattr(self, 'sentiment_excel_path'):
                file_path = self.sentiment_excel_path
                
            if not file_path or not os.path.exists(file_path):
                print("심리 데이터 파일이 지정되지 않았거나 존재하지 않습니다.")
                return None
                
            print(f"심리 데이터 파일 경로: {file_path}")
            
            # 엑셀 파일 로드
            wb = load_workbook(file_path, data_only=True)
            
            # '7.매수매도' 시트 선택
            try:
                sheet = wb['7.매수매도']
                print("'7.매수매도' 시트를 찾았습니다.")
            except KeyError:
                print("'7.매수매도' 시트를 찾을 수 없습니다.")
                return None
                
            # 지역 목록 가져오기 (2행)
            regions = []
            for col in range(2, sheet.max_column + 1, 3):  # 각 지역은 3개 열 단위로 구분됨
                region_name = sheet.cell(row=2, column=col).value
                if region_name:
                    # 각 지역별로 세 가지 데이터 열 위치 저장
                    regions.append({
                        'name': region_name,
                        'sell_col': col,     # 매도자많음 열
                        'buy_col': col + 1,  # 매수자많음 열
                        'index_col': col + 2 # 매수우위지수 열
                    })
                    
            print(f"발견된 지역: {', '.join([r['name'] for r in regions])}")
            
            # 날짜와 심리 데이터 추출
            sentiment_data = {}
            for row in range(3, sheet.max_row + 1):
                date_cell = sheet.cell(row=row, column=1).value
                if not date_cell:
                    continue
                    
                # 날짜 처리
                if isinstance(date_cell, datetime):
                    date = date_cell
                else:
                    try:
                        # 문자열을 날짜로 변환 시도
                        date = datetime.strptime(str(date_cell), '%Y-%m-%d')
                    except ValueError:
                        try:
                            date = datetime.strptime(str(date_cell), '%Y%m%d')
                        except ValueError:
                            print(f"날짜 변환 실패: {date_cell}")
                            continue
                
                date_str = date.strftime('%Y-%m-%d')
                sentiment_data[date_str] = {}
                
                # 각 지역의 심리 데이터 추출
                for region in regions:
                    region_name = region['name']
                    sell_value = sheet.cell(row=row, column=region['sell_col']).value
                    buy_value = sheet.cell(row=row, column=region['buy_col']).value
                    
                    # 각 지역별로 매도자/매수자 데이터 저장
                    sentiment_data[date_str][region_name] = {
                        'sell': float(sell_value) if sell_value is not None else None,
                        'buy': float(buy_value) if buy_value is not None else None
                    }
                        
            print(f"심리 데이터 로드 완료: {len(sentiment_data)} 개 날짜에 대한 데이터")
            return {'regions': [r['name'] for r in regions], 'data': sentiment_data}
            
        except Exception as e:
            print(f"심리 데이터 로드 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def update_widget_fonts(self, widget):
        """모든 위젯의 폰트를 재귀적으로 업데이트"""
        try:
            widget_class = widget.winfo_class()
            if widget_class in ['TLabel', 'Label']:
                widget.configure(font=self.font_normal)
            elif widget_class in ['TButton', 'Button']:
                widget.configure(font=self.font_button)  # 버튼용 폰트 적용
            elif widget_class == 'TEntry':
                widget.configure(font=self.font_normal)
            elif widget_class == 'Listbox':
                widget.configure(font=self.font_normal)
            elif widget_class == 'TCombobox':
                widget.configure(font=self.font_normal)
            elif widget_class == 'Treeview':
                style = ttk.Style()
                style.configure('Treeview', font=self.font_normal)
                style.configure('Treeview.Heading', font=self.font_normal)
            
            # 제목 레이블은 큰 폰트 사용
            if isinstance(widget, ttk.Label) and widget.cget('text') == "KB부동산 시세 + 실거래가 차트":
                widget.configure(font=self.font_title)
                
        except:
            pass
        
        # 자식 위젯들도 폰트 업데이트
        try:
            for child in widget.winfo_children():
                self.update_widget_fonts(child)
        except:
            pass


    

    def load_lawdong_file(self):
        """법정동 코드 파일 로드"""
        try:
            if not os.path.exists(self.lawdong_path):
                messagebox.showerror("오류", "법정동 코드 파일이 존재하지 않습니다.")
                return False
                
            for encoding in ['cp949', 'euc-kr', 'utf-8']:
                try:
                    with open(self.lawdong_path, 'r', encoding=encoding) as file:
                        for line in file:
                            parts = line.strip().split('\t')
                            if len(parts) < 2:
                                continue
                            
                            code = parts[0].strip()
                            name = parts[1].strip()
                            
                            # 폐지된 동 필터링 (행 전체에서 '폐지' 검사)
                            if any('폐지' in part for part in parts):
                                continue
                                
                            sigungu_code = code[:5]
                            
                            if not code.endswith('00000'):
                                names = name.split()
                                if len(names) >= 2:
                                    sido = names[0]
                                    remaining = names[1:]
                                    
                                    for i in range(len(remaining)-1, -1, -1):
                                        if any(remaining[i].endswith(suffix) for suffix in ['동', '읍', '면', '가']):
                                            dong = remaining[i]
                                            sigungu = ' '.join(remaining[:i])
                                            
                                            if not sigungu:
                                                continue
                                                
                                            if sido not in self.sido_list:
                                                self.sido_list.append(sido)
                                                self.sigungu_dict[sido] = []
                                                
                                            if sigungu not in self.sigungu_dict[sido]:
                                                self.sigungu_dict[sido].append(sigungu)
                                                self.dong_dict[sigungu] = []
                                                
                                            if dong not in self.dong_dict[sigungu]:
                                                self.dong_dict[sigungu].append(dong)
                                                
                                            self.region_codes[(sido, sigungu, dong)] = (code, sigungu_code)
                                            break
                    return True
                except UnicodeDecodeError:
                    continue
            
            messagebox.showerror("오류", "법정동 코드 파일을 읽을 수 없습니다. 인코딩을 확인해주세요.")
            return False
                    
        except Exception as e:
            messagebox.showerror("오류", f"법정동 코드 파일 로드 중 오류: {str(e)}")
            return False


    def load_history(self):
        history_list = []
        if os.path.exists(self.history_path):
            try:
                # 모든 파일을 가져옴
                all_files = os.listdir(self.history_path)
                
                # 단일 분석용 파일과 비교 분석용 파일 분리
                for file in all_files:
                    file_path = os.path.join(self.history_path, file)
                    try:
                        if file.startswith('history_compare_'):
                            # 비교 분석용 파일 처리
                            wb = load_workbook(file_path)
                            ws = wb.active
                            
                            apt_names = ws['B1'].value  # "단지1 vs 단지2" 형식
                            analysis_type = ws['B2'].value  # 분석유형
                            image_path = ws['B5'].value  # 이미지 파일 경로
                            
                            history_list.append({
                                'file_path': file_path,
                                'apt_name': f"[비교] {apt_names}",
                                'area': analysis_type,
                                'search_date': os.path.getmtime(file_path),
                                'max_trade': "비교분석",
                                'type': 'compare'
                            })
                            wb.close()
                            
                        elif file.startswith('history_'):
                            # 단일 분석용 파일 처리
                            df_info = pd.read_excel(file_path, nrows=8)
                            apt_name = str(df_info.iloc[2, 1])
                            area = str(df_info.iloc[5, 1])
                            
                            wb = load_workbook(file_path)
                            ws = wb.active
                            max_trade = ws['M1'].value if ws['M1'].value else "정보없음"
                            wb.close()
                            
                            history_list.append({
                                'file_path': file_path,
                                'apt_name': apt_name,
                                'area': area,
                                'search_date': os.path.getmtime(file_path),
                                'max_trade': max_trade,
                                'type': 'single'
                            })
                            
                    except Exception as e:
                        print(f"파일 처리 중 오류 ({file}): {str(e)}")
                        continue
                        
            except Exception as e:
                print(f"히스토리 로드 중 오류: {str(e)}")
                
        return sorted(history_list, key=lambda x: x['search_date'], reverse=True)
        
    # def load_history(self):
    #     """저장된 히스토리 목록 로드"""
    #     history_list = []
    #     if os.path.exists(self.history_path):
    #         # 엑셀 파일 처리 (단일 분석용)
    #         excel_files = {f: os.path.getmtime(os.path.join(self.history_path, f)) 
    #                       for f in os.listdir(self.history_path) 
    #                       if f.startswith('history_') and f.endswith('.xlsx')}
            
    #         # 이미지 파일 처리 (비교 분석용)
    #         image_files = {f: os.path.getmtime(os.path.join(self.history_path, f)) 
    #                       for f in os.listdir(self.history_path) 
    #                       if f.startswith('비교_') and f.endswith('.jpg')}
    
    #         # 단일 분석 히스토리 처리
    #         for excel_file in excel_files:
    #             file_path = os.path.join(self.history_path, excel_file)
    #             try:
    #                 # 엑셀 파일 읽기
    #                 df_info = pd.read_excel(file_path, nrows=8)
    #                 apt_name = str(df_info.iloc[2, 1])
    #                 area = str(df_info.iloc[5, 1])
                    
    #                 # 최고거래가 정보 읽기
    #                 wb = load_workbook(file_path)
    #                 ws = wb.active
    #                 max_trade = ws['M1'].value if ws['M1'].value else "정보없음"
    #                 wb.close()
                    
    #                 history_list.append({
    #                     'file_path': file_path,
    #                     'apt_name': apt_name,
    #                     'area': area,
    #                     'search_date': excel_files[excel_file],
    #                     'max_trade': max_trade,
    #                     'type': 'single'  # 단일 분석 표시
    #                 })
    #             except Exception as e:
    #                 print(f"단일 분석 파일 처리 중 오류 ({excel_file}): {str(e)}")
    
    #         # 비교 분석 히스토리 처리
    #         for image_file in image_files:
    #             try:
    #                 # 파일명에서 정보 추출 (비교_단지1_단지2_분석유형_날짜시간.jpg)
    #                 parts = image_file.replace('.jpg', '').split('_')
    #                 if len(parts) >= 5:
    #                     apt_names = f"{parts[1]} vs {parts[2]}"
    #                     compare_type = parts[3]
                        
    #                     history_list.append({
    #                         'file_path': os.path.join(self.history_path, image_file),
    #                         'apt_name': f"[비교] {apt_names}",
    #                         'area': compare_type,
    #                         'search_date': image_files[image_file],
    #                         'max_trade': "비교분석",
    #                         'type': 'compare'  # 비교 분석 표시
    #                     })
    #             except Exception as e:
    #                 print(f"비교 분석 파일 처리 중 오류 ({image_file}): {str(e)}")
    
    #     return sorted(history_list, key=lambda x: x['search_date'], reverse=True)

    def save_to_history(self, file_path, trades_data=None):

            # 이전 파일과 같은 이름의 파일이 있다면 삭제
        apt_info = pd.read_excel(file_path, nrows=8)
        apt_name = str(apt_info.iloc[2, 1])
        area = str(apt_info.iloc[5, 1])
        
        for file in os.listdir(self.download_path):
            if apt_name in file and area in file and file.endswith('.xlsx'):
                old_path = os.path.join(self.download_path, file)
                if old_path != file_path:  # 현재 파일이 아닌 경우만 삭제
                    os.remove(old_path)
        print("\n=== 히스토리 저장 시작 ===")
        if file_path and os.path.exists(file_path):
            try:
                # 파일명 생성 (시간 포함)
                new_filename = f"history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                new_path = os.path.join(self.history_path, new_filename)
                print(f"새 히스토리 경로: {new_path}")
                
                # 파일 복사
                shutil.copy2(file_path, new_path)
                print("파일 복사 완료")
                
                # 실거래가 데이터를 이용해 최고거래가 정보 저장
                if trades_data:
                    print("실거래 데이터 확인됨")
                    print(f"실거래 데이터 개수: {len(trades_data)}")
                    try:
                        max_trade = max(trades_data, key=lambda x: x['price'])
                        max_trade_info = f"{max_trade['price']:,}만원"
                        print(f"최고거래가 정보: {max_trade_info}")
                        
                        wb = load_workbook(new_path)
                        ws = wb.active
                        ws['M1'] = max_trade_info
                        wb.save(new_path)
                        print("최고거래가 정보 저장 완료")
                    except Exception as e:
                        print(f"최고거래가 정보 저장 중 오류: {str(e)}")
                else:
                    print("실거래 데이터 없음")
                
                # 히스토리 리스트 갱신
                self.history_list = self.load_history()
                print("히스토리 리스트 갱신 완료")
                
                # GUI 업데이트 강제 실행
                self.update_history_display()
                self.root.update_idletasks()
                print("GUI 업데이트 완료")
                
            except Exception as e:
                print(f"히스토리 저장 중 오류: {str(e)}")
                
    def check_chrome_installation(self):
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        ]
        if not any(os.path.exists(path) for path in chrome_paths):
            messagebox.showerror("오류", "Chrome 브라우저가 설치되어 있지 않습니다.\nChrome을 먼저 설치해주세요.")
            sys.exit(1)

    
    
    def find_exact_apartment(self, driver, search_input):
        try:
            self.update_progress(40, "아파트 검색 중...")
            
            # 검색어 입력 및 엔터
            search_input_element = WebDriverWait(driver, 10).until(
               EC.presence_of_element_located((By.CSS_SELECTOR, "input.form-control"))
            )
            search_input_element.clear() 
            search_input_element.send_keys(search_input)
            time.sleep(1)
            search_input_element.send_keys(Keys.RETURN)
            time.sleep(2)
            
            # 먼저 검색 결과가 있는지 확인
            try:
               WebDriverWait(driver, 3).until(
                   EC.any_of(
                       EC.presence_of_element_located((By.CLASS_NAME, "widthTypeSelect")),
                       EC.presence_of_element_located((By.CLASS_NAME, "item-search-poi"))
                   )
               )
            except:
               # 검색 실패 시 처리
               messagebox.showerror("오류", "KB부동산에서 단지명을 확인해 주세요.")
               self.search_button.config(state="normal")  # 검색 버튼 활성화
               self.update_progress(0, "")  # 프로그레스 바 초기화
               if hasattr(self, 'driver'):
                   self.driver.quit()  # 브라우저 종료
               return False
    
            # 검색 결과 확인
            try:
                # 단일 단지로 바로 연결되는 경우
                WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "widthTypeSelect"))
                )
                return True
                
            except:
                # 여러 검색 결과가 있는 경우 기존 팝업 처리
                popup = tk.Toplevel(self.root)
                popup.title("검색 결과")
                popup.attributes('-topmost', True)
                
                width = 500
                height = 300
                screen_width = popup.winfo_screenwidth()
                screen_height = popup.winfo_screenheight()
                x = (screen_width - width) // 2
                y = (screen_height - height) // 2
                popup.geometry(f"{width}x{height}+{x}+{y}")
                
                # 커스텀 리스트박스 프레임 생성
                frame = ttk.Frame(popup)
                frame.pack(fill='both', expand=True, padx=10, pady=5)
                
                scrollbar = ttk.Scrollbar(frame)
                scrollbar.pack(side='right', fill='y')
                
                # 리스트박스 생성 (기본 옵션만 사용)
                listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, 
                                    font=self.font_normal if hasattr(self, 'font_normal') else ('KoPubWorld Dotum Medium', 10), 
                                    height=10,
                                    selectmode='single')
                listbox.pack(side='left', fill='both', expand=True)
                scrollbar.config(command=listbox.yview)
                
                apt_items = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.item-search-poi"))
                )
                
                items = []
                
                # 검색 결과가 하나만 있는 경우
                if len(apt_items) == 1:
                    popup.destroy()
                    search_poi = apt_items[0].find_element(By.CSS_SELECTOR, "span.search-poi")
                    driver.execute_script("arguments[0].click();", search_poi)
                    time.sleep(2)
                    return True
                
                # 검색 결과가 여러 개인 경우 - 간격을 위해 빈 줄 추가
                for i, apt_item in enumerate(apt_items):
                    try:
                        name = apt_item.find_element(By.CSS_SELECTOR, "span.text").text.strip()
                        location = apt_item.find_element(By.CSS_SELECTOR, "span.date").text.strip()
                        type_element = apt_item.find_element(By.CSS_SELECTOR, "span.ico-poi").text.strip()
                        
                        # 첫 번째 항목이 아닌 경우 빈 줄 추가
                        if i > 0:
                            listbox.insert(tk.END, "")  # 빈 줄로 간격 생성
                        
                        display_text = f"{type_element} | {name} | {location}"
                        listbox.insert(tk.END, display_text)
                        
                        search_poi = apt_item.find_element(By.CSS_SELECTOR, "span.search-poi")
                        items.append(search_poi)
                        
                    except Exception as e:
                        print(f"항목 처리 중 오류: {str(e)}")
                        continue
                
                def on_select(event=None):
                    if listbox.curselection():
                        idx = listbox.curselection()[0]
                        # 빈 줄을 고려한 실제 인덱스 계산
                        actual_idx = idx // 2  # 빈 줄이 있으므로 2로 나눔
                        try:
                            if actual_idx < len(items):
                                search_poi = items[actual_idx]
                                driver.execute_script("arguments[0].click();", search_poi)
                                time.sleep(2)
                                popup.destroy()
                        except Exception as e:
                            print(f"클릭 실패: {str(e)}")
                
                listbox.bind('<Double-Button-1>', on_select)
                ttk.Button(popup, text="선택", command=on_select).pack(pady=10)
                
                self.root.wait_window(popup)
                return True
                
        except Exception as e:
            messagebox.showerror("오류", f"아파트 검색 중 오류 발생: {str(e)}")
            self.search_button.config(state="normal")
            self.update_progress(0, "")
            if hasattr(self, 'driver'):
                self.driver.quit()
            return False

            
    
    def check_area_dropdown_exists(self, driver):
        try:
            dropdown = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CLASS_NAME, "widthTypeSelect"))
            )
            return True
        except:
            return False
    
 

    def select_area(self, driver, target_area):
        try:
            self.update_progress(60, "전용면적 선택 중...")
            
            time.sleep(3)
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
    
            # 전체 세대수 정보 찾기
            try:
                print("\n=== 전체 세대수 찾기 시작 ===")
                
                # 여러 선택자 시도
                selectors = [
                    "span[data-v-5ee451de][data-v-22e380f2]",
                    "span:contains('세대')",
                    ".household-info span",
                    ".complex-info span:contains('세대')"
                ]
                
                households_element = None
                for selector in selectors:
                    try:
                        if ':contains' in selector:
                            # jQuery 스타일 선택자는 XPath로 변환
                            elements = driver.find_elements(By.XPATH, "//span[contains(text(), '세대')]")
                            if elements:
                                households_element = elements[0]
                                print(f"XPath로 세대수 요소 찾기 성공")
                                break
                        else:
                            households_element = driver.find_element(By.CSS_SELECTOR, selector)
                            if households_element:
                                print(f"선택자 {selector}로 세대수 요소 찾기 성공")
                                break
                    except:
                        continue
                
                if households_element:
                    households_text = households_element.text
                    print(f"찾은 텍스트: '{households_text}'")
                    
                    # "1,148세대" 또는 "1148세대" 형식에서 숫자 추출
                    import re
                    match = re.search(r'([\d,]+)\s*세대', households_text)
                    if match:
                        self.total_households = int(match.group(1).replace(',', ''))
                        print(f"★ 웹에서 추출한 전체 세대수: {self.total_households}")
                    else:
                        print(f"세대수 패턴 매칭 실패: '{households_text}'")
                        self.total_households = None
                else:
                    print("웹에서 세대수 요소를 찾을 수 없음")
                    self.total_households = None
                    
            except Exception as e:
                print(f"웹에서 전체 세대수 찾기 오류: {str(e)}")
                self.total_households = None
            
            # 준공년 찾기 (기존 코드)
            try:
                print("\n=== 준공년 찾기 시작 ===")
                selector = "div.summaryType span:last-child"
                print(f"선택자 사용: {selector}")
                
                completion_element = WebDriverWait(driver, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                
                if completion_element:
                    completion_text = completion_element.text
                    print(f"찾은 텍스트: {completion_text}")
                    
                    if '년차' in completion_text:
                        year_parts = completion_text.split('(')[0].strip().split('.')
                        if len(year_parts[0]) == 2:
                            year_num = int(year_parts[0])
                            if year_num > 50:
                                self.completion_year = '19' + str(year_num).zfill(2)
                            else:
                                self.completion_year = '20' + str(year_num).zfill(2)
                            print(f"설정된 준공년: {self.completion_year}")
                else:
                    print("준공년 요소를 찾을 수 없음")
                    self.completion_year = None
            
            except Exception as e:
                print(f"준공년 찾기 오류: {str(e)}")
                self.completion_year = None
                logging.warning("준공년 정보를 찾을 수 없습니다.")
            
            # 면적 선택 처리
            return self.select_area_size(driver, target_area)
        
        except Exception as e:
            messagebox.showerror("오류", f"전용면적 선택 중 오류 발생: {str(e)}")
            return False
    
    def select_area_size(self, driver, target_area):
        try:
            print(f"\n=== 전용면적 선택 시작 ===")
            print(f"목표 전용면적: {target_area}")
            
            # 드롭다운 버튼 찾기 - 여러 선택자 시도
            dropdown_selectors = [
                "div.widthTypeSelect",
                ".widthTypeSelect",
                "button.widthTypeSelect"
            ]
            
            dropdown_button = None
            for selector in dropdown_selectors:
                try:
                    dropdown_button = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    print(f"드롭다운 버튼 찾기 성공: {selector}")
                    break
                except:
                    print(f"드롭다운 선택자 {selector} 실패")
                    continue
            
            if not dropdown_button:
                print("드롭다운 버튼을 찾을 수 없습니다.")
                raise Exception("면적 선택 드롭다운을 찾을 수 없습니다.")
            
            # 드롭다운 클릭
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_button)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", dropdown_button)
            print("드롭다운 클릭 완료")
            time.sleep(2)
            
            # 면적 목록 가져오기 - 새로운 구조에 맞춰 수정
            area_rows = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.tbody-tr"))
            )
            print(f"찾은 면적 행 수: {len(area_rows)}")
            
            target_area = target_area.replace("m²", "").replace("㎡", "").strip()
            target_area_int = target_area.split('.')[0]
            print(f"목표 전용면적 (정수): {target_area_int}")
            
            # 일치하는 전용면적을 가진 모든 행 찾기
            matching_rows = []
            
            for idx, row in enumerate(area_rows):
                try:
                    # 전용면적 정보 추출 - 새로운 구조에 맞춰 수정
                    # span.tdbox 내의 전용면적 정보 찾기
                    area_elements = row.find_elements(By.CSS_SELECTOR, "span.tdbox span.tbarea-point em")
                    
                    for elem in area_elements:
                        text = elem.text
                        if '전용' in text:
                            # "전용 59.96m²" 형식에서 숫자 추출
                            import re
                            match = re.search(r'전용\s*([\d.]+)', text)
                            if match:
                                actual_area = match.group(1)
                                actual_area_int = actual_area.split('.')[0]
                                print(f"행 {idx}: 전용면적 = {actual_area}m² (정수: {actual_area_int})")
                                
                                if actual_area_int == target_area_int:
                                    print(f"★ 일치하는 면적 발견!")
                                    
                                    # 매매가 추출 - 두 번째 tdbox의 첫 번째 intd 내의 tdblue
                                    try:
                                        tdbox_elements = row.find_elements(By.CSS_SELECTOR, "span.tdbox")
                                        if len(tdbox_elements) >= 2:
                                            price_box = tdbox_elements[1]
                                            price_elements = price_box.find_elements(By.CSS_SELECTOR, "span.intd em.tdblue")
                                            if price_elements:
                                                price_text = price_elements[0].text.strip()
                                                print(f"  매매가: {price_text}")
                                                
                                                # 가격 값 추출
                                                price_value = self.extract_price_value(price_text)
                                                
                                                matching_rows.append({
                                                    'row': row,
                                                    'price': price_value,
                                                    'price_text': price_text,
                                                    'area': actual_area
                                                })
                                    except Exception as e:
                                        print(f"  매매가 추출 오류: {str(e)}")
                                        matching_rows.append({
                                            'row': row,
                                            'price': 0,
                                            'price_text': "추출 오류",
                                            'area': actual_area
                                        })
                                break
                    
                except Exception as e:
                    print(f"행 {idx} 처리 중 오류: {str(e)}")
                    continue
            
            if not matching_rows:
                print(f"오류: 면적 {target_area}㎡를 찾을 수 없습니다.")
                # 사용 가능한 면적 목록 출력
                print("\n사용 가능한 전용면적 목록:")
                for idx, row in enumerate(area_rows[:5]):  # 처음 5개만 표시
                    try:
                        area_elements = row.find_elements(By.CSS_SELECTOR, "span.tdbox span.tbarea-point em")
                        for elem in area_elements:
                            if '전용' in elem.text:
                                print(f"  - {elem.text}")
                                break
                    except:
                        continue
                raise ValueError(f"면적 {target_area}㎡를 찾을 수 없습니다.")
            
            # 매매가가 가장 높은 행 선택
            highest_price_row = max(matching_rows, key=lambda x: x['price'])
            print(f"\n선택된 행: 전용면적 {highest_price_row['area']}m², 매매가 {highest_price_row['price_text']}")
            
            # 선택된 행 클릭
            selected_row = highest_price_row['row']
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", selected_row)
            time.sleep(0.5)
            
            # 클릭 시도 - 여러 방법 시도
            try:
                # 1. JavaScript 클릭
                driver.execute_script("arguments[0].click();", selected_row)
                print("JavaScript 클릭 성공")
            except:
                try:
                    # 2. 일반 클릭
                    selected_row.click()
                    print("일반 클릭 성공")
                except:
                    # 3. Actions 클릭
                    from selenium.webdriver.common.action_chains import ActionChains
                    actions = ActionChains(driver)
                    actions.move_to_element(selected_row).click().perform()
                    print("Actions 클릭 성공")
            
            time.sleep(1)
            print("=== 전용면적 선택 완료 ===\n")
            return True
                
        except Exception as e:
            print(f"면적 선택 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
            
            target_area = target_area.replace("m²", "").strip()
            target_area_int = target_area.split('.')[0]
            
            # 일치하는 전용면적을 가진 모든 행 찾기
            matching_rows = []
            
            for row in area_rows:
                try:
                    # 전용면적 추출
                    tdbold_element = row.find_element(By.CLASS_NAME, "tdbold")
                    area_text = tdbold_element.text
                    
                    if '/' in area_text:
                        actual_area = area_text.split('/')[1].strip().replace("m²", "").strip()
                        actual_area_int = actual_area.split('.')[0]
                        
                        if actual_area_int == target_area_int:
                            # 매매가 추출
                            try:
                                # 두 번째 tdbox 내의 첫 번째 intd 내의 tdblue 클래스 요소 찾기
                                price_element = row.find_elements(By.CLASS_NAME, "tdbox")[1]
                                price_intd = price_element.find_elements(By.CLASS_NAME, "intd")[0]
                                price_text = price_intd.find_element(By.CLASS_NAME, "tdblue").text.strip()
                                
                                # 매매가 텍스트를 숫자로 변환
                                price_value = self.extract_price_value(price_text)
                                
                                # 행, 가격 값, 원 텍스트 저장 (디버깅용)
                                matching_rows.append({
                                    'row': row,
                                    'price': price_value,
                                    'price_text': price_text
                                })
                                logging.info(f"일치하는 면적 발견: {area_text}, 매매가: {price_text}")
                            except Exception as e:
                                logging.warning(f"매매가 추출 중 오류: {str(e)}")
                                # 오류가 발생해도 행은 후보에 포함 (기본 가격 0)
                                matching_rows.append({
                                    'row': row,
                                    'price': 0,
                                    'price_text': "추출 오류"
                                })
                except Exception as e:
                    logging.warning(f"면적 행 처리 중 오류: {str(e)}")
                    continue
            
            if not matching_rows:
                raise ValueError(f"면적 {target_area}㎡를 찾을 수 없습니다.")
            
            # 매매가가 가장 높은 행 선택
            highest_price_row = max(matching_rows, key=lambda x: x['price'])
            logging.info(f"선택된 행: 매매가 {highest_price_row['price_text']}")
            
            # 선택된 행 클릭
            selected_row = highest_price_row['row']
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", selected_row)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", selected_row)
            time.sleep(1)
            
            return True
                
        except Exception as e:
            logging.error(f"면적 선택 중 오류: {str(e)}")
            return False
    
    def extract_price_value(self, price_text):
        """가격 텍스트(예: '25억', '21억 5,000만')에서 만원 단위 숫자로 변환"""
        try:
            # 가격 문자열 정제
            price_text = price_text.replace(',', '').strip()
            total_price = 0
            
            # 억 단위 처리
            if '억' in price_text:
                parts = price_text.split('억')
                # 앞부분(억 단위) - 예: "25억"에서 "25"
                if parts[0].strip():
                    total_price += int(float(parts[0].strip())) * 10000
                
                # 뒷부분(만원 단위) - 예: "5,000만"에서 "5000"
                if len(parts) > 1 and parts[1].strip() and '만' in parts[1]:
                    man_part = parts[1].split('만')[0].strip()
                    if man_part:
                        total_price += int(float(man_part))
            
            # 억 단위가 없는 경우
            elif '만' in price_text:
                man_part = price_text.split('만')[0].strip()
                if man_part:
                    total_price = int(float(man_part))
            
            return total_price
        except Exception as e:
            logging.warning(f"가격 변환 중 오류: {str(e)}. 원본 텍스트: {price_text}")
            return 0
    
    
    def download_files(self, driver):
        try:
            self.update_progress(80, "파일 다운로드 중...")
            logging.info("=== 파일 다운로드 시작 ===")
            
            # 다운로드 전 폴더 내 파일 목록 확인
            before_files = set(os.listdir(self.download_path))
            
            # KB시세 파일 다운로드 - 정확한 선택자 사용
            logging.info("KB시세 파일 다운로드 시도...")
            
            kb_button_selectors = [
                "#시세 > div:nth-child(2) > div.f-row2-gap8 > button:nth-child(2)",  # 정확한 선택자
                "div.f-row2-gap8 > button:nth-child(2)",  # 단순화된 버전
                "button.btn-land-sqlinebx",  # 클래스명
                "//button[contains(text(), 'KB시세 다운로드')]"  # XPath
            ]
            
            kb_download_button = None
            for selector in kb_button_selectors:
                try:
                    if selector.startswith("//"):
                        # XPath
                        kb_download_button = WebDriverWait(driver, 3).until(
                            EC.element_to_be_clickable((By.XPATH, selector))
                        )
                        logging.info(f"KB시세 버튼 찾기 성공 (XPath): {selector}")
                    else:
                        # CSS Selector
                        kb_download_button = WebDriverWait(driver, 3).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                        )
                        logging.info(f"KB시세 버튼 찾기 성공: {selector}")
                    break
                except Exception as e:
                    logging.info(f"선택자 {selector} 시도 실패: {str(e)}")
                    continue
            
            if not kb_download_button:
                # 버튼을 못 찾았을 때 페이지의 모든 버튼 확인 (디버깅용)
                all_buttons = driver.find_elements(By.TAG_NAME, "button")
                logging.info(f"페이지의 전체 버튼 수: {len(all_buttons)}")
                for idx, btn in enumerate(all_buttons[:10]):  # 처음 10개만
                    try:
                        btn_text = btn.text
                        btn_class = btn.get_attribute("class")
                        logging.info(f"버튼 {idx}: 텍스트='{btn_text}', 클래스='{btn_class}'")
                    except:
                        pass
                
                raise Exception("KB시세 다운로드 버튼을 찾을 수 없습니다.")
            
            # 버튼이 화면에 보이도록 스크롤
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", kb_download_button)
            time.sleep(1)
            
            # 클릭 전 버튼 정보 확인
            btn_text = kb_download_button.text
            logging.info(f"클릭할 버튼 텍스트: '{btn_text}'")
            
            # JavaScript 클릭 (가장 안정적)
            driver.execute_script("arguments[0].click();", kb_download_button)
            logging.info("KB시세 버튼 클릭 성공")
            time.sleep(3)
            
            # 실거래가 파일 다운로드 (기존 코드 유지)
            logging.info("실거래가 파일 다운로드 시도...")
            history_download_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn-rounded"))
            )
            history_download_button.click()
            time.sleep(3)
            
            # 파일 다운로드 완료 대기
            max_wait = 30
            while max_wait > 0:
                current_files = set(os.listdir(self.download_path))
                new_files = current_files - before_files
                if len(new_files) >= 1:  # 최소 1개 파일이 다운로드되었는지 확인
                    # 새로 다운로드된 파일들 중 엑셀 파일만 처리
                    excel_files = [f for f in new_files if f.endswith('.xlsx')]
                    for excel_file in excel_files:
                        # 파일 처리 로직 (기존 코드)
                        old_path = os.path.join(self.download_path, excel_file)
                        
                        try:
                            # 엑셀 파일에서 정보 읽기
                            df = pd.read_excel(old_path, nrows=8)
                            apt_name = str(df.iloc[2, 1])
                            area_text = str(df.iloc[5, 1])
                            
                            # 전용면적 추출
                            if '/' in area_text:
                                area = area_text.split('/')[1].strip()
                                area = area.replace('m²', '').strip()
                            else:
                                area = area_text.replace('m²', '').strip()
                            
                            area_clean = str(int(float(''.join(c for c in area if c.isdigit() or c == '.'))))
                            apt_name_clean = ''.join(char for char in apt_name if char.isalnum() or char.isspace())
                            apt_name_clean = apt_name_clean.replace(' ', '_')
                            
                            # 새 파일명 생성
                            new_filename = f"{apt_name_clean}_{area_clean}m2.xlsx"
                            new_path = os.path.join(self.download_path, new_filename)
                            
                            # 파일 이름 변경
                            if os.path.exists(new_path):
                                os.remove(new_path)
                            os.rename(old_path, new_path)
                            logging.info(f"파일 이름 변경: {excel_file} -> {new_filename}")
                            
                        except Exception as e:
                            logging.error(f"파일 처리 중 오류: {str(e)}")
                            continue
                    
                    # Chrome 브라우저 닫기
                    logging.info("다운로드 완료 - Chrome 브라우저 종료")
                    try:
                        driver.quit()
                        logging.info("Chrome 브라우저 정상 종료됨")
                    except Exception as e:
                        logging.warning(f"Chrome 종료 중 경고: {str(e)}")
                    
                    return True
                
                time.sleep(1)
                max_wait -= 1
                
            if max_wait == 0:
                logging.error("파일 다운로드 시간 초과")
                # 타임아웃 시에도 브라우저 닫기
                try:
                    driver.quit()
                    logging.info("Chrome 브라우저 종료됨 (타임아웃)")
                except:
                    pass
                return False
                
        except Exception as e:
            logging.error(f"다운로드 중 오류 발생: {str(e)}")
            # 오류 발생 시에도 브라우저 닫기
            try:
                driver.quit()
                logging.info("Chrome 브라우저 종료됨 (오류)")
            except:
                pass
            messagebox.showerror("오류", f"파일 다운로드 중 오류 발생: {str(e)}")
            return False




    def format_naver_excel(self, excel_path):
        """네이버 매물 엑셀 파일의 형식을 개선합니다"""
        try:
            print(f"\n=== 네이버 매물 엑셀 파일 형식 개선 시작 ===")
            print(f"파일 경로: {excel_path}")
            
            # Excel 애플리케이션 직접 사용 (win32com 활용)
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False  # 이 부분에서 오류 발생
                excel.DisplayAlerts = False
            except AttributeError:
                # COM 오류 발생 시 pandas로 대체 처리
                print("Excel COM 인터페이스 오류 - pandas로 대체 처리")
                return self.format_naver_excel_pandas(excel_path)
            
            try:
                # 나머지 기존 코드...
                wb = excel.Workbooks.Open(excel_path)
                ws = wb.Worksheets(1)
                
                # 사용된 범위 확인
                used_range = ws.UsedRange
                last_row = used_range.Rows.Count
                last_col = used_range.Columns.Count
                print(f"데이터 범위: {last_row}행 x {last_col}열")
                
                # 필터 적용 (A1부터 마지막 열까지)
                last_col_letter = chr(64 + last_col) if last_col <= 26 else chr(64 + last_col // 26) + chr(64 + last_col % 26)
                filter_range = f"A1:{last_col_letter}{last_row}"
                print(f"필터 적용 범위: {filter_range}")
                ws.Range(filter_range).AutoFilter()
                
                # 헤더 서식 설정
                header_range = ws.Range(f"A1:{last_col_letter}1")
                header_range.Font.Bold = True
                header_range.Font.Size = 11
                header_range.Interior.Color = 0xF2E1D9  # 연한 파란색 (RGB 순서 반대)
                
                # 가격 관련 열 식별 및 서식 적용
                price_columns = ['매매가', '최저가매물', '보증금', '월세', 
                                '최소매매가', '최대매매가', '최소보증금', '최대보증금',
                                '최소월세', '최대월세', '최소프리미엄', '최대프리미엄']
                
                # 각 헤더를 순회하며 가격 관련 열 찾기
                for col in range(1, last_col + 1):
                    header_value = ws.Cells(1, col).Value
                    print(f"열 {col}: {header_value}")
                    
                    if header_value in price_columns:
                        print(f"가격 열 발견: {col} ({header_value})")
                        # 열 전체에 천 단위 구분 서식 적용
                        column_range = ws.Range(ws.Cells(2, col), ws.Cells(last_row, col))
                        column_range.NumberFormat = "#,##0"
                
                # 열 너비 자동 조정
                ws.Columns.AutoFit()
                
                # 저장 및 종료
                wb.Save()
                print("엑셀 파일 형식 변경 완료")
                return True
                
          
            finally:
                # 항상 엑셀 종료
                try:
                    wb.Close(SaveChanges=True)
                    excel.Quit()
                except:
                    pass
                    
        except Exception as e:
            print(f"엑셀 파일 형식 개선 중 오류 발생: {str(e)}")
            return False
    
    def format_naver_excel_pandas(self, excel_path):
        """pandas를 사용한 대체 방법"""
        try:
            print("pandas를 사용하여 엑셀 파일 처리")
            # 기본적인 데이터 읽기만 수행
            df = pd.read_excel(excel_path)
            print(f"엑셀 파일 읽기 완료: {len(df)}행")
            return True
        except Exception as e:
            print(f"pandas 처리 중 오류: {str(e)}")
            return False
 
    def start_search(self):
        try:
            if hasattr(self, 'search_address'):
                apt_name = self.search_address
                area = self.area.get().strip()
                delattr(self, 'search_address')
            else:
                apt_name = self.apt_name.get().strip()
                area = self.area.get().strip()
                
                # 신축단지 정보 가져오기
                self.new_apt_info = {
                    'name': self.new_apt_name.get().strip(),
                    'price': self.new_apt_price.get().strip()
                }
                
                if not apt_name or not area:
                    raise ValueError("아파트 이름과 전용면적을 모두 입력해주세요.")
                if not area.replace('.','').isdigit():
                    raise ValueError("전용면적은 숫자만 입력해주세요.")
                # 분양가가 입력된 경우 숫자 확인
                if self.new_apt_info['price'] and not self.new_apt_info['price'].isdigit():
                    raise ValueError("분양가는 숫자만 입력해주세요.")
    
            # 버튼 비활성화 추가
            self.search_button.config(state="disabled")
            if hasattr(self, 'multi_analysis_button'):
                self.multi_analysis_button.config(state="disabled")
                
            self.browser_search(apt_name)
            
            try:
                df = self.analyze_prices()
                if os.path.exists(self.image_path):
                    os.startfile(self.image_path)
                self.history_list = self.load_history()
                self.update_history_display()
                self.update_progress(100, "분석 완료!")                                                  
            except Exception as e:
                messagebox.showerror("오류", f"분석 중 오류 발생: {str(e)}")
        except ValueError as ve:
            messagebox.showwarning("입력 오류", str(ve))
        except Exception as e:
            messagebox.showerror("오류", f"예상치 못한 오류가 발생했습니다: {str(e)}")
        finally:
            self.search_button.config(state="normal")
            if hasattr(self, 'multi_analysis_button'):
                self.multi_analysis_button.config(state="normal")

    def get_naver_min_jeonse_price(self, apt_name=None, target_area=None):
        """네이버 부동산 매물 중 전세 최저가 정보를 가져오는 함수"""
        try:
            print("\n=== 네이버 전세매물 정보 가져오기 시작 ===")
            
            # 1. 단지명 결정 (이전 코드와 동일)
            if apt_name is None:
                apt_name = self.apt_name.get().strip()
                print(f"GUI에서 단지명 가져오기: {apt_name}")
                
                # KB부동산 엑셀에서 단지명 가져오기 (GUI 입력값이 없는 경우)
                if not apt_name and hasattr(self, 'latest_excel_file'):
                    try:
                        df_info = pd.read_excel(self.latest_excel_file, nrows=8)
                        apt_name = str(df_info.iloc[3, 1])
                        print(f"KB부동산 엑셀에서 단지명 추출: {apt_name}")
                    except Exception as e:
                        print(f"KB부동산 엑셀에서 단지명 추출 실패: {str(e)}")
                        return None
            
            # 전용면적 확인 및 변환 (이전 코드와 동일)
            if target_area is None:
                target_area = self.area.get().strip()
                print(f"GUI에서 전용면적 가져오기: {target_area}")
                
                if not target_area:
                    print("전용면적 정보가 없습니다.")
                    return None
            
            try:
                target_area = float(target_area.replace('㎡', '').replace('m²', ''))
                print(f"변환된 전용면적: {target_area}㎡")
            except ValueError:
                print(f"전용면적 형식 오류: {target_area}")
                return None
            
            # 네이버 매물 데이터 가져오기 (이전 코드와 동일)
            today = datetime.now().strftime('%Y%m%d')
            excel_filename = os.path.join(self.download_path, f'{apt_name}_네이버매물_{today}.xlsx')
            
            if not os.path.exists(excel_filename):
                complex_id = self.search_naver_complex(apt_name)
                if not complex_id:
                    print("네이버 부동산에서 단지를 찾을 수 없습니다.")
                    return None
                
                excel_path = self.download_naver_data(complex_id, apt_name)
                if not excel_path or not os.path.exists(excel_path):
                    print("네이버 매물 데이터를 다운로드할 수 없습니다.")
                    return None
            else:
                excel_path = excel_filename
            
            # 엑셀 파일에서 데이터 읽기
            print(f"\n=== 네이버 매물 데이터 파일 분석 시작 ===")
            print(f"파일 경로: {excel_path}")
            df = pd.read_excel(excel_path)
            
            # 열 인덱스와 열 이름 출력 (디버깅용)
            print("\n=== 엑셀 열 정보 ===")
            for i, col_name in enumerate(df.columns):
                col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
                print(f"{col_letter}열({i+1}): {col_name}")
            
            # 전세 매물만 필터링 - query() 메서드 사용
            # 0원 초과 조건 추가 (보증금 > 0)
            df = df.query('보증금 > 0 and (월세 == 0 or 월세 == "" or 월세.isnull())').copy()
            if df.empty:
                print(f"전세 매물이 없습니다.")
                return None
            
            # 전용면적으로 필터링 - query() 사용
            area_tolerance = 1.0  # 면적 허용 오차
            area_filter = f"전용면적 >= {target_area - area_tolerance} and 전용면적 <= {target_area + area_tolerance}"
            area_filtered = df.query(area_filter).copy()
            print(f"면적 필터링 결과: {len(area_filtered)}개 매물")
            
            if area_filtered.empty:
                print(f"{target_area}㎡ 면적의 전세 매물이 없습니다.")
                return None
            
            # 층수 정보 처리 (이전 코드와 동일)
            def is_high_floor(floor_info):
                if pd.isna(floor_info):
                    return False
                
                floor_str = str(floor_info).lower()
                
                # '중' 또는 '고'를 포함하면 즉시 True 반환
                if '중' in floor_str or '고' in floor_str:
                    return True
                
                # '저' 포함 시 즉시 False 반환
                if '저' in floor_str:
                    return False
                
                # 숫자로 된 층수 추출 시도
                try:
                    # 'n층/m층' 형식에서 n 추출
                    if '/' in floor_str:
                        first_part = floor_str.split('/')[0].strip()
                        # 숫자인지 확인
                        if first_part.isdigit():
                            floor_num = int(first_part)
                            return floor_num >= 5
                    # 단일 숫자인 경우
                    elif floor_str.isdigit():
                        return int(floor_str) >= 5
                except Exception:
                    pass
                
                return False
            
            # 각 매물의 층수와 고층 여부 출력 (디버깅)
            print("\n=== 고층 전세매물 필터링 시작 ===")
            area_filtered.loc[:, 'isHighFloor'] = area_filtered['층/전체층'].apply(is_high_floor)
            
            for i, row in area_filtered.head(10).iterrows():  # 첫 10개만 출력
                print(f"전세매물 층수: {row['층/전체층']}, 고층여부: {row['isHighFloor']}")
            
            # 고층 매물만 필터링 - query() 사용
            high_floor_df = area_filtered.query('isHighFloor == True').copy()
            print(f"고층 전세매물 필터링 결과: {len(high_floor_df)}개 매물")
            
            if high_floor_df.empty:
                print(f"{target_area}㎡ 면적의 5층 이상/중층/고층 전세 매물이 없습니다.")
                return None
            
            # 보증금과 최소보증금 숫자로 변환하는 함수
            def safe_convert_to_number(value):
                if pd.isna(value) or value == "":
                    return float('inf')  # 값이 없으면 무한대로 설정하여 최소값에서 제외
                
                # 문자열로 변환 후 처리
                value_str = str(value)
                # 쉼표 제거 및 '만원' 같은 단위 제거
                clean_value = value_str.replace(',', '').replace('만원', '').strip()
                
                try:
                    number_value = float(clean_value)
                    # 0원 값은 무한대로 처리하여 최소값에서 제외
                    return float('inf') if number_value == 0 else number_value
                except ValueError:
                    return float('inf')  # 변환 실패 시 무한대
            
            # 보증금과 최소보증금 컬럼 숫자로 변환
            high_floor_df.loc[:, '보증금_num'] = high_floor_df['보증금'].apply(safe_convert_to_number)
            
            # 최소보증금 컬럼이 있는지 확인하고 처리
            if '최소보증금' in high_floor_df.columns:
                high_floor_df.loc[:, '최소보증금_num'] = high_floor_df['최소보증금'].apply(safe_convert_to_number)
                
                # 최저가 전세 컬럼 생성: 보증금과 최소보증금 중 작은 값 선택 (0원 제외)
                high_floor_df.loc[:, '최저가전세'] = high_floor_df.apply(
                    lambda row: min(row['보증금_num'], row['최소보증금_num']), 
                    axis=1
                )
                
                # 변환 결과 디버깅
                print("\n=== 보증금 변환 결과 (최초 10개) ===")
                for i, row in high_floor_df.head(10).iterrows():
                    print(f"매물 {i}: 보증금={row['보증금']}→{row['보증금_num']:.0f}, "
                          f"최소보증금={row['최소보증금']}→{row['최소보증금_num']:.0f}, "
                          f"최저가전세={row['최저가전세']:.0f}만원")
            else:
                # 최소보증금 컬럼이 없으면 보증금을 그대로 최저가전세로 사용
                high_floor_df.loc[:, '최저가전세'] = high_floor_df['보증금_num']
                
                # 변환 결과 디버깅
                print("\n=== 보증금 변환 결과 (최초 10개) ===")
                for i, row in high_floor_df.head(10).iterrows():
                    print(f"매물 {i}: 보증금={row['보증금']}→{row['보증금_num']:.0f}, "
                          f"최저가전세={row['최저가전세']:.0f}만원")
            
            # 0원 값이 있는지 체크
            zero_values = high_floor_df[high_floor_df['최저가전세'] == 0]
            if not zero_values.empty:
                print(f"경고: {len(zero_values)}개의 0원 매물이 있습니다. 이 매물들은 제외됩니다.")
                high_floor_df = high_floor_df.query('최저가전세 > 0').copy()
                
            if high_floor_df.empty:
                print(f"유효한 가격(0원 초과)의 5층 이상/중층/고층 전세 매물이 없습니다.")
                return None
            
            # 최저가 매물 선택 부분까지 동일...
            min_price_row = high_floor_df.sort_values('최저가전세').iloc[0]
            min_price = min_price_row['최저가전세']
            floor_info = min_price_row['층/전체층']
            
            # 해당 평형 전체 매물 건수 계산
            property_count = len(area_filtered)
            
            # 해당 매물의 동 정보 추출
            dong_info = min_price_row.get('동', '정보없음')
            
            # 가격을 10000으로 나누어 만원 단위로 변환
            converted_price = min_price / 10000
            
            print(f"\n=== 네이버 전세 최저가 추출 완료 ===")
            print(f"전세 최저가(원본): {min_price}")
            print(f"전세 최저가(변환): {converted_price:.0f}만원 ({floor_info})")
            print(f"매물 건수: 총 {property_count}건")
            print(f"동 정보: {dong_info}")
            
            return {
                'price': converted_price,  # 만원 단위로 변환
                'price_str': f"{converted_price:.0f}만원",
                'floor': floor_info,
                'area': min_price_row['전용면적'],
                'property_count': property_count,
                'dong_info': dong_info
            }
        
        except Exception as e:
            print(f"네이버 전세 최저가 조회 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def get_naver_min_price(self, apt_name=None, target_area=None):
        """네이버 부동산 매물 중 최저가 정보를 가져오는 함수"""
        try:
            print("\n=== 네이버 매물 정보 가져오기 시작 ===")
            
            # 1. 단지명 결정 (GUI 입력값 또는 KB엑셀 파일에서 추출)
            if apt_name is None:
                apt_name = self.apt_name.get().strip()
                print(f"GUI에서 단지명 가져오기: {apt_name}")
                
                # KB부동산 엑셀에서 단지명 가져오기 (GUI 입력값이 없는 경우)
                if not apt_name and hasattr(self, 'latest_excel_file'):
                    try:
                        df_info = pd.read_excel(self.latest_excel_file, nrows=8)
                        apt_name = str(df_info.iloc[3, 1])  # B열 4행 (0-based index이므로 3행)
                        print(f"KB부동산 엑셀에서 단지명 추출: {apt_name}")
                    except Exception as e:
                        print(f"KB부동산 엑셀에서 단지명 추출 실패: {str(e)}")
                        return None
            
            # 전용면적 확인
            if target_area is None:
                target_area = self.area.get().strip()
                print(f"GUI에서 전용면적 가져오기: {target_area}")
                
                if not target_area:
                    print("전용면적 정보가 없습니다.")
                    return None
            
            # 전용면적을 숫자로 변환
            try:
                target_area = float(target_area.replace('㎡', '').replace('m²', ''))
                print(f"변환된 전용면적: {target_area}㎡")
            except ValueError:
                print(f"전용면적 형식 오류: {target_area}")
                return None
            
            print(f"네이버 부동산에서 '{apt_name}' 단지, {target_area}㎡ 매물 검색 시작")
            
            # 네이버 부동산 검색 및 매물 데이터 수집
            print("\n=== 네이버 부동산 단지 검색 시작 ===")
            complex_id = self.search_naver_complex(apt_name)
            if not complex_id:
                print("네이버 부동산에서 단지를 찾을 수 없습니다.")
                return None
            
            print(f"검색된 단지번호: {complex_id}")
            
            # 매물 데이터 다운로드
            print("\n=== 네이버 매물 데이터 다운로드 시작 ===")
            excel_path = self.download_naver_data(complex_id, apt_name)
            if not excel_path or not os.path.exists(excel_path):
                print("네이버 매물 데이터를 다운로드할 수 없습니다.")
                return None
            
            # 엑셀 파일에서 데이터 읽기
            print(f"\n=== 네이버 매물 데이터 파일 분석 시작 ===")
            print(f"파일 경로: {excel_path}")
            df = pd.read_excel(excel_path)
            print(f"데이터 크기: {len(df)}행 × {len(df.columns)}열")
            print(f"포함된 열: {', '.join(df.columns.tolist())}")
            
            # 중요: 매매가 > 0인 매물만 필터링 (전세, 월세 제외)
            # query() 메서드 사용
            df = df.query('매매가 > 0').copy()  # .copy()로 경고 방지
            if df.empty:
                print(f"매매가 > 0인 매물이 없습니다.")
                return None
            
            # 전용면적으로 필터링 - query() 사용
            area_tolerance = 1.0  # 면적 허용 오차
            area_filter = f"전용면적 >= {target_area - area_tolerance} and 전용면적 <= {target_area + area_tolerance}"
            print(f"면적 필터링 조건: {area_filter}")
            
            area_filtered = df.query(area_filter).copy()  # .copy()로 경고 방지
            print(f"면적 필터링 결과: {len(area_filtered)}개 매물")
            
            if area_filtered.empty:
                print(f"{target_area}㎡ 면적의 매물이 없습니다.")
                return None
            
            # 층수 정보 추출 및 필터링 - 5층 이상 또는 중층/고층
            def is_high_floor(floor_info):
                """층수 정보에서 고층 여부를 판단하는 함수"""
                if pd.isna(floor_info):
                    return False
                
                floor_str = str(floor_info).lower()
                
                # '중' 또는 '고'를 포함하면 즉시 True 반환
                if '중' in floor_str or '고' in floor_str:
                    return True
                
                # '저' 포함 시 즉시 False 반환
                if '저' in floor_str:
                    return False
                
                # 숫자로 된 층수 추출 시도
                try:
                    # 'n층/m층' 형식에서 n 추출
                    if '/' in floor_str:
                        first_part = floor_str.split('/')[0].strip()
                        # 숫자인지 확인
                        if first_part.isdigit():
                            floor_num = int(first_part)
                            return floor_num >= 5
                    # 단일 숫자인 경우
                    elif floor_str.isdigit():
                        return int(floor_str) >= 5
                except Exception:
                    # 오류 메시지 출력 없이 조용히 실패
                    pass
                
                return False  # 판단할 수 없는 경우 False 반환
            
            # isHighFloor 열 추가 - SettingWithCopyWarning 방지를 위해 .loc 사용
            print("\n=== 고층 매물 필터링 시작 ===")
            area_filtered.loc[:, 'isHighFloor'] = area_filtered['층/전체층'].apply(is_high_floor)
            
            # 각 매물의 층수와 고층 여부 출력 (디버깅)
            for i, row in area_filtered.head(10).iterrows():  # 첫 10개만 출력
                print(f"매물 층수: {row['층/전체층']}, 고층여부: {row['isHighFloor']}")           
    
            # 고층 매물만 필터링 - query() 사용
            high_floor_df = area_filtered.query('isHighFloor == True').copy()  # .copy()로 경고 방지
            print(f"고층 매물 필터링 결과: {len(high_floor_df)}개 매물")
            
            # 아래 부분이 수정된 부분: 고층 매물이 없는 경우 저층 매물로 대체
            if high_floor_df.empty:
                print(f"{target_area}㎡ 면적의 5층 이상 매물이 없습니다. 저층 매물을 사용합니다.")
                
                # 저층 매물 필터링 (isHighFloor == False)
                low_floor_df = area_filtered.query('isHighFloor == False').copy()
                print(f"저층 매물 필터링 결과: {len(low_floor_df)}개 매물")
                
                if low_floor_df.empty:
                    print(f"{target_area}㎡ 면적의 저층 매물도 없습니다.")
                    return None
                    
                # 저층 매물 중 최저가 찾기
                min_price_row = low_floor_df.sort_values('최저가매물').iloc[0]
                min_price = min_price_row['최저가매물']
                floor_info = min_price_row['층/전체층']
                
                # 저층 매물 여부 표시
                is_low_floor = True
            else:
                # 고층 매물 중 최저가 찾기 (기존 코드)
                min_price_row = high_floor_df.sort_values('최저가매물').iloc[0]
                min_price = min_price_row['최저가매물']
                floor_info = min_price_row['층/전체층']
                
                # 고층 매물 여부 표시
                is_low_floor = False
            
            # 가격을 10000으로 나누어 만원 단위로 변환
            converted_price = min_price / 10000
            
            # 해당 평형 전체 매물 건수 계산
            property_count = len(area_filtered)
            
            # 해당 매물의 동 정보 추출 (단일 동)
            dong_info = min_price_row.get('동', '정보없음')
                
            print(f"\n=== 네이버 매물 최저가 추출 완료 ===")
            print(f"최저가: {converted_price:.0f}만원 ({floor_info})")
            print(f"매물 건수: 총 {property_count}건")
            print(f"동 정보: {dong_info}")
            print(f"저층 매물 여부: {is_low_floor}")
                    
            return {
                'price': converted_price,
                'price_str': f"{converted_price:.0f}만원",
                'floor': floor_info,
                'area': min_price_row['전용면적'],
                'property_count': property_count,  # 총 매물 수
                'dong_info': dong_info,  # 해당 매물의 동 정보
                'is_low_floor': is_low_floor  # 저층 매물 여부 추가
            }
        
        except Exception as e:
            print(f"네이버 매물 최저가 조회 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
        
    def search_naver_complex(self, apt_name):
        """네이버 부동산 API로 단지 검색 및 단지번호 추출"""
        try:
            print(f"\n네이버 부동산에서 '{apt_name}' 단지 검색 중...")
            
            import urllib.parse
            encoded_name = urllib.parse.quote(apt_name)
            
            search_url = f"https://fin.land.naver.com/front-api/v1/search/autocomplete/complexes?keyword={encoded_name}&size=10&page=0"
            print(f"검색 URL: {search_url}")
            
            # 참고 코드와 동일한 완전한 헤더 설정
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
                "Accept-Encoding": "gzip, deflate, br",
                "Origin": "https://fin.land.naver.com",
                "Referer": "https://fin.land.naver.com/",
                "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
                "Sec-Ch-Ua-Mobile": "?0",
                "Sec-Ch-Ua-Platform": '"Windows"',
                "Sec-Fetch-Dest": "empty",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Site": "same-origin",
                "Cache-Control": "no-cache",
                "Pragma": "no-cache"
            }
            
            # 세션 사용으로 쿠키 유지 (참고 코드와 동일)
            if not hasattr(self, 'naver_session'):
                self.naver_session = requests.Session()
                # 먼저 메인 페이지 방문하여 쿠키 획득
                self.naver_session.get("https://fin.land.naver.com/", headers=headers)
                time.sleep(0.5)
            
            # 재시도 로직 추가 (참고 코드와 동일)
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = self.naver_session.get(search_url, headers=headers, timeout=10)
                    
                    if response.status_code == 429:
                        # Rate limit 에러 시 대기 후 재시도
                        wait_time = (attempt + 1) * 2  # 2, 4, 6초 대기
                        print(f"요청 제한 발생. {wait_time}초 대기 후 재시도... (시도 {attempt+1}/{max_retries})")
                        time.sleep(wait_time)
                        continue
                    
                    if response.status_code != 200:
                        print(f"API 요청 실패: 상태 코드 {response.status_code}")
                        continue
                    
                    data = response.json()
                    
                    # result.list에서 단지 목록 가져오기
                    complexes = data.get('result', {}).get('list', [])
                    print(f"검색된 단지 수: {len(complexes)}")
                    
                    if not complexes:
                        print("검색 결과가 없습니다.")
                        return None
                    
                    # 검색 결과가 1개면 바로 반환
                    if len(complexes) == 1:
                        complex_id = str(complexes[0].get('complexNumber'))
                        complex_name = complexes[0].get('complexName', '')
                        print(f"단지 선택: {complex_name} (ID: {complex_id})")
                        return complex_id
                    
                    # 검색 결과가 여러 개면 선택 대화상자 표시
                    print(f"여러 개의 단지가 검색되어 선택 대화상자를 표시합니다.")
                    return self.show_complex_selection_dialog(complexes)
                    
                except requests.exceptions.RequestException as e:
                    print(f"네트워크 오류 (시도 {attempt+1}/{max_retries}): {str(e)}")
                    if attempt < max_retries - 1:
                        time.sleep(2)
                    else:
                        return None
            
            return None
                
        except Exception as e:
            print(f"네이버 단지 검색 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def show_complex_selection_dialog(self, complexes):
        """여러 단지 중 선택하는 대화상자"""
        dialog = tk.Toplevel(self.root)
        dialog.title("단지 선택")
        dialog.attributes('-topmost', True)
        
        width = 700
        height = 400
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        selected_complex_id = None
        
        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text="검색된 단지를 선택해주세요:",
                 font=self.font_normal if hasattr(self, 'font_normal') else None).pack(pady=(0, 10))
        
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox = tk.Listbox(list_frame, 
                            yscrollcommand=scrollbar.set,
                            font=self.font_normal if hasattr(self, 'font_normal') else ('Malgun Gothic', 10))
        listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        
        complex_map = {}
        for i, complex_info in enumerate(complexes):
            complex_name = complex_info.get('complexName', '')
            address = complex_info.get('legalDivisionName', '')
            complex_type = complex_info.get('type', '')
            # 타입 변환 (A01=아파트, A06=빌라/다세대 등)
            type_text = "아파트" if complex_type == "A01" else "빌라/다세대" if complex_type == "A06" else complex_type
            
            display_text = f"{complex_name} | {type_text} | {address}"
            listbox.insert(tk.END, display_text)
            complex_map[i] = str(complex_info.get('complexNumber'))
        
        def on_select():
            nonlocal selected_complex_id
            if listbox.curselection():
                idx = listbox.curselection()[0]
                selected_complex_id = complex_map[idx]
                dialog.destroy()
        
        listbox.bind('<Double-Button-1>', lambda e: on_select())
        
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill='x', pady=(10, 0))
        
        ttk.Button(button_frame, text="선택", command=on_select).pack(side='right', padx=(5, 0))
        ttk.Button(button_frame, text="취소", command=dialog.destroy).pack(side='right')
        
        self.root.wait_window(dialog)
        
        return selected_complex_id
        
    def download_naver_data(self, complex_id, apt_name):
        """네이버 부동산 API 페이지를 직접 열어서 데이터 수집"""
        try:
            print(f"\n=== 네이버 부동산 웹페이지로 '{apt_name}' 단지(ID: {complex_id}) 데이터 수집 중... ===")
            
            from selenium import webdriver
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.common.by import By
            from webdriver_manager.chrome import ChromeDriverManager
            import json
            import time
            
            # Chrome 옵션 설정
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # 백그라운드 실행
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # User-Agent 설정
            chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # 자동화 감지 방지
            driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                'source': '''
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    })
                '''
            })
            
            all_properties = []
            
            try:
                # 먼저 메인 페이지 접속 (쿠키 및 세션 설정)
                main_url = f"https://fin.land.naver.com/complexes/{complex_id}"
                print(f"메인 페이지 접속: {main_url}")
                driver.get(main_url)
                time.sleep(2)  # 페이지 로딩 및 쿠키 설정 대기
                
                # 페이지별로 데이터 수집
                page = 0
                has_more_data = True
                max_pages = 50  # 최대 페이지 수 제한
                
                while has_more_data and page < max_pages:
                    # API URL 직접 접속
                    api_url = f"https://fin.land.naver.com/front-api/v1/complex/article/list?complexNumber={complex_id}&dateDescending=false&userChannelType=PC&page={page}"
                    print(f"\n페이지 {page} 데이터 수집 중...")
                    print(f"URL: {api_url}")
                    
                    # 페이지 열기
                    driver.get(api_url)
                    time.sleep(1.5)  # 페이지 로딩 대기
                    
                    try:
                        # 페이지 소스에서 JSON 데이터 추출
                        page_source = driver.page_source
                        
                        # <pre> 태그 안의 JSON 데이터 찾기
                        pre_element = driver.find_element(By.TAG_NAME, "pre")
                        json_text = pre_element.text
                        
                        # JSON 파싱
                        data = json.loads(json_text)
                        
                        # 응답 확인
                        if 'result' in data and 'list' in data['result']:
                            property_list = data['result']['list']
                            
                            if not property_list:
                                print(f"페이지 {page}에 매물이 없습니다. 수집 완료.")
                                has_more_data = False
                                break
                            
                            print(f"페이지 {page}에서 {len(property_list)}개의 매물 발견")
                            
                            # 매물 정보 추출
                            for item in property_list:
                                property_data = self.extract_property_data_from_json(item, page)
                                if property_data:
                                    all_properties.append(property_data)
                            
                            # 다음 페이지 확인
                            has_more_data = data['result'].get('hasNextPage', False)
                            
                            if has_more_data:
                                print(f"다음 페이지 존재 - 계속 수집...")
                            else:
                                print(f"페이지 {page}가 마지막 페이지입니다.")
                        else:
                            print(f"페이지 {page}에서 유효한 데이터가 없습니다.")
                            has_more_data = False
                            
                    except json.JSONDecodeError as e:
                        print(f"JSON 파싱 오류: {str(e)}")
                        # HTML 응답인 경우 (에러 페이지 등)
                        if "<!DOCTYPE" in page_source or "<html" in page_source:
                            print("HTML 응답 감지 - API 접근 차단됨")
                            # 대기 후 재시도
                            time.sleep(3)
                            driver.get(main_url)  # 메인 페이지 재접속
                            time.sleep(2)
                            continue
                        break
                        
                    except Exception as e:
                        print(f"데이터 추출 중 오류: {str(e)}")
                        break
                    
                    # 다음 페이지로
                    page += 1
                    
                    # 너무 빠른 요청 방지
                    time.sleep(1)
                    
            finally:
                driver.quit()
                print("Chrome 드라이버 종료")
            
            print(f"\n총 {len(all_properties)}개 매물 데이터 수집 완료")
            
            if not all_properties:
                print("수집된 매물이 없습니다.")
                return None
            
            # 데이터프레임 생성
            df = pd.DataFrame(all_properties)
            
            # 엑셀 파일로 저장
            today = datetime.now().strftime('%Y%m%d')
            excel_filename = os.path.join(self.download_path, f'{apt_name}_네이버매물_{today}.xlsx')
            
            df.to_excel(excel_filename, index=False)
            print(f"네이버 매물 데이터가 '{excel_filename}' 파일로 저장되었습니다.")
            
            # 엑셀 파일 형식 개선
            self.format_naver_excel(excel_filename)
            
            return excel_filename
            
        except Exception as e:
            print(f"네이버 매물 데이터 수집 중 오류: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def extract_property_data_from_json(self, item, page):
        """JSON 데이터에서 매물 정보 추출"""
        try:
            rep_info = item.get('representativeArticleInfo', {})
            
            property_data = {
                '단지명': rep_info.get('complexName', ''),
                '동': rep_info.get('dongName', ''),
                '거래유형': rep_info.get('tradeType', ''),
                '전용면적': 0,
                '타입구분': '',
                '층/전체층': '',
                '방향': '',
                '매물특징': '',
                '보증금': 0,
                '월세': 0,
                '매매가': 0,
                '프리미엄': 0,
                '가격변동상태': 0,
                '중개사명': '',
                'VR노출여부': False,
                '중개사 수': 0,
                '최소매매가': 0,
                '최대매매가': 0,
                '최소보증금': 0,
                '최대보증금': 0,
                '최소월세': 0,
                '최대월세': 0,
                '최소프리미엄': 0,
                '최대프리미엄': 0,
                '페이지번호': page
            }
            
            # 공간 정보
            if 'spaceInfo' in rep_info:
                property_data['전용면적'] = rep_info['spaceInfo'].get('exclusiveSpace', 0)
                property_data['타입구분'] = rep_info['spaceInfo'].get('nameType', '')
            
            # 상세 정보
            if 'articleDetail' in rep_info:
                property_data['층/전체층'] = rep_info['articleDetail'].get('floorInfo', '')
                property_data['방향'] = rep_info['articleDetail'].get('direction', '')
                property_data['매물특징'] = rep_info['articleDetail'].get('articleFeatureDescription', '')
            
            # 가격 정보
            if 'priceInfo' in rep_info:
                property_data['보증금'] = rep_info['priceInfo'].get('warrantyPrice', 0)
                property_data['월세'] = rep_info['priceInfo'].get('rentPrice', 0)
                property_data['매매가'] = rep_info['priceInfo'].get('dealPrice', 0)
                property_data['프리미엄'] = rep_info['priceInfo'].get('premiumPrice', 0)
                property_data['가격변동상태'] = rep_info['priceInfo'].get('priceChangeStatus', 0)
            
            # 중개사 정보
            if 'brokerInfo' in rep_info:
                property_data['중개사명'] = rep_info['brokerInfo'].get('brokerageName', '')
            
            # VR 정보
            if 'articleMediaDto' in rep_info and rep_info['articleMediaDto']:
                property_data['VR노출여부'] = rep_info['articleMediaDto'].get('isVrExposed', False)
            
            # 중복 매물 정보 처리
            if 'duplicatedArticlesInfo' in item and item['duplicatedArticlesInfo']:
                dup_info = item['duplicatedArticlesInfo']
                property_data['중개사 수'] = dup_info.get('realtorCount', 0)
                
                if 'representativePriceInfo' in dup_info:
                    price_info = dup_info['representativePriceInfo']
                    
                    if 'dealPrice' in price_info:
                        property_data['최소매매가'] = price_info['dealPrice'].get('minPrice', 0)
                        property_data['최대매매가'] = price_info['dealPrice'].get('maxPrice', 0)
                    
                    if 'warrantyPrice' in price_info:
                        property_data['최소보증금'] = price_info['warrantyPrice'].get('minPrice', 0)
                        property_data['최대보증금'] = price_info['warrantyPrice'].get('maxPrice', 0)
                    
                    if 'rentPrice' in price_info:
                        property_data['최소월세'] = price_info['rentPrice'].get('minPrice', 0)
                        property_data['최대월세'] = price_info['rentPrice'].get('maxPrice', 0)
                    
                    if 'premiumPrice' in price_info:
                        property_data['최소프리미엄'] = price_info['premiumPrice'].get('minPrice', 0)
                        property_data['최대프리미엄'] = price_info['premiumPrice'].get('maxPrice', 0)
            
            # 거래유형 변환
            trade_type_map = {
                'A1': '매매',
                'B1': '전세',
                'B2': '월세',
                'B3': '단기임대'
            }
            property_data['거래유형'] = trade_type_map.get(property_data['거래유형'], property_data['거래유형'])
            
            # 최저가매물 계산
            if property_data['최소매매가'] > 0 and property_data['최소매매가'] < property_data['매매가']:
                property_data['최저가매물'] = property_data['최소매매가']
            else:
                property_data['최저가매물'] = property_data['매매가']
            
            return property_data
            
        except Exception as e:
            print(f"매물 데이터 추출 오류: {str(e)}")
            return None



    
            
    
  
            
    
    # KB부동산과 네이버 부동산에서 공통으로 사용할 driver 변수 추가
    def browser_search(self, apt_name):
        # start_search에서 이미 비활성화하므로 여기서는 중복 호출 방지
        if not hasattr(self, '_search_in_progress'):
            self.search_button.config(state="disabled")
            if hasattr(self, 'multi_analysis_button'):
                self.multi_analysis_button.config(state="disabled")
        
        self.update_progress(0)
        
        # 다운로드 경로 확인 및 생성
        os.makedirs(self.download_path, exist_ok=True)
        print(f"\n=== 브라우저 검색 시작 ===")
        print(f"다운로드 경로: {self.download_path}")
    
        # 기존 엑셀 파일 정리
        try:
            for file in os.listdir(self.download_path):
                if file.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(self.download_path, file)
                    try:
                        os.remove(file_path)
                        print(f"기존 파일 삭제: {file}")
                    except PermissionError:
                        print(f"파일 삭제 실패 (사용 중): {file}")
                    except Exception as e:
                        print(f"파일 삭제 중 오류: {str(e)}")
        except Exception as e:
            print(f"파일 정리 중 오류: {str(e)}")
     
        # Chrome driver 초기화 코드 - headless 모드 복원
        # Chrome driver 초기화 코드 - 디버깅을 위해 headless 모드 비활성화
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        
        # ===== 디버깅을 위해 headless 모드 비활성화 =====
        chrome_options.add_argument("--headless")  # 주석 처리하여 브라우저가 보이도록 함
        
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        
        # 브라우저가 보이는 모드에서도 안정적으로 작동하도록 옵션 유지
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')

        
        # 다운로드 설정
        absolute_download_path = os.path.abspath(self.download_path)
        prefs = {
            'profile.default_content_setting_values': {
                'images': 2,  # 이미지 로딩 비활성화
                'plugins': 2,  # 플러그인 비활성화
                'javascript': 1  # 자바스크립트는 필수이므로 활성화
            },
            'profile.default_content_setting_values.popups': 0,
            'profile.default_content_settings.popups': 0,
            'download.default_directory': absolute_download_path,
            'download.prompt_for_download': False,
            'download.directory_upgrade': True,
            'safebrowsing.enabled': False
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.page_load_strategy = 'eager'
    
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
        # headless 모드에서 자동화 감지 방지
        self.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                })
            '''
        })
     
        try:
            # 다운로드 경로 생성 확인 
            os.makedirs(self.download_path, exist_ok=True)
            print(f"다운로드 경로 설정: {absolute_download_path}")
            
            # 다운로드 동작 설정
            self.driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
            params = {
                'cmd': 'Page.setDownloadBehavior',
                'params': {
                    'behavior': 'allow',
                    'downloadPath': absolute_download_path
                }
            }
            self.driver.execute("send_command", params)
            self.driver.set_page_load_timeout(30)
            print("Chrome 다운로드 설정 완료")
    
            # KB부동산 검색
            self.driver.get("https://kbland.kr/map")
            
            # 페이지 로딩 대기
            time.sleep(5)
            
            # 팝업이 있는지 확인하고 닫기 (headless 모드에서도 작동)
            try:
                popup_close_selectors = [
                    "button.close",
                    "button.popup-close",
                    "button.modal-close",
                    "a.close",
                    ".close-button",
                    "[aria-label='close']",
                    "[aria-label='닫기']"
                ]
                
                for selector in popup_close_selectors:
                    try:
                        close_buttons = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        for button in close_buttons:
                            if button.is_displayed():
                                self.driver.execute_script("arguments[0].click();", button)
                                print(f"팝업 닫기 버튼 클릭: {selector}")
                                time.sleep(1)
                    except:
                        continue
            except Exception as e:
                print(f"팝업 확인/닫기 중 오류 (무시): {str(e)}")
    
            # 주소 검색 버튼 클릭
            # 주소 검색 버튼 클릭
            try:
                # 검색창 버튼 클릭 - 새로운 선택자 추가
                search_button_selectors = [
                    "#app > div > div.mapsearch-wrap > div > button",  # 제공된 선택자
                    "div.mapsearch-wrap button",  # 단순화된 선택자
                    "button.btn-land-search",  # 기존 대체 선택자
                    "strong[aria-label='주소로 찾기 검색탭']",  # 기존 선택자
                    "strong[role='button']",  # 기존 선택자
                ]
                
                search_button_clicked = False
                for selector in search_button_selectors:
                    try:
                        print(f"시도 중: {selector}")
                        search_button = WebDriverWait(self.driver, 3).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        
                        # JavaScript로 강제 클릭 시도
                        self.driver.execute_script("arguments[0].click();", search_button)
                        print(f"검색 버튼 클릭 성공: {selector}")
                        search_button_clicked = True
                        time.sleep(2)  # 클릭 후 대기
                        break
                    except Exception as e:
                        print(f"선택자 {selector} 시도 실패: {str(e)}")
                        continue
                
                # 검색창이 나타났는지 확인
                if search_button_clicked:
                    try:
                        # 검색 입력창 대기
                        search_input = WebDriverWait(self.driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder*='단지'], input[placeholder*='주소'], input.form-control"))
                        )
                        print("검색창 표시 확인")
                        time.sleep(1)
                    except:
                        print("검색창을 찾을 수 없습니다. 다시 시도...")
                        # 검색창이 안 나타났으면 다시 클릭 시도
                        for selector in search_button_selectors[:2]:  # 처음 두 개만 재시도
                            try:
                                button = self.driver.find_element(By.CSS_SELECTOR, selector)
                                self.driver.execute_script("arguments[0].click();", button)
                                time.sleep(2)
                                break
                            except:
                                continue
                else:
                    print("검색 버튼을 찾을 수 없습니다.")
                    raise Exception("KB부동산 검색 버튼을 찾을 수 없습니다.")
                    
            except Exception as e:
                print(f"주소 검색 버튼 처리 중 오류: {str(e)}")
                raise
                
                # XPath로 추가 시도
                if not address_button:
                    try:
                        elements = self.driver.find_elements(By.XPATH, "//strong[@role='button'][contains(text(), '구') or contains(text(), '동')]")
                        if elements:
                            address_button = elements[0]
                            print(f"주소 버튼 찾기 성공 (XPath)")
                    except:
                        pass
                
                if address_button:
                    # JavaScript로 클릭 (headless 모드에서 더 안정적)
                    self.driver.execute_script("arguments[0].click();", address_button)
                    print("주소 버튼 클릭 성공")
                    time.sleep(2)
                    
                    # 검색창이 나타날 때까지 대기
                    search_input = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder*='단지']"))
                    )
                    print("검색창 표시 확인")
                    time.sleep(1)
                else:
                    print("주소 버튼을 찾을 수 없습니다. 대체 방법 시도...")
                    
                    # 대체 방법들
                    alternative_selectors = [
                        "button.btn-land-search",
                        "button[variant='adrSearch']",
                        "button[type='button']"
                    ]
                    
                    for alt_selector in alternative_selectors:
                        try:
                            search_button = WebDriverWait(self.driver, 3).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, alt_selector))
                            )
                            self.driver.execute_script("arguments[0].click();", search_button)
                            print(f"대체 검색 버튼 클릭 성공: {alt_selector}")
                            time.sleep(2)
                            break
                        except:
                            continue
                            
            except Exception as e:
                print(f"주소 검색 버튼 처리 중 오류: {str(e)}")
                raise
    
            # 아파트를 찾지 못하면 여기서 중단
            if not self.find_exact_apartment(self.driver, apt_name):
                self.driver.quit()
                self.search_button.config(state="normal")
                self.update_progress(0)
                return False
    
            time.sleep(2)
    
            if self.check_area_dropdown_exists(self.driver):
                if self.select_area(self.driver, self.area.get()):
                    try:
                        if self.download_files(self.driver):
                            self.update_progress(90, "분석을 시작합니다.")
                            excel_file = self.get_latest_excel()
                            self.latest_excel_file = excel_file
                            df_info = pd.read_excel(excel_file, nrows=8)
                            kb_apt_name = str(df_info.iloc[2, 1])
                            
                            return True
                    except Exception as e:
                        print(f"파일 다운로드 중 오류: {str(e)}")
                        raise
    
            self.driver.quit()
            return False
    
        except Exception as e:
            print(f"브라우저 검색 중 오류: {str(e)}")
            self.driver.quit()
            raise e
    
        finally:
            if not hasattr(self, '_search_in_progress'):
                self.search_button.config(state="normal")
                if hasattr(self, 'multi_analysis_button'):
                    self.multi_analysis_button.config(state="normal")
    
    def setup_gui(self):
        """GUI를 설정하는 메서드"""
        # 위젯 업데이트 최적화
        def delayed_update():
            if hasattr(self, '_update_timer'):
                self.root.after_cancel(self._update_timer)
            self._update_timer = self.root.after(100, self.update_history_display)
        
        # 이미지 캐싱
        self.image_cache = {}
        
        # 스타일 설정
        style = ttk.Style()
        if hasattr(self, 'font_button'):
            style.configure("Custom.TButton", font=self.font_button, padding=5)
        else:
            style.configure("Custom.TButton", padding=5)
    
        # 폰트가 설정된 후에 적용할 스타일들
        if hasattr(self, 'font_normal'):
            style.configure('TLabel', font=self.font_normal)
            style.configure('TButton', font=self.font_normal)
            style.configure('TEntry', font=self.font_normal)
            style.configure('Treeview', font=self.font_normal)
            style.configure('Treeview.Heading', font=self.font_normal)
        
        # 메인 프레임을 그리드로 구성 - 좌측 컨텐츠, 우측 히스토리
        self.root.columnconfigure(0, weight=3)  # 좌측 컨텐츠 영역 (더 넓게)
        self.root.columnconfigure(1, weight=1)  # 우측 히스토리 영역
        self.root.rowconfigure(0, weight=1)
    
        # 좌측 메인 컨텐츠 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        # 우측 히스토리 프레임
        history_frame = ttk.LabelFrame(self.root, text="검색 히스토리", padding="10")
        history_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        
        # 제목 레이블에 사용자 정의 폰트 적용
        title_label = ttk.Label(
            main_frame,
            text="KB부동산 시세 + 실거래가 차트",
            font=self.font_title  # 여기서 self.font_title을 명시적으로 사용
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 10))
    
        # guide_label도 폰트 적용
        # guide_label 설정 부분 수정
        self.guide_label = ttk.Label(
            main_frame, 
            wraplength=350,
            foreground="gray",
            font=self.font_normal if hasattr(self, 'font_normal') else ('KoPubWorld Dotum Medium', 9)  # Helvetica를 KoPubWorld로 변경
        )
        self.guide_label.grid(row=1, column=0, columnspan=2, pady=5)
        
        # 상단에 설정 버튼 추가
        settings_frame = ttk.Frame(main_frame)
        settings_frame.grid(row=0, column=1, sticky="e", padx=5)
        
        settings_button = ttk.Button(
            settings_frame,
            text="⚙",  # 톱니바퀴 이모지
            width=3,
            command=self.show_settings_dialog,
            style="Custom.TButton"
        )
        settings_button.pack(side="right")
        
        # 프로그레스 바와 상태 레이블 먼저 생성 (중요)
        self.progress = ttk.Progressbar(
            main_frame,
            orient="horizontal",
            length=300,
            mode="determinate"
        )
        self.progress.grid(row=4, column=0, columnspan=2, pady=10, padx=5, sticky="ew")
    
        # 상태 표시 레이블 (먼저 생성)
        self.status_label = ttk.Label(
            main_frame,
            text="",
            wraplength=500  # 가로 길이 증가
        )
        self.status_label.grid(row=5, column=0, columnspan=2, pady=5, sticky="ew")
        
        # 검색 방법 선택을 위한 노트북(탭) 추가
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=5)
        
        # 직접 검색 탭
        direct_frame = ttk.LabelFrame(self.notebook, text="직접 검색", padding=10)
        self.notebook.add(direct_frame, text="직접 검색")
        
        # 입력 필드를 담을 프레임
        # 입력 필드를 담을 프레임
        input_frame = ttk.Frame(direct_frame)
        input_frame.pack(fill='x', expand=True)
        
        # 왼쪽 영역 (라벨과 입력 필드)
        left_frame = ttk.Frame(input_frame)
        left_frame.pack(side='left', fill='x', expand=True)
        
        # 아파트 이름 입력 행에 다중 분석 버튼 추가
        apt_name_frame = ttk.Frame(left_frame)
        apt_name_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=5)
        apt_name_frame.columnconfigure(1, weight=1)  # 입력 필드가 확장되도록
        
        ttk.Label(apt_name_frame, text="아파트 이름:").grid(row=0, column=0, sticky="w", padx=(0, 5))
        self.apt_name = ttk.Entry(apt_name_frame, width=25)
        self.apt_name.grid(row=0, column=1, sticky="ew", padx=5)
        
        # 다중 분석 버튼 추가
        self.multi_analysis_button = ttk.Button(
            apt_name_frame,
            text="다중 분석",
            command=self.show_multi_analysis_dialog,
            style="Custom.TButton",
            width=10
        )
        self.multi_analysis_button.grid(row=0, column=2, padx=(5, 0))
        
        ttk.Label(left_frame, text="전용면적(m²):").grid(row=1, column=0, sticky="w", pady=5)
        self.area = ttk.Entry(left_frame, width=30)
        self.area.grid(row=1, column=1, padx=5, pady=5)
        
        # 신축단지 정보 입력 필드
        ttk.Label(left_frame, text="신축단지명:").grid(row=2, column=0, sticky="w", pady=5)
        self.new_apt_name = ttk.Entry(left_frame, width=30)
        self.new_apt_name.grid(row=2, column=1, padx=5, pady=5)
        
        ttk.Label(left_frame, text="분양가(만원):").grid(row=3, column=0, sticky="w", pady=5)
        self.new_apt_price = ttk.Entry(left_frame, width=30)
        self.new_apt_price.grid(row=3, column=1, padx=5, pady=5)
        
        # 오른쪽 영역 (검색 버튼)
        right_frame = ttk.Frame(input_frame)
        right_frame.pack(side='right', fill='y', padx=(10, 0))
        
        # 검색 버튼
        self.search_button = ttk.Button(
            right_frame,
            text="검색 시작",
            command=self.start_search,
            style="Custom.TButton"
        )
        self.search_button.pack(expand=True, padx=5)
        
        # 힌트 텍스트
        # 힌트 텍스트
        hint_label = ttk.Label(
            direct_frame, 
            text="※ 신축단지 정보는 선택 입력사항입니다",
            foreground="gray",
            font=self.font_normal if hasattr(self, 'font_normal') else ('KoPubWorld Dotum Medium', 8)
        )
        hint_label.pack(pady=(5,0), anchor='w')
    
        # 가격 데이터 선택 프레임 추가
        price_select_frame = ttk.LabelFrame(direct_frame, text="가격 데이터 선택", padding=5)
        price_select_frame.pack(fill='x', pady=10, padx=5, anchor='w')
        
        # 가격 선택 변경 함수
        def on_price_type_changed():
            """가격 유형 선택 변경 시 호출되는 함수"""
            # 매매가 선택 확인
            sale_types = [t for t, var in self.sale_price_types.items() if var.get()]
            lease_types = [t for t, var in self.lease_price_types.items() if var.get()]
            
            # 각 그룹에서 최소 1개, 최대 2개 선택되도록 제한
            if len(sale_types) > 2:
                # 가장 최근에 선택된 2개만 유지
                for type_name in ["low", "normal", "high"]:
                    if type_name not in sale_types[-2:]:
                        self.sale_price_types[type_name].set(False)
            elif len(sale_types) == 0:
                # 아무것도 선택되지 않았으면 normal 자동 선택
                self.sale_price_types["normal"].set(True)
                
            if len(lease_types) > 2:
                # 가장 최근에 선택된 2개만 유지
                for type_name in ["low", "normal", "high"]:
                    if type_name not in lease_types[-2:]:
                        self.lease_price_types[type_name].set(False)
            elif len(lease_types) == 0:
                # 아무것도 선택되지 않았으면 normal 자동 선택
                self.lease_price_types["normal"].set(True)
                
            # 현재 선택 상태 확인 (디버깅용)
            sale_types = [t for t, var in self.sale_price_types.items() if var.get()]
            lease_types = [t for t, var in self.lease_price_types.items() if var.get()]
            
            print(f"현재 매매가 선택: {sale_types}")
            print(f"현재 전세가 선택: {lease_types}")
            
            # status_label이 있으면 업데이트
            if hasattr(self, 'status_label'):
                type_map = {"low": "하위평균", "normal": "일반평균", "high": "상위평균"}
                sale_txt = ", ".join([type_map[t] for t in sale_types])
                lease_txt = ", ".join([type_map[t] for t in lease_types])
                self.status_label.config(text=f"가격 선택: 매매가({sale_txt}), 전세가({lease_txt})")
        
        # 매매가 선택 프레임
        sale_frame = ttk.Frame(price_select_frame)
        sale_frame.pack(fill='x', pady=5)
        ttk.Label(sale_frame, text="매매가:").pack(side="left", padx=(5, 10))
        
        # 체크박스로 변경
        cb_sale_low = ttk.Checkbutton(
            sale_frame, text="하위 평균", variable=self.sale_price_types["low"], 
            command=on_price_type_changed
        )
        cb_sale_low.pack(side="left", padx=10)
        
        cb_sale_normal = ttk.Checkbutton(
            sale_frame, text="일반 평균", variable=self.sale_price_types["normal"], 
            command=on_price_type_changed
        )
        cb_sale_normal.pack(side="left", padx=10)
        
        cb_sale_high = ttk.Checkbutton(
            sale_frame, text="상위 평균", variable=self.sale_price_types["high"], 
            command=on_price_type_changed
        )
        cb_sale_high.pack(side="left", padx=10)
        
        # 전세가 선택 프레임
        lease_frame = ttk.Frame(price_select_frame)
        lease_frame.pack(fill='x', pady=5)
        ttk.Label(lease_frame, text="전세가:").pack(side="left", padx=(5, 10))
        
        # 체크박스로 변경
        cb_lease_low = ttk.Checkbutton(
            lease_frame, text="하위 평균", variable=self.lease_price_types["low"], 
            command=on_price_type_changed
        )
        cb_lease_low.pack(side="left", padx=10)
        
        cb_lease_normal = ttk.Checkbutton(
            lease_frame, text="일반 평균", variable=self.lease_price_types["normal"], 
            command=on_price_type_changed
        )
        cb_lease_normal.pack(side="left", padx=10)
        
        cb_lease_high = ttk.Checkbutton(
            lease_frame, text="상위 평균", variable=self.lease_price_types["high"], 
            command=on_price_type_changed
        )
        cb_lease_high.pack(side="left", padx=10)
        
        # 초기 상태 설정 (자동으로 설정됨)
        on_price_type_changed()
        
        # 초기 상태 설정 (명시적으로 선택)
        on_price_type_changed()
        
        # 심리차트 설정 프레임 추가
        sentiment_frame = ttk.LabelFrame(direct_frame, text="차트 표시 설정", padding=5)
        sentiment_frame.pack(fill='x', pady=10, padx=5, anchor='w')
        
        # 체크박스를 담을 내부 프레임
        sentiment_checkbox_frame = ttk.Frame(sentiment_frame)
        sentiment_checkbox_frame.pack(fill='x', pady=5)
        
        # 유나심리차트 체크박스 (기존 두 개를 하나로 통합)
        una_sentiment_check = ttk.Checkbutton(
            sentiment_checkbox_frame, 
            text="유나심리차트 표시", 
            variable=self.show_una_sentiment,
            command=self.update_charts
        )
        una_sentiment_check.pack(side="left", padx=10)
        
        # PIR 차트 체크박스 추가
        pir_chart_check = ttk.Checkbutton(
            sentiment_checkbox_frame, 
            text="PIR 표시", 
            variable=self.show_pir,
            command=self.update_charts
        )
        pir_chart_check.pack(side="left", padx=10)


        # 데이터 선택 프레임 추가
        data_select_frame = ttk.LabelFrame(direct_frame, text="데이터 표시 선택", padding=5)
        data_select_frame.pack(fill='x', pady=10, padx=5, anchor='w')
        
        # 데이터 체크박스를 담을 내부 프레임
        data_checkbox_frame = ttk.Frame(data_select_frame)
        data_checkbox_frame.pack(fill='x', pady=5)
        
        # KB시세 체크박스
        kb_price_check = ttk.Checkbutton(
            data_checkbox_frame, 
            text="KB시세", 
            variable=self.show_kb_price,
            command=self.update_charts
        )
        kb_price_check.pack(side="left", padx=10)
        
        # 네이버매물 체크박스
        naver_deal_check = ttk.Checkbutton(
            data_checkbox_frame, 
            text="네이버매물", 
            variable=self.show_naver_deal,
            command=self.update_charts
        )
        naver_deal_check.pack(side="left", padx=10)
        
        # 실거래가 체크박스
        real_trade_check = ttk.Checkbutton(
            data_checkbox_frame, 
            text="실거래가", 
            variable=self.show_real_trade,
            command=self.update_charts
        )
        real_trade_check.pack(side="left", padx=10)
        
        # 지역 검색 탭
        region_frame = ttk.LabelFrame(self.notebook, text="지역 검색", padding=10)
        self.notebook.add(region_frame, text="지역 검색")
        
        # 지역 선택 콤보박스들
        ttk.Label(region_frame, text="시/도:").grid(row=0, column=0, sticky="w", pady=5)
        self.sido_combobox = ttk.Combobox(region_frame, width=20)
        self.sido_combobox['values'] = sorted(self.sido_list)
        self.sido_combobox.set("시/도 선택")
        self.sido_combobox.grid(row=0, column=1, padx=5, pady=5)
        self.sido_combobox.bind('<<ComboboxSelected>>', self.on_sido_selected)
        
        ttk.Label(region_frame, text="시/군/구:").grid(row=1, column=0, sticky="w", pady=5)
        self.sigungu_combobox = ttk.Combobox(region_frame, width=20)
        self.sigungu_combobox.set("시/군/구 선택")
        self.sigungu_combobox.grid(row=1, column=1, padx=5, pady=5)
        self.sigungu_combobox.bind('<<ComboboxSelected>>', self.on_sigungu_selected)
        
        ttk.Label(region_frame, text="읍/면/동:").grid(row=2, column=0, sticky="w", pady=5)
        self.dong_combobox = ttk.Combobox(region_frame, width=20)
        self.dong_combobox.set("읍/면/동 선택")
        self.dong_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.dong_combobox.bind('<<ComboboxSelected>>', self.on_dong_selected)
        
        ttk.Button(region_frame, text="아파트 목록 조회", 
                   command=self.show_apt_list).grid(row=3, column=0, columnspan=2, pady=10)
    
        # region_frame에 검색 시작 버튼 추가
        self.region_search_button = ttk.Button(
            region_frame,
            text="검색 시작",
            command=self.start_search,
            style="Custom.TButton"
        )
        self.region_search_button.grid(row=5, column=0, columnspan=2, pady=(5,10))
        
        # 비교 분석 탭 추가 (새로운 탭으로 이동)
        compare_frame = ttk.LabelFrame(self.notebook, text="비교 분석", padding=10)
        self.notebook.add(compare_frame, text="비교 분석")  # 세 번째 탭으로 추가
    
        # 첫 번째 단지 입력
        apt1_frame = ttk.LabelFrame(compare_frame, text="단지 1", padding=5)
        apt1_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    
        ttk.Label(apt1_frame, text="아파트 이름:").grid(row=0, column=0, sticky="w", pady=5)
        self.apt1_name = ttk.Entry(apt1_frame, width=30)
        self.apt1_name.grid(row=0, column=1, padx=5, pady=5)
    
        ttk.Label(apt1_frame, text="전용면적(m²):").grid(row=1, column=0, sticky="w", pady=5)
        self.apt1_area = ttk.Entry(apt1_frame, width=30)
        self.apt1_area.grid(row=1, column=1, padx=5, pady=5)
    
        # 두 번째 단지 입력
        apt2_frame = ttk.LabelFrame(compare_frame, text="단지 2", padding=5)
        apt2_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
    
        ttk.Label(apt2_frame, text="아파트 이름:").grid(row=0, column=0, sticky="w", pady=5)
        self.apt2_name = ttk.Entry(apt2_frame, width=30)
        self.apt2_name.grid(row=0, column=1, padx=5, pady=5)
    
        ttk.Label(apt2_frame, text="전용면적(m²):").grid(row=1, column=0, sticky="w", pady=5)
        self.apt2_area = ttk.Entry(apt2_frame, width=30)
        self.apt2_area.grid(row=1, column=1, padx=5, pady=5)
    
        # 버튼 프레임
        button_frame = ttk.Frame(compare_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
    
        # 매매가 비교 버튼
        self.compare_button = ttk.Button(
            button_frame,
            text="매매가 비교",
            command=lambda: self.start_comparison("매매가"),
            style="Custom.TButton"
        )
        self.compare_button.pack(side="left", padx=5)
    
        # 전세가 비교 버튼
        self.jeonse_compare_button = ttk.Button(
            button_frame,
            text="전세가 비교",
            command=lambda: self.start_comparison("전세가"),
            style="Custom.TButton"
        )
        self.jeonse_compare_button.pack(side="left", padx=5)
        
        # 클릭 가능한 블로그 링크 생성
        def open_blog(event):
            import webbrowser
            webbrowser.open('https://blog.naver.com/landlover333')
            
        # 하단 블로그 링크 (같은 폰트를 사용)
        blog_label = ttk.Label(
            main_frame,
            text="부태리의 블로그",
            foreground="blue",
            cursor="hand2",
            font=self.font_normal  # 혹은 self.font_title 등 원하는 크기로 지정
        )
        blog_label.grid(row=6, column=0, columnspan=2, pady=(10, 5), sticky="s")
        blog_label.bind("<Button-1>", open_blog)
        
        # 기존 footer_label의 위치를 약간 조정
        footer_label = ttk.Label(
            main_frame,
            text="만든 사람: 부태리(v1.0)",   # 원하는 문구로 수정
            font=self.font_normal,           # KoPubWorld Dotum Medium
            foreground="black"
        )
        footer_label.grid(row=7, column=0, columnspan=2, pady=(0, 5), sticky="s")
        # 히스토리 영역 구성 -------------------------------------------
        
        # 히스토리 버튼 프레임 추가
        history_btn_frame = ttk.Frame(history_frame)
        history_btn_frame.pack(side="bottom", fill="x", pady=(5, 0))
        
        # 폴더 열기 버튼 추가
        open_folder_btn = ttk.Button(
            history_btn_frame,
            text="저장 폴더 열기",
            command=self.open_download_folder,
            style="Custom.TButton"
        )
        open_folder_btn.pack(side="left", padx=5)
        
        # 기존 삭제 버튼들
        delete_btn = ttk.Button(
            history_btn_frame,
            text="선택 삭제",
            command=self.delete_selected_history,
            style="Custom.TButton"
        )
        delete_btn.pack(side="right", padx=5)
        
        delete_all_btn = ttk.Button(
            history_btn_frame,
            text="전체 삭제",
            command=self.delete_all_history,
            style="Custom.TButton"
        )
        delete_all_btn.pack(side="right", padx=5)
        
        # 히스토리 트리뷰
        self.history_tree = ttk.Treeview(history_frame, 
                                      columns=("date", "apt", "area", "max_trade"),
                                      show="headings",
                                      height=24)  # 높이 증가
        
        # 컬럼 헤더 설정
        self.history_tree.heading("date", text="검색일시")
        self.history_tree.heading("apt", text="아파트")
        self.history_tree.heading("area", text="면적")
        self.history_tree.heading("max_trade", text="최근 최고거래가")
        
        # 컬럼 너비 조정
        self.history_tree.column("date", width=100)
        self.history_tree.column("apt", width=130)
        self.history_tree.column("area", width=60)
        self.history_tree.column("max_trade", width=100)
        
        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(history_frame, orient="vertical", 
                                 command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        
        self.history_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 히스토리 클릭 이벤트 바인딩
        self.history_tree.bind('<Double-1>', self.on_history_select)
        
        # 초기 히스토리 표시
        self.update_history_display()
    
    # update_charts 메서드를 다음과 같이 수정합니다 (약 950줄 근처)
    def update_charts(self):
        """체크박스 상태가 변경되었을 때 호출되는 메서드"""
        pir_status = "표시" if self.show_pir.get() else "숨김"
        una_status = "표시" if self.show_una_sentiment.get() else "숨김"
        
        # 상태 메시지 업데이트
        # 데이터 표시 상태 추가
        kb_status = "표시" if self.show_kb_price.get() else "숨김"
        naver_status = "표시" if self.show_naver_deal.get() else "숨김"
        trade_status = "표시" if self.show_real_trade.get() else "숨김"
        
        # 상태 메시지 업데이트
        self.status_label.config(text=f"PIR: {pir_status}, 유나심리: {una_status}\n"
                                     f"KB시세: {kb_status}, 네이버매물: {naver_status}, 실거래가: {trade_status}")
        
        # 이미 검색한 결과가 있다면 (최신 엑셀 파일이 있다면) 그래프 다시 그리기
        try:
            # 최신 엑셀 파일이 있는지 확인
            if hasattr(self, 'latest_excel_file') and os.path.exists(self.latest_excel_file):
                # 사용자에게 그래프 다시 그릴지 물어보기
                if messagebox.askyesno("그래프 다시 그리기", 
                                      "설정이 변경되었습니다. 그래프를 다시 그리시겠습니까?"):
                    # 현재 다운로드 경로에 있는 관련 그래프 파일 찾기 (이름 패턴 체크)
                    excel_filename = os.path.basename(self.latest_excel_file)
                    apt_info = os.path.splitext(excel_filename)[0]  # 확장자 제거
                    
                    # analyze_prices 메서드 호출해서 그래프 다시 그리기
                    df = self.analyze_prices(self.latest_excel_file)
                    
                    # 그래프 파일이 생성되었다면 열기
                    if os.path.exists(self.image_path):
                        os.startfile(self.image_path)
                    
                    # 상태 메시지 업데이트
                    self.status_label.config(text=f"PIR 차트: {pir_status}, 유나심리차트: {una_status} - 그래프가 다시 생성되었습니다.")
                else:
                    # 다음 검색부터 적용 메시지
                    self.status_label.config(text=f"PIR 차트: {pir_status}, 유나심리차트: {una_status} - 다음 검색부터 적용됩니다.")
            else:
                # 데이터가 없으면 다음 검색부터 적용된다는 메시지
                self.status_label.config(text=f"PIR 차트: {pir_status}, 유나심리차트: {una_status} - 다음 검색부터 적용됩니다.")
        except Exception as e:
            print(f"그래프 다시 그리기 실패: {str(e)}")
            self.status_label.config(text=f"PIR 차트: {pir_status}, 유나심리차트: {una_status} - 다음 검색부터 적용됩니다.")    
  
    
    
    # 그래프는 업데이트하지 않음 - 다음 검색부터 적용됨
    def create_comparison_graph(self, df1, df2, apt1_name, apt2_name, apt1_area, apt2_area, compare_type, apt1_year=None, apt2_year=None):
        try:
            # 시작할 때 모든 그래프 및 세션 초기화
            print("비교 그래프 생성 시작 - 기존 그래프 정리")
            plt.close('all')  # 모든 열려있는 그래프 창 닫기
            
            # 폰트 설정
            plt.rcParams['font.family'] = 'Malgun Gothic'
            plt.rcParams['axes.unicode_minus'] = False
    
            plt.figure(figsize=(16, 9), clear=True)  # clear=True로 기존 그래프 초기화
            
            # 데이터 준비
            df1['date'] = pd.to_datetime(df1['date'])
            df2['date'] = pd.to_datetime(df2['date'])
        
            # 데이터 정렬
            df1 = df1.sort_values('date')
            df2 = df2.sort_values('date')
        
            # 현재가 데이터 추출
            latest_date1 = df1['date'].max()
            latest_date2 = df2['date'].max()
            
            # 현재가 값 추출 수정
            latest_mask1 = df1['date'] == latest_date1
            latest_mask2 = df2['date'] == latest_date2
            
            if latest_mask1.any() and latest_mask2.any():
                latest_value1 = df1.loc[latest_mask1, compare_type].iloc[0]
                latest_value2 = df2.loc[latest_mask2, compare_type].iloc[0]
            else:
                raise ValueError("현재가 데이터를 찾을 수 없습니다.")
    
            # 그래프 설정
            plt.gca().spines['right'].set_visible(False)
            plt.gca().spines['top'].set_visible(False)
            plt.gca().spines['left'].set_visible(True)
            plt.gca().spines['bottom'].set_visible(True)
            plt.grid(False)
    
            # 각 비교 유형별 색상 및 스타일 설정
            # 각 비교 유형별 색상 및 스타일 설정 부분 수정
            # 각 비교 유형별 색상 및 스타일 설정 부분 수정
            if compare_type == "매매가":
                # 매매가 비교 그래프
                line1_color = '#0066CC'  # 진한 파란색 유지
                line2_color = '#8B00FF'  # 보라색
                line1_style = '-'        # 실선
                line2_style = (0, (1, 1))  # 더 작은 점선 (점 1px, 간격 1px)
            else:  # 전세가
                # 전세가 비교 그래프
                line1_color = '#FF6B00'  # 주황색
                line2_color = '#FF0000'  # 빨간색
                line1_style = '-'        # 실선
                line2_style = (0, (1, 1))  # 더 작은 점선 (점 1px, 간격 1px)
                
            # 그래프 생성 부분 수정 (필요 시)
            line1, = plt.plot(df1['date'], df1[compare_type], label=f'{apt1_name}', 
                            color=line1_color, linewidth=4, linestyle=line1_style)
            line2, = plt.plot(df2['date'], df2[compare_type], label=f'{apt2_name}', 
                            color=line2_color, linewidth=3, linestyle=line2_style)  # 작은 점선을 위해 선 굵기 약간 감소
            # 현재가 마커 추가
            plt.scatter([latest_date1], [latest_value1], color='white', 
                        edgecolor=line1_color, s=100, linewidth=2, zorder=5)
            plt.scatter([latest_date2], [latest_value2], color='white', 
                        edgecolor=line2_color, s=100, linewidth=2, zorder=5)
            
            # 한번만 현재가 레이블 추가 - 값에 따라 위치 조정
            if latest_value1 >= latest_value2:
                plt.annotate(f'{latest_value1:,.0f}\n(현재가)',
                            xy=(latest_date1, latest_value1),
                            xytext=(10, 10),  # 위로 이동
                            textcoords='offset points',
                            bbox=dict(boxstyle='round,pad=0.5', fc='white', alpha=0.5),
                            fontsize=10,
                            zorder=6)
                
                plt.annotate(f'{latest_value2:,.0f}\n(현재가)',
                            xy=(latest_date2, latest_value2),
                            xytext=(10, -30),  # 아래로 이동
                            textcoords='offset points',
                            bbox=dict(boxstyle='round,pad=0.5', fc='white', alpha=0.5),
                            fontsize=10,
                            zorder=6)
            else:
                plt.annotate(f'{latest_value1:,.0f}\n(현재가)',
                            xy=(latest_date1, latest_value1),
                            xytext=(10, -30),  # 아래로 이동
                            textcoords='offset points',
                            bbox=dict(boxstyle='round,pad=0.5', fc='white', alpha=0.5),
                            fontsize=10,
                            zorder=6)
                
                plt.annotate(f'{latest_value2:,.0f}\n(현재가)',
                            xy=(latest_date2, latest_value2),
                            xytext=(10, 10),  # 위로 이동
                            textcoords='offset points',
                            bbox=dict(boxstyle='round,pad=0.5', fc='white', alpha=0.5),
                            fontsize=10,
                            zorder=6)
            title = f'{apt1_name} vs {apt2_name}\n{compare_type} 비교 분석'
            
            # 준공년도가 있으면 제목에 포함
            if apt1_year and apt2_year:
                title = f'{apt1_year}년 {apt1_name} ({apt1_area}㎡) vs {apt2_year}년 {apt2_name} ({apt2_area}㎡)\n{compare_type} 비교 분석'
            elif apt1_year:
                title = f'{apt1_year}년 {apt1_name} ({apt1_area}㎡) vs {apt2_name} ({apt2_area}㎡)\n{compare_type} 비교 분석'
            elif apt2_year:
                title = f'{apt1_name} ({apt1_area}㎡) vs {apt2_year}년 {apt2_name} ({apt2_area}㎡)\n{compare_type} 비교 분석'
            else:
                title = f'{apt1_name} ({apt1_area}㎡) vs {apt2_name} ({apt2_area}㎡)\n{compare_type} 비교 분석'

    
            # 그래프 제목 추가 (폰트 크기와 여백 조정)
            plt.title(title, pad=20, fontsize=20, fontweight='bold')
            plt.xlabel('날짜', fontsize=12)
            plt.ylabel('가격(만원)', fontsize=12)
            
            # 범례 설정
            plt.legend(loc='lower right',     
                      bbox_to_anchor=(0.98, 0.02),  
                      ncol=1,               
                      fontsize=14,          
                      markerscale=2,        
                      framealpha=0.9,       
                      frameon=True,         
                      edgecolor='gray',     
                      handlelength=3,       
                      handleheight=2,       
                      labelspacing=1.5,     
                      borderpad=1,          
                      prop={'size': 14, 'weight': 'bold'})
    
            # x축 날짜 형식 설정
            plt.gca().xaxis.set_major_locator(mdates.YearLocator())
            plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            plt.xticks(rotation=45)
    
            # 여백 조정
            plt.tight_layout()
            
            # 파일 저장
            # 연식 정보가 있으면 파일명에 포함
            if apt1_year and apt2_year:
                compare_filename = f"비교_{apt1_year}년_{apt1_name}_{apt2_year}년_{apt2_name}_{compare_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            elif apt1_year:
                compare_filename = f"비교_{apt1_year}년_{apt1_name}_{apt2_name}_{compare_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            elif apt2_year:
                compare_filename = f"비교_{apt1_name}_{apt2_year}년_{apt2_name}_{compare_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            else:
                compare_filename = f"비교_{apt1_name}_{apt2_name}_{compare_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            
            compare_filename = ''.join(char for char in compare_filename if char.isalnum() or char in ['_', '.'])     
            # 원본 이미지는 download 폴더에 저장
            self.image_path = os.path.join(self.download_path, compare_filename)
            plt.savefig(self.image_path, bbox_inches='tight', dpi=600, pad_inches=0.3)
            
            # 그래프를 저장한 후 명시적으로 닫기
            print(f"그래프 저장 완료: {self.image_path}")
            plt.close()  # 현재 그래프 닫기
            # 히스토리용 이미지는 history 폴더에 복사
            history_image_path = os.path.join(self.history_path, compare_filename)
            shutil.copy2(self.image_path, history_image_path)
            
            # 비교 분석용 엑셀 파일 생성
            history_excel_filename = f"history_compare_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            history_excel_path = os.path.join(self.history_path, history_excel_filename)
            
            wb = Workbook()
            ws = wb.active
            
            # 기본 정보 입력 - 연식 정보 포함
            ws['A1'] = '비교분석'
            if apt1_year and apt2_year:
                ws['B1'] = f"{apt1_year}년 {apt1_name} vs {apt2_year}년 {apt2_name}"
            elif apt1_year:
                ws['B1'] = f"{apt1_year}년 {apt1_name} vs {apt2_name}" 
            elif apt2_year:
                ws['B1'] = f"{apt1_name} vs {apt2_year}년 {apt2_name}"
            else:
                ws['B1'] = f"{apt1_name} vs {apt2_name}"
                
            ws['A2'] = '분석유형'
            ws['B2'] = compare_type
            ws['A3'] = '단지명1'
            ws['B3'] = apt1_name
            ws['A4'] = '단지명2'
            ws['B4'] = apt2_name
            ws['M1'] = "비교분석"
            
            # 이미지 파일 경로도 저장
            ws['A5'] = 'image_path'
            ws['B5'] = history_image_path
            
            wb.save(history_excel_path)
            
            # 히스토리 목록에 비교분석 항목 추가
            self.history_list.append({
                'file_path': history_excel_path,
                'apt_name': f"[비교] {apt1_name} vs {apt2_name}",
                'area': f"{compare_type}",
                'search_date': datetime.now().timestamp(),
                'max_trade': "비교분석",
                'type': 'compare'  # type 필드 추가
            })
            
            # 히스토리 화면 갱신
            self.update_history_display()
            
            plt.close()
    
            
            # 생성된 그래프 파일 팝업으로 표시
            os.startfile(self.image_path)
    
        except Exception as e:
            print(f"그래프 생성 중 오류: {str(e)}")
            messagebox.showerror("오류", f"그래프 생성 중 오류 발생: {str(e)}")



    def start_comparison(self, compare_type):
        try:
            apt1_name = self.apt1_name.get().strip()
            apt1_area = self.apt1_area.get().strip()
            apt2_name = self.apt2_name.get().strip()
            apt2_area = self.apt2_area.get().strip()

            # 준공년도를 저장할 변수 추가
            apt1_year = None
            apt2_year = None

            print("\n=== 비교분석 시작 ===")
            print(f"다운로드 경로: {self.download_path}")

            if not all([apt1_name, apt1_area, apt2_name, apt2_area]):
                raise ValueError("모든 입력값을 채워주세요.")

            # 버튼 비활성화
            self.compare_button.config(state="disabled")
            self.jeonse_compare_button.config(state="disabled")

            df1 = None
            df2 = None
            key1 = f"{apt1_name}_{apt1_area}"
            key2 = f"{apt2_name}_{apt2_area}"

            # 단지1 처리
            print("\n=== 단지1 처리 시작 ===")
            if key1 in self.downloaded_data:
                print(f"단지1: 메모리에 저장된 데이터 사용 - {key1}")
                df1 = self.downloaded_data[key1]
                # 캐시된 준공년도 가져오기
                if f"{key1}_year" in self.downloaded_data:
                    apt1_year = self.downloaded_data[f"{key1}_year"]
                    print(f"단지1 준공년도: {apt1_year}년")
            else:
                print(f"단지1: 새로 다운로드 시작")
                self.update_progress(20, f"{apt1_name} 데이터 수집 중...")
                self.apt_name.delete(0, tk.END)
                self.apt_name.insert(0, apt1_name)
                self.area.delete(0, tk.END)
                self.area.insert(0, apt1_area)
                
                if not self.browser_search(apt1_name):
                    raise ValueError("첫 번째 단지 검색 실패")
                latest_file = self.get_latest_excel()
                df1 = self.analyze_prices(latest_file)
                self.downloaded_data[key1] = df1
                # 준공년도 저장 (다음 단지 검색 전에)
                apt1_year = self.completion_year
                if apt1_year:
                    self.downloaded_data[f"{key1}_year"] = apt1_year
                    print(f"단지1 준공년도: {apt1_year}년")

            # 단지2 처리
            print("\n=== 단지2 처리 시작 ===")
            if key2 in self.downloaded_data:
                print(f"단지2: 메모리에 저장된 데이터 사용 - {key2}")
                df2 = self.downloaded_data[key2]
                # 캐시된 준공년도 가져오기
                if f"{key2}_year" in self.downloaded_data:
                    apt2_year = self.downloaded_data[f"{key2}_year"]
                    print(f"단지2 준공년도: {apt2_year}년")

                
            else:
                print(f"단지2: 새로 다운로드 시작")
                self.update_progress(60, f"{apt2_name} 데이터 수집 중...")
                self.apt_name.delete(0, tk.END)
                self.apt_name.insert(0, apt2_name)
                self.area.delete(0, tk.END)
                self.area.insert(0, apt2_area)
                
                if not self.browser_search(apt2_name):
                    raise ValueError("두 번째 단지 검색 실패")
                latest_file = self.get_latest_excel()
                df2 = self.analyze_prices(latest_file)
                self.downloaded_data[key2] = df2

                # 준공년도 저장
                apt2_year = self.completion_year
                if apt2_year:
                    self.downloaded_data[f"{key2}_year"] = apt2_year
                    print(f"단지2 준공년도: {apt2_year}년")

            if df1 is None or df2 is None:
                raise ValueError("데이터 수집 실패")

            print("\n=== 비교 그래프 생성 시작 ===")
            self.update_progress(90, "비교 그래프 생성 중...")
            # 준공년도 정보를 create_comparison_graph에 전달
            self.create_comparison_graph(df1, df2, apt1_name, apt2_name, apt1_area, apt2_area, compare_type, apt1_year, apt2_year)

                
            self.update_progress(100, "비교 분석 완료!")

        except Exception as e:
            print(f"\n=== 오류 발생 ===\n{str(e)}")
            messagebox.showerror("오류", str(e))
        
        finally:
            # 작업 완료 후 버튼 상태 복구
            self.compare_button.config(state="normal")
            self.jeonse_compare_button.config(state="normal")
            self.update_progress(0, "")
            
    # 기존 설정 창에 심리 데이터 엑셀 파일 경로 설정 추가
    def show_settings_dialog(self):
        settings = tk.Toplevel(self.root)
        settings.title("설정")
        settings.attributes('-topmost', True)
        
        # 창 크기와 위치 설정
        width = 900
        height = 300  # 네이버 관련 항목 제거로 높이 줄임
        screen_width = settings.winfo_screenwidth()
        screen_height = settings.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        settings.geometry(f"{width}x{height}+{x}+{y}")
        
        # 저장 경로 프레임
        path_frame = ttk.LabelFrame(settings, text="저장 경로 설정", padding=10)
        path_frame.pack(fill="x", padx=10, pady=5)
        
        # 다운로드 경로
        download_frame = ttk.Frame(path_frame)
        download_frame.pack(fill="x", pady=5)
        ttk.Label(download_frame, text="다운로드 경로:").pack(side="left")
        
        download_path_var = tk.StringVar(value=self.download_path)
        download_entry = ttk.Entry(download_frame, textvariable=download_path_var, width=50)
        download_entry.pack(side="left", padx=5)
        
        # 법정동 코드 파일 경로
        lawdong_frame = ttk.Frame(path_frame)
        lawdong_frame.pack(fill="x", pady=5)
        ttk.Label(lawdong_frame, text="법정동 코드 파일:").pack(side="left")
        
        lawdong_path_var = tk.StringVar(value=self.lawdong_path)
        lawdong_entry = ttk.Entry(lawdong_frame, textvariable=lawdong_path_var, width=50)
        lawdong_entry.pack(side="left", padx=5)
        
        # 심리 데이터 엑셀 경로
        sentiment_frame = ttk.Frame(path_frame)
        sentiment_frame.pack(fill="x", pady=5)
        ttk.Label(sentiment_frame, text="심리 데이터 엑셀:").pack(side="left")
        
        sentiment_path = getattr(self, 'sentiment_excel_path', '')
        sentiment_path_var = tk.StringVar(value=sentiment_path)
        sentiment_entry = ttk.Entry(sentiment_frame, textvariable=sentiment_path_var, width=50)
        sentiment_entry.pack(side="left", padx=5)
    
        def select_download_path():
            settings.attributes('-topmost', False)
            path = filedialog.askdirectory(initialdir=self.download_path)
            settings.attributes('-topmost', True)
            if path:
                download_path_var.set(path)
                    
        def select_lawdong_path():
            settings.attributes('-topmost', False)
            path = filedialog.askopenfilename(
                initialdir=os.path.dirname(lawdong_path_var.get()),
                title="법정동 코드 파일 선택",
                filetypes=[("Text files", "*.txt")]
            )
            settings.attributes('-topmost', True)
            if path:
                lawdong_path_var.set(path)
                
        # 심리 데이터 엑셀 선택 함수
        def select_sentiment_path():
            settings.attributes('-topmost', False)
            path = filedialog.askopenfilename(
                initialdir=os.path.dirname(sentiment_path_var.get()) if sentiment_path else os.getcwd(),
                title="심리 데이터 엑셀 파일 선택",
                filetypes=[("Excel files", "*.xlsx")]
            )
            settings.attributes('-topmost', True)
            if path:
                sentiment_path_var.set(path)
    
        ttk.Button(download_frame, text="찾아보기", command=select_download_path).pack(side="left")
        ttk.Button(lawdong_frame, text="찾아보기", command=select_lawdong_path).pack(side="left")
        ttk.Button(sentiment_frame, text="찾아보기", command=select_sentiment_path).pack(side="left")
      
        # 저장 함수 
        def save_settings_wrapper():
            try:
                # 다운로드 경로 설정 및 생성
                download_path = download_path_var.get()
                os.makedirs(download_path, exist_ok=True)
                
                # history 폴더는 다운로드 경로 하위에 자동 생성
                history_path = os.path.join(download_path, "history")
                os.makedirs(history_path, exist_ok=True)
                
                # 법정동 코드 파일 존재 확인
                lawdong_path = lawdong_path_var.get()
                if not os.path.exists(lawdong_path):
                    raise FileNotFoundError("법정동 코드 파일을 찾을 수 없습니다.")
                    
                # 심리 데이터 엑셀 파일 확인
                sentiment_path = sentiment_path_var.get()
                if sentiment_path and not os.path.exists(sentiment_path):
                    raise FileNotFoundError("심리 데이터 엑셀 파일을 찾을 수 없습니다.")
                
                # 경로 변경
                self.download_path = download_path
                self.history_path = history_path
                self.lawdong_path = lawdong_path
                self.sentiment_excel_path = sentiment_path
                
                # 설정 저장
                settings_data = {
                    'download_path': self.download_path,
                    'history_path': self.history_path,
                    'lawdong_path': self.lawdong_path,
                    'sentiment_excel_path': self.sentiment_excel_path
                }
                
                with open('settings.json', 'w', encoding='utf-8') as f:
                    json.dump(settings_data, f, ensure_ascii=False, indent=2)
                    
                # GUI 업데이트
                self.history_list = self.load_history()
                self.update_history_display()
                
                # 성공 메시지
                success_dialog = tk.Toplevel(settings)
                success_dialog.title("알림")
                success_dialog.attributes('-topmost', True)
                success_dialog.transient(settings)
                
                dialog_width = 250
                dialog_height = 115
                dialog_x = settings.winfo_x() + (settings.winfo_width() - dialog_width) // 2
                dialog_y = settings.winfo_y() + (settings.winfo_height() - dialog_height) // 2
                success_dialog.geometry(f"{dialog_width}x{dialog_height}+{dialog_x}+{dialog_y}")
                
                ttk.Label(success_dialog, text="설정이 저장되었습니다.", padding=20).pack()
                
                def close_dialogs():
                    success_dialog.destroy()
                    settings.destroy()
                    
                ttk.Button(success_dialog, text="확인", command=close_dialogs).pack(pady=10)
                
                success_dialog.grab_set()
                
            except Exception as e:
                error_dialog = tk.Toplevel(settings)
                error_dialog.title("오류")
                error_dialog.attributes('-topmost', True)
                error_dialog.transient(settings)
                
                dialog_width = 300
                dialog_height = 150
                dialog_x = settings.winfo_x() + (settings.winfo_width() - dialog_width) // 2
                dialog_y = settings.winfo_y() + (settings.winfo_height() - dialog_height) // 2
                error_dialog.geometry(f"{dialog_width}x{dialog_height}+{dialog_x}+{dialog_y}")
                
                ttk.Label(error_dialog, 
                         text=f"설정 저장 중 오류 발생:\n{str(e)}", 
                         padding=20, 
                         wraplength=250).pack()
                ttk.Button(error_dialog, text="확인", command=error_dialog.destroy).pack(pady=10)
                
                error_dialog.grab_set()
        
        # 버튼 프레임
        button_frame = ttk.Frame(settings)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="저장", command=save_settings_wrapper).pack(side="left", padx=5)
        ttk.Button(button_frame, text="취소", command=settings.destroy).pack(side="left", padx=5)

    def open_download_folder(self):
        """저장 폴더 열기"""
        try:
            import os
            import platform
            
            # OS별 폴더 열기 명령어 실행
            if platform.system() == "Windows":
                os.startfile(self.download_path)
            elif platform.system() == "Darwin":  # macOS
                os.system(f"open {self.download_path}")
            else:  # Linux
                os.system(f"xdg-open {self.download_path}")
                
        except Exception as e:
            messagebox.showerror("오류", f"폴더를 여는 중 오류가 발생했습니다: {str(e)}")
    
    # 삭제 관련 메서드 추가
    def delete_selected_history(self):
        selection = self.history_tree.selection()
        if not selection:
            error_dialog = tk.Toplevel(self.root)
            error_dialog.title("알림")
            error_dialog.attributes('-topmost', True)
            error_dialog.transient(self.root)
            
            ttk.Label(error_dialog, text="삭제할 항목을 선택해주세요.", 
                     padding=20, font=self.font_normal).pack()
            ttk.Button(error_dialog, text="확인", command=error_dialog.destroy).pack(pady=10)
            error_dialog.grab_set()
            return
    
        confirm_dialog = tk.Toplevel(self.root)
        confirm_dialog.title("확인")
        confirm_dialog.attributes('-topmost', True)
        confirm_dialog.transient(self.root)
        
        ttk.Label(confirm_dialog, text="선택한 히스토리를 삭제하시겠습니까?\n(엑셀 파일과 그래프 이미지가 모두 삭제됩니다)", 
                 padding=20, font=self.font_normal).pack()
    
        def confirm_delete():
            try:
                for item in selection:
                    idx = self.history_tree.index(item)
                    history_item = self.history_list[idx]
                    
                    # 비교분석인 경우
                    if '[비교]' in history_item['apt_name']:
                        # 히스토리 폴더의 비교분석 이미지 파일 삭제
                        if os.path.exists(history_item['file_path']):
                            os.remove(history_item['file_path'])
                        
                        # 다운로드 폴더의 동일 이미지 파일 삭제
                        download_image_path = os.path.join(self.download_path, 
                            os.path.basename(history_item['file_path']))
                        if os.path.exists(download_image_path):
                            os.remove(download_image_path)
        
                    else:  # 단일 분석인 경우 - 기존 코드 유지
                        history_path = history_item['file_path']
                        apt_name = history_item['apt_name']
                        area = history_item['area']
                        
                        # 면적 정보 처리 (기존 코드 유지)
                        try:
                            if '/' in area:
                                area = area.split('/')[-1].strip()
                            area = area.replace('m²', '').replace('m2', '').strip()
                            area = str(int(float(''.join(c for c in area if c.isdigit() or c == '.'))))
                        except:
                            area = ''.join(c for c in area if c.isdigit())
                        
                        apt_name_clean = ''.join(char for char in apt_name if char.isalnum() or char.isspace())
                        apt_name_clean = apt_name_clean.replace(' ', '_')
                        
                        excel_filename = f"{apt_name_clean}_{area}m2.xlsx"
                        graph_filename = f"{apt_name_clean}_{area}m2.jpg"
                        
                        excel_path = os.path.join(self.download_path, excel_filename)
                        graph_path = os.path.join(self.download_path, graph_filename)
                        
                        # 파일 삭제
                        if os.path.exists(history_path):
                            os.remove(history_path)
                        if os.path.exists(excel_path):
                            os.remove(excel_path)
                        if os.path.exists(graph_path):
                            os.remove(graph_path)
        
                # 히스토리 목록 갱신
                self.history_list = self.load_history()
                self.update_history_display()
        
                # 성공 다이얼로그 (기존 코드 유지)
                success_dialog = tk.Toplevel(confirm_dialog)
                success_dialog.title("알림")
                success_dialog.attributes('-topmost', True)
                success_dialog.transient(confirm_dialog)
                
                ttk.Label(success_dialog, text="선택한 히스토리가 삭제되었습니다.", 
                         padding=20, font=self.font_normal).pack()
                ttk.Button(success_dialog, text="확인", 
                          command=lambda: [success_dialog.destroy(), confirm_dialog.destroy()]).pack(pady=10)
                success_dialog.grab_set()
    
            except Exception as e:
                error_dialog = tk.Toplevel(confirm_dialog)
                error_dialog.title("오류")
                error_dialog.attributes('-topmost', True)
                error_dialog.transient(confirm_dialog)
                
                ttk.Label(error_dialog, text=f"히스토리 삭제 중 오류가 발생했습니다:\n{str(e)}", 
                         padding=20, font=self.font_normal).pack()
                ttk.Button(error_dialog, text="확인", command=error_dialog.destroy).pack(pady=10)
                error_dialog.grab_set()
    
        button_frame = ttk.Frame(confirm_dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="확인", command=confirm_delete).pack(side='left', padx=5)
        ttk.Button(button_frame, text="취소", command=confirm_dialog.destroy).pack(side='left', padx=5)
        
        confirm_dialog.grab_set()

    def delete_all_history(self):
        if not self.history_list:
            error_dialog = tk.Toplevel(self.root)
            error_dialog.title("알림")
            error_dialog.attributes('-topmost', True)
            error_dialog.transient(self.root)
            
            # 기본 폰트 설정
            try:
                font_style = self.font_normal
            except AttributeError:
                font_style = ('Helvetica', 9)
            
            ttk.Label(error_dialog, text="삭제할 히스토리가 없습니다.", 
                     padding=20, font=font_style).pack()
            ttk.Button(error_dialog, text="확인", 
                      command=error_dialog.destroy).pack(pady=10)
            error_dialog.grab_set()
            return
        
        confirm_dialog = tk.Toplevel(self.root)
        confirm_dialog.title("확인")
        confirm_dialog.attributes('-topmost', True)
        confirm_dialog.transient(self.root)
        
        ttk.Label(confirm_dialog, 
                 text="모든 히스토리와 관련 파일을 삭제하시겠습니까?\n이 작업은 되돌릴 수 없습니다.", 
                 padding=20, font=self.font_normal).pack()
    
        def confirm_delete():
            try:
                # 히스토리 폴더의 모든 파일 삭제
                for file in os.listdir(self.history_path):
                    file_path = os.path.join(self.history_path, file)
                    if os.path.exists(file_path):
                        os.remove(file_path)
        
                # 다운로드 폴더의 파일 삭제
                for file in os.listdir(self.download_path):
                    # 비교분석 파일 또는 일반 분석 파일 체크
                    if (file.startswith('비교_') or 
                        (file.endswith('.xlsx') or file.endswith('.jpg'))):
                        file_path = os.path.join(self.download_path, file)
                        if os.path.exists(file_path):
                            os.remove(file_path)
        
                # 히스토리 목록 초기화 (기존 코드 유지)
                self.history_list = []
                self.update_history_display()
        
                # 성공 다이얼로그 (기존 코드 유지)
                success_dialog = tk.Toplevel(confirm_dialog)
                success_dialog.title("알림")
                success_dialog.attributes('-topmost', True)
                success_dialog.transient(confirm_dialog)
                
                ttk.Label(success_dialog, text="모든 히스토리가 삭제되었습니다.", 
                         padding=20, font=self.font_normal).pack()
                ttk.Button(success_dialog, text="확인", 
                          command=lambda: [success_dialog.destroy(), confirm_dialog.destroy()]).pack(pady=10)
                success_dialog.grab_set()
    
            except Exception as e:
                error_dialog = tk.Toplevel(confirm_dialog)
                error_dialog.title("오류")
                error_dialog.attributes('-topmost', True)
                error_dialog.transient(confirm_dialog)
                
                ttk.Label(error_dialog, text=f"히스토리 삭제 중 오류가 발생했습니다:\n{str(e)}", 
                         padding=20, font=self.font_normal).pack()
                ttk.Button(error_dialog, text="확인", command=error_dialog.destroy).pack(pady=10)
                error_dialog.grab_set()
    
        button_frame = ttk.Frame(confirm_dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="확인", command=confirm_delete).pack(side='left', padx=5)
        ttk.Button(button_frame, text="취소", command=confirm_dialog.destroy).pack(side='left', padx=5)
        
        confirm_dialog.grab_set()
    
    def update_history_display(self):
        # 기존 항목 삭제
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        
        # 최신 순으로 정렬 (timestamp 기준 내림차순)
        sorted_history = sorted(self.history_list, key=lambda x: x['search_date'], reverse=True)   # reverse=False로 변경
        
        # 새로운 항목 추가
        for item in sorted_history:
            search_date = datetime.fromtimestamp(item['search_date'])
            self.history_tree.insert("", "end", values=(
                search_date.strftime("%Y-%m-%d %H:%M"),
                item['apt_name'],
                item['area'],
                item['max_trade']
            ))

            
    def get_graph_filename(self, apt_name, area):
        """아파트명과 면적으로 그래프 파일명 생성"""
        # 특수문자 제거 및 공백을 언더스코어로 변경
        apt_name = ''.join(char for char in apt_name if char.isalnum() or char.isspace())
        apt_name = apt_name.replace(' ', '_')
        
        # 면적에서 소수점 제거 (예: 84.51 -> 84)
        try:
            # 숫자 추출 및 정수화
            area = ''.join(c for c in area if c.isdigit() or c == '.')
            area = str(int(float(area)))
        except:
            area = area.split('.')[0]  # 변환 실패시 소수점 앞자리만 사용
        
        return f"{apt_name}_{area}m2.jpg"

        
    def on_history_select(self, event):
        selection = self.history_tree.selection()
        if not selection:
            return
                
        self.history_list = self.load_history()
        idx = self.history_tree.index(selection[0])
        item = self.history_list[idx]
        
        try:
            if '[비교]' in item['apt_name']:  # 비교분석인 경우
                # 기존 비교분석 코드 그대로 유지
                try:
                    # 비교 분석 이미지는 항상 history 폴더에 저장됨
                    # 엑셀 파일이름에서 파일명 생성 (history_compare_YYYYMMDD_HHMMSS.xlsx -> 비교_...)
                    excel_name = os.path.basename(item['file_path'])
                    timestamp = excel_name.replace('history_compare_', '').replace('.xlsx', '')
                    
                    # 비교 정보에서 아파트 이름과 비교 유형 추출
                    apt_info = item['apt_name'].replace('[비교] ', '')
                    compare_type = item['area']  # 매매가 또는 전세가
                    
                    # 이미지 파일 찾기 - 더 유연한 검색 방법으로 수정
                    found = False
                    for folder in [self.download_path, self.history_path]:
                        files = os.listdir(folder)
                        # 준공년도 포함된 형식과 포함되지 않은 형식 모두 검색
                        for file in files:
                            # 비교_ 로 시작하고 timestamp가 포함된 모든 jpg 파일 검색
                            if (file.startswith('비교_') and 
                                file.endswith('.jpg') and 
                                timestamp in file):
                                image_path = os.path.join(folder, file)
                                if os.path.exists(image_path):
                                    os.startfile(image_path)
                                    found = True
                                    break
                            # 또는 단지명이 포함된 jpg 파일 검색 (더 느슨한 검색)
                            elif (file.startswith('비교_') and 
                                  file.endswith('.jpg') and
                                  all(apt in file for apt in apt_info.split(' vs '))):
                                image_path = os.path.join(folder, file)
                                if os.path.exists(image_path):
                                    os.startfile(image_path)
                                    found = True
                                    break
                        if found:
                            break
                    
                    # 비교분석 이미지 못 찾을 경우 처리 (기존 코드 그대로)
                    if not found:
                        # 파일을 찾을 수 없는 경우 시간정보만 사용하여 다시 검색
                        for folder in [self.download_path, self.history_path]:
                            files = os.listdir(folder)
                            for file in files:
                                if (file.startswith('비교_') and 
                                    file.endswith('.jpg') and
                                    timestamp[:8] in file):  # 날짜만 비교
                                    image_path = os.path.join(folder, file)
                                    if os.path.exists(image_path):
                                        os.startfile(image_path)
                                        found = True
                                        break
                            if found:
                                break
                    
                    if not found:
                        # 그래도 못 찾으면 비교 구조체를 만들어 새로 그려준다
                        messagebox.showinfo("파일 재생성", "원본 비교 그래프를 찾을 수 없어 새로 생성합니다.")
                        try:
                            # 엑셀에서 정보 읽기
                            wb = load_workbook(item['file_path'])
                            ws = wb.active
                            apt1_name = ws['B3'].value
                            apt2_name = ws['B4'].value
                            compare_type = ws['B2'].value
                            
                            # 단지 데이터 확인
                            key1 = f"{apt1_name}_{self.area.get()}"
                            key2 = f"{apt2_name}_{self.area.get()}"
                            
                            if key1 in self.downloaded_data and key2 in self.downloaded_data:
                                df1 = self.downloaded_data[key1]
                                df2 = self.downloaded_data[key2]
                                apt1_year = self.downloaded_data.get(f"{key1}_year")
                                apt2_year = self.downloaded_data.get(f"{key2}_year")
                                
                                # 그래프 새로 생성
                                self.create_comparison_graph(df1, df2, apt1_name, apt2_name, 
                                                            self.apt1_area.get(), self.apt2_area.get(), 
                                                            compare_type, apt1_year, apt2_year)
    
                                if os.path.exists(self.image_path):
                                    os.startfile(self.image_path)
                            else:
                                messagebox.showwarning("알림", 
                                                     "비교 단지 데이터가 메모리에 없습니다. 두 단지를 다시 검색해주세요.")
                        except Exception as e:
                            messagebox.showerror("오류", f"비교 그래프 재생성 중 오류: {str(e)}")
                        
                except Exception as e:
                    logging.error(f"비교분석 파일 처리 중 오류: {str(e)}")
                    messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다: {str(e)}")
                    
            else:  # 일반 분석인 경우 - 설정에 따른 파일 찾기를 변경
                try:
                    df_info = pd.read_excel(item['file_path'], nrows=8)
                    apt_name = str(df_info.iloc[2, 1])
                    area = str(df_info.iloc[5, 1])
                    
                    if '/' in area:
                        area = area.split('/')[-1].strip()
                    area = area.replace('m²', '').replace('m2', '').strip()
                    area = str(int(float(''.join(c for c in area if c.isdigit() or c == '.'))))
                    
                    apt_name_clean = ''.join(char for char in apt_name if char.isalnum() or char.isspace())
                    apt_name_clean = apt_name_clean.replace(' ', '_')
                    
                    # 여기가 수정되는 부분: 현재 설정에 맞는 파일명으로 그래프 찾기
                    pir_suffix = "pir" if self.show_pir.get() else "no_pir"
                    una_suffix = "una" if self.show_una_sentiment.get() else "no_una"
                    settings_suffix = f"{pir_suffix}_{una_suffix}"
                    
                    # 현재 설정에 맞는 그래프 파일명
                    current_graph_filename = f"{apt_name_clean}_{area}m2_{settings_suffix}.jpg"
                    
                    # 먼저 현재 설정에 맞는 파일 찾기
                    found = False
                    for folder in [self.download_path, self.history_path]:
                        graph_path = os.path.join(folder, current_graph_filename)
                        if os.path.exists(graph_path):
                            os.startfile(graph_path)
                            found = True
                            break
                    
                    # 현재 설정 파일이 없으면 다른 모든 설정 파일 찾기
                    if not found:
                        # 모든 가능한 설정 조합
                        all_settings = [
                            "pir_una", "pir_no_una", 
                            "no_pir_una", "no_pir_no_una"
                        ]
                        
                        # 설정 조합별 파일 찾기
                        for setting in all_settings:
                            alt_filename = f"{apt_name_clean}_{area}m2_{setting}.jpg"
                            for folder in [self.download_path, self.history_path]:
                                graph_path = os.path.join(folder, alt_filename)
                                if os.path.exists(graph_path):
                                    # 파일 발견: 사용자에게 물어보기
                                    if messagebox.askyesno("다른 설정의 그래프 발견", 
                                                         f"현재 설정과 다른 그래프 파일이 발견되었습니다.\n"
                                                         f"설정: {setting}\n"
                                                         f"이 파일을 열겠습니까? '아니오'를 선택하면 새 그래프를 생성합니다."):
                                        os.startfile(graph_path)
                                        found = True
                                        break
                                    else:
                                        # 새 그래프 생성
                                        df = self.analyze_prices(item['file_path'])
                                        if os.path.exists(self.image_path):
                                            os.startfile(self.image_path)
                                        found = True
                                        break
                            if found:
                                break
                    
                    # 어떤 파일도 없으면 새로 그래프 생성
                    if not found:
                        confirmation = messagebox.askyesno(
                            "그래프 파일 없음", 
                            "그래프 파일을 찾을 수 없습니다. 새로 생성하시겠습니까?"
                        )
                        if confirmation:
                            # 메모리에 있는 데이터 지우기 (강제로 다시 계산하도록)
                            key = f"{apt_name}_{area}"
                            if key in self.downloaded_data:
                                del self.downloaded_data[key]
                            
                            # 분석 다시 실행
                            df = self.analyze_prices(item['file_path'])
                            if os.path.exists(self.image_path):
                                os.startfile(self.image_path)
                            
                except Exception as e:
                    logging.error(f"일반분석 파일 처리 중 오류: {str(e)}")
                    messagebox.showerror("오류", f"파일 처리 중 오류가 발생했습니다: {str(e)}")
                    
        except Exception as e:
            logging.error(f"히스토리 항목 열기 실패: {str(e)}")
            messagebox.showerror("오류", f"파일 열기 실패: {str(e)}")
    
    
            
    def show_saved_graph(self, graph_path):
        """저장된 그래프 이미지 표시"""
        if os.path.exists(graph_path):
            img = Image.open(graph_path)
            
            width = 1280
            height = int(width * 9/16)
            
            popup = tk.Toplevel()
            popup.title("KB부동산 시세 분석 결과")
            
            screen_width = popup.winfo_screenwidth()
            screen_height = popup.winfo_screenheight()
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            popup.geometry(f"{width}x{height}+{x}+{y}")
            
            img = img.resize((width, height), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            label = tk.Label(popup, image=photo)
            label.image = photo
            label.pack(fill='both', expand=True)

    def show_apt_list(self):
        sido = self.sido_combobox.get()
        sigungu = self.sigungu_combobox.get()
        dong = self.dong_combobox.get()
        
        if "선택" in [sido, sigungu, dong]:
            messagebox.showerror("오류", "지역을 모두 선택해주세요.")
            return
            
        region_code = self.region_codes.get((sido, sigungu, dong))
        if not region_code:
            messagebox.showerror("오류", "해당 지역의 코드를 찾을 수 없습니다.")
            return
    
        try:
            self.search_button.config(state="disabled")
            self.update_progress(20, "아파트 목록 검색 중...")
            
            apt_list = self.get_apt_list_from_api(region_code[1], dong)
            if apt_list:
                dialog = AptSelectDialog(
                    self.root, 
                    apt_list,
                    self.service_key,
                    region_code[1],
                    dong,
                    title=f"{dong} 아파트 목록"
                )
                self.root.wait_window(dialog.top)
                
                if dialog.result:
                    jibun_addr, apt_name, area, trades = dialog.result
                    self.apt_name.delete(0, tk.END)
                    self.apt_name.insert(0, apt_name)
                    self.area.delete(0, tk.END)
                    self.area.insert(0, area)
                    self.trades_data = trades
                    self.search_address = jibun_addr
                    # browser_search 직접 호출 대신 start_search 호출
                    self.start_search()
            else:
                messagebox.showinfo("알림", f"{dong}에 거래 내역이 있는 아파트가 없습니다.")
        finally:
            self.search_button.config(state="normal")



    def on_sido_selected(self, event):
        sido = self.sido_combobox.get()
        if sido in self.sigungu_dict:
            self.sigungu_combobox['values'] = sorted(self.sigungu_dict[sido])
            self.sigungu_combobox.set("시/군/구 선택")
            self.dong_combobox.set("읍/면/동 선택")
        
    def on_sigungu_selected(self, event):
        sigungu = self.sigungu_combobox.get()
        if sigungu in self.dong_dict:
            self.dong_combobox['values'] = sorted(self.dong_dict[sigungu])
            self.dong_combobox.set("읍/면/동 선택")
        
    def on_dong_selected(self, event):
        sido = self.sido_combobox.get()
        sigungu = self.sigungu_combobox.get()
        dong = self.dong_combobox.get()
        
        if "선택" not in [sido, sigungu, dong]:
            region_code = self.region_codes.get((sido, sigungu, dong))
            if region_code:
                _, sigungu_code = region_code
                apt_list = self.get_apt_list_from_api(sigungu_code, dong)
                if apt_list:
                    self.status_label.config(text=f"{dong}의 아파트 {len(apt_list)}개가 검색되었습니다.")
                else:
                    self.status_label.config(text=f"{dong}의 검색 가능한 아파트가 없습니다.")


    def get_apt_list_from_api(self, sigungu_code, dong):
       apt_info = {}
       current_date = datetime.now()
       
       logging.info(f"\n=== {dong} 아파트 목록 검색 시작 ===")
       
       for i in range(3):  # 최근 3개월 데이터 검색
           search_date = current_date - timedelta(days=30*i)
           deal_ymd = search_date.strftime("%Y%m")
           
           url = (f"http://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
                  f"?serviceKey={self.service_key}"
                  f"&LAWD_CD={sigungu_code}"
                  f"&DEAL_YMD={deal_ymd}"
                  f"&numOfRows=1000")
           
           try:
               response = requests.get(url)
               if response.status_code == 200:
                   root = ET.fromstring(response.text)
                   items = root.findall('.//item')
                   
                   for item in items:
                       item_dong = item.findtext('umdNm', '').strip()
                       if item_dong == dong:
                           apt_name = item.findtext('aptNm', '').strip()
                           logging.info(f"발견된 아파트: {apt_name}")
                           
                           if apt_name and apt_name not in apt_info:
                               jibun = item.findtext('jibun', '').strip()
                               jibun_addr = f"{dong} {jibun}"
                               
                               road = item.findtext('roadName', '').strip()
                               road_main = item.findtext('roadNameBonbun', '').strip()
                               road_sub = item.findtext('roadNameBubun', '').strip()
                               road_addr = f"{road} {road_main}"
                               if road_sub:
                                   road_addr += f"-{road_sub}"
                               
                               apt_info[apt_name] = {
                                   'jibun_addr': jibun_addr,
                                   'road_addr': road_addr
                               }
                               logging.info(f"주소정보: {jibun_addr} / {road_addr}")
    
           except Exception as e:
               logging.error(f"API 호출 중 오류: {str(e)}")
               continue
    
       apt_list = [f"{apt_name} [{info['road_addr']} / {info['jibun_addr']}]" 
                   for apt_name, info in sorted(apt_info.items())]
    
       logging.info(f"총 {len(apt_list)}개 아파트 발견")
       for apt in apt_list:
           logging.info(f"- {apt}")
    
       return apt_list
    


            





    def update_progress(self, value, message=""):
        self.progress["value"] = value
        if message:
            self.status_label.config(text=message)
        self.root.update_idletasks()



    def get_trade_data(self, sido, sigungu, dong, apt_name, target_area):
        _, sigungu_code = self.region_codes[(sido, sigungu, dong)]
        logging.info(f"\n=== 실거래가 데이터 요청 시작 ===")
        
        apt_list = self.get_apt_list_from_api(sigungu_code, dong)
        found = False
        
        for apt in apt_list:
            if apt_name == apt.split('[')[0].strip():
                found = True
                # 선택된 단지 정보 저장
                self.selected_apt_info = {
                    'sido': sido,
                    'sigungu': sigungu,
                    'dong': dong,
                    'apt_name': apt_name,
                    'sigungu_code': sigungu_code
                }
                break
        
        if not found:
            dialog = AptSelectDialog(
                self.root, 
                apt_list,
                self.service_key,
                sigungu_code,
                dong,
                title=f"{dong} 아파트 목록"
            )
            self.root.wait_window(dialog.top)
            if dialog.result:
                jibun_addr, selected_apt_name, area, trades = dialog.result
                apt_name = selected_apt_name
                # 선택된 단지 정보 저장
                self.selected_apt_info = {
                    'sido': sido,
                    'sigungu': sigungu,
                    'dong': dong,
                    'apt_name': selected_apt_name,
                    'sigungu_code': sigungu_code
                }
        
        return self.get_trade_details(sigungu_code, dong, apt_name, target_area)


    def get_trade_details(self, sigungu_code, dong, apt_name, target_area):
        trades = []
        current_date = datetime.now()
        
        # 데이터 수집 기간 설정 (5년)
        years_to_fetch = 20
        months_to_fetch = years_to_fetch * 12
        
        for i in range(months_to_fetch):  # 60개월(5년) 동안 반복
            search_date = current_date - timedelta(days=30*i)
            deal_ymd = search_date.strftime("%Y%m")
            url = (f"http://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
                   f"?serviceKey={self.service_key}"
                   f"&LAWD_CD={sigungu_code}"
                   f"&DEAL_YMD={deal_ymd}"
                   f"&numOfRows=1000")
            
            response = requests.get(url)
            if response.status_code == 200:
                root = ET.fromstring(response.text)
                items = root.findall('.//item')
                
                for item in items:
                    if (item.findtext('umdNm') == dong and 
                        item.findtext('aptNm').strip() == apt_name):
                        area = float(item.findtext('excluUseAr', '0'))
                        if abs(int(area) - target_area) <= 1:
                            # dealDay 필드 추가로 읽어오기
                            trade = {
                                'date': datetime(
                                    int(item.findtext('dealYear')), 
                                    int(item.findtext('dealMonth')), 
                                    int(item.findtext('dealDay', '1'))  # dealDay 추가
                                ),
                                'price': int(item.findtext('dealAmount').replace(',', '')),
                                'floor': int(item.findtext('floor'))
                            }
                            trades.append(trade)
        return trades

    def get_lease_data(self, sido, sigungu, dong, apt_name, target_area):
        print("\n=== 전세 실거래가 데이터 조회 시작 ===")
        
        # 이미 선택된 단지 정보가 있는지 확인
        if hasattr(self, 'selected_apt_info') and self.selected_apt_info:
            # 매매 데이터 조회 시 선택한 단지 정보 사용
            apt_info = self.selected_apt_info
            sigungu_code = apt_info['sigungu_code']
            dong = apt_info['dong']
            apt_name = apt_info['apt_name']
            print(f"매매 조회에서 선택한 단지 정보 사용: {apt_name}")
        else:
            # 기존 방식대로 처리
            _, sigungu_code = self.region_codes[(sido, sigungu, dong)]
        
        print(f"조회 정보: {sido} {sigungu} {dong}")
        print(f"단지명: {apt_name}")
        print(f"전용면적: {target_area}")
        print(f"법정동코드: {sigungu_code}")
        
        # 기존 전세 데이터 조회 코드
        current_date = datetime.now()
        leases = []
    
        # 데이터 수집 기간 설정 (5년)
        years_to_fetch = 20
        months_to_fetch = years_to_fetch * 12
        
        for i in range(months_to_fetch):  # 60개월(5년) 동안 반복
            search_date = current_date - timedelta(days=30*i)
            deal_ymd = search_date.strftime("%Y%m")
            
            url = (f"http://apis.data.go.kr/1613000/RTMSDataSvcAptRent/getRTMSDataSvcAptRent"
                   f"?serviceKey={self.service_key}"
                   f"&LAWD_CD={sigungu_code}"
                   f"&DEAL_YMD={deal_ymd}"
                   f"&numOfRows=1000")
            
            try:
                response = requests.get(url)
                if response.status_code == 200:
                    root = ET.fromstring(response.text)
                    items = root.findall('.//item')
                    
                    for item in items:
                        if (item.findtext('umdNm') == dong and 
                            item.findtext('aptNm').strip() == apt_name):
                            area = float(item.findtext('excluUseAr', '0'))
                            
                            # 월세 체크 - monthlyRent가 0이 아닌 경우 제외
                            monthly_rent = item.findtext('monthlyRent', '0')
                            try:
                                monthly_rent = int(monthly_rent.replace(',', ''))
                                if monthly_rent > 0:
                                    continue
                            except:
                                continue
                                
                            if abs(int(area) - target_area) <= 1:
                                try:
                                    deposit = item.findtext('deposit', '0')  # 보증금
                                    price = int(deposit.replace(',', ''))
                                    
                                    if price > 0:  # 가격이 0보다 큰 경우만 처리
                                        lease = {
                                            'date': datetime(
                                                int(item.findtext('dealYear')), 
                                                int(item.findtext('dealMonth')), 
                                                int(item.findtext('dealDay', '1'))
                                            ),
                                            'price': price,
                                            'floor': int(item.findtext('floor', '0'))
                                        }
                                        leases.append(lease)
                                except ValueError as e:
                                    print(f"가격 변환 오류: {str(e)}")
                                    continue
    
            except Exception as e:
                print(f"데이터 처리 중 오류: {str(e)}")
                continue
    
        print(f"\n총 수집된 전세 실거래 건수: {len(leases)}")
        return leases


    
    def analyze_prices(self, file_path=None):
        # 변수 초기화 (메서드 시작 부분에 추가)
        sido = ""
        sigungu = ""
        dong = ""
        apt_name = ""
        target_area = None
        type_households = None  # 해당 타입 세대수 변수 추가
        
        # 기존 폰트 설정을 custom 폰트 설정으로 대체
        font_path = os.path.join(os.path.dirname(self.lawdong_path), "KoPubWorld Dotum Medium.ttf")
        plt.rcParams['axes.prop_cycle'] = plt.cycler(color=['#0066CC', '#FF6B00', '#AADD00', '#008800', 'blue', 'red'])
       
        try:
            if file_path is None:
                file_path = self.get_latest_excel()
            
            print(f"\n=== KB시세 데이터 읽기 시작 ===")
            print(f"파일 경로: {file_path}")
            
            # KB시세 데이터 읽기 - 원본 데이터 보존
            original_df = pd.read_excel(file_path, skiprows=14)
            
            # 선택된 가격 유형 확인
            sale_types = [t for t, var in self.sale_price_types.items() if var.get()]
            lease_types = [t for t, var in self.lease_price_types.items() if var.get()]
            
            print("\n=== 가격 데이터 선택 정보 ===")
            print(f"선택된 매매가 유형: {sale_types}")
            print(f"선택된 전세가 유형: {lease_types}")
            
            # 열 매핑 정보
            sale_column_map = {
                "low": 1,     # B열 (하위 평균 매매가)
                "normal": 2,  # C열 (일반 평균 매매가) 
                "high": 3     # D열 (상위 평균 매매가)
            }
            
            lease_column_map = {
                "low": 4,     # E열 (하위 평균 전세가)
                "normal": 5,  # F열 (일반 평균 전세가)
                "high": 6     # G열 (상위 평균 전세가)
            }
            
            # 가격 유형 매핑 정보 (레이블 생성용)
            price_type_labels = {
                "low": "하위평균",
                "normal": "일반평균",
                "high": "상위평균"
            }
            
            # 색상 매핑 정보
            sale_colors = {
                "low": "#3399FF",   # 연한 파란색
                "normal": "#0066CC", # 표준 파란색
                "high": "#003366"   # 진한 파란색
            }
            
            lease_colors = {
                "low": "#FFAA80",   # 연한 주황색
                "normal": "#FF6B00", # 표준 주황색
                "high": "#CC5500"   # 진한 주황색
            }
            
            # 선택된 열 인덱스 가져오기
            sale_column_indices = [sale_column_map[t] for t in sale_types]
            lease_column_indices = [lease_column_map[t] for t in lease_types]
            
            print(f"\n=== 선택된 열 인덱스 정보 ===")
            print(f"매매가 열 인덱스: {sale_column_indices}")
            print(f"전세가 열 인덱스: {lease_column_indices}")
            
            # 새로운 데이터프레임 생성
            df = pd.DataFrame()
            
            # 날짜 열 추가
            df['date'] = original_df.iloc[:, 0]  # 첫 번째 열(A열)은 항상 날짜
            df['date'] = pd.to_datetime(df['date'].astype(str).apply(lambda x: f"{x[:4]}-{x[4:6]}-01"))
            
            # 우선순위에 따라 매매가, 전세가 열 덮어쓰기
            sale_types = [t for t, var in self.sale_price_types.items() if var.get()]
            lease_types = [t for t, var in self.lease_price_types.items() if var.get()]
            
            # 우선순위에 따라 정렬 (상위 > 일반 > 하위)
            priority_order = {"high": 0, "normal": 1, "low": 2}
            sale_types = sorted(sale_types, key=lambda x: priority_order[x])
            lease_types = sorted(lease_types, key=lambda x: priority_order[x])
            
            # 매매가, 전세가 열 직접 추가
            for price_type in sale_types:
                col_idx = sale_column_map[price_type]
                col_name = f"매매가_{price_type}"
                df[col_name] = original_df.iloc[:, col_idx]
            
            for price_type in lease_types:
                col_idx = lease_column_map[price_type]
                col_name = f"전세가_{price_type}"
                df[col_name] = original_df.iloc[:, col_idx]
            
            # 우선순위 높은 유형의 데이터로 기본 열 설정
            if sale_types:
                top_sale_type = sale_types[0]
                df['매매가'] = df[f'매매가_{top_sale_type}'].copy()
                print(f"매매가 열을 우선순위가 가장 높은 {top_sale_type} 유형으로 설정")
                sale_col_name = f'매매가_{top_sale_type}'
            else:
                sale_col_name = '매매가'
            
            if lease_types:
                top_lease_type = lease_types[0]
                df['전세가'] = df[f'전세가_{top_lease_type}'].copy()
                print(f"전세가 열을 우선순위가 가장 높은 {top_lease_type} 유형으로 설정")
                lease_col_name = f'전세가_{top_lease_type}'
            else:
                lease_col_name = '전세가'
            
            # 데이터 확인 (디버깅용)
            print("\n=== 데이터프레임 구성 확인 ===")
            print(f"데이터프레임 열: {df.columns.tolist()}")
            
            # KB시세 엑셀 파일명을 그대로 사용하여 그래프 파일명 생성
            excel_filename = os.path.basename(file_path)
            graph_base = excel_filename.replace('.xlsx', '')
            
            # 차트 제목을 위한 정보 읽기
            if file_path and file_path.startswith(self.history_path):
                # 히스토리에서 불러온 경우
                try:
                    df_info = pd.read_excel(file_path, nrows=8)
                    apt_name = str(df_info.iloc[2, 1])
                    area = str(df_info.iloc[5, 1])
                    if '/' in area:
                        area = area.split('/')[-1].strip()
                        area = area.replace('m2', '').strip()
                    area = str(int(float(''.join(c for c in area if c.isdigit() or c == '.'))))
                except Exception as e:
                    print(f"파일 읽기 오류: {str(e)}")
                    return None
            else:
                # 직접 검색한 경우
                apt_name = self.apt_name.get().strip()
                area = self.area.get().strip()
                area = str(int(float(area)))
            
            # 그래프 생성 코드 시작
            plt.rcParams['font.family'] = 'Malgun Gothic'
            plt.rcParams['axes.unicode_minus'] = False
            
            # 단지 정보 읽기
            # 단지 정보 읽기 부분
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                
                wb = excel.Workbooks.Open(file_path)
                sheet = wb.ActiveSheet
                
                # 엑셀에서 정보 추출
                complex_details = {
                    '단지명': str(sheet.Cells(4, 2).Value or '정보없음'),
                    '대표번지': str(sheet.Cells(5, 2).Value or '정보없음'),
                    '공급/전용면적': str(sheet.Cells(7, 2).Value or '정보없음'),
                    '타입': str(sheet.Cells(8, 2).Value or '정보없음'),
                    '세대수': str(sheet.Cells(9, 2).Value or '정보없음'),
                    '방/욕실수': str(sheet.Cells(10, 2).Value or '정보없음'),
                    '현관구조': str(sheet.Cells(11, 2).Value or '정보없음')
                }
                
                # 세대수 추출 로직 개선
                # 세대수 추출 로직 개선 (analyze_prices 메서드 내)
                print("\n=== 세대수 추출 디버그 ===")
                households_info = complex_details['세대수']
                print(f"엑셀에서 읽은 세대수 원본: '{households_info}'")
                
                type_households = None
                
                # 1. 먼저 웹에서 가져온 세대수 확인
                if hasattr(self, 'total_households') and self.total_households:
                    type_households = self.total_households
                    print(f"★ 웹에서 가져온 전체 세대수 사용: {type_households}")
                
                # 2. 웹에서 못 가져왔으면 엑셀에서 추출
                else:
                    import re
                    
                    # "65 세대(총 3002 세대)" 형식을 위한 패턴 추가
                    patterns = [
                        r'총\s*(\d+)\s*세대',  # "총 3002 세대" 부분 우선 매칭
                        r'\(.*?(\d{4,})\s*세대\)',  # 괄호 안의 4자리 이상 숫자
                        r'(\d+)\s*세대.*?\(.*?(\d+)\s*세대\)',  # "65 세대(총 3002 세대)" 전체 패턴
                        r'(\d+)\s*세대\s*/\s*(\d+)\s*세대',  # "162세대/4932세대" 형식
                        r'(\d+)\s*/\s*(\d+)',  # "162/4932" 형식
                        r'전체\s*(\d+)',  # "전체 4932" 형식
                    ]
                    
                    for pattern in patterns:
                        match = re.search(pattern, households_info)
                        if match:
                            if pattern == r'(\d+)\s*세대.*?\(.*?(\d+)\s*세대\)':
                                # "65 세대(총 3002 세대)" 형식인 경우
                                type_h = int(match.group(1))  # 65
                                total = int(match.group(2))   # 3002
                                type_households = total  # 전체 세대수 사용
                                print(f"패턴 매칭: 타입세대={type_h}, 전체세대={total}")
                                print(f"★ 엑셀에서 추출한 전체 세대수 사용: {type_households}")
                                break
                            elif len(match.groups()) >= 2:
                                # "전체/타입" 형식인 경우
                                total = int(match.group(1))
                                type_h = int(match.group(2))
                                # 더 큰 숫자를 전체 세대수로 사용
                                type_households = max(total, type_h)
                                print(f"패턴 '{pattern}' 매칭: 값1={total}, 값2={type_h}")
                                print(f"★ 엑셀에서 추출한 전체 세대수: {type_households}")
                                break
                            else:
                                # 단일 숫자인 경우
                                type_households = int(match.group(1))
                                print(f"패턴 '{pattern}' 매칭: {type_households}")
                                print(f"★ 엑셀에서 추출한 세대수: {type_households}")
                                break
                    
                    if not type_households:
                        print("엑셀에서 세대수 추출 실패")
                
                print(f"=== 최종 사용할 세대수: {type_households} ===\n")
                
                # ... 나머지 코드
                
                # 주소 정보 파싱
                addr = str(sheet.Cells(5, 2).Value or '')
                addr_parts = addr.split()
                sido = ""
                sigungu = ""
                dong = ""
                
                # 주소 파싱 로직
                for i in range(len(addr_parts)-1, -1, -1):
                    if any(addr_parts[i].endswith(suffix) for suffix in ['동', '읍', '면', '가']):
                        dong = addr_parts[i]
                        for j in range(i-1, -1, -1):
                            if addr_parts[j].endswith(('구', '군')):
                                sigungu = addr_parts[j]
                                if addr_parts[0].endswith('도'):
                                    for k in range(j-1, -1, -1):
                                        if addr_parts[k].endswith('시'):
                                            sigungu = addr_parts[k] + ' ' + sigungu
                                            sido = addr_parts[0]
                                            break
                                else:
                                    sido = addr_parts[0]
                                break
                            elif addr_parts[j].endswith('시'):
                                sigungu = addr_parts[j]
                                sido = addr_parts[0]
                                break
                        break
                
                print(f"파싱된 주소: {sido} {sigungu} {dong}")
                
                # 전용면적 추출
                target_area = None
                area_text = str(sheet.Cells(7, 2).Value or '')
                if '/' in area_text:
                    try:
                        area_str = area_text.split('/')[-1].strip()
                        area_str = ''.join(c for c in area_str if c.isdigit() or c == '.')
                        target_area = int(float(area_str))
                    except ValueError:
                        pass
                
                wb.Close(SaveChanges=False)
                excel.Quit()
                
                # 실거래 데이터 가져오기
                apt_name = complex_details['단지명']
                trades = None
                if self.show_real_trade.get():
                    print("매매 실거래가 조회 활성화")
                    if hasattr(self, 'trades_data'):
                        trades = self.trades_data
                        delattr(self, 'trades_data')
                    elif all([sido, sigungu, dong, apt_name]):
                        trades = self.get_trade_data(sido, sigungu, dong, apt_name, target_area)
                else:
                    print("매매 실거래가 조회 비활성화")
                    trades = None
            
            except Exception as e:
                print(f"단지 정보 읽기 오류: {e}")
                complex_details = {'단지명': '정보없음', '대표번지': '정보없음', '공급/전용면적': '정보없음',
                                 '타입': '정보없음', '세대수': '정보없음', '방/욕실수': '정보없음', '현관구조': '정보없음'}
                trades = None
                type_households = None
            
            # 연도별 소득 데이터
            yearly_income = {
                2004: 1150.2, 2005: 1201.0, 2006: 1252.9, 2007: 1327.2, 2008: 1428.1,
                2009: 1460.4, 2010: 1545.3, 2011: 1622.8, 2012: 1683.8, 2013: 1749.5,
                2014: 1826.1, 2015: 1937.4, 2016: 1982.5, 2017: 2037.8, 2018: 2119.7,
                2019: 2205.9, 2020: 2280.0, 2021: 2364.4, 2022: 2494.5, 2023: 2545.3,
                2024: 2647.112, 2025: 2752.99
            }
            
            # PIR 계산
            df['year'] = df['date'].dt.year
            df['yearly_income'] = df['year'].map(yearly_income)
            df['PIR'] = df['매매가'] / (df['yearly_income'])
            
            # PIR 누적평균 계산
            df = df.sort_values('date')
            pir_values = df['PIR'].values
            running_avg = []
            total = 0
            
            for i in range(len(pir_values)):
                total += pir_values[i]
                avg = total / (i + 1)
                running_avg.append(avg)
            
            df['PIR_running_avg'] = running_avg
            
            # 전저점 찾기
            old_low_point = None
            df_2011_2020 = df[(df['date'].dt.year >= 2011) & (df['date'].dt.year <= 2020)]
            
            if not df_2011_2020.empty:
                min_pir_idx = df_2011_2020['PIR'].idxmin()
                min_pir_date = df_2011_2020.loc[min_pir_idx, 'date']
                start_date = min_pir_date - timedelta(days=365)
                end_date = min_pir_date + timedelta(days=365)
                period_mask = (df['date'] >= start_date) & (df['date'] <= end_date)
                period_df = df[period_mask]
                
                if not period_df.empty:
                    old_low_point = period_df['매매가'].idxmin()
            
            logging.info(f"전저점 발견: {old_low_point}")
            if old_low_point:
                logging.info(f"전저점 가격: {df.loc[old_low_point, '매매가']:,.0f}")
                logging.info(f"전저점 날짜: {df.loc[old_low_point, 'date']}")
            
            # 첫 번째로 0이 아닌 매매가 찾기
            first_nonzero_sale_idx = df[df['매매가'] > 0].index[0]
            first_nonzero_sale_date = df.loc[first_nonzero_sale_idx, 'date']
            sale_start = df.loc[first_nonzero_sale_idx, '매매가']
            
            # 첫 번째로 0이 아닌 전세가 찾기
            first_nonzero_lease_idx = df[df['전세가'] > 0].index[0]
            first_nonzero_lease_date = df.loc[first_nonzero_lease_idx, 'date']
            lease_start = df.loc[first_nonzero_lease_idx, '전세가']
            
            # 주요 지표 계산
            latest_date = df['date'].max()
            max_sale_idx = df['매매가'].idxmax()
            max_lease_idx = df['전세가'].idxmax()
            max_pir_idx = df['PIR'].idxmax()
            current_idx = df[df['date'] == latest_date].index[0]
            
            # price_info 계산
            if df.loc[current_idx, sale_col_name] >= df.loc[max_sale_idx, '매매가']:
                price_info = {
                    'max': df.loc[current_idx, sale_col_name],
                    'max_date': latest_date,
                    'current': df.loc[current_idx, sale_col_name],
                    'current_date': latest_date,
                    'recent_low': None,
                    'recent_low_date': None
                }
            else:
                period_df = df.loc[max_sale_idx:current_idx]
                min_sale_idx = period_df['매매가'].idxmin()
                
                if min_sale_idx == current_idx:
                    price_info = {
                        'max': df.loc[max_sale_idx, '매매가'],
                        'max_date': df.loc[max_sale_idx, 'date'],
                        'current': df.loc[current_idx, sale_col_name],
                        'current_date': latest_date,
                        'recent_low': None,
                        'recent_low_date': None
                    }
                else:
                    price_info = {
                        'max': df.loc[max_sale_idx, '매매가'],
                        'max_date': df.loc[max_sale_idx, 'date'],
                        'current': df.loc[current_idx, sale_col_name],
                        'current_date': latest_date,
                        'recent_low': df.loc[min_sale_idx, '매매가'],
                        'recent_low_date': df.loc[min_sale_idx, 'date']
                    }
            
            # 매매가/전세가 연복리 계산
            sale_years = (latest_date - first_nonzero_sale_date).days / 365.25
            sale_end = df.loc[current_idx, sale_col_name]
            sale_cagr = (np.power(sale_end/sale_start, 1/sale_years) - 1) * 100 if sale_start > 0 else 0
            
            lease_years = (latest_date - first_nonzero_lease_date).days / 365.25
            lease_end = df.loc[current_idx, '전세가']
            lease_cagr = (np.power(lease_end/lease_start, 1/lease_years) - 1) * 100 if lease_start > 0 else 0
            
            # 거래 회전율 및 패턴 분석 함수
            # 거래 회전율 및 패턴 분석 함수
            def analyze_trade_patterns(trades, type_households):
                """거래 회전율 및 패턴 분석"""
                print("\n=== 거래 패턴 분석 시작 ===")
                print(f"입력값: 거래건수={len(trades) if trades else 0}, 세대수={type_households}")
                
                trade_info = {
                    'recent_turnover': None,
                    'avg_turnover': None,
                    'most_traded_month': None,
                    'most_traded_month_avg': None,
                    'recent_3month_avg': None,
                    'total_trades': 0
                }
                
                if not trades:
                    print("거래 데이터 없음 - 분석 중단")
                    return trade_info
                
                try:
                    # 현재 날짜
                    current_date = datetime.now()
                    one_year_ago = current_date - timedelta(days=365)
                    three_months_ago = current_date - timedelta(days=90)
                    
                    # 전체 거래 건수 저장
                    trade_info['total_trades'] = len(trades)
                    print(f"전체 거래 건수: {len(trades)}")
                    
                    # 최근 3개월 평균 거래량 계산
                    recent_3month_trades = [t for t in trades if t['date'] >= three_months_ago]
                    recent_3month_avg = len(recent_3month_trades) / 3  # 3개월 평균
                    trade_info['recent_3month_avg'] = recent_3month_avg
                    print(f"최근 3개월 거래: {len(recent_3month_trades)}건, 월평균: {recent_3month_avg:.1f}건")
                    
                    # 세대수가 있는 경우만 회전율 계산
                    if type_households and type_households > 0:
                        # 1. 최근 연간 회전율 계산
                        recent_trades = [t for t in trades if t['date'] >= one_year_ago]
                        recent_turnover = (len(recent_trades) / type_households) * 100
                        trade_info['recent_turnover'] = recent_turnover
                        
                        print(f"최근 1년 거래 계산:")
                        print(f"  - 기준일: {one_year_ago.strftime('%Y-%m-%d')} ~ {current_date.strftime('%Y-%m-%d')}")
                        print(f"  - 최근 1년 거래: {len(recent_trades)}건")
                        print(f"  - 전체 세대수: {type_households}")
                        print(f"  - 회전율 계산: {len(recent_trades)} / {type_households} * 100 = {recent_turnover:.2f}%")
                        
                        # 2. 연간 평균 회전율 계산
                        first_trade_date = min(trades, key=lambda x: x['date'])['date']
                        years_diff = (current_date - first_trade_date).days / 365.25
                        if years_diff > 0:
                            avg_turnover = (len(trades) / type_households / years_diff) * 100
                            trade_info['avg_turnover'] = avg_turnover
                            
                            print(f"연간 평균 회전율 계산:")
                            print(f"  - 첫 거래일: {first_trade_date.strftime('%Y-%m-%d')}")
                            print(f"  - 기간: {years_diff:.2f}년")
                            print(f"  - 계산: {len(trades)} / {type_households} / {years_diff:.2f} * 100 = {avg_turnover:.2f}%")
                    else:
                        print("세대수 정보 없음 - 회전율 계산 불가")
                    
                    # 3. 월별 거래량 분석 (수정된 코드)
                    # 3. 월별 거래량 분석 (수정된 코드)
                    month_counts = {}  # 월별 총 거래 건수
                    month_year_data = {}  # 월별 연도별 거래 데이터
                    
                    # 첫 거래일과 마지막 거래일 찾기
                    if trades:
                        first_trade_date = min(trades, key=lambda x: x['date'])['date']
                        last_trade_date = max(trades, key=lambda x: x['date'])['date']
                        
                        # 첫 해와 마지막 해 계산
                        first_year = first_trade_date.year
                        last_year = last_trade_date.year
                        
                        # 월별 연도별 거래 집계
                        for trade in trades:
                            month = trade['date'].month
                            year = trade['date'].year
                            
                            # 월별 총 카운트
                            if month not in month_counts:
                                month_counts[month] = 0
                            month_counts[month] += 1
                            
                            # 월별 연도별 데이터
                            if month not in month_year_data:
                                month_year_data[month] = {}
                            if year not in month_year_data[month]:
                                month_year_data[month][year] = 0
                            month_year_data[month][year] += 1
                        
                        # 월별 거래량 디버그 출력
                        print("\n=== 월별 매매 거래량 ===")
                        print(f"데이터 기간: {first_trade_date.strftime('%Y-%m-%d')} ~ {last_trade_date.strftime('%Y-%m-%d')}")
                        print(f"총 연도 범위: {first_year}년 ~ {last_year}년 ({last_year - first_year + 1}년간)\n")
                        
                        month_names = {1:'1월', 2:'2월', 3:'3월', 4:'4월', 5:'5월', 6:'6월',
                                      7:'7월', 8:'8월', 9:'9월', 10:'10월', 11:'11월', 12:'12월'}
                        
                        for month in sorted(month_counts.keys()):
                            # 실제 거래가 있었던 연도 수
                            actual_years = len(month_year_data[month])
                            
                            # 해당 월이 포함될 수 있는 전체 연도 수 계산
                            # 첫 해의 경우: 해당 월이 첫 거래일 이후인지 확인
                            # 마지막 해의 경우: 해당 월이 마지막 거래일 이전인지 확인
                            potential_years = 0
                            for year in range(first_year, last_year + 1):
                                # 첫 해 체크
                                if year == first_year:
                                    if month >= first_trade_date.month:
                                        potential_years += 1
                                # 마지막 해 체크
                                elif year == last_year:
                                    if month <= last_trade_date.month:
                                        potential_years += 1
                                # 중간 연도들
                                else:
                                    potential_years += 1
                            
                            # 평균 계산 (전체 가능 연도 기준)
                            avg_trades = month_counts[month] / potential_years if potential_years > 0 else 0
                            
                            # 실제 거래 연도 리스트
                            years_with_trades = sorted(month_year_data[month].keys())
                            
                            print(f"{month_names[month]}: 총 {month_counts[month]}건")
                            print(f"  - 거래 발생 연도: {years_with_trades}")
                            print(f"  - 가능 연도 수: {potential_years}년")
                            print(f"  - 연평균: {avg_trades:.1f}건/년")
                        
                        # 가장 많이 거래되는 월 찾기
                        if month_counts:
                            # 최다 거래월 찾기
                            max_month = max(month_counts.keys(), key=lambda x: month_counts[x])
                            max_count = month_counts[max_month]
                            
                            trade_info['most_traded_month'] = max_month
                            
                            # 해당 월의 정확한 평균 거래량 계산
                            # 해당 월이 포함될 수 있는 연도 수 계산
                            potential_years_for_max = 0
                            for year in range(first_year, last_year + 1):
                                if year == first_year:
                                    if max_month >= first_trade_date.month:
                                        potential_years_for_max += 1
                                elif year == last_year:
                                    if max_month <= last_trade_date.month:
                                        potential_years_for_max += 1
                                else:
                                    potential_years_for_max += 1
                            
                            avg_trades_for_most = max_count / potential_years_for_max if potential_years_for_max > 0 else 0
                            trade_info['most_traded_month_avg'] = avg_trades_for_most
                            
                            print(f"\n★ 가장 많이 거래되는 월: {month_names[max_month]}")
                            print(f"  - 총 거래: {max_count}건")
                            print(f"  - 가능 연도: {potential_years_for_max}년")
                            print(f"  - 연평균: {avg_trades_for_most:.1f}건/년")
                    else:
                        print("거래 데이터 없음")
                    
                    print("=== 거래 패턴 분석 완료 ===\n")
                    
                except Exception as e:
                    print(f"거래 패턴 분석 중 오류: {str(e)}")
                    import traceback
                    traceback.print_exc()
                
                return trade_info
            
            # 거래 패턴 분석 실행 - trades 데이터만 있으면 실행
            trade_patterns = None
            if trades:  # type_households 조건 제거
                if type_households:
                    trade_patterns = analyze_trade_patterns(trades, type_households)
                    print(f"거래 패턴 분석 완료: {trade_patterns}")
                else:
                    print("세대수 정보가 없어 회전율을 계산할 수 없습니다.")
                    # 세대수 없이 거래 패턴만 분석
                    trade_patterns = {
                        'recent_turnover': None,
                        'avg_turnover': None,
                        'most_traded_month': None,
                        'total_trades': len(trades) if trades else 0
                    }
                    # 월별 분석은 세대수 없이도 가능
                    if trades:
                        month_counts = {}
                        for trade in trades:
                            month = trade['date'].month
                            month_counts[month] = month_counts.get(month, 0) + 1
                        if month_counts:
                            most_traded = max(month_counts.items(), key=lambda x: x[1])
                            trade_patterns['most_traded_month'] = most_traded[0]
            else:
                print("실거래 데이터가 없습니다.")
            
            def find_previous_peak(df, max_pir_idx, current_idx, max_sale_idx):
                current_date = df.loc[current_idx, 'date']
                max_sale_date = df.loc[max_sale_idx, 'date']
                months_diff = (current_date.year - max_sale_date.year) * 12 + current_date.month - max_sale_date.month
                
                # 전고점 표시 여부 확인
                show_max_peak = months_diff > 6  # 6개월 초과면 전고점이 그래프에 표시됨
                
                # 전고점이 그래프에 표시되지 않을 때만 이전 고점을 찾음
                if not show_max_peak:
                    # PIR 최고점 전후 1년 기간 설정
                    pir_peak_date = df.loc[max_pir_idx, 'date']
                    start_date = pir_peak_date - timedelta(days=365)
                    end_date = pir_peak_date + timedelta(days=365)
                    
                    # 해당 기간의 데이터 추출
                    mask = (df['date'] >= start_date) & (df['date'] <= end_date)
                    period_df = df[mask]
                    
                    if not period_df.empty:
                        prev_peak_idx = period_df['매매가'].idxmax()
                        prev_peak_date = df.loc[prev_peak_idx, 'date']
                        
                        # 이전 고점이 현재와 6개월 초과 차이나는지 확인
                        months_to_current = ((current_date.year - prev_peak_date.year) * 12 + 
                                         current_date.month - prev_peak_date.month)
                        
                        if months_to_current > 6:  # 6개월 초과인 경우만 이전 고점으로 표시
                            return {
                                'date': prev_peak_date,
                                'price': df.loc[prev_peak_idx, '매매가'],
                                'type': 'prev_peak'
                            }
                
                return None



        
            # 그래프 생성 시작
            # 그래프 생성 직전에 데이터 확인 (디버깅용)
            print("\n=== 그래프 생성 직전 데이터 확인 ===")
            print(f"현재 데이터프레임 열: {df.columns.tolist()}")
            if len(df) > 0:
                print(f"매매가 데이터 처음 3개: {df['매매가'].head(3).tolist()}")
                print(f"전세가 데이터 처음 3개: {df['전세가'].head(3).tolist()}")
            
            # 가격 유형 텍스트 및 레이블 설정 (명확한 조건문 사용)
            if self.sale_price_type.get() == "low":
                sale_type_text = "하위평균"
                sale_label = "매매가(하위평균)"
            elif self.sale_price_type.get() == "high":
                sale_type_text = "상위평균"
                sale_label = "매매가(상위평균)"
            else:  # "normal" 또는 기타
                sale_type_text = "일반평균"
                sale_label = "매매가(일반평균)"
            
            if self.lease_price_type.get() == "low":
                lease_type_text = "하위평균"
                lease_label = "전세가(하위평균)"
            elif self.lease_price_type.get() == "high":
                lease_type_text = "상위평균"
                lease_label = "전세가(상위평균)"
            else:  # "normal" 또는 기타
                lease_type_text = "일반평균"
                lease_label = "전세가(일반평균)"
            
            print(f"매매가 유형: {sale_type_text}, 레이블: {sale_label}")
            print(f"전세가 유형: {lease_type_text}, 레이블: {lease_label}")
            
            # 그래프 생성 코드
            # 그래프 생성 코드
            fig = plt.figure(figsize=(16, 9))  # 원래 크기로
            ax1 = plt.subplot2grid((10, 1), (0, 0), rowspan=7)  # 원래 비율로
            ax2 = ax1.twinx()  # ax2 정의 - 심리지수 축
            
            # ax2를 ax1보다 뒤에 배치
            ax2.set_zorder(ax1.get_zorder() - 1)
            ax1.patch.set_visible(False)  # ax1의 배경을 투명하게 설정
            
            # 심리지수 축 범위 설정 (0-100)
            ax2.set_ylim(0, 100)
            ax2.set_ylabel('심리지수')
            
            # 그래프 생성 부분 수정 (각 유형별로 라인 생성)
            # 그래프 생성 부분 수정 (각 유형별로 라인 생성)
            # 그래프 생성 부분 수정 (각 유형별로 라인 생성)
            lines = []  # 모든 라인을 여기에 저장
            line_labels = []  # 라인 레이블 저장
            
            # 가격 유형 매핑 정보 (레이블 생성용)
            price_type_labels = {
                "low": "하위평균",
                "normal": "일반평균",
                "high": "상위평균"
            }
            
            # 색상 매핑 정보
            sale_colors = {
                "low": "#3399FF",   # 연한 파란색
                "normal": "#0066CC", # 표준 파란색
                "high": "#003366"   # 진한 파란색
            }
            
            lease_colors = {
                "low": "#FFAA80",   # 연한 주황색
                "normal": "#FF6B00", # 표준 주황색
                "high": "#CC5500"   # 진한 주황색
            }
            
            # 선택된 가격 유형 확인
            sale_types = [t for t, var in self.sale_price_types.items() if var.get()]
            lease_types = [t for t, var in self.lease_price_types.items() if var.get()]
            
            # 우선순위에 따라 정렬 (상위 > 일반 > 하위)
            priority_order = {"high": 0, "normal": 1, "low": 2}
            sale_types = sorted(sale_types, key=lambda x: priority_order[x])
            lease_types = sorted(lease_types, key=lambda x: priority_order[x])
            
            print(f"우선순위 정렬 후 매매가 유형: {sale_types}")
            print(f"우선순위 정렬 후 전세가 유형: {lease_types}")
            # KB시세 표시 여부 확인
            if self.show_kb_price.get():
                print("KB시세 그래프 표시 활성화")
                # 매매가 그래프 (각 유형별로)
                for idx, sale_type in enumerate(sale_types):
                    col_name = f"매매가_{sale_type}"
                    label = f"매매가({price_type_labels[sale_type]})"
                    color = sale_colors[sale_type]
                    
                    # 해당 열이 데이터프레임에 존재하는지 확인
                    if col_name in df.columns:
                        line_sale, = ax1.plot(df['date'], df[col_name], color=color, label=label, 
                                            linewidth=3, zorder=4)
                        lines.append(line_sale)
                        line_labels.append(label)
                        print(f"매매가 그래프 추가: {label}")
                
                # 전세가 그래프 (각 유형별로)
                for idx, lease_type in enumerate(lease_types):
                    col_name = f"전세가_{lease_type}"
                    label = f"전세가({price_type_labels[lease_type]})"
                    color = lease_colors[lease_type]
                    
                    # 해당 열이 데이터프레임에 존재하는지 확인
                    if col_name in df.columns:
                        line_lease, = ax1.plot(df['date'], df[col_name], color=color, label=label, 
                                            linewidth=3, zorder=4)
                        lines.append(line_lease)
                        line_labels.append(label)
                        print(f"전세가 그래프 추가: {label}")
            else:
                print("KB시세 그래프 표시 비활성화 - KB시세 그래프 생략")
                # KB시세를 표시하지 않아도 line_sale과 line_lease 변수는 필요 (다른 코드에서 참조)
                line_sale = None
                line_lease = None
            # 그래프 생성 직후에 데이터 확인 (디버깅용)
            print("\n=== 그래프 생성 직후 데이터 확인 ===")
            print(f"매매가 그래프 데이터 처음 3개: {line_sale.get_ydata()[:3]}")
            print(f"전세가 그래프 데이터 처음 3개: {line_lease.get_ydata()[:3]}")

        
            print(f"결정된 매매가 레이블: {sale_label}")
            print(f"결정된 전세가 레이블: {lease_label}")
            
            # # 매매가 그래프 (선택된 유형 레이블 적용)
            # line_sale, = ax1.plot(df['date'], df['매매가'], color='#0066CC', label=sale_label, 
            #                      linewidth=3, zorder=4)
            
            # # 전세가 그래프 (선택된 유형 레이블 적용)
            # line_lease, = ax1.plot(df['date'], df['전세가'], color='#FF6B00', label=lease_label, 
            #                       linewidth=3, zorder=4)
            # PIR 값 정규화하여 ax1에 표시 (가격 축에 맞춤)
            # PIR 값 정규화하여 ax1에 표시 (가격 축에 맞춤)
            # 수정 후
            # PIR 값 정규화하여 ax1에 표시 (가격 축에 맞춤)
            y_min, y_max = ax1.get_ylim()
            y_range = y_max - y_min
                        
            # PIR 범위 찾기
            pir_min = df['PIR'].min()
            pir_max = df['PIR'].max()
            pir_range = pir_max - pir_min
                        
            # PIR 값 정규화
            normalized_pir = [y_min + y_range * ((val - pir_min) / pir_range * 0.7) for val in df['PIR']]
            normalized_pir_avg = [y_min + y_range * ((val - pir_min) / pir_range * 0.7) for val in df['PIR_running_avg']]
                        
            # 변수 초기화 - 중요! 이 부분 추가
            line_pir = None
            line_pir_avg = None
                        
            # 정규화된 PIR 그래프 - PIR 표시 체크박스 상태에 따라 조건부 실행
            if hasattr(self, 'show_pir') and self.show_pir.get():
                line_pir, = ax1.plot(df['date'], normalized_pir, color='#9C27B0', label='PIR', 
                                    linewidth=2, zorder=1)
                line_pir_avg, = ax1.plot(df['date'], normalized_pir_avg, color='#9C27B0', 
                                        linestyle='--', label='PIR 누적평균', linewidth=1.5, zorder=1)          
            # 범례 설정
            lines = [line_sale, line_lease]  # 기본 라인만 추가
                        
            # PIR 선이 그려졌을 경우에만 범례에 추가
            if hasattr(self, 'show_pir') and self.show_pir.get() and line_pir is not None and line_pir_avg is not None:
                lines.append(line_pir)
                lines.append(line_pir_avg)
                        
            labels = [l.get_label() for l in lines]
            valid_lines = [l for l in lines if l is not None]
                        
            if valid_lines:  # 유효한 선이 있을 때만 범례 생성
                ax1.legend(valid_lines, labels, 
                          loc='upper left',
                          bbox_to_anchor=(0.02, 0.98),
                          ncol=1,               
                          fontsize=9,           
                          framealpha=0.8)
                        
            # ... 이후 코드 ...
            # PIR 값 주요 포인트에 실제 PIR 값 표시하는 어노테이션 추가
            # (max_pir_idx와 current_idx 위치에 실제 PIR 값 표시)
            # ax1.annotate(f'PIR: {df.loc[max_pir_idx, "PIR"]:.1f}',
            #             xy=(df.loc[max_pir_idx, 'date'], normalized_pir[df.index.get_loc(max_pir_idx)]),
            #             xytext=(-40, 20),
            #             textcoords='offset points',
            #             bbox=dict(boxstyle='round,pad=0.5', fc='#FF9999', alpha=0.7),
            #             arrowprops=dict(arrowstyle='-', connectionstyle='angle3', color='#9C27B0'))
            
            # ax1.annotate(f'PIR: {df.loc[current_idx, "PIR"]:.1f}',
            #             xy=(latest_date, normalized_pir[df.index.get_loc(current_idx)]),
            #             xytext=(20, -40),
            #             textcoords='offset points',
            #             bbox=dict(boxstyle='round,pad=0.5', fc='#FF9999', alpha=0.7),
            #             arrowprops=dict(arrowstyle='-', connectionstyle='angle3', color='#9C27B0'))


            # 네이버 매물 정보 가져오기 (try-except로 감싸기)
            naver_deal_info = None
            if self.show_naver_deal.get():  # 네이버매물 체크박스가 선택된 경우만
                print("네이버 매매 매물 조회 활성화")
                try:
                    print("\n=== 네이버 매물 정보 가져오기 시작 ===")
                    naver_deal_info = self.get_naver_min_price()
                    print(f"네이버 매물 정보 결과: {naver_deal_info}")
                except Exception as e:
                    print(f"네이버 매물 정보 가져오기 실패: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    naver_deal_info = None
            else:
                print("네이버 매매 매물 조회 비활성화")

            # 네이버 전세 매물 정보 가져오기
            naver_jeonse_info = None
            if self.show_naver_deal.get():  # 네이버매물 체크박스가 선택된 경우만
                print("네이버 전세 매물 조회 활성화")
                try:
                    print("\n=== 네이버 전세 매물 정보 가져오기 시작 ===")
                    naver_jeonse_info = self.get_naver_min_jeonse_price()
                    print(f"네이버 전세 매물 정보 결과: {naver_jeonse_info}")
                except Exception as e:
                    print(f"네이버 전세 매물 정보 가져오기 실패: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    naver_jeonse_info = None
            else:
                print("네이버 전세 매물 조회 비활성화")                
            
            # 네이버 매물 최저가 표시 - None 체크 추가
            # 네이버 매물 최저가 표시 - None 체크 추가
            # 네이버 매물 최저가 표시 부분에 추가
            # 네이버 매물 최저가 표시 부분 수정
            if self.show_naver_deal.get() and naver_deal_info is not None and isinstance(naver_deal_info, dict) and 'price' in naver_deal_info and 'floor' in naver_deal_info:
                print("네이버 매물 정보를 그래프에 추가합니다.")
                
                # mid_point 변수를 명확하게 정의
                date_points = df['date'].values
                if len(date_points) > 0:
                    mid_index = len(date_points) // 4
                    mid_point = date_points[mid_index] if mid_index < len(date_points) else date_points[-1]
                else:
                    # 날짜 데이터가 없는 경우 현재 날짜 사용
                    mid_point = datetime.now()
                
                # 저층 매물인 경우 다른 스타일/색상 사용
                line_color = '#FFA07A' if naver_deal_info.get('is_low_floor', False) else '#66B3FF'
                line_style = ':' if naver_deal_info.get('is_low_floor', False) else '--'
                
                floor_label = f"네이버 최저매물 ({naver_deal_info['floor']})"
                if naver_deal_info.get('is_low_floor', False):
                    floor_label += " [저층]"
                
                ax1.axhline(y=naver_deal_info['price'], 
                           color=line_color,
                           linestyle=line_style, 
                           linewidth=2,
                           alpha=0.7,
                           label=floor_label,
                           zorder=9)
                
                # 매물 정보 텍스트에도 저층 표시 추가
                property_info = f"네이버 최저매물 {naver_deal_info['price']:,.0f}만원 ({naver_deal_info['floor']})"
                if naver_deal_info.get('is_low_floor', False):
                    property_info += " [저층]"
                
                # 동 정보 추가 (해당 매물의 동만)
                if 'dong_info' in naver_deal_info and naver_deal_info['dong_info']:
                    property_info += f"\n{naver_deal_info['dong_info']}"
                
                # 매물 수 정보 추가
                if 'property_count' in naver_deal_info:
                    property_info += f"\n매물 {naver_deal_info['property_count']}건"
                
                ax1.text(mid_point, naver_deal_info['price'],
                        property_info,
                        color=line_color,
                        fontsize=9,
                        horizontalalignment='center',
                        verticalalignment='bottom',
                        bbox=dict(facecolor='white', edgecolor='none', alpha=0.7),
                        zorder=10)
                
                # lines 변수 문제 해결 (기존 코드와 동일하게 유지)
                
                # lines 변수 문제 해결 (기존 코드와 동일하게 유지)
                
                # lines 변수 문제 해결 - 여기서 lines는 ax1.plot()로 생성된 객체의 리스트여야 합니다
                # 기존에 lines 변수가 있는지 확인하고, 없으면 빈 리스트로 초기화
                try:
                    # lines가 이미 정의되어 있다고 가정하고 네이버 매물 데이터 추가
                    naver_line = plt.Line2D([0], [0], color='#66B3FF', linestyle='--', 
                                          label=f"네이버 최저매물 ({naver_deal_info['floor']})")
                    lines.append(naver_line)
                except NameError:
                    # lines가 정의되지 않은 경우 무시하고 진행
                    print("lines 변수가 정의되지 않아 네이버 매물 정보를 범례에 추가할 수 없습니다.")
            else:
                print("네이버 매물 정보가 없거나 형식이 잘못되어 그래프에 추가하지 않습니다.")    


            # 네이버 전세 매물 최저가 표시
            # 네이버 전세 매물 최저가 표시
            if self.show_naver_deal.get() and naver_jeonse_info is not None and isinstance(naver_jeonse_info, dict) and 'price' in naver_jeonse_info and 'floor' in naver_jeonse_info:
                print("네이버 전세 매물 정보를 그래프에 추가합니다.")
                mid_point = df['date'].iloc[len(df)//4]  # 매매보다 약간 다른 위치에 텍스트 표시
                ax1.axhline(y=naver_jeonse_info['price'], 
                           color='#FF9B4D',  # 전세는 주황색 계열
                           linestyle='--', 
                           linewidth=2,
                           alpha=0.7,
                           label=f"네이버 전세 최저가 ({naver_jeonse_info['floor']})",
                           zorder=9)
                
                # 전세 매물 정보 텍스트 - 전체 매물 건수와 해당 매물 동만 표시
                jeonse_info = f"네이버 전세 최저가 {naver_jeonse_info['price']:,.0f}만원 ({naver_jeonse_info['floor']})"
                # 동 정보 추가 (해당 매물의 동만)
                if 'dong_info' in naver_jeonse_info and naver_jeonse_info['dong_info']:
                    jeonse_info += f"\n{naver_jeonse_info['dong_info']}"

                
                # 매물 수 정보 추가
                if 'property_count' in naver_jeonse_info:
                    jeonse_info += f"\n매물 {naver_jeonse_info['property_count']}건"
                

                
                ax1.text(mid_point, naver_jeonse_info['price'],
                        jeonse_info,
                        color='#FF9B4D',
                        fontsize=9,
                        horizontalalignment='center',
                        verticalalignment='bottom',
                        bbox=dict(facecolor='white', edgecolor='none', alpha=0.7),
                        zorder=10)
                
                # 범례에 네이버 전세 매물 추가 (기존 코드와 동일하게 유지)
                
                # 범례에 네이버 전세 매물 추가 (기존 코드와 동일하게 유지)
                
                # 범례에 네이버 전세 매물 추가
                try:
                    # lines가 이미 정의되어 있다고 가정
                    naver_jeonse_line = plt.Line2D([0], [0], color='#FF9B4D', linestyle='--', 
                                                 label=f"네이버 전세 최저가 ({naver_jeonse_info['floor']})")
                    lines.append(naver_jeonse_line)
                except NameError:
                    print("lines 변수가 정의되지 않아 네이버 전세 매물 정보를 범례에 추가할 수 없습니다.")
            else:
                print("네이버 전세 매물 정보가 없거나 형식이 잘못되어 그래프에 추가하지 않습니다.")
            
            
            # 네이버 매물 매매 최저가 표시
            # 네이버 매물 매매 최저가 표시
            # 네이버 매물 매매 최저가 표시
            
            # if hasattr(self, 'trade_price') and self.trade_price and self.complex_id:  # complex_id 체크 추가
            #     mid_point = df['date'].iloc[len(df)//4]
            #     ax1.axhline(y=self.trade_price, 
            #                 color='#66B3FF',
            #                 linestyle='--', 
            #                 linewidth=2,
            #                 alpha=0.7,
            #                 label=f'매매 최저매물 ({self.trade_floor}층)',
            #                 zorder=9)
                
            #     ax1.text(mid_point, self.trade_price,
            #              f'매매 최저매물 {self.trade_price:,.0f}만원 ({self.trade_floor}층)\n매물 {len(self.trade_counts)}건',
            #              color='#66B3FF',
            #              fontsize=9,
            #              horizontalalignment='center',
            #              verticalalignment='bottom',
            #              bbox=dict(facecolor='white', edgecolor='none', alpha=0.7),
            #              zorder=10)
            
            # # 네이버 매물 전세 최저가 표시 (조건 추가)
            # if hasattr(self, 'lease_price') and self.lease_price and self.complex_id:  # complex_id 체크 추가
            #     mid_point = df['date'].iloc[len(df)//4]
            #     ax1.axhline(y=self.lease_price, 
            #                 color='#FF9B4D',
            #                 linestyle='--', 
            #                 linewidth=2,
            #                 alpha=0.7,
            #                 label=f'전세 최저매물 ({self.lease_floor}층)',
            #                 zorder=9)
                
            #     ax1.text(mid_point, self.lease_price,
            #              f'전세 최저매물 {self.lease_price:,.0f}만원 ({self.lease_floor}층)\n매물 {len(self.lease_counts)}건',
            #              color='#FF9B4D',
            #              fontsize=9,
            #              horizontalalignment='center',
            #              verticalalignment='top',
            #              bbox=dict(facecolor='white', edgecolor='none', alpha=0.7),
            #              zorder=10)
            
                        # 범례 업데이트를 위한 lines와 labels 수정
            # 실거래가 점도표를 위한 색상 정의를 가장 앞에서 해야 함
            high_floor_color = '#69B1FF'  # 파란색 점
            low_floor_color = '#C0C0C0'   # 회색 점

            # 네이버 매물 표시 후 범례 lines 추가 부분을 이렇게 수정
            lines = [line_sale]  # 우선 매매가 라인만 추가
            
            # # 네이버 매물 매매/전세가 추가
            # if hasattr(self, 'trade_price') and self.trade_price:
            #     lines.append(plt.Line2D([0], [0], color='#66B3FF', linestyle='--', 
            #                            label=f'매매 최저매물 ({self.trade_floor}층)'))
            # if hasattr(self, 'lease_price') and self.lease_price:
            #     lines.append(plt.Line2D([0], [0], color='#FF9B4D', linestyle='--', 
            #                            label=f'전세 최저매물 ({self.lease_floor}층)'))
            
            # 전세가 그래프 그린 후
            line_lease, = ax1.plot(df['date'], df['전세가'], color='#FF6B00', label='전세가', 
                                  linewidth=3, zorder=4)
            lines.append(line_lease)  # 전세가 라인 추가
            
            # # PIR 그래프 그린 후
            # line_pir, = ax2.plot(df['date'], df['PIR'], color='#9C27B0', label='PIR', 
            #                      linewidth=2, zorder=1)
            # line_pir_avg, = ax2.plot(df['date'], df['PIR_running_avg'], color='#9C27B0', 
            #                         linestyle='--', label='PIR 누적평균', linewidth=1.5, zorder=1)
            lines.extend([line_pir, line_pir_avg])  # PIR 라인들 추가
            
            # 실거래가 점도표 추가
            if trades and any(t['floor'] > 4 for t in trades):
                lines.append(plt.Line2D([0], [0], color=high_floor_color, marker='o', 
                                       linestyle='None', markersize=8, alpha=0.6,
                                       label='실거래(5층↑)'))
            if trades and any(t['floor'] <= 4 for t in trades):
                lines.append(plt.Line2D([0], [0], color=low_floor_color, marker='o',
                                       linestyle='None', markersize=8, alpha=0.6,
                                       label='실거래(4층↓)'))
            
            
            # 어노테이션 함수 정의
            def create_annotation_with_line(ax, text, xy, xytext, color='black', bbox_fc='yellow'):
                return ax.annotate(
                    text,
                    xy=xy,
                    xytext=xytext,
                    textcoords='offset points',
                    bbox=dict(boxstyle='round,pad=0.5', fc=bbox_fc, alpha=0.7),
                    arrowprops=dict(
                        arrowstyle='-',
                        connectionstyle='angle3',
                        color=color,
                        lw=1
                    ),
                    fontsize=9,
                    zorder=6
                )

            #
            # 최근 저점 찾기
            def find_recent_low(df, current_idx, points_to_mark):
                current_date = df.loc[current_idx, 'date']
                
                # points_to_mark가 비어있는지 먼저 체크
                if not points_to_mark:
                    return None
                    
                # 전고점이나 이전 고점 찾기
                peak_point = None
                for point in points_to_mark:
                    if point['type'] in ['max', 'prev_peak']:
                        peak_point = point
                        break
                        
                if peak_point:
                    # 고점과 현재 사이의 기간 설정
                    peak_date = peak_point['date']
                    
                    # 해당 기간의 데이터에서 최저점 찾기
                    mask = (df['date'] >= peak_date) & (df['date'] <= current_date)
                    period_df = df[mask]
                    
                    if not period_df.empty:  # period_df가 비어있지 않은지 확인
                        min_idx = period_df['매매가'].idxmin()
                        low_date = df.loc[min_idx, 'date']
                        
                        # 최저점이 현재가가 아니고, 현재와 6개월 초과 차이나는 경우에만 표시
                        if min_idx != current_idx:
                            months_to_current = (current_date.year - low_date.year) * 12 + current_date.month - low_date.month
                            if months_to_current > 6:
                                return {
                                    'date': low_date,
                                    'price': df.loc[min_idx, '매매가'],
                                    'type': 'recent_low'
                                }
                
                return None
            
            # 매매가 포인트 처리 블록 시작
            # 매매가 포인트 처리 블록
            # 매매가 포인트 처리 블록 시작
            # 매매가 포인트 처리 블록
            points_to_mark = []
            
            
            # 1. 전저점 추가 (가장 먼저 추가)
            if old_low_point is not None:
                points_to_mark.append({
                    'date': df.loc[old_low_point, 'date'],
                    'price': df.loc[old_low_point, '매매가'],
                    'type': 'old_low'
                })
            
            # 1. 현재가
            points_to_mark.append({
                'date': latest_date,
                'price': price_info["current"],
                'type': 'current'
            })
            
            # 2. 전고점 처리
            months_diff = (latest_date.year - price_info["max_date"].year) * 12 + latest_date.month - price_info["max_date"].month
            if months_diff > 6:
                points_to_mark.append({
                    'date': price_info["max_date"],
                    'price': price_info["max"],
                    'type': 'max'
                })
            else:
                # 전고점이 없을 때 이전 고점 찾기
                prev_peak = find_previous_peak(df, max_pir_idx, current_idx, max_sale_idx)
                if prev_peak:
                    points_to_mark.append(prev_peak)
            
            # 3. 최근 저점 찾기
            recent_low = find_recent_low(df, current_idx, points_to_mark)
            if recent_low:
                points_to_mark.append(recent_low)
                # 최근저점 정보를 price_info에도 추가하여 정보텍스트와 어노테이션 일치시키기
                price_info['recent_low'] = recent_low['price']
                price_info['recent_low_date'] = recent_low['date']
            
            # 마커와 레이블 처리 (모든 어노테이션을 여기서 처리)
            # 마커와 레이블 처리 (모든 어노테이션을 여기서 처리)
            if points_to_mark:
                dates_sale = [point['date'] for point in points_to_mark]
                prices_sale = [point['price'] for point in points_to_mark]
                ax1.scatter(dates_sale, prices_sale, 
                           color='white',
                           edgecolor='#0099FF',
                           s=100,
                           linewidth=2,
                           zorder=5)
                
                for point in points_to_mark:
                    if point['type'] == 'old_low':  # 전저점 처리 추가
                        create_annotation_with_line(
                            ax1,
                            f'{point["price"]:,.0f}\n(전저점 {point["date"].strftime("%Y-%m")})',
                            xy=(point['date'], point['price']),
                            xytext=(-70, -30),
                            bbox_fc='#0099FF'
                        )
                    elif point['type'] == 'current':
                        ax1.annotate(f'{point["price"]:,.0f}\n(현재가 {point["date"].strftime("%Y-%m")})',
                                    xy=(point['date'], point['price']),
                                    xytext=(10, -10),
                                    textcoords='offset points',
                                    bbox=dict(boxstyle='round,pad=0.5', fc='white', alpha=0.5),
                                    fontsize=10,
                                    zorder=6)
                    elif point['type'] == 'max':
                        create_annotation_with_line(
                            ax1,
                            f'{point["price"]:,.0f}\n(전고점 {point["date"].strftime("%Y-%m")})',
                            xy=(point['date'], point['price']),
                            xytext=(-90, 0),
                            bbox_fc='#0099FF'
                        )
                    elif point['type'] == 'prev_peak':
                        create_annotation_with_line(
                            ax1,
                            f'{point["price"]:,.0f}\n(이전고점 {point["date"].strftime("%Y-%m")})',
                            xy=(point['date'], point['price']),
                            xytext=(-70, 30),
                            bbox_fc='#0099FF'
                        )
                    elif point['type'] == 'recent_low':
                        create_annotation_with_line(
                            ax1,
                            f'{point["price"]:,.0f}\n(최근저점 {point["date"].strftime("%Y-%m")})',
                            xy=(point['date'], point['price']),
                            xytext=(-150, 0),
                            bbox_fc='#0099FF'
                        )
                                    



            # 전세가 그래프
            line_lease, = ax1.plot(df['date'], df['전세가'], color='#FF6B00', label='전세가', 
                                 linewidth=3, zorder=4)
    
            # 전세가 주요 포인트에 마커 추가
            important_points_lease = []
    
            # 현재 전세가 마커
            
            current_lease = (latest_date, df.loc[current_idx, '전세가'])
            important_points_lease.append(current_lease)
    
            # 전세가 전고점 마커
            max_lease = (df.loc[max_lease_idx, 'date'], df.loc[max_lease_idx, '전세가'])
            important_points_lease.append(max_lease)
    
            # 전세가 마커 일괄 추가
            dates_lease = [point[0] for point in important_points_lease]
            prices_lease = [point[1] for point in important_points_lease]
            ax1.scatter(dates_lease, prices_lease,
                       color='white',
                       edgecolor='#FF6B00',
                       s=100,
                       linewidth=2,
                       zorder=5)
    
    
            
            # # PIR 그래프
            # line_pir, = ax2.plot(df['date'], df['PIR'], color='#9C27B0', label='PIR', 
            #                     linewidth=2, zorder=1)  # 선은 중간 레이어
            
            # line_pir_avg, = ax2.plot(df['date'], df['PIR_running_avg'], color='#9C27B0', 
            #                         linestyle='--', label='PIR 누적평균', linewidth=1.5, zorder=1)
    
       

            # 매매가 그래프 그린 후, 심리 데이터 추가
            # 매매가 그래프 그린 후, 심리 데이터 추가
            # 심리 데이터 시각화 코드 수정 부분 (analyze_prices 메서드 내부에 삽입)
            # 심리 데이터 시각화 코드 수정 부분 (analyze_prices 메서드 내부에 삽입)
            try:
                # 유나심리차트 체크박스 확인 - 체크되지 않았으면 전체 처리 스킵
                if not self.show_una_sentiment.get():
                    print("유나심리차트 표시 옵션이 꺼져 있어 차트를 그리지 않습니다.")
                else:
                    # KB부동산 주소 추출
                    kb_address = None
                    
                    # Excel에서 주소 정보 추출 시도
                    try:
                        location_df = pd.read_excel(file_path, nrows=8)
                        # 주소 정보가 있는 셀 찾기 (보통 2행 또는 3행의 B열)
                        for i in range(2, 5):
                            if i < len(location_df):
                                if '주소' in str(location_df.iloc[i, 0]) or '대표번지' in str(location_df.iloc[i, 0]):
                                    kb_address = str(location_df.iloc[i, 1])
                                    break
                                
                        if not kb_address:
                            # 백업 - 특정 셀 직접 확인
                            kb_address = str(location_df.iloc[3, 1])  # 보통 4행 B열에 주소 정보
                    except Exception as e:
                        print(f"주소 정보 추출 중 오류: {str(e)}")
                        
                    print(f"추출된 주소: {kb_address}")
                    
                    # 심리 데이터 로드
                    if hasattr(self, 'sentiment_excel_path') and self.sentiment_excel_path:
                        sentiment_info = self.load_sentiment_data()
                        
                        # 심리 데이터를 그래프에 추가하는 부분 수정
                        if sentiment_info and kb_address:
                            # 지역 매칭
                            matched_region = self.match_region_from_address(kb_address, sentiment_info['regions'])
                            
                            if matched_region:
                                print(f"매칭된 지역 데이터로 심리 그래프를 그립니다: {matched_region}")
                                
                                # KB부동산 데이터의 날짜 범위 확인
                                kb_start_date = df['date'].min()
                                kb_end_date = df['date'].max()
                                
                                # 날짜별 심리 데이터 추출
                                sentiment_dates = []
                                sell_values = []
                                buy_values = []
                                
                                for date_str, regions_data in sentiment_info['data'].items():
                                    if matched_region in regions_data:
                                        try:
                                            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                                            
                                            # KB부동산 데이터 기간 내의 데이터만 사용
                                            if date_obj < kb_start_date or date_obj > kb_end_date:
                                                continue
                                                
                                            region_data = regions_data[matched_region]
                                            
                                            # 날짜와 값 저장
                                            sentiment_dates.append(date_obj)
                                            
                                            # 매도자많음 데이터
                                            sell_value = region_data.get('sell')
                                            sell_values.append(sell_value if sell_value is not None else np.nan)
                                            
                                            # 매수자많음 데이터
                                            buy_value = region_data.get('buy')
                                            buy_values.append(buy_value if buy_value is not None else np.nan)
                                            
                                        except Exception as e:
                                            print(f"심리 데이터 처리 중 오류 ({date_str}): {str(e)}")
                                
                                # 색상 강조 함수 정의 (수정: 모든 구간 동일한 스타일 적용)
                                def get_segment_colors(dates, values, threshold, base_color, is_higher=True):
                                    # 색상, 투명도, 선 굵기를 저장할 리스트
                                    segment_colors = []
                                    segment_alphas = []
                                    segment_linewidths = []

                                    for i in range(len(values)-1):
                                        # 모든 구간에 동일한 스타일 적용
                                        segment_colors.append(base_color)
                                        segment_alphas.append(0.6)  # 동일한 투명도
                                        segment_linewidths.append(2)  # 동일한 선 굵기

                                    return segment_colors, segment_alphas, segment_linewidths
                                
                                # 매도자많음 구간별 색상 생성 (90% 이상 강조)
                                if len(sentiment_dates) > 1 and len(sell_values) > 1:
                                    sell_colors, sell_alphas, sell_linewidths = get_segment_colors(
                                        sentiment_dates, sell_values, 90, 'blue', True
                                    )
                                    
                                    # 매도심리 선 그리기 (구간별 스타일 적용)
                                    for i in range(len(sentiment_dates)-1):
                                        if not (np.isnan(sell_values[i]) or np.isnan(sell_values[i+1])):
                                            ax2.plot(
                                                [sentiment_dates[i], sentiment_dates[i+1]],
                                                [sell_values[i], sell_values[i+1]],
                                                color=sell_colors[i],
                                                alpha=sell_alphas[i],
                                                linewidth=sell_linewidths[i],
                                                solid_capstyle='round',
                                                zorder=2
                                            )
                                    
                                    # 매도심리 범례용 더미 라인
                                    sell_line, = ax2.plot(
                                        [], [], color='blue', alpha=0.6, linestyle='-',
                                        linewidth=2, label='매도심리', zorder=2
                                    )
                                
                                # 매수자많음 구간별 색상 생성 (20% 이상 강조)
                                if len(sentiment_dates) > 1 and len(buy_values) > 1:
                                    buy_colors, buy_alphas, buy_linewidths = get_segment_colors(
                                        sentiment_dates, buy_values, 20, 'red', True
                                    )
                                    
                                    # 매수심리 선 그리기 (구간별 스타일 적용)
                                    for i in range(len(sentiment_dates)-1):
                                        if not (np.isnan(buy_values[i]) or np.isnan(buy_values[i+1])):
                                            ax2.plot(
                                                [sentiment_dates[i], sentiment_dates[i+1]],
                                                [buy_values[i], buy_values[i+1]],
                                                color=buy_colors[i],
                                                alpha=buy_alphas[i],
                                                linewidth=buy_linewidths[i],
                                                solid_capstyle='round',
                                                zorder=2
                                            )
                                    
                                    # 매수심리 범례용 더미 라인
                                    buy_line, = ax2.plot(
                                        [], [], color='red', alpha=0.6, linestyle='-',
                                        linewidth=2, label='매수심리', zorder=2
                                    )
                                
                                # # 중요 신호선 표시 (매도 90%, 매수 20%)
                                # ax2.axhline(y=90, color='blue', linestyle='--', alpha=0.5, linewidth=1)
                                # ax2.axhline(y=20, color='red', linestyle='--', alpha=0.5, linewidth=1)
                                
                                # 차트 제목 즉시 업데이트
                                ax1.set_title(chart_title + " + 유나심리차트", pad=20, fontsize=20, fontweight='bold')
                                
                                # 범례에 매수/매도 심리 추가
                                if 'sell_line' in locals():
                                    lines.append(sell_line)
                                if 'buy_line' in locals():
                                    lines.append(buy_line)
                                
                                print("유나심리차트 추가 완료")
                                
                            else:
                                print("지역 매칭 실패")
                        else:
                            print("심리 데이터 로드 실패 또는 주소 정보 없음")          
                    else:
                        print("심리 데이터 엑셀 경로가 설정되지 않았습니다.")
                    
            except Exception as e:
                print(f"심리 그래프 추가 중 오류: {str(e)}")
                import traceback
                traceback.print_exc()

        
            # 범례 설정
            # 범례 설정
            lines = [line_sale, line_lease]  # 기본 라인만 추가
            
            # PIR 선이 그려졌을 경우에만 범례에 추가
            if hasattr(self, 'show_pir') and self.show_pir.get() and line_pir is not None and line_pir_avg is not None:
                lines.append(line_pir)
                lines.append(line_pir_avg)
                
            labels = [l.get_label() for l in lines]
            valid_lines = [l for l in lines if l is not None]
            
            if valid_lines:  # 유효한 선이 있을 때만 범례 생성
                ax1.legend(valid_lines, labels, 
                          loc='upper left',
                          bbox_to_anchor=(0.02, 0.98),
                          ncol=1,               
                          fontsize=9,           
                          framealpha=0.8)
            
         
            # 수정 후
            valid_lines = [l for l in lines if l is not None]
            valid_labels = [l.get_label() for l in valid_lines]
            if valid_lines:
                ax1.legend(valid_lines, valid_labels, loc='best')
            # 신축단지 PIR 처리 (약 2900줄 근처, PIR 그래프 그리는 부분)
            # 신축단지 PIR 처리
            # 신축단지 PIR 처리
            # 신축단지 PIR 처리 및 그래프 표시 추가
            if hasattr(self, 'new_apt_info') and self.new_apt_info['name'] and self.new_apt_info['price']:
                try:
                    # 신축단지 분양가
                    new_apt_price = float(self.new_apt_info['price'])
                    
                    # 검색단지 현재 정보
                    current_price = df.loc[current_idx, sale_col_name]  # 현재 매매가
                    current_pir = df.loc[current_idx, 'PIR']      # 현재 PIR
                    
                    # 연식 계산 (준공년도로부터)
                    if self.completion_year:
                        current_year = datetime.now().year
                        years_old = current_year - int(self.completion_year)
                        
                        # 시간 경과에 따른 가치 상승 반영
                        value_appreciation_rate = years_old * 0.03  # 매년 3%의 가치 상승
                        adjusted_price = current_price + (current_price * value_appreciation_rate)
                        
                        # 신축단지 PIR 계산
                        new_apt_pir = current_pir * (new_apt_price / adjusted_price)
                        
                        print(f"신축단지 PIR 계산 내역:")
                        print(f"- 신축단지 분양가: {new_apt_price:,.0f}")
                        print(f"- 검색단지 현재가: {current_price:,.0f}")
                        print(f"- 검색단지 연식: {years_old}년")
                        print(f"- 가치 상승률: {value_appreciation_rate:.2f} (연 3%)")
                        print(f"- 신축 기준 검색단지 가격: {adjusted_price:,.0f}")
                        print(f"- 검색단지 현재 PIR: {current_pir:.1f}")
                        print(f"- 계산된 신축단지 PIR: {new_apt_pir:.1f}")
                        
                        # 신축단지 PIR을 그래프에 표시하기 위한 값 계산
                        # PIR 범위를 찾아서 PIR 값을 y축 값으로 변환
                        y_min, y_max = ax1.get_ylim()
                        y_range = y_max - y_min
                        pir_min = df['PIR'].min()
                        pir_max = df['PIR'].max()
                        pir_range = pir_max - pir_min
                        
                        # PIR 값을 정규화하여 ax1에 표시할 y 값 계산
                        normalized_new_apt_pir = y_min + y_range * ((new_apt_pir - pir_min) / pir_range * 0.7)
                        
                        # 신축단지 PIR 수평선 그리기
                        ax1.axhline(
                            y=normalized_new_apt_pir,
                            color='#FF1493',  # 진한 핑크색
                            linestyle='-.',   # 점선
                            linewidth=2,
                            alpha=0.7,
                            label=f'신축단지 PIR ({new_apt_pir:.1f})',
                            zorder=5
                        )
                        
                        # 수평선 레이블 추가 - 위치를 왼쪽으로 이동
                        # 날짜 범위의 처음 1/4 지점에 텍스트 배치 (중간의 1/3 대신)
                        date_range = df['date'].max() - df['date'].min()
                        left_position_date = df['date'].min() + (date_range * 0.1)  # 더 왼쪽으로 이동
                        
                        ax1.text(
                            left_position_date,  # 왼쪽으로 이동된 위치
                            normalized_new_apt_pir + (y_range * 0.03), # 약간 위에 텍스트 표시
                            f"{self.new_apt_info['name']} 신축단지 PIR: {new_apt_pir:.1f}",
                            color='#FF1493',
                            fontsize=10,
                            bbox=dict(facecolor='white', edgecolor='#FF1493', alpha=0.7, boxstyle='round,pad=0.3')
                        )
                        
                        # 범례 라인에 신축단지 PIR 추가
                        if 'lines' in locals():
                            new_apt_pir_line = plt.Line2D(
                                [0], [0],
                                color='#FF1493',
                                linestyle='-.',
                                linewidth=2,
                                label=f'신축단지 PIR ({new_apt_pir:.1f})'
                            )
                            lines.append(new_apt_pir_line)
                        
                except Exception as e:
                    print(f"신축단지 PIR 처리 중 오류: {str(e)}")
        
        



            
    
    
    
            # 먼저 함수 정의
            # 먼저 함수 정의
            def create_annotation_with_line(ax, text, xy, xytext, color='black', bbox_fc='yellow'):
                return ax.annotate(
                    text,
                    xy=xy,             # 데이터 포인트 위치
                    xytext=xytext,     # 텍스트 위치
                    textcoords='offset points',
                    bbox=dict(boxstyle='round,pad=0.5', fc=bbox_fc, alpha=0.7),
                    arrowprops=dict(
                        arrowstyle='-',        # 화살표 없는 직선
                        connectionstyle='angle3',  # 꺾인 선 스타일
                        color=color,
                        lw=1
                    ),
                    fontsize=9,
                    zorder=6
                )
            
            # 그 다음 매매가 주요 포인트에 마커와 레이블 추가하는 부분
            prev_peak = find_previous_peak(df, max_pir_idx, current_idx, max_sale_idx)

            important_points_sale = []
            
            # 현재가 마커
            current_sale = (latest_date, df.loc[current_idx, sale_col_name])
            important_points_sale.append(current_sale)
            
            # 전고점 마커
            max_sale = (df.loc[max_sale_idx, 'date'], df.loc[max_sale_idx, '매매가'])
            important_points_sale.append(max_sale)
            
            # 매매가 마커 일괄 추가
            dates_sale = [point[0] for point in important_points_sale]
            prices_sale = [point[1] for point in important_points_sale]
            ax1.scatter(dates_sale, prices_sale,
                       color='white',
                       edgecolor='#0099FF',
                       s=100,
                       linewidth=2,
                       zorder=5)
            
            # 매매가 어노테이션
            ax1.annotate(f'{df.loc[current_idx, "매매가"]:,.0f}\n(현재가)',
                        xy=(latest_date, df.loc[current_idx, sale_col_name]),
                        xytext=(10, -10),
                        textcoords='offset points',
                        bbox=dict(boxstyle='round,pad=0.5', fc='white', alpha=0.5),
                        fontsize=10,
                        zorder=6)
        
            # 실거래가 점도표 부분 수정
            if trades:
                high_trades = [t for t in trades if t['floor'] > 4]
                low_trades = [t for t in trades if t['floor'] <= 4]
                
                # 최근 3개월 데이터만 필터링
                three_months_ago = latest_date - timedelta(days=45)
                recent_trades = [t for t in trades if t['date'] >= three_months_ago]
                
                # 최근 5층 이상 최저가 거래 찾기
                recent_high_trades = [t for t in recent_trades if t['floor'] > 4]
                if recent_high_trades:
                    min_recent_high = min(recent_high_trades, key=lambda x: x['price'])
                    create_annotation_with_line(
                        ax1,
                        f"최근 최저거래가\n"
                        f"{min_recent_high['price']:,}만원\n"
                        f"{min_recent_high['date'].strftime('%Y-%m-%d')}\n"
                        f"{min_recent_high['floor']}층",
                        xy=(min_recent_high['date'], min_recent_high['price']),
                        xytext=(30, -80),  # 위치 조정
                        bbox_fc='lightblue'
                    )

            # analyze_prices 메서드 내부의 실거래가 점도표 부분을 수정
            
            # 색상 정의
            high_floor_color = '#69B1FF'  # 매매 5층 이상 (파란색)
            low_floor_color = '#C0C0C0'   # 매매 5층 미만 (회색)
            lease_high_floor_color = '#FFB366'  # 전세 5층 이상 (연한 주황색)
            lease_low_floor_color = '#FFE5B4'   # 전세 5층 미만 (노란색)
            
            # 매매 실거래가 표시
            if self.show_real_trade.get() and trades:
                print(f"매매 실거래가 그래프 표시: {len(trades)}건")
                high_trades = [t for t in trades if t['floor'] > 4]
                low_trades = [t for t in trades if t['floor'] <= 4]
                
                if high_trades:
                    dates_high = [t['date'] for t in high_trades]
                    prices_high = [t['price'] for t in high_trades]
                    ax1.scatter(dates_high, prices_high, color=high_floor_color,
                               alpha=0.4, s=50, edgecolors='white', linewidths=0.5,
                               label='매매실거래(5층↑)', zorder=3)

                if low_trades:
                    dates_low = [t['date'] for t in low_trades]
                    prices_low = [t['price'] for t in low_trades]
                    ax1.scatter(dates_low, prices_low, color=low_floor_color,
                               alpha=0.4, s=50, edgecolors='white', linewidths=0.5,
                               label='매매실거래(4층↓)', zorder=3)
            
            else:
                print("매매 실거래가 그래프 표시 생략")   
            # 전세 실거래가 데이터 가져오기
            leases = None
            if self.show_real_trade.get():
                print("전세 실거래가 조회 활성화")
            
                try:
                    # sido, sigungu, dong이 모두 유효한 경우에만 전세 데이터 조회
                    if all([sido and sido.strip(), sigungu and sigungu.strip(), dong and dong.strip(), apt_name]):
                        print(f"전세 실거래가 조회 시작: {sido} {sigungu} {dong} {apt_name}")
                        leases = self.get_lease_data(sido, sigungu, dong, apt_name, target_area)
                    else:
                        print(f"전세 실거래가 조회 생략 - 주소 정보 불완전: sido='{sido}', sigungu='{sigungu}', dong='{dong}'")
                        leases = []
                except Exception as e:
                    print(f"전세 실거래가 조회 중 오류: {str(e)}")
                    leases = []
            else:
                print("전세 실거래가 조회 비활성화")
                leases = None            
      
            # 전세 실거래가 표시

            if self.show_real_trade.get() and leases:
                print(f"전세 실거래가 그래프 표시: {len(leases)}건")
            
                high_leases = [t for t in leases if t['floor'] > 4]
                low_leases = [t for t in leases if t['floor'] <= 4]
                
                if high_leases:
                    dates_high = [t['date'] for t in high_leases]
                    prices_high = [t['price'] for t in high_leases]
                    ax1.scatter(dates_high, prices_high, color=lease_high_floor_color,
                               alpha=0.4, s=50, edgecolors='white', linewidths=0.5,
                               label='전세실거래(5층↑)', zorder=3)

                if low_leases:
                    dates_low = [t['date'] for t in low_leases]
                    prices_low = [t['price'] for t in low_leases]
                    ax1.scatter(dates_low, prices_low, color=lease_low_floor_color,
                               alpha=0.4, s=50, edgecolors='white', linewidths=0.5,
                               label='전세실거래(4층↓)', zorder=3)

            else:
                print("전세 실거래가 그래프 표시 생략")                    
            
            # 범례에 전세 실거래 추가
            if leases and any(t['floor'] > 4 for t in leases):
                lines.append(plt.Line2D([0], [0], color=lease_high_floor_color, marker='o',
                                       linestyle='None', markersize=8, alpha=0.6,
                                       label='전세실거래(5층↑)'))
            if leases and any(t['floor'] <= 4 for t in leases):
                lines.append(plt.Line2D([0], [0], color=lease_low_floor_color, marker='o',
                                       linestyle='None', markersize=8, alpha=0.6,
                                       label='전세실거래(4층↓)'))
##############################
                
                # 최고거래가 처리
             

                # 최고거래가 처리
    
                # 새로운 검색인 경우만 히스토리에 저장 (파일이 히스토리 경로에 없는 경우)
                if file_path and not file_path.startswith(self.history_path):
                    # 최고거래가 처리
                
                    # 최고거래가 처리
                    if trades:
                        try:
                            # 역대 최고거래가 찾기
                            all_time_max_trade = max(trades, key=lambda x: x['price'])
                            max_trade_info = f"{all_time_max_trade['price']:,}만원"
                            
                            # 최근 3개월 거래 필터링
                            three_months_ago = latest_date - timedelta(days=90)
                            recent_trades = [t for t in trades if t['date'] >= three_months_ago]
                            
                            # 최고거래가를 현재 파일에 저장
                            try:
                                if os.path.exists(file_path):
                                    wb = load_workbook(file_path)
                                    try:
                                        ws = wb.active
                                        ws['M1'] = max_trade_info
                                        wb.save(file_path)
                                    finally:
                                        wb.close()  # 파일 사용 후 반드시 닫기
                            except Exception as e:
                                print(f"엑셀 파일 처리 중 오류: {str(e)}")
                            
                            # 약간의 딜레이를 주어 파일 처리 완료 보장
                            time.sleep(0.5)
                            # 히스토리에 저장 (최고거래가 정보 포함)
                            self.save_to_history(file_path, trades)
                    
                            # 최근 3개월 내 최고거래가 처리
                            if recent_trades:
                                recent_max_trade = max(recent_trades, key=lambda x: x['price'])
                                
                                # 3개월 내 최고거래가와 역대 최고거래가 비교
                                if recent_max_trade['price'] == all_time_max_trade['price']:
                                    # 최근 최고거래가가 역대 최고거래가와 같은 경우 - "최고거래가"로만 표시
                                    create_annotation_with_line(
                                        ax1,
                                        f"최고거래가\n"
                                        f"{all_time_max_trade['price']:,}만원\n"
                                        f"{all_time_max_trade['date'].strftime('%Y-%m-%d')}\n"
                                        f"{all_time_max_trade['floor']}층",
                                        xy=(all_time_max_trade['date'], all_time_max_trade['price']),
                                        xytext=(-50, 50),
                                        bbox_fc='gold'
                                    )
                                else:
                                    # 두 값이 다른 경우 - 둘 다 표시
                                    # 역대 최고거래가
                                    create_annotation_with_line(
                                        ax1,
                                        f"최고거래가\n"
                                        f"{all_time_max_trade['price']:,}만원\n"
                                        f"{all_time_max_trade['date'].strftime('%Y-%m-%d')}\n"
                                        f"{all_time_max_trade['floor']}층",
                                        xy=(all_time_max_trade['date'], all_time_max_trade['price']),
                                        xytext=(-50, 50),
                                        bbox_fc='gold'
                                    )
                                    
                                    # 최근 최고거래가
                                    create_annotation_with_line(
                                        ax1,
                                        f"최근 최고거래가\n"
                                        f"{recent_max_trade['price']:,}만원\n"
                                        f"{recent_max_trade['date'].strftime('%Y-%m-%d')}\n"
                                        f"{recent_max_trade['floor']}층",
                                        xy=(recent_max_trade['date'], recent_max_trade['price']),
                                        xytext=(0, 50),  # 위치 조정 (겹치지 않도록)
                                        bbox_fc='lightsalmon'  # 다른 색상으로 구분
                                    )
                            else:
                                # 최근 3개월 거래가 없는 경우 - 역대 최고거래가만 표시
                                create_annotation_with_line(
                                    ax1,
                                    f"최고거래가\n"
                                    f"{all_time_max_trade['price']:,}만원\n"
                                    f"{all_time_max_trade['date'].strftime('%Y-%m-%d')}\n"
                                    f"{all_time_max_trade['floor']}층",
                                    xy=(all_time_max_trade['date'], all_time_max_trade['price']),
                                    xytext=(-50, 50),
                                    bbox_fc='gold'
                                )
                            
                        except Exception as e:
                            print(f"최고거래가 처리 중 오류: {str(e)}")
                    else:
                        # 실거래 데이터가 없는 경우에도 히스토리는 저장
                        self.save_to_history(file_path, None)
                



                
                # 기존 실거래가 점도표 그리기
                if high_trades:
                    dates_high = [t['date'] for t in high_trades]
                    prices_high = [t['price'] for t in high_trades]
                    ax1.scatter(dates_high, prices_high, color=high_floor_color, alpha=0.4,
                               s=50, edgecolors='white', linewidths=0.5,
                               label='실거래(5층↑)', zorder=1)

                if low_trades:
                    dates_low = [t['date'] for t in low_trades]
                    prices_low = [t['price'] for t in low_trades]
                    ax1.scatter(dates_low, prices_low, color=low_floor_color, alpha=0.4,
                               s=50, edgecolors='white', linewidths=0.5,
                               label='실거래(4층↓)', zorder=1)
        
            # 범례 설정
            # 수정 후
            # 범례 설정
            lines = [line_sale, line_lease]  # 기본 라인만 추가
                        
            # PIR 선이 그려졌을 경우에만 범례에 추가
            if line_pir is not None and line_pir_avg is not None:
                lines.append(line_pir)
                lines.append(line_pir_avg)
                        
            # None이 아닌 선만 필터링
            valid_lines = [l for l in lines if l is not None]
            valid_labels = [l.get_label() for l in valid_lines]
                        
            if valid_lines:  # 유효한 선이 있을 때만 범례 생성
                ax1.legend(valid_lines, valid_labels, loc='best')
            else:
                ax1.legend(lines, [l.get_label() for l in lines if l is not None], loc='best')
            
            # 실거래가 범례도 정확히 같은 색상으로 추가
            if trades and any(t['floor'] > 4 for t in trades):
                scatter_high = plt.Line2D([0], [0], color=high_floor_color, marker='o',
                                        linestyle='None', markersize=8, alpha=0.4,
                                        markeredgecolor='white', markeredgewidth=0.5,
                                        label='실거래(5층↑)')
                lines.append(scatter_high)

            if trades and any(t['floor'] <= 4 for t in trades):
                scatter_low = plt.Line2D([0], [0], color=low_floor_color, marker='o',
                                        linestyle='None', markersize=8, alpha=0.4,
                                        markeredgecolor='white', markeredgewidth=0.5,
                                        label='실거래(4층↓)')
                lines.append(scatter_low)
            
            # y축 단위를 만원으로 표시
            ax = plt.gca()
            yticks = ax.get_yticks()
            ax.set_yticks(yticks)
            ax.set_yticklabels([f'{int(x):,}' for x in yticks])
    
    
    
    
            # 우선순위가 가장 높은 매매가/전세가 유형 결정
            if sale_types:
                top_sale_type = sale_types[0] # 이미 우선순위로 정렬됨
                sale_col_name = f'매매가_{top_sale_type}'
            else:
                sale_col_name = '매매가'  # 기본값
            
            if lease_types:
                top_lease_type = lease_types[0] # 이미 우선순위로 정렬됨
                lease_col_name = f'전세가_{top_lease_type}'
            else:
                lease_col_name = '전세가'  # 기본값
            
            # 매매가 정보 계산 - 우선순위 높은 유형 사용
            # 매매가 정보 계산 - 개선된 최근 저점 찾기 로직 추가
            # 매매가 정보 계산 - 우선순위 높은 유형 사용
            # 매매가 정보 계산 - 개선된 최근 저점 찾기 로직 추가
            if df.loc[current_idx, sale_col_name] >= df[sale_col_name].max():
                # 현재가가 최고가인 경우
                price_info = {
                    'max': df.loc[current_idx, sale_col_name],
                    'max_date': latest_date,
                    'current': df.loc[current_idx, sale_col_name],
                    'current_date': latest_date,
                    'recent_low': None,
                    'recent_low_date': None
                }
                
                # 최근 12개월 데이터에서 저점을 찾기
                one_year_ago = latest_date - timedelta(days=365)
                period_mask = (df['date'] >= one_year_ago) & (df['date'] <= latest_date)
                recent_year_df = df[period_mask]
                
                if not recent_year_df.empty:
                    min_sale_idx = recent_year_df[sale_col_name].idxmin()
                    low_date = df.loc[min_sale_idx, 'date']
                    
                    # 최저점이 현재가가 아니고, 현재와 최소 1개월 차이나는 경우에만 표시
                    if min_sale_idx != current_idx:
                        months_to_current = (latest_date.year - low_date.year) * 12 + latest_date.month - low_date.month
                        if months_to_current >= 1:
                            price_info['recent_low'] = df.loc[min_sale_idx, sale_col_name]
                            price_info['recent_low_date'] = low_date
                    
                # 백업: 최근 저점이 아직도 없으면 전체 데이터에서 찾기
                if price_info['recent_low'] is None:
                    # 최근 2년 데이터에서 저점을 찾기
                    two_years_ago = latest_date - timedelta(days=730)
                    period_mask = (df['date'] >= two_years_ago) & (df['date'] <= latest_date)
                    period_df = df[period_mask]
                    
                    if not period_df.empty:
                        min_sale_idx = period_df[sale_col_name].idxmin()
                        low_date = df.loc[min_sale_idx, 'date']
                        
                        # 최저점이 현재가가 아니고 유의미한 가격 차이가 있는 경우에만 표시
                        if min_sale_idx != current_idx:
                            # 최고가와 최저가 차이가 5% 이상인 경우만 표시
                            if (price_info['max'] - df.loc[min_sale_idx, sale_col_name]) / price_info['max'] > 0.05:
                                price_info['recent_low'] = df.loc[min_sale_idx, sale_col_name]
                                price_info['recent_low_date'] = low_date
            else:
                # 기존 로직: 현재가가 최고가가 아닌 경우
                max_sale_idx_priority = df[sale_col_name].idxmax()
                period_df = df.loc[max_sale_idx_priority:current_idx]
                
                if period_df.empty:
                    price_info = {
                        'max': df.loc[max_sale_idx_priority, sale_col_name],
                        'max_date': df.loc[max_sale_idx_priority, 'date'],
                        'current': df.loc[current_idx, sale_col_name],
                        'current_date': latest_date,
                        'recent_low': None,
                        'recent_low_date': None
                    }
                else:
                    min_sale_idx = period_df[sale_col_name].idxmin()
                    
                    if min_sale_idx == current_idx:
                        price_info = {
                            'max': df.loc[max_sale_idx_priority, sale_col_name],
                            'max_date': df.loc[max_sale_idx_priority, 'date'],
                            'current': df.loc[current_idx, sale_col_name],
                            'current_date': latest_date,
                            'recent_low': None,
                            'recent_low_date': None
                        }
                    else:
                        price_info = {
                            'max': df.loc[max_sale_idx_priority, sale_col_name],
                            'max_date': df.loc[max_sale_idx_priority, 'date'],
                            'current': df.loc[current_idx, sale_col_name],
                            'current_date': latest_date,
                            'recent_low': df.loc[min_sale_idx, sale_col_name],
                            'recent_low_date': df.loc[min_sale_idx, 'date']
                        }
    
    
            # 어노테이션 처리 시
            def create_annotation_with_line(ax, text, xy, xytext, color='black', bbox_fc='yellow'):
                return ax.annotate(
                    text,
                    xy=xy,             # 데이터 포인트 위치
                    xytext=xytext,     # 텍스트 위치
                    textcoords='offset points',
                    bbox=dict(boxstyle='round,pad=0.5', fc=bbox_fc, alpha=0.7),
                    arrowprops=dict(
                        arrowstyle='-',        # 화살표 없는 직선
                        connectionstyle='angle3',  # 꺾인 선 스타일
                        color=color,
                        lw=1
                    ),
                    zorder=6
                )


                # 전고점과 이전 고점 표시 여부 확인
                # 매매가 최고점/현재가 어노테이션
                # 필요하다면 레이블에 유형 표시 추가
                if price_info["current"] == price_info["max"]:
                    create_annotation_with_line(
                        ax1,
                        f'{price_info["max"]:,.0f}\n(최고가-{sale_type_text})',
                        xy=(price_info["max_date"], price_info["max"]),
                        xytext=(-100, 0),
                        bbox_fc='#0099FF'
                    )
                else:
                    # 전고점
                    create_annotation_with_line(
                        ax1,
                        f'{price_info["max"]:,.0f}\n(전고점-{sale_type_text})',
                        xy=(price_info["max_date"], price_info["max"]),
                        xytext=(-90, 0),
                        bbox_fc='#0099FF'
                    )
                
                # 현재가 어노테이션
                if price_info["current"] != price_info["max"]:
                    ax1.annotate(f'{price_info["current"]:,.0f}\n(현재가-{sale_type_text})',
                                xy=(latest_date, price_info["current"]),
                                xytext=(10, -10), textcoords='offset points',
                                bbox=dict(boxstyle='round,pad=0.5', fc='white', alpha=0.5),
                                fontsize=10, zorder=6)
    
    
    
    
            # 전세가 어노테이션
            ax1.annotate(f'{df.loc[max_lease_idx, "전세가"]:,.0f}',
                        xy=(df.loc[max_lease_idx, 'date'], df.loc[max_lease_idx, '전세가']),
                        xytext=(10, 10), textcoords='offset points',
                        bbox=dict(boxstyle='round,pad=0.5', fc='#FF6B00', alpha=0.5),
                        fontsize=10, zorder=6)
                        
            ax1.annotate(f'{df.loc[current_idx, "전세가"]:,.0f}',
                        xy=(latest_date, df.loc[current_idx, '전세가']),
                        xytext=(10, -10), textcoords='offset points',
                        bbox=dict(boxstyle='round,pad=0.5', fc='#FF6B00', alpha=0.5),
                        fontsize=10, zorder=6)
            
            # PIR 어노테이션
            # create_annotation_with_line(
            #     ax2,
            #     f'PIR: {df.loc[max_pir_idx, "PIR"]:.1f}',
            #     xy=(df.loc[max_pir_idx, 'date'], df.loc[max_pir_idx, 'PIR']),
            #     xytext=(-300, 0),
            #     bbox_fc='#FF9999'
            # )
            
            # create_annotation_with_line(
            #     ax2,
            #     f'PIR: {df.loc[current_idx, "PIR"]:.1f}',
            #     xy=(latest_date, df.loc[current_idx, 'PIR']),
            #     xytext=(-40, -40),
            #     bbox_fc='#FF9999'
            # )
    
            # 정보 텍스트 영역 설정 - 2행으로 변경
            # 정보 텍스트 영역 설정 - 1행 6열로 단순화
            # 정보 텍스트 영역 설정 (원래대로 4x6 그리드)
            # 정보 텍스트 영역 설정
            ax_info_complex = plt.subplot2grid((4, 6), (3, 0))
            ax_info_price1 = plt.subplot2grid((4, 6), (3, 1))
            ax_info_price2 = plt.subplot2grid((4, 6), (3, 2))
            ax_info_lease = plt.subplot2grid((4, 6), (3, 3))
            ax_info_pir = plt.subplot2grid((4, 6), (3, 4))
            ax_info_trade = plt.subplot2grid((4, 6), (3, 5))  # 거래 패턴을 6번째 박스로 추가
            
            for ax in [ax_info_complex, ax_info_price1, ax_info_price2, 
                      ax_info_lease, ax_info_pir, ax_info_trade]:  # ax_info_trade 추가
                ax.axis('off')
    
            # 단지 정보 텍스트
            complex_info_text = "단지 정보\n\n"
            for key, value in complex_details.items():
                complex_info_text += f"{key}: {value}\n"
    
            # 매매가 정보 1 (전저점/전고점)
            # 매매가 정보 1 (전저점/전고점/이전고점)
            price_info1_text = "매매가 정보\n\n"
            
            # 현재가와 전고점 시점의 차이 계산
            months_diff = (latest_date.year - price_info["max_date"].year) * 12 + latest_date.month - price_info["max_date"].month
            show_max_peak = months_diff > 6  # 전고점 표시 여부
            
            if old_low_point is not None:
                if show_max_peak:
                    # 전고점이 있는 경우 - 전저점 대비 전고점 상승률
                    old_to_max_rate = ((price_info['max'] - df.loc[old_low_point, '매매가']) / 
                                     df.loc[old_low_point, '매매가'] * 100)
                    price_info1_text += f"전저점: {df.loc[old_low_point, '매매가']:,.0f}\n"
                    price_info1_text += f"({df.loc[old_low_point, 'date'].strftime('%Y-%m')})\n"
                    price_info1_text += f"전고점 상승률: [{old_to_max_rate:.1f}%]\n\n"
                else:
                    # 전고점이 없는 경우 - 전저점 대비 이전고점 상승률
                    prev_peak = find_previous_peak(df, max_pir_idx, current_idx, max_sale_idx)
                    if prev_peak:
                        old_to_prev_rate = ((prev_peak['price'] - df.loc[old_low_point, '매매가']) / 
                                          df.loc[old_low_point, '매매가'] * 100)
                        price_info1_text += f"전저점: {df.loc[old_low_point, '매매가']:,.0f}\n"
                        price_info1_text += f"({df.loc[old_low_point, 'date'].strftime('%Y-%m')})\n"
                        price_info1_text += f"이전고점 상승률: [{old_to_prev_rate:.1f}%]\n\n"
            
            # 전고점/이전고점 정보 추가
            if show_max_peak:
                label = "최고가" if price_info["current"] == price_info["max"] else "전고점"
                price_info1_text += f"{label}: {price_info['max']:,.0f}\n"
                price_info1_text += f"({price_info['max_date'].strftime('%Y-%m')})"
            else:
                # 전고점이 없을 때는 이전고점 정보 표시
                prev_peak = find_previous_peak(df, max_pir_idx, current_idx, max_sale_idx)
                if prev_peak:
                    price_info1_text += f"이전고점: {prev_peak['price']:,.0f}\n"
                    price_info1_text += f"({prev_peak['date'].strftime('%Y-%m')})"
    
    
    
            # 매매가 정보 2 (최근저점/현재가)
            # 1. 매매가 정보 2 (최근저점/현재가) 텍스트 부분 수정
            # 매매가 정보 2 (최근저점/현재가)
            # 1. 매매가 정보 2 (최근저점/현재가) 텍스트 부분 수정
            price_info2_text = "매매가 정보\n\n"
            
            # points_to_mark에서 recent_low 타입의 포인트 찾기
            recent_low_point = None
            for point in points_to_mark:
                if point['type'] == 'recent_low':
                    recent_low_point = point
                    break
            
            # 어노테이션의 최근 저점 정보 활용
            if recent_low_point:
                # 기준점 설정 (현재가가 최고가일 경우 표현 방식 변경)
                reference_point = '현재가' if price_info['current'] == price_info['max'] else '전고점'
                
                # 최근저점과 비교 기준점(현재가 또는 전고점) 사이의 비율 계산
                compare_price = price_info['current'] if reference_point == '현재가' else price_info['max']
                low_drop_rate = ((recent_low_point['price'] - compare_price) / compare_price * 100)
                
                price_info2_text += f"최근저점: {recent_low_point['price']:,.0f}\n"
                price_info2_text += f"({recent_low_point['date'].strftime('%Y-%m')})\n"
                price_info2_text += f"{reference_point} 대비 [{low_drop_rate:.1f}%]"
                
                # 최근 저점 대비 현재 상승률과 상승 금액 추가
                if price_info['current'] > recent_low_point['price']:
                    # 상승 금액 계산
                    price_increase = price_info['current'] - recent_low_point['price']
                    # 상승률 계산
                    current_vs_low_rate = (price_increase / recent_low_point['price'] * 100)
                    
                    price_info2_text += f"\n최근저점 대비 상승금액 [+{price_increase:,.0f}]\n"
                    price_info2_text += f"최근저점 대비 상승률 [+{current_vs_low_rate:.1f}%]"
            # price_info의 recent_low 사용 (어노테이션에서 얻은 것이 없을 경우 대비)
            elif price_info.get('recent_low'):
                # 기준점 설정 (현재가가 최고가일 경우 표현 방식 변경)
                reference_point = '현재가' if price_info['current'] == price_info['max'] else '전고점'
                
                # 최근저점과 비교 기준점(현재가 또는 전고점) 사이의 비율 계산
                compare_price = price_info['current'] if reference_point == '현재가' else price_info['max']
                low_drop_rate = ((price_info['recent_low'] - compare_price) / compare_price * 100)
                
                price_info2_text += f"최근저점: {price_info['recent_low']:,.0f}\n"
                price_info2_text += f"({price_info['recent_low_date'].strftime('%Y-%m')})\n"
                price_info2_text += f"{reference_point} 대비 [{low_drop_rate:.1f}%]"
                
                # 최근 저점 대비 현재 상승률과 상승 금액 추가
                if price_info['current'] > price_info['recent_low']:
                    # 상승 금액 계산
                    price_increase = price_info['current'] - price_info['recent_low']
                    # 상승률 계산
                    current_vs_low_rate = (price_increase / price_info['recent_low'] * 100)
                    
                    price_info2_text += f"\n최근저점 대비 상승금액 [+{price_increase:,.0f}]\n"
                    price_info2_text += f"최근저점 대비 상승률 [+{current_vs_low_rate:.1f}%]"
            
            # 현재가 정보 (항상 표시하도록 수정)
            # 현재가가 최고가가 아닌 경우
            if price_info['current'] != price_info['max']:
                drop_rate = ((price_info['current'] - price_info['max']) / 
                            price_info['max'] * 100)
                price_info2_text += f"\n\n현재가: {price_info['current']:,.0f}\n"
                price_info2_text += f"({price_info['current_date'].strftime('%Y-%m')})\n"
                price_info2_text += f"전고점 대비 [{drop_rate:.1f}%]"
            # 현재가가 최고가인 경우에도 정보 표시
            else:
                # 최근저점 정보가 없을 경우에만 새 줄 추가 (있으면 이미 위에서 표시)
                if not price_info.get('recent_low'):
                    price_info2_text += "\n"
                else:
                    price_info2_text += "\n\n"
                price_info2_text += f"현재가(최고가): {price_info['current']:,.0f}\n"
                price_info2_text += f"({price_info['current_date'].strftime('%Y-%m')})"
            # 전세가 정보 
            center_info_text = (
                f"전세가 정보\n\n"
                f"최고가: {df.loc[max_lease_idx, '전세가']:,.0f}\n"
                f"({df.loc[max_lease_idx, 'date'].strftime('%Y-%m')})\n\n"
                f"현재가: {df.loc[current_idx, '전세가']:,.0f}\n"
                f"({latest_date.strftime('%Y-%m')})\n"
                f"전고점 대비 [{((df.loc[current_idx, '전세가'] - df.loc[max_lease_idx, '전세가']) / df.loc[max_lease_idx, '전세가'] * 100):.1f}%]"
            )
    
            # PIR 정보
            # PIR 정보
            # PIR 정보 (연복리 정보 포함)
            right_info_text = (
                f"PIR / 연복리\n\n"
                f"PIR 최고: {df.loc[max_pir_idx, 'PIR']:.1f}\n"
                f"({df.loc[max_pir_idx, 'date'].strftime('%Y-%m')})\n"
                f"PIR 현재: {df.loc[current_idx, 'PIR']:.1f}\n"
                f"({latest_date.strftime('%Y-%m')})\n"
                f"전고점 대비 [{((df.loc[current_idx, 'PIR'] - df.loc[max_pir_idx, 'PIR']) / df.loc[max_pir_idx, 'PIR'] * 100):.1f}%]\n\n"
                f"매매 연복리: {sale_cagr:.1f}%\n"
                f"전세 연복리: {lease_cagr:.1f}%"
            )
            

            # 거래 패턴 정보 텍스트
            # 거래 패턴 정보 텍스트
            trade_info_text = "거래 패턴\n\n"
            if trade_patterns:
                # 최근 3개월 평균 거래량 추가
                if trade_patterns.get('recent_3month_avg') is not None:
                    trade_info_text += f"최근3개월평균  {trade_patterns['recent_3month_avg']:.1f}건/월\n\n"
                
                if trade_patterns['recent_turnover'] is not None:
                    trade_info_text += f"최근연간회전율  {trade_patterns['recent_turnover']:.1f}%\n\n"
                else:
                    trade_info_text += f"최근연간회전율  세대수 정보없음\n\n"
                
                if trade_patterns['avg_turnover'] is not None:
                    trade_info_text += f"평균연간회전율  {trade_patterns['avg_turnover']:.1f}%\n\n"
                else:
                    trade_info_text += f"평균연간회전율  세대수 정보없음\n\n"
                
                if trade_patterns['most_traded_month'] is not None:
                    month_names = {1:'1월', 2:'2월', 3:'3월', 4:'4월', 5:'5월', 6:'6월',
                                  7:'7월', 8:'8월', 9:'9월', 10:'10월', 11:'11월', 12:'12월'}
                    month_text = month_names[trade_patterns['most_traded_month']]
                    
                    # 평균 거래량 추가
                    if trade_patterns.get('most_traded_month_avg'):
                        trade_info_text += f"최다거래월  {month_text}\n({trade_patterns['most_traded_month_avg']:.1f}건/년)"
                    else:
                        trade_info_text += f"최다거래월  {month_text}"
                else:
                    trade_info_text += f"최다거래월  정보없음"
            else:
                trade_info_text += "실거래\n데이터없음"
            
            # 연복리 정보는 별도로 처리 (기존 코드 유지)
            cagr_info_text = (
                f"연복리 정보\n\n"
                f"매매가\n"
                f"연복리: [{sale_cagr:.1f}%]\n\n"
                f"전세가\n"
                f"연복리: [{lease_cagr:.1f}%]"
            )
            
            # 텍스트 스타일 설정
            text_style = dict(
                ha='center', 
                va='center',
                fontsize=8,
                fontweight='bold',
                linespacing=1.5,
                bbox=dict(
                    facecolor='white',
                    edgecolor='#6366F1',
                    linewidth=2,
                    alpha=1,
                    boxstyle='round4,pad=0.6,rounding_size=0.2'
                )
            )
            
            # 연복리 텍스트 스타일 (가로로 긴 박스)
            cagr_style = dict(
                ha='center',
                va='center', 
                fontsize=9,
                fontweight='bold',
                bbox=dict(
                    facecolor='#F0F0F0',
                    edgecolor='#808080',
                    linewidth=1,
                    alpha=1,
                    boxstyle='round4,pad=0.5,rounding_size=0.1'
                )
            )
            
            # 각 정보 텍스트 표시
            # 각 정보 텍스트 표시
            # 각 정보 텍스트 표시
            ax_info_complex.text(0.5, 0.5, complex_info_text, transform=ax_info_complex.transAxes, **text_style)
            ax_info_price1.text(0.5, 0.5, price_info1_text, transform=ax_info_price1.transAxes, **text_style)
            ax_info_price2.text(0.5, 0.5, price_info2_text, transform=ax_info_price2.transAxes, **text_style)
            ax_info_lease.text(0.5, 0.5, center_info_text, transform=ax_info_lease.transAxes, **text_style)
            ax_info_pir.text(0.5, 0.5, right_info_text, transform=ax_info_pir.transAxes, **text_style)
            ax_info_trade.text(0.5, 0.5, trade_info_text, transform=ax_info_trade.transAxes, **text_style)  # 추가
            # ax_info_cagr 관련 줄 삭제
            
            # 거래 패턴 정보 텍스트 생성

    
    

            # analyze_prices 메서드 내
            location_df = pd.read_excel(file_path, nrows=7)
            area1 = location_df.iloc[2, 1]
            area2 = location_df.iloc[5, 1]
            
            # 준공년 포함하여 차트 제목 생성
            # 매매가/전세가 유형 텍스트 추가
            # analyze_prices 메서드 내 그래프 제목 설정 부분을 더 명확하게 수정
            # 이 코드를 약 3300줄 근처의 차트 제목 설정 부분에 적용하세요
            
            # 그래프 제목 설정 코드 부분 (약 3300줄 근처)
            print("\n=== 그래프 제목 설정 ===")
            
            # 선택된 가격 유형 확인
            sale_types = [t for t, var in self.sale_price_types.items() if var.get()]
            lease_types = [t for t, var in self.lease_price_types.items() if var.get()]
            
            # 우선순위에 따라 정렬 (상위 > 일반 > 하위)
            priority_order = {"high": 0, "normal": 1, "low": 2}
            sale_types = sorted(sale_types, key=lambda x: priority_order[x])
            lease_types = sorted(lease_types, key=lambda x: priority_order[x])
            
            print(f"우선순위 정렬 후 매매가 유형: {sale_types}")
            print(f"우선순위 정렬 후 전세가 유형: {lease_types}")
            
            # 가격 유형 매핑 정보 (레이블 생성용)
            price_type_labels = {
                "low": "하위평균",
                "normal": "일반평균",
                "high": "상위평균"
            }
            
            # 그래프 제목에 선택한 가격 유형 표시 - 모든 선택된 유형 표시
            sale_type_text = ", ".join([price_type_labels[t] for t in sale_types])
            lease_type_text = ", ".join([price_type_labels[t] for t in lease_types])
            
            print(f"결정된 매매가 유형 텍스트: {sale_type_text}")
            print(f"결정된 전세가 유형 텍스트: {lease_type_text}")
            
            # 차트 제목 설정
            if hasattr(self, 'completion_year') and self.completion_year:
                chart_title = f"{self.completion_year}년 {area1} {area2} 매전차트\n매매가: {sale_type_text}, 전세가: {lease_type_text}"
            else:
                chart_title = f"{area1} {area2} 매전차트\n매매가: {sale_type_text}, 전세가: {lease_type_text}"
            
            print(f"차트 제목: {chart_title}")




            ax1.set_title(chart_title, pad=20, fontsize=20, fontweight='bold')
            
            ax1.set_title(chart_title, pad=20, fontsize=20, fontweight='bold')
            # y축 레이블 설정 부분 수정
            ax1.set_xlabel('')
            ax1.set_ylabel('가격(만원)')
            ax2.set_ylabel('심리지수(0-100)')

            # 그리드 추가 (가독성 개선)
            ax1.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, color='gray', axis='both')
            ax1.set_axisbelow(True)  # 그리드를 그래프 아래로

            ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=12))
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
            plt.setp(ax1.get_xticklabels(), rotation=-90, ha='center')
            
            # 범례

            # 색상 정의 부분은 그대로 유지
            lease_high_floor_color = '#FFB366'  # 전세 5층 이상 (연한 주황색)
            lease_low_floor_color = '#FFE5B4'   # 전세 5층 미만 (노란색)
            # 범례 설정 부분
            # 수정 후
            # 범례 설정
            lines = [line_sale, line_lease]  # 기본 라인만 추가
                        
            # PIR 선이 그려졌을 경우에만 범례에 추가
            if line_pir is not None and line_pir_avg is not None:
                lines.append(line_pir)
                lines.append(line_pir_avg)
                        
            # None이 아닌 선만 필터링
            valid_lines = [l for l in lines if l is not None]
            valid_labels = [l.get_label() for l in valid_lines]
                        
            if valid_lines:  # 유효한 선이 있을 때만 범례 생성
                ax1.legend(valid_lines, valid_labels, loc='best')
            else:
                ax1.legend(lines, [l.get_label() for l in lines if l is not None], loc='best')

            # 유나심리차트 선들 추가 (show_una_sentiment가 켜져 있는 경우)
            # 유나심리차트 선들 추가 (show_una_sentiment가 켜져 있는 경우)
            if hasattr(self, 'show_una_sentiment') and self.show_una_sentiment.get():
                # 'sell_line'과 'buy_line' 변수가 있는지 확인
                if 'sell_line' in locals() and sell_line is not None:
                    lines.append(sell_line)
                if 'buy_line' in locals() and buy_line is not None:
                    lines.append(buy_line)
            # 실거래가가 있는 경우 범례에 추가
 
            if self.show_real_trade.get():
                if trades and any(t['floor'] > 4 for t in trades):
                    lines.append(plt.Line2D([0], [0], color=high_floor_color, marker='o',
                                           linestyle='None', markersize=8, alpha=0.4,
                                           markeredgecolor='white', markeredgewidth=0.5,
                                           label='매매실거래(5층↑)'))

                if trades and any(t['floor'] <= 4 for t in trades):
                    lines.append(plt.Line2D([0], [0], color=low_floor_color, marker='o',
                                           linestyle='None', markersize=8, alpha=0.4,
                                           markeredgecolor='white', markeredgewidth=0.5,
                                           label='매매실거래(4층↓)'))

                # 전세 실거래가 범례 추가
                if leases and any(t['floor'] > 4 for t in leases):
                    lines.append(plt.Line2D([0], [0], color=lease_high_floor_color, marker='o',
                                           linestyle='None', markersize=8, alpha=0.4,
                                           markeredgecolor='white', markeredgewidth=0.5,
                                           label='전세실거래(5층↑)'))

                if leases and any(t['floor'] <= 4 for t in leases):
                    lines.append(plt.Line2D([0], [0], color=lease_low_floor_color, marker='o',
                                           linestyle='None', markersize=8, alpha=0.4,
                                           markeredgecolor='white', markeredgewidth=0.5,
                                           label='전세실거래(4층↓)'))
            
            # 범례 위치를 좌측 하단으로 설정
            # 범례 위치를 우측 하단으로 설정
            # 범례 위치 수정 부분 (약 3300줄 근처)
            ax1.legend(lines, [l.get_label() for l in lines], 
                      loc='upper left',     # 'lower right'에서 'upper left'로 변경
                      bbox_to_anchor=(0.02, 0.98),  # 좌측 상단에 위치하도록 조정
                      ncol=1,               
                      fontsize=9,           
                      framealpha=0.8)

            ax1.text(0, 1.02,  # x=0은 왼쪽 끝, y=1.02는 그래프 위쪽
                     '만든사람 부태리\nhttps://blog.naver.com/landlover333',
                     fontsize=8,
                     color='gray',
                     alpha=0.6,
                     transform=ax1.transAxes,
                     verticalalignment='bottom')
            
            # 워터마크 추가 (그래프 저장 직전)
            # plt.savefig 호출 직전에 추가
            ax1.text(0.02, 0.02,  # 그래프 영역 내 좌표 (0~1 사이 값)
                     '만든사람 부태리\nhttps://blog.naver.com/landlover333',
                     fontsize=8,
                     color='gray',
                     alpha=0.6,
                     transform=ax1.transAxes,  # 그래프 좌표계 사용
                     verticalalignment='bottom')
            
            # 여백 조정
            plt.subplots_adjust(
                top=0.91,
                bottom=0.15,
                left=0.08,
                right=0.92,
                hspace=0.4
            )
    
    
        # 파일명 생성을 위해 GUI 입력값 가져오기
        # 파일명 생성을 위해 GUI 입력값 가져오기

        
            # 특수문자 제거 및 공백을 언더스코어로 변경
            # 특수문자 제거 및 공백을 언더스코어로 변경
            apt_name = ''.join(char for char in apt_name if char.isalnum() or char.isspace())
            apt_name = apt_name.replace(' ', '_')
            
            # 설정 정보를 파일명에 추가
            pir_suffix = "pir" if self.show_pir.get() else "no_pir"
            una_suffix = "una" if self.show_una_sentiment.get() else "no_una"
            settings_suffix = f"{pir_suffix}_{una_suffix}"
            
            # 파일명 생성 (설정 정보 포함)
            filename = f"{apt_name}_{area}m2_{settings_suffix}.jpg"
            self.image_path = os.path.join(self.download_path, filename)
            
            # 이미 그래프 파일이 존재하면 분석만 하고 그래프는 생성하지 않음
            if os.path.exists(self.image_path):
                plt.close()
                return df
            
    
            plt.savefig(self.image_path, bbox_inches='tight', dpi=600, pad_inches=0.3)
            plt.close()
            
            # # 새로운 검색인 경우에만 히스토리에 저장
            # if file_path and not file_path.startswith(self.history_path):
            #     self.save_to_history(file_path)
                
            return df
    
        except Exception as e:
            messagebox.showerror("오류", f"분석 중 오류 발생: {str(e)}")
            return None


    def get_latest_excel(self):
        print(f"\n=== 최신 엑셀 파일 검색 ===")
        print(f"검색 경로: {self.download_path}")
        
        excel_files = [(f, os.path.getmtime(os.path.join(self.download_path, f))) 
                       for f in os.listdir(self.download_path) if f.endswith(('.xlsx', '.xls'))]
        
        if not excel_files:
            print("엑셀 파일을 찾을 수 없음")
            raise FileNotFoundError("엑셀 파일을 찾을 수 없습니다")
        
        latest_file = max(excel_files, key=lambda x: x[1])[0]
        file_path = os.path.join(self.download_path, latest_file)
        print(f"발견된 최신 파일: {file_path}")
        print(f"파일 수정 시간: {datetime.fromtimestamp(os.path.getmtime(file_path))}")
        
        return file_path


    def show_graph(self):
        if os.path.exists(self.image_path):
            # Windows의 경우 기본 이미지 뷰어로 열기
            os.startfile(self.image_path)

    def show_multi_analysis_dialog(self):
        """다중 단지 분석을 위한 입력 대화상자"""
        self.multi_dialog = tk.Toplevel(self.root)
        self.multi_dialog.title("다중 단지 분석")
        self.multi_dialog.attributes('-topmost', True)
        
        # 창 크기와 위치 설정
        width = 600
        height = 500
        screen_width = self.multi_dialog.winfo_screenwidth()
        screen_height = self.multi_dialog.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.multi_dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # 메인 프레임
        main_frame = ttk.Frame(self.multi_dialog, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        # 상단 설명
        title_label = ttk.Label(
            main_frame, 
            text="다중 단지 분석 - 여러 단지의 매전차트를 일괄 생성",
            font=self.font_title if hasattr(self, 'font_title') else ('Arial', 12, 'bold')
        )
        title_label.pack(pady=(0, 10))
        
        # 입력 영역
        input_frame = ttk.LabelFrame(main_frame, text="단지 정보 입력", padding=10)
        input_frame.pack(fill='x', pady=(0, 10))
        
        # 단일 입력 행
        single_frame = ttk.Frame(input_frame)
        single_frame.pack(fill='x', pady=5)
        
        ttk.Label(single_frame, text="아파트명:").grid(row=0, column=0, sticky='w', padx=(0, 5))
        self.multi_apt_name = ttk.Entry(single_frame, width=20)
        self.multi_apt_name.grid(row=0, column=1, padx=5)
        
        ttk.Label(single_frame, text="전용면적:").grid(row=0, column=2, sticky='w', padx=(10, 5))
        self.multi_area = ttk.Entry(single_frame, width=10)
        self.multi_area.grid(row=0, column=3, padx=5)
        
        # 추가 버튼
        add_button = ttk.Button(
            single_frame,
            text="추가",
            command=self.add_to_multi_list,
            style="Custom.TButton"
        )
        add_button.grid(row=0, column=4, padx=(10, 0))
        
        # 벌크 입력 영역
        bulk_frame = ttk.Frame(input_frame)
        bulk_frame.pack(fill='x', pady=(10, 0))
        
        ttk.Label(bulk_frame, text="일괄 입력 (한 줄에 하나씩: 아파트명,전용면적):").pack(anchor='w')
        
        bulk_text_frame = ttk.Frame(bulk_frame)
        bulk_text_frame.pack(fill='x', pady=5)
        
        self.bulk_text = tk.Text(bulk_text_frame, height=4, width=50)
        bulk_scrollbar = ttk.Scrollbar(bulk_text_frame, orient="vertical", command=self.bulk_text.yview)
        self.bulk_text.configure(yscrollcommand=bulk_scrollbar.set)
        
        self.bulk_text.pack(side='left', fill='both', expand=True)
        bulk_scrollbar.pack(side='right', fill='y')
        
        # 일괄 추가 버튼
        bulk_add_button = ttk.Button(
            bulk_frame,
            text="일괄 추가",
            command=self.add_bulk_to_multi_list,
            style="Custom.TButton"
        )
        bulk_add_button.pack(pady=5)
        
        # 단지 목록 영역
        list_frame = ttk.LabelFrame(main_frame, text="분석 대상 단지 목록", padding=10)
        list_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        # 트리뷰로 목록 표시
        columns = ('apt_name', 'area')
        self.multi_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=8)
        
        self.multi_tree.heading('apt_name', text='아파트명')
        self.multi_tree.heading('area', text='전용면적(㎡)')
        
        self.multi_tree.column('apt_name', width=300)
        self.multi_tree.column('area', width=100)
        
        # 스크롤바
        multi_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.multi_tree.yview)
        self.multi_tree.configure(yscrollcommand=multi_scrollbar.set)
        
        self.multi_tree.pack(side='left', fill='both', expand=True)
        multi_scrollbar.pack(side='right', fill='y')
        
        # 목록 관리 버튼
        list_button_frame = ttk.Frame(list_frame)
        list_button_frame.pack(side='bottom', fill='x', pady=(5, 0))
        
        delete_button = ttk.Button(
            list_button_frame,
            text="선택 삭제",
            command=self.delete_from_multi_list,
            style="Custom.TButton"
        )
        delete_button.pack(side='left', padx=(0, 5))
        
        clear_button = ttk.Button(
            list_button_frame,
            text="전체 삭제",
            command=self.clear_multi_list,
            style="Custom.TButton"
        )
        clear_button.pack(side='left')
        
        # 하단 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x')
        
        # 분석 시작 버튼
        self.multi_start_button = ttk.Button(
            button_frame,
            text="다중 분석 시작",
            command=self.start_multi_analysis,
            style="Custom.TButton"
        )
        self.multi_start_button.pack(side='right', padx=(5, 0))
        
        # 취소 버튼
        cancel_button = ttk.Button(
            button_frame,
            text="취소",
            command=self.multi_dialog.destroy,
            style="Custom.TButton"
        )
        cancel_button.pack(side='right')
        
        # 엔터키 바인딩
        self.multi_apt_name.bind('<Return>', lambda e: self.multi_area.focus())
        self.multi_area.bind('<Return>', lambda e: self.add_to_multi_list())

    def add_to_multi_list(self):
        """단일 입력을 목록에 추가"""
        apt_name = self.multi_apt_name.get().strip()
        area = self.multi_area.get().strip()
        
        if not apt_name or not area:
            messagebox.showwarning("입력 오류", "아파트명과 전용면적을 모두 입력해주세요.")
            return
        
        # 면적 유효성 검사
        try:
            float(area)
        except ValueError:
            messagebox.showwarning("입력 오류", "전용면적은 숫자만 입력해주세요.")
            return
        
        # 중복 검사
        for item in self.multi_tree.get_children():
            values = self.multi_tree.item(item, 'values')
            if values[0] == apt_name and values[1] == area:
                messagebox.showwarning("중복", "이미 추가된 단지입니다.")
                return
        
        # 목록에 추가
        self.multi_tree.insert('', 'end', values=(apt_name, area))
        
        # 입력 필드 초기화
        self.multi_apt_name.delete(0, tk.END)
        self.multi_area.delete(0, tk.END)
        self.multi_apt_name.focus()

    def add_bulk_to_multi_list(self):
        """일괄 입력을 목록에 추가"""
        bulk_text = self.bulk_text.get('1.0', tk.END).strip()
        
        if not bulk_text:
            messagebox.showwarning("입력 오류", "일괄 입력 내용이 비어있습니다.")
            return
        
        lines = bulk_text.split('\n')
        added_count = 0
        error_lines = []
        
        for i, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue
                
            parts = line.split(',')
            if len(parts) != 2:
                error_lines.append(f"라인 {i}: 형식 오류 (쉼표로 구분)")
                continue
                
            apt_name = parts[0].strip()
            area = parts[1].strip()
            
            if not apt_name or not area:
                error_lines.append(f"라인 {i}: 빈 값")
                continue
                
            try:
                float(area)
            except ValueError:
                error_lines.append(f"라인 {i}: 면적이 숫자가 아님")
                continue
            
            # 중복 검사
            duplicate = False
            for item in self.multi_tree.get_children():
                values = self.multi_tree.item(item, 'values')
                if values[0] == apt_name and values[1] == area:
                    duplicate = True
                    break
            
            if duplicate:
                error_lines.append(f"라인 {i}: 중복")
                continue
            
            # 목록에 추가
            self.multi_tree.insert('', 'end', values=(apt_name, area))
            added_count += 1
        
        # 결과 메시지
        if added_count > 0:
            message = f"{added_count}개 단지가 추가되었습니다."
            if error_lines:
                message += f"\n\n오류 라인:\n" + "\n".join(error_lines)
            messagebox.showinfo("처리 완료", message)
            
            # 텍스트 영역 초기화
            self.bulk_text.delete('1.0', tk.END)
        else:
            if error_lines:
                messagebox.showerror("오류", "추가된 단지가 없습니다.\n\n오류 라인:\n" + "\n".join(error_lines))
            else:
                messagebox.showwarning("알림", "추가할 유효한 단지가 없습니다.")

    def delete_from_multi_list(self):
        """선택된 항목을 목록에서 삭제"""
        selection = self.multi_tree.selection()
        if not selection:
            messagebox.showwarning("선택 오류", "삭제할 항목을 선택해주세요.")
            return
        
        for item in selection:
            self.multi_tree.delete(item)

    def clear_multi_list(self):
        """목록 전체 삭제"""
        if not self.multi_tree.get_children():
            messagebox.showwarning("알림", "삭제할 항목이 없습니다.")
            return
        
        if messagebox.askyesno("확인", "모든 항목을 삭제하시겠습니까?"):
            for item in self.multi_tree.get_children():
                self.multi_tree.delete(item)

    def start_multi_analysis(self):
        """다중 분석 시작"""
        # 목록 확인
        items = self.multi_tree.get_children()
        if not items:
            messagebox.showwarning("목록 오류", "분석할 단지를 추가해주세요.")
            return
        
        # 분석 대상 목록 생성
        analysis_list = []
        for item in items:
            values = self.multi_tree.item(item, 'values')
            analysis_list.append({
                'apt_name': values[0],
                'area': values[1]
            })
        
        # 확인 대화상자
        if not messagebox.askyesno("확인", 
                                  f"총 {len(analysis_list)}개 단지를 분석하시겠습니까?\n"
                                  f"시간이 오래 걸릴 수 있습니다."):
            return
        
        # 다중 분석 대화상자 닫기
        self.multi_dialog.destroy()
        
        # 분석 실행
        self.execute_multi_analysis(analysis_list)

    def execute_multi_analysis(self, analysis_list):
        """다중 분석 실행"""
        try:
            # 버튼 비활성화
            self.search_button.config(state="disabled")
            self.multi_analysis_button.config(state="disabled")
            
            # 다중 분석 진행 중 플래그 설정
            self._search_in_progress = True
            
            total_count = len(analysis_list)
            success_count = 0
            error_list = []
            
            for i, item in enumerate(analysis_list, 1):
                try:
                    apt_name = item['apt_name']
                    area = item['area']
                    
                    # 진행률 표시
                    progress_percent = (i - 1) / total_count * 100
                    self.update_progress(progress_percent, 
                                       f"분석 중... ({i}/{total_count}) {apt_name}")
                    
                    # 기존 입력 필드에 값 설정
                    self.apt_name.delete(0, tk.END)
                    self.apt_name.insert(0, apt_name)
                    self.area.delete(0, tk.END)
                    self.area.insert(0, area)
                    
                    # 브라우저 검색 실행
                    if self.browser_search(apt_name):
                        # 분석 실행
                        try:
                            df = self.analyze_prices()
                            if df is not None:
                                success_count += 1
                                print(f"✓ {apt_name} ({area}㎡) 분석 완료")
                            else:
                                error_list.append(f"{apt_name} ({area}㎡): 분석 실패")
                        except Exception as e:
                            error_list.append(f"{apt_name} ({area}㎡): 분석 오류 - {str(e)}")
                    else:
                        error_list.append(f"{apt_name} ({area}㎡): 검색 실패")
                    
                    # 잠시 대기 (서버 부하 방지)
                    time.sleep(1)
                    
                except Exception as e:
                    error_list.append(f"{apt_name} ({area}㎡): 처리 오류 - {str(e)}")
                    print(f"✗ {apt_name} ({area}㎡) 처리 중 오류: {str(e)}")
            
            # 완료 메시지
            self.update_progress(100, "다중 분석 완료!")
            
            # 결과 대화상자
            self.show_multi_analysis_result(total_count, success_count, error_list)
            
            # 히스토리 갱신
            self.history_list = self.load_history()
            self.update_history_display()
            
        except Exception as e:
            messagebox.showerror("오류", f"다중 분석 중 오류 발생: {str(e)}")
            
        finally:
            # 다중 분석 진행 중 플래그 해제
            if hasattr(self, '_search_in_progress'):
                delattr(self, '_search_in_progress')
                
            # 버튼 활성화
            self.search_button.config(state="normal")
            self.multi_analysis_button.config(state="normal")
            self.update_progress(0, "")

    def show_multi_analysis_result(self, total_count, success_count, error_list):
        """다중 분석 결과 표시"""
        result_dialog = tk.Toplevel(self.root)
        result_dialog.title("다중 분석 결과")
        result_dialog.attributes('-topmost', True)
        
        # 창 크기와 위치 설정
        width = 500
        height = 400
        screen_width = result_dialog.winfo_screenwidth()
        screen_height = result_dialog.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        result_dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # 메인 프레임
        main_frame = ttk.Frame(result_dialog, padding=10)
        main_frame.pack(fill='both', expand=True)
        
        # 결과 요약
        summary_text = f"다중 분석 완료\n\n" \
                       f"총 단지 수: {total_count}개\n" \
                       f"성공: {success_count}개\n" \
                       f"실패: {len(error_list)}개"
        
        summary_label = ttk.Label(
            main_frame, 
            text=summary_text,
            font=self.font_large if hasattr(self, 'font_large') else ('Arial', 10, 'bold')
        )
        summary_label.pack(pady=(0, 10))
        
        # 오류 목록 (있는 경우)
        if error_list:
            error_frame = ttk.LabelFrame(main_frame, text="실패 목록", padding=5)
            error_frame.pack(fill='both', expand=True, pady=(0, 10))
            
            error_text = tk.Text(error_frame, height=10, wrap=tk.WORD)
            error_scrollbar = ttk.Scrollbar(error_frame, orient="vertical", command=error_text.yview)
            error_text.configure(yscrollcommand=error_scrollbar.set)
            
            for error in error_list:
                error_text.insert(tk.END, f"• {error}\n")
            
            error_text.config(state=tk.DISABLED)
            
            error_text.pack(side='left', fill='both', expand=True)
            error_scrollbar.pack(side='right', fill='y')
        
        # 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x')
        
        # 폴더 열기 버튼
        open_folder_button = ttk.Button(
            button_frame,
            text="결과 폴더 열기",
            command=self.open_download_folder,
            style="Custom.TButton"
        )
        open_folder_button.pack(side='left')
        
        # 확인 버튼
        ok_button = ttk.Button(
            button_frame,
            text="확인",
            command=result_dialog.destroy,
            style="Custom.TButton"
        )
        ok_button.pack(side='right')
        
        # 성공한 경우 폴더 자동 열기 옵션
        if success_count > 0:
            if messagebox.askyesno("폴더 열기", 
                                  f"{success_count}개 단지 분석이 완료되었습니다.\n"
                                  f"결과 폴더를 열어보시겠습니까?",
                                  parent=result_dialog):
                self.open_download_folder()

    def run(self):
        self.root.mainloop()


class AptSelectDialog:
    def __init__(self, parent, apt_list, service_key, sigungu_code, dong, title="아파트 선택"):
        self.parent = parent
        self.service_key = service_key
        self.sigungu_code = sigungu_code
        self.dong = dong
        self.apt_list = apt_list
        self.result = None
        self.selected_apt = None
        self.trade_data = {}
        
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.attributes('-topmost', True)
        
        # 창 크기와 위치 설정
        width = 800
        height = 500
        screen_width = self.top.winfo_screenwidth()
        screen_height = self.top.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.top.geometry(f"{width}x{height}+{x}+{y}")
        
        # 부모의 폰트 가져오기
        if hasattr(parent, 'font_normal'):
            self.font_normal = parent.font_normal
            self.font_large = parent.font_large
            self.font_title = parent.font_title
            self.font_button = parent.font_button
        
        # 검색창 프레임
        search_frame = ttk.Frame(self.top, padding="5")
        search_frame.pack(fill='x', padx=5, pady=5)
        
        # 폰트 적용
        ttk.Label(search_frame, text="검색:", 
                 font=self.font_normal if hasattr(self, 'font_normal') else None).pack(side='left')
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_apartments)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=40,
                               font=self.font_normal if hasattr(self, 'font_normal') else None)
        search_entry.pack(side='left', fill='x', expand=True, padx=5)
        
        # 리스트박스 프레임
        list_frame = ttk.Frame(self.top, padding="5")
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        # 리스트박스에 폰트 적용
        self.listbox = tk.Listbox(list_frame, 
                                yscrollcommand=scrollbar.set,
                                font=self.font_normal if hasattr(self, 'font_normal') else ('KoPubWorld Dotum Medium', 10))
        self.listbox.pack(fill='both', expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        # 아파트 목록 초기화
        self.update_listbox(apt_list)
        
        self.listbox.bind('<ButtonRelease-1>', self.on_select)

    def update_listbox(self, items):
        self.listbox.delete(0, tk.END)
        for item in items:
            self.listbox.insert(tk.END, item)
    
    def filter_apartments(self, *args):
        search_text = self.search_var.get().lower()
        filtered_list = [apt for apt in self.apt_list if search_text in apt.lower()]
        self.update_listbox(filtered_list)
    
    def on_select(self, event):
        if self.listbox.curselection():
            full_text = self.listbox.get(self.listbox.curselection())
            # '[' 와 ']' 사이의 주소 정보 추출
            address_info = full_text[full_text.find('[')+1:full_text.find(']')]
            # 도로명 주소와 지번 주소 분리
            jibun_addr = address_info.split(' / ')[1]  # 지번 주소만 사용
            # 지번 주소에서 동과 번지만 추출
            self.simple_addr = ' '.join(jibun_addr.split()[-2:])  # 동 번지
            self.selected_apt = full_text.split('[')[0].strip()
            self.show_area_dialog()
    
    def show_area_dialog(self):
        area_list = self.get_areas_for_apt(self.selected_apt)
        
        if not area_list:
            error_dialog = tk.Toplevel(self.top)
            error_dialog.title("알림")
            error_dialog.attributes('-topmost', True)
            error_dialog.transient(self.top)
            
            ttk.Label(error_dialog, 
                     text="해당 아파트의 전용면적 정보를 찾을 수 없습니다.", 
                     padding=20,
                     font=self.font_normal if hasattr(self, 'font_normal') else None).pack()
            
            ttk.Button(error_dialog, 
                      text="확인", 
                      command=error_dialog.destroy,
                      style='Custom.TButton' if hasattr(self, 'font_button') else None).pack(pady=10)
            
            error_dialog.geometry(f"+{self.top.winfo_x() + 50}+{self.top.winfo_y() + 50}")
            error_dialog.grab_set()
            return

        area_dialog = tk.Toplevel(self.top)
        area_dialog.title(f"{self.selected_apt} - 전용면적 선택")
        area_dialog.attributes('-topmost', True)
        
        width = 300
        height = 200
        x = self.top.winfo_x() + 50
        y = self.top.winfo_y() + 50
        area_dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        list_frame = ttk.Frame(area_dialog, padding="5")
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        # show_area_dialog 메서드 내 listbox
        listbox = tk.Listbox(list_frame, 
                          yscrollcommand=scrollbar.set,
                          font=self.font_normal if hasattr(self, 'font_normal') else ('KoPubWorld Dotum Medium', 12))  # Helvetica를 KoPubWorld로 변경
        listbox.pack(fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        
        for area in sorted(area_list, key=lambda x: float(x)):
            listbox.insert(tk.END, f"{area}㎡")
        
        def on_area_select(event):
            if listbox.curselection():
                selected_area = listbox.get(listbox.curselection())
                area_value = selected_area.replace('㎡', '').strip()
                search_text = f"{self.simple_addr}"
                trades = self.trade_data.get(area_value, [])
                self.result = (search_text, self.selected_apt, area_value, trades)
                area_dialog.destroy()
                self.top.destroy()
        
        listbox.bind('<ButtonRelease-1>', on_area_select)

        
    def get_areas_for_apt(self, apt_name):
        areas = set()
        trades = []  # 실거래 데이터 수집
        current_date = datetime.now()
        
        for i in range(24):
            search_date = current_date - timedelta(days=30*i)
            deal_ymd = search_date.strftime("%Y%m")
            
            url = (f"http://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
                   f"?serviceKey={self.service_key}"
                   f"&LAWD_CD={self.sigungu_code}"
                   f"&DEAL_YMD={deal_ymd}"
                   f"&numOfRows=1000")
            
            try:
                response = requests.get(url)
                if response.status_code == 200:
                    root = ET.fromstring(response.text)
                    items = root.findall('.//item')
                    
                    for item in items:
                        item_apt = item.findtext('aptNm', '').strip()
                        item_dong = item.findtext('umdNm', '').strip()
                        if item_apt == apt_name and item_dong == self.dong:
                            area = float(item.findtext('excluUseAr', '0'))
                            if area > 0:
                                areas.add(str(int(area)))
                                # 실거래 정보 저장
                                trade = {
                                    'date': datetime(
                                        int(item.findtext('dealYear')),
                                        int(item.findtext('dealMonth')),
                                        1
                                    ),
                                    'price': int(item.findtext('dealAmount').replace(',', '')),
                                    'floor': int(item.findtext('floor', '0')),
                                    'area': int(area)
                                }
                                trades.append(trade)
            except Exception as e:
                print(f"API 호출 중 오류: {str(e)}")
                continue
        
        # 면적별로 실거래 데이터 정리
        self.trade_data = {}
        for trade in trades:
            area_key = str(trade['area'])
            if area_key not in self.trade_data:
                self.trade_data[area_key] = []
            self.trade_data[area_key].append(trade)
        
        return sorted(list(areas), key=float)


class SearchResultDialog:
    def __init__(self, parent, search_results, title="검색 결과"):
        self.top = tk.Toplevel(parent)
        self.top.title(title)
        self.top.attributes('-topmost', True)  # 추가
        
        
        # 창 크기와 위치 설정
        width = 500
        height = 400
        screen_width = self.top.winfo_screenwidth()
        screen_height = self.top.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.top.geometry(f"{width}x{height}+{x}+{y}")
        
        # 리스트박스 생성
        frame = ttk.Frame(self.top, padding="10")
        
        frame.pack(fill='both', expand=True)
        
        # 설명 레이블
        ttk.Label(frame, text="아래 목록에서 검색할 단지를 선택하세요", 
                 wraplength=400).pack(pady=(0, 10))
        
        # 리스트박스와 스크롤바
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, 
                                 font=('KoPubWorld Dotum Medium', 10))  
        self.listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.listbox.yview)
        
        # 검색 결과 표시
        for item in search_results:
            display_text = f"{item['name']} ({item['type']}) - {item['location']}"
            self.listbox.insert(tk.END, display_text)            
        # 선택 버튼
        ttk.Button(frame, text="선택", 
                  command=self.on_select).pack(pady=10)
        
        # 더블클릭 이벤트 바인딩
        self.listbox.bind('<Double-1>', lambda e: self.on_select())
        
    def on_select(self):
        if self.listbox.curselection():
            self.top.destroy()


if __name__ == "__main__":
    app = KBLandApp()
    app.run()