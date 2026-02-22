import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.font_manager as fm
from datetime import datetime
import numpy as np
import json
import os
import openpyxl
import warnings
import re  # 기존 import 구문들과 함께 추가
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from PIL import Image, ImageTk
import pandas as pd
import numpy as np
import os
import json
import requests
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
import logging
import shutil
from matplotlib import font_manager
import time
import concurrent.futures
import threading
# 추가할 임포트
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# openpyxl 경고 억제
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

class ApartmentPriceAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("아파트 매매가 상승/하락률 분석")
        self.root.geometry("1200x800")
        
        # 설정 파일 경로
        self.config_file = "apartment_analyzer_config.json"
        
        # 한글 폰트 설정
        plt.rcParams['font.family'] = ['Malgun Gothic', 'AppleGothic', 'Noto Sans CJK KR']
        plt.rcParams['axes.unicode_minus'] = False
        
        self.df = None
        self.create_widgets()
        
        # 프로그램 종료 시 설정 저장
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 시작 시 이전 설정 로드
        self.load_settings()
        
    def create_widgets(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 파일 선택 프레임
        file_frame = ttk.LabelFrame(main_frame, text="엑셀 파일 선택", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path_var, width=60, state="readonly").grid(row=0, column=0, padx=(0, 10))
        ttk.Button(file_frame, text="파일 선택", command=self.select_file).grid(row=0, column=1)
        ttk.Button(file_frame, text="데이터 로드", command=self.load_data).grid(row=0, column=2, padx=(10, 0))
        
        # 지역 선택 프레임
        region_frame = ttk.LabelFrame(main_frame, text="지역 선택 및 차트 옵션", padding="10")
        region_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 지역 선택 방식 선택
        ttk.Label(region_frame, text="선택 방식:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.selection_mode = tk.StringVar(value="single")
        ttk.Radiobutton(region_frame, text="단일 선택", variable=self.selection_mode, value="single", 
                       command=self.on_selection_mode_change).grid(row=1, column=0, sticky=tk.W)
        ttk.Radiobutton(region_frame, text="다중 선택", variable=self.selection_mode, value="multiple",
                       command=self.on_selection_mode_change).grid(row=2, column=0, sticky=tk.W)
        
        # 차트 유형 선택 추가
        ttk.Label(region_frame, text="차트 유형:").grid(row=3, column=0, sticky=tk.W, pady=(10, 5))
        self.chart_type = tk.StringVar(value="change_rate")
        ttk.Radiobutton(region_frame, text="상승/하락률 (%)", variable=self.chart_type, value="change_rate").grid(row=4, column=0, sticky=tk.W)
        ttk.Radiobutton(region_frame, text="절대가격 (만원/평)", variable=self.chart_type, value="absolute_price").grid(row=5, column=0, sticky=tk.W)
        
        # 기준점 선택 추가
        ttk.Label(region_frame, text="기준점 (상승/하락률 계산용):").grid(row=6, column=0, sticky=tk.W, pady=(10, 5))
        self.base_point = tk.StringVar(value="first")
        ttk.Radiobutton(region_frame, text="첫 번째 데이터", variable=self.base_point, value="first").grid(row=7, column=0, sticky=tk.W)
        ttk.Radiobutton(region_frame, text="이전 기간 대비", variable=self.base_point, value="previous").grid(row=8, column=0, sticky=tk.W)
        ttk.Radiobutton(region_frame, text="특정 연도 기준", variable=self.base_point, value="year").grid(row=9, column=0, sticky=tk.W)
        
        # 기준 연도 입력
        year_frame = ttk.Frame(region_frame)
        year_frame.grid(row=10, column=0, sticky=(tk.W, tk.E), pady=(5, 10))
        ttk.Label(year_frame, text="기준 연도:").pack(side=tk.LEFT)
        self.base_year_var = tk.StringVar(value="2020")
        year_entry = ttk.Entry(year_frame, textvariable=self.base_year_var, width=8)
        year_entry.pack(side=tk.LEFT, padx=(5, 0))
        
        # 단일 선택용 콤보박스
        ttk.Label(region_frame, text="지역:").grid(row=11, column=0, sticky=tk.W, pady=(10, 5))
        self.region_var = tk.StringVar()
        self.region_combo = ttk.Combobox(region_frame, textvariable=self.region_var, state="readonly", width=30)
        self.region_combo.grid(row=12, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 다중 선택용 리스트박스
        ttk.Label(region_frame, text="지역 목록 (Ctrl+클릭으로 다중 선택):").grid(row=13, column=0, sticky=tk.W, pady=(10, 5))
        
        # 리스트박스 프레임
        listbox_frame = ttk.Frame(region_frame)
        listbox_frame.grid(row=14, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 리스트박스와 스크롤바
        self.region_listbox = tk.Listbox(listbox_frame, selectmode=tk.EXTENDED, height=8, width=30)
        scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.region_listbox.yview)
        self.region_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.region_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        listbox_frame.columnconfigure(0, weight=1)
        listbox_frame.rowconfigure(0, weight=1)
        
        # 선택된 지역 표시
        ttk.Label(region_frame, text="선택된 지역:").grid(row=15, column=0, sticky=tk.W, pady=(10, 5))
        self.selected_regions_var = tk.StringVar(value="없음")
        selected_label = ttk.Label(region_frame, textvariable=self.selected_regions_var, wraplength=300)
        selected_label.grid(row=16, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 리스트박스 선택 이벤트 바인딩
        self.region_listbox.bind('<<ListboxSelect>>', self.on_listbox_select)
        
        # 그래프 그리기 버튼
        ttk.Button(region_frame, text="그래프 그리기", command=self.plot_graph).grid(row=17, column=0, pady=10)
        
        # 초기에는 단일 선택 모드로 설정
        self.on_selection_mode_change()
        
        # 그래프 프레임
        self.graph_frame = ttk.LabelFrame(main_frame, text="시계열 그래프", padding="10")
        self.graph_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(10, 0))
        
        # 그리드 가중치 설정
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
            
    def calculate_change_rate(self, data, base_point_type="first", base_year=None):
        """상승/하락률 계산"""
        if len(data) == 0:
            return pd.Series(dtype=float), None
        
        base_date = None
        
        if base_point_type == "first":
            # 첫 번째 데이터 대비 변화율
            base_value = data.iloc[0]
            base_date = data.index[0]
            if base_value == 0 or pd.isna(base_value):
                return pd.Series([0] * len(data), index=data.index), base_date
            change_rate = ((data - base_value) / base_value * 100)
        elif base_point_type == "previous":
            # 이전 기간 대비 변화율
            change_rate = data.pct_change() * 100
            change_rate.iloc[0] = 0  # 첫 번째 값은 0으로 설정
            base_date = None
        else:  # year
            # 특정 연도 기준 변화율
            try:
                base_year_int = int(base_year) if base_year else 2020
                
                # 해당 연도의 첫 번째 데이터를 찾기
                year_mask = data.index.year == base_year_int
                if not year_mask.any():
                    # 해당 연도가 없으면 가장 가까운 연도 찾기
                    available_years = data.index.year.unique()
                    closest_year = min(available_years, key=lambda x: abs(x - base_year_int))
                    year_mask = data.index.year == closest_year
                    print(f"기준 연도 {base_year_int}가 없어서 {closest_year}년을 사용합니다.")
                
                base_idx = data[year_mask].index[0]
                base_value = data.loc[base_idx]
                base_date = base_idx
                
                if base_value == 0 or pd.isna(base_value):
                    return pd.Series([0] * len(data), index=data.index), base_date
                
                # 기준 연도 이후 데이터만 계산
                data_after_base = data[data.index >= base_idx]
                change_rate = ((data_after_base - base_value) / base_value * 100)
                
            except (ValueError, IndexError):
                # 오류 발생 시 첫 번째 데이터 기준으로 fallback
                base_value = data.iloc[0]
                base_date = data.index[0]
                if base_value == 0 or pd.isna(base_value):
                    return pd.Series([0] * len(data), index=data.index), base_date
                change_rate = ((data - base_value) / base_value * 100)
        
        return change_rate, base_date
    
    def read_monthly_data_with_openpyxl(self, file_path, sheet_name):
        """openpyxl을 사용해서 월별 데이터 직접 추출 - 개선된 버전"""
        import openpyxl
        
        print("=" * 60)
        print("월별 데이터 직접 추출 시작 (개선된 버전)")
        print("=" * 60)
        
        # 엑셀 파일 열기
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook[sheet_name]
        
        print(f"워크시트 정보: {worksheet.max_row}행 x {worksheet.max_column}열")
        
        # 2~4행에서 지역명 추출
        regions = []
        
        print("\n지역명 추출 중...")
        for col in range(2, min(worksheet.max_column + 1, 50)):  # B열부터 최대 50개 컬럼
            # 2행: 시도
            sido_cell = worksheet.cell(row=2, column=col)
            sido = str(sido_cell.value).strip() if sido_cell.value else ""
            
            # 3행: 시군구  
            sigungu_cell = worksheet.cell(row=3, column=col)
            sigungu = str(sigungu_cell.value).strip() if sigungu_cell.value else ""
            
            # 4행: 추가 정보 (있으면)
            extra_cell = worksheet.cell(row=4, column=col)
            extra = str(extra_cell.value).strip() if extra_cell.value else ""
            
            # 지역명 조합
            if sido and sigungu:
                if extra and extra not in sido and extra not in sigungu and len(extra) < 10:
                    region_name = f"{sido} {sigungu} {extra}".strip()
                else:
                    region_name = f"{sido} {sigungu}".strip()
                
                # 유효한 지역명인지 확인
                exclude_terms = ['구분', '지역', '시점', '년도', '평균', '계', 'total']
                if not any(term in region_name.lower() for term in exclude_terms):
                    regions.append((col, region_name))
                    
                    if len(regions) <= 10:  # 처음 10개만 출력
                        print(f"  컬럼 {col}: {region_name}")
        
        print(f"총 {len(regions)}개 지역 발견")
        
        if len(regions) == 0:
            raise ValueError("유효한 지역을 찾을 수 없습니다.")
        
        # 5행부터 데이터 끝까지 모든 행을 확인
        monthly_data = []
        data_start_row = 5
        
        # 2013년 4월부터 시작
        current_year = 2013
        current_month = 4
        
        print(f"\n월별 데이터 읽기 시작 ({data_start_row}행부터)...")
        print("각 행별 데이터 확인:")
        
        total_months = 0
        consecutive_empty_rows = 0
        max_empty_rows = 5  # 연속 빈 행이 5개 이상이면 중단
        
        for row in range(data_start_row, worksheet.max_row + 1):
            # 현재 날짜 생성 - 매월 1일로 설정
            current_date = pd.to_datetime(f"{current_year}-{current_month:02d}-01")
            
            # 해당 행에 실제 가격 데이터가 있는지 확인
            valid_prices_in_row = 0
            row_data = []
            
            for col, region_name in regions:
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        price_str = str(cell.value).replace(',', '').strip()
                        # 숫자인지 확인 (소수점 포함)
                        if price_str and price_str.replace('.', '').replace('-', '').isdigit():
                            price_value = float(price_str)
                            if price_value > 0 and price_value < 100000:  # 현실적인 범위
                                price_per_pyeong = price_value * 3.3  # 평당 가격
                                row_data.append({
                                    '시점': current_date,
                                    '지역': region_name,
                                    '가격': price_per_pyeong
                                })
                                valid_prices_in_row += 1
                    except (ValueError, TypeError):
                        continue
            
            # 해당 행에 유효한 데이터가 있으면 추가
            if valid_prices_in_row > 0:
                monthly_data.extend(row_data)
                total_months += 1
                consecutive_empty_rows = 0
                
                # 진행상황 출력 (매 12개월마다)
                if total_months % 12 == 0:
                    years = total_months // 12
                    print(f"  {years}년차 완료 ({current_date.strftime('%Y-%m')}) - 행 {row}, 유효 데이터 {valid_prices_in_row}개")
                elif total_months <= 5 or total_months % 6 == 0:  # 처음 5개월과 6개월마다
                    print(f"  {current_date.strftime('%Y-%m')} - 행 {row}, 유효 데이터 {valid_prices_in_row}개")
            else:
                consecutive_empty_rows += 1
                if consecutive_empty_rows <= 3:  # 처음 몇 개 빈 행만 출력
                    print(f"  {current_date.strftime('%Y-%m')} - 행 {row}, 데이터 없음")
            
            # 연속으로 빈 행이 너무 많으면 중단
            if consecutive_empty_rows >= max_empty_rows:
                print(f"연속 빈 행 {consecutive_empty_rows}개 발견 - 데이터 끝으로 판단")
                break
            
            # 다음 달로 이동
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1
        
        workbook.close()
        
        print(f"\n추출 결과:")
        print(f"처리된 월수: {total_months}개월")
        print(f"총 데이터 포인트: {len(monthly_data)}개")
        if total_months > 0:
            last_year = current_year - 1 if current_month == 1 else current_year
            last_month = 12 if current_month == 1 else current_month - 1
            print(f"기간: 2013-04 ~ {last_year}-{last_month:02d}")
        
        if not monthly_data:
            raise ValueError("월별 데이터를 추출할 수 없습니다.")
        
        # DataFrame 생성
        self.df = pd.DataFrame(monthly_data)
        
        # 지역 목록 생성
        self.region_list = sorted(list(set([item[1] for item in regions])))
        
        # 시점순 정렬
        self.df = self.df.sort_values(['지역', '시점']).reset_index(drop=True)
        
        print(f"\n최종 결과:")
        print(f"DataFrame shape: {self.df.shape}")
        print(f"지역 개수: {len(self.region_list)}")
        print(f"시점 범위: {self.df['시점'].min().strftime('%Y-%m')} ~ {self.df['시점'].max().strftime('%Y-%m')}")
        
        # 월별 데이터 확인
        unique_dates = sorted(self.df['시점'].unique())
        print(f"고유 시점 개수: {len(unique_dates)}개월")
        
        if len(unique_dates) >= 24:
            print(f"첫 12개월: {[d.strftime('%Y-%m') for d in unique_dates[:12]]}")
            print(f"마지막 12개월: {[d.strftime('%Y-%m') for d in unique_dates[-12:]]}")
        else:
            print(f"모든 시점: {[d.strftime('%Y-%m') for d in unique_dates]}")
        
        # 시점 간격 확인
        if len(unique_dates) > 1:
            diff = unique_dates[1] - unique_dates[0]
            print(f"시점 간격: {diff.days}일 (월별이면 약 30일)")
            if 25 <= diff.days <= 35:
                print("✓ 월별 데이터 확인됨")
            else:
                print("⚠ 월별이 아닐 수 있음")
        
        # 한 지역의 샘플 데이터 (24개월)
        if len(self.df) > 0:
            sample_region = self.region_list[0]
            sample_data = self.df.query('지역 == @sample_region').head(50)  # 더 많은 데이터 확인
            print(f"\n샘플 데이터 ({sample_region} 첫 50개월):")
            print("시점별 데이터:")
            for i, (_, row) in enumerate(sample_data.iterrows()):
                if i < 24:  # 첫 24개월 출력
                    print(f"  {row['시점'].strftime('%Y-%m')}: {row['가격']:,.0f}만원/평")
                elif i == 24:
                    print("  ...")
            print(f"총 {len(sample_data)}개월 데이터")
            
            # 월별 데이터 연속성 확인
            sample_dates = sample_data['시점'].dt.to_period('M').unique()
            print(f"해당 지역 월별 기간: {len(sample_dates)}개월")
            if len(sample_dates) >= 10:
                print(f"첫 10개월: {sample_dates[:10].tolist()}")
                print(f"마지막 10개월: {sample_dates[-10:].tolist()}")
        
        # 전체 데이터의 월별 분포 확인
        monthly_counts = self.df.groupby(self.df['시점'].dt.to_period('M')).size()
        print(f"\n전체 월별 데이터 분포:")
        print(f"총 {len(monthly_counts)}개월")
        print(f"각 월별 지역 수: 평균 {monthly_counts.mean():.1f}개, 최소 {monthly_counts.min()}개, 최대 {monthly_counts.max()}개")
        
        # price_columns 설정 (호환성을 위해)
        self.price_columns = ['가격']
        
        print("=" * 60)
        print("월별 데이터 추출 완료!")
        print("=" * 60)
        
        # 데이터 확인을 위한 디버깅 코드
        print("\n" + "=" * 40)
        print("디버깅: 전체 데이터 확인")
        print("=" * 40)
        print("전체 데이터 샘플:")
        print(self.df.head(20))
        print("\n고유 시점들:")
        unique_times = sorted(self.df['시점'].unique())
        print(f"총 {len(unique_times)}개 시점")
        for i, time in enumerate(unique_times[:20]):
            print(f"  {i+1}: {time.strftime('%Y-%m-%d')}")
        
        # 각 지역별 데이터 개수 확인
        region_counts = self.df['지역'].value_counts()
        print(f"\n지역별 데이터 개수:")
        for region, count in region_counts.head(10).items():
            print(f"  {region}: {count}개월")
        
        print("=" * 40)
    
    def plot_graph(self):
        """시계열 그래프 그리기"""
        if self.df is None:
            messagebox.showerror("오류", "먼저 데이터를 로드해주세요.")
            return
        
        mode = self.selection_mode.get()
        selected_regions = []
        
        if mode == "single":
            selected_region = self.region_var.get()
            if not selected_region:
                messagebox.showerror("오류", "지역을 선택해주세요.")
                return
            selected_regions = [selected_region]
        else:
            selected_indices = self.region_listbox.curselection()
            if not selected_indices:
                messagebox.showerror("오류", "하나 이상의 지역을 선택해주세요.")
                return
            selected_regions = [self.region_listbox.get(i) for i in selected_indices]
            
            if len(selected_regions) > 10:
                messagebox.showwarning("경고", "성능상 최대 10개 지역까지만 선택 가능합니다.")
                selected_regions = selected_regions[:10]
        
        try:
            # 기존 그래프 제거
            for widget in self.graph_frame.winfo_children():
                widget.destroy()
                
            # 새 그래프 생성
            fig, ax = plt.subplots(figsize=(14, 8))
            
            # 색상 팔레트 설정
            colors = plt.cm.tab10(np.linspace(0, 1, len(selected_regions)))
            
            price_col = '가격'
            chart_type = self.chart_type.get()
            base_point_type = self.base_point.get()
            base_year = self.base_year_var.get() if hasattr(self, 'base_year_var') else None
            all_plot_data = []
            base_dates = []  # 각 지역의 기준점 날짜 저장
            
            # 각 지역별로 그래프 그리기
            for i, region in enumerate(selected_regions):
                # 지역 데이터 필터링 - query() 사용
                filtered_data = self.df.query('지역 == @region').copy()
                
                if filtered_data.empty:
                    print(f"경고: {region} 지역의 데이터가 없습니다.")
                    continue
                    
                # 데이터 정렬 및 정리
                filtered_data = filtered_data.sort_values('시점')
                plot_data = filtered_data.dropna(subset=['시점', price_col])
                
                print(f"\n{region} 데이터 확인:")
                print(f"  필터링된 데이터: {len(filtered_data)}개")
                print(f"  정리된 데이터: {len(plot_data)}개")
                print(f"  시점 범위: {plot_data['시점'].min()} ~ {plot_data['시점'].max()}")
                
                if plot_data.empty:
                    print(f"경고: {region} 지역의 유효한 데이터가 없습니다.")
                    continue
                
                # 월별 데이터 확인
                monthly_data_count = len(plot_data)
                expected_months = (plot_data['시점'].max() - plot_data['시점'].min()).days // 30 + 1
                print(f"  실제 월 수: {monthly_data_count}개, 예상 월 수: {expected_months}개")
                
                # 시점을 인덱스로 설정 (calculate_change_rate에서 사용)
                plot_data_indexed = plot_data.set_index('시점')[price_col]
                
                # 차트 유형에 따른 데이터 처리
                if chart_type == "change_rate":
                    # 상승/하락률 계산
                    y_data, base_date = self.calculate_change_rate(plot_data_indexed, base_point_type, base_year)
                    y_label = "상승/하락률 (%)"
                    y_format = lambda x, p: f'{x:+.1f}%'
                    
                    # 기준점 날짜 저장
                    if base_date:
                        base_dates.append(base_date)
                    
                    # 특정 연도 기준인 경우 기준 연도 이후 데이터만 사용
                    if base_point_type == "year" and base_date:
                        # plot_data도 기준 연도 이후로 필터링
                        plot_data = plot_data[plot_data['시점'] >= base_date].copy()
                        
                else:
                    # 절대가격
                    y_data = plot_data[price_col]
                    y_label = "평균평단가 (만원/평)"
                    y_format = lambda x, p: f'{x:,.0f}'
                
                all_plot_data.append((region, plot_data, y_data))
                
                # 선 그래프 그리기
                color = colors[i]
                if chart_type == "change_rate":
                    # 상승/하락률 차트의 경우 인덱스(시점) 사용
                    line = ax.plot(y_data.index, y_data.values, 
                                 marker='o', linewidth=2, markersize=3, 
                                 label=region, color=color)
                else:
                    # 절대가격 차트
                    line = ax.plot(plot_data['시점'], y_data, 
                                 marker='o', linewidth=2, markersize=3, 
                                 label=region, color=color)
                
                # 마지막 데이터 포인트에 지역명과 값 표시
                if len(plot_data) > 0:
                    if chart_type == "change_rate":
                        last_date = y_data.index[-1]
                        last_y = y_data.iloc[-1]
                    else:
                        last_point = plot_data.iloc[-1]
                        last_date = last_point['시점']
                        last_y = y_data.iloc[-1]
                    
                    # 지역명을 짧게 표시
                    short_name = region.split()[-1] if ' ' in region else region
                    
                    # 항상 평균평단가를 표시
                    last_price = plot_data.iloc[-1][price_col]
                    price_text = f'{last_price:,.0f}'
                    
                    ax.annotate(f'{short_name} {price_text}', 
                               xy=(last_date, last_y),
                               xytext=(5, 0), 
                               textcoords='offset points',
                               bbox=dict(boxstyle='round,pad=0.2', facecolor=color, alpha=0.7),
                               fontsize=8,
                               fontweight='bold',
                               ha='left',
                               va='center',
                               clip_on=False)
            
            if not all_plot_data:
                messagebox.showerror("오류", "선택된 지역들에 유효한 데이터가 없습니다.")
                return
            
            # 상승/하락률 차트인 경우 0% 기준선 추가
            if chart_type == "change_rate":
                ax.axhline(y=0, color='gray', linestyle='--', alpha=0.7, linewidth=1.5, label='0% 기준선')
                
                # 특정 연도 기준인 경우 기준점에 수직선 추가
                if base_point_type == "year" and base_dates:
                    # 모든 지역의 기준점이 같다고 가정하고 첫 번째 사용
                    base_date = base_dates[0]
                    ax.axvline(x=base_date, color='red', linestyle=':', alpha=0.7, linewidth=1.5, 
                              label=f'기준점 ({base_date.strftime("%Y-%m")})')
            
            # 그래프 설정
            if len(selected_regions) == 1:
                if chart_type == "change_rate":
                    if base_point_type == "first":
                        base_text = "첫 데이터 대비"
                    elif base_point_type == "previous":
                        base_text = "이전 기간 대비"
                    else:
                        base_text = f"{base_year}년 기준"
                    title = f'{selected_regions[0]} 상승/하락률 ({base_text})'
                else:
                    title = f'{selected_regions[0]} 평균평단가'
            else:
                if chart_type == "change_rate":
                    if base_point_type == "first":
                        base_text = "첫 데이터 대비"
                    elif base_point_type == "previous":
                        base_text = "이전 기간 대비"
                    else:
                        base_text = f"{base_year}년 기준"
                    title = f'지역별 상승/하락률 비교 ({base_text}, {len(selected_regions)}개 지역)'
                else:
                    title = f'지역별 평균평단가 비교 ({len(selected_regions)}개 지역)'
            
            ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
            ax.set_xlabel('시점', fontsize=12)
            ax.set_ylabel(y_label, fontsize=12)
            ax.grid(True, alpha=0.3)
            
            # y축 포맷 설정
            ax.yaxis.set_major_formatter(plt.FuncFormatter(y_format))
            
            # x축 날짜 포맷 설정 - 모든 월별 데이터 표시
            import matplotlib.dates as mdates
            
            # 모든 월별 데이터가 표시되도록 설정
            ax.xaxis.set_major_locator(mdates.YearLocator())  # 연도별 major tick
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))  # 연도 표시
            ax.xaxis.set_minor_locator(mdates.MonthLocator())  # 모든 월 minor tick
            
            # 데이터 포인트 확인을 위한 디버깅
            if len(all_plot_data) > 0:
                sample_region, sample_data, sample_y = all_plot_data[0]
                print(f"\n그래프 데이터 확인 ({sample_region}):")
                print(f"총 데이터 포인트: {len(sample_data)}개")
                print(f"시점 범위: {sample_data['시점'].min()} ~ {sample_data['시점'].max()}")
                print(f"첫 10개 시점: {sample_data['시점'].head(10).tolist()}")
            
            # 날짜 라벨 회전
            fig.autofmt_xdate(rotation=45)
            
            # 범례 설정
            if len(selected_regions) > 1 or chart_type == "change_rate":
                ax.legend(loc='upper left', bbox_to_anchor=(0, 1), ncol=min(3, len(selected_regions)))
            
            # 여백 조정 - 오른쪽 여백을 충분히 확보
            plt.subplots_adjust(left=0.1, right=0.85, top=0.9, bottom=0.15)
            
            # 그래프를 GUI에 추가
            canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # 통계 정보 표시
            if all_plot_data:
                if chart_type == "change_rate":
                    self.show_change_rate_statistics(all_plot_data, base_point_type)
                else:
                    first_region, first_data, first_y = all_plot_data[0]
                    self.show_statistics(first_data[price_col], first_region)
            
            # 선택된 지역 저장
            self.save_settings()
            
        except Exception as e:
            messagebox.showerror("오류", f"그래프 생성 중 오류가 발생했습니다: {str(e)}")
            import traceback
            print(f"상세 오류: {traceback.format_exc()}")


    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            
    def load_data(self):
        file_path = self.file_path_var.get()
        if not file_path:
            messagebox.showerror("오류", "먼저 엑셀 파일을 선택해주세요.")
            return
            
        try:
            # 파일 경로 정규화 및 권한 확인
            import os
            if not os.path.exists(file_path):
                messagebox.showerror("오류", "파일이 존재하지 않습니다. 파일 경로를 확인해주세요.")
                return
                
            if not os.access(file_path, os.R_OK):
                messagebox.showerror("오류", "파일 읽기 권한이 없습니다.")
                return
            
            # 엑셀 파일의 시트 목록 확인
            try:
                xl_file = pd.ExcelFile(file_path)
                sheet_names = xl_file.sheet_names
                print(f"사용 가능한 시트: {sheet_names}")
            except Exception as e:
                messagebox.showerror("오류", f"엑셀 파일을 열 수 없습니다: {str(e)}")
                return
            
            # '47.㎡당아파트평균매매' 시트 확인
            target_sheet = None
            for sheet in sheet_names:
                if '47' in sheet and '㎡' in sheet and '아파트' in sheet:
                    target_sheet = sheet
                    break
                    
            if target_sheet is None:
                # 시트가 없으면 사용자에게 선택하게 함
                sheet_selection = self.select_sheet(sheet_names)
                if sheet_selection:
                    target_sheet = sheet_selection
                else:
                    return
            
            # openpyxl로 직접 월별 데이터 읽기
            print("월별 데이터 직접 추출 시작...")
            self.read_monthly_data_with_openpyxl(file_path, target_sheet)
            
            # 지역 콤보박스 업데이트
            self.update_region_combos()
            
            messagebox.showinfo("완료", f"월별 데이터 로드 완료!\n시트: {target_sheet}\n총 {len(self.df)}개 월별 데이터")
            
            # 파일 경로 저장
            self.save_settings()
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 읽기 중 오류가 발생했습니다: {str(e)}")
    
    def select_sheet(self, sheet_names):
        """시트 선택 다이얼로그"""
        sheet_window = tk.Toplevel(self.root)
        sheet_window.title("시트 선택")
        sheet_window.geometry("400x300")
        sheet_window.grab_set()
        
        ttk.Label(sheet_window, text="분석할 시트를 선택해주세요:").pack(pady=10)
        
        selected_sheet = tk.StringVar()
        
        # 시트 목록 표시
        listbox = tk.Listbox(sheet_window, height=10)
        for sheet in sheet_names:
            listbox.insert(tk.END, sheet)
        listbox.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
        
        def on_select():
            selection = listbox.curselection()
            if selection:
                selected_sheet.set(sheet_names[selection[0]])
                sheet_window.destroy()
        
        def on_cancel():
            selected_sheet.set("")
            sheet_window.destroy()
        
        button_frame = ttk.Frame(sheet_window)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="선택", command=on_select).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="취소", command=on_cancel).pack(side=tk.LEFT, padx=5)
        
        # 더블클릭으로도 선택 가능
        listbox.bind('<Double-1>', lambda e: on_select())
        
        sheet_window.wait_window()
        return selected_sheet.get() if selected_sheet.get() else None
            
    def preprocess_data(self):
        """데이터 전처리"""
        print("원본 데이터 구조:")
        print(f"Shape: {self.df.shape}")
        print(f"첫 5행:\n{self.df.head()}")
        
        if len(self.df) < 4:
            raise ValueError("데이터가 부족합니다. 최소 4행이 필요합니다.")
        
        # 원본 엑셀 파일에서 병합 정보 읽기
        file_path = self.file_path_var.get()
        region_mapping = self.get_merged_cell_mapping(file_path)
        
        # 실제 데이터는 4행부터 (인덱스 3부터)
        data_df = self.df.iloc[3:].copy()
        data_df.reset_index(drop=True, inplace=True)
        
        # 첫 번째 컬럼은 시점 정보
        time_column = data_df.iloc[:, 0]
        
        # 지역별 데이터 구조 만들기
        processed_data = []
        region_list = []
        
        for col_idx, (sido_clean, sigungu_clean) in region_mapping.items():
            # 지역명 조합 (시도 + 시군구)
            region_name = f"{sido_clean} {sigungu_clean}"
            region_list.append(region_name)
            
            # 해당 지역의 가격 데이터
            if col_idx < len(data_df.columns):
                price_data = data_df.iloc[:, col_idx]
                
                # 각 시점별로 데이터 생성
                for time_idx, (time_val, price_val) in enumerate(zip(time_column, price_data)):
                    # 시점 데이터 정리
                    if pd.isna(time_val) or str(time_val).strip() == '':
                        continue
                    
                    # 가격 데이터 정리
                    if pd.isna(price_val):
                        continue
                        
                    try:
                        # 가격 데이터를 숫자로 변환하고 3.3을 곱해서 평당 가격으로 변환
                        price_clean = str(price_val).replace(',', '').replace(' ', '').strip()
                        if price_clean == '' or price_clean == 'nan':
                            continue
                        price_numeric = float(price_clean) * 3.3  # 1평(3.3㎡) 기준으로 변환
                        
                        processed_data.append({
                            '시점': time_val,
                            '지역': region_name,
                            '가격': price_numeric
                        })
                    except (ValueError, TypeError):
                        continue
        
        if not processed_data:
            raise ValueError("처리할 수 있는 유효한 데이터가 없습니다.")
        
        # 새로운 DataFrame 생성
        self.df = pd.DataFrame(processed_data)
        
        # 지역 목록 저장
        self.region_list = region_list
        
        print(f"처리된 데이터 shape: {self.df.shape}")
        print(f"지역 목록: {region_list}")
        print(f"샘플 데이터:\n{self.df.head()}")
        
        # 시점 데이터 처리
        def parse_date(date_val):
            if pd.isna(date_val):
                return None
                
            date_str = str(date_val).strip()
            
            # 2013.4 형식 처리
            if '.' in date_str and len(date_str.split('.')) == 2:
                try:
                    year, month = date_str.split('.')
                    return pd.to_datetime(f"{year}-{month.zfill(2)}-01")
                except:
                    pass
            
            # 다른 형식들 시도
            formats_to_try = [
                '%Y.%m', '%Y-%m', '%Y/%m',
                '%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d',
                '%Y%m',
            ]
            
            for fmt in formats_to_try:
                try:
                    return pd.to_datetime(date_str, format=fmt)
                except:
                    continue
            
            try:
                return pd.to_datetime(date_str, errors='coerce')
            except:
                return None
        
        self.df['시점'] = self.df['시점'].apply(parse_date)
        
        # 유효한 시점 데이터만 유지
        self.df = self.df.dropna(subset=['시점']).reset_index(drop=True)
        
        # 시점순 정렬
        self.df = self.df.sort_values('시점').reset_index(drop=True)
        
        print(f"최종 처리된 데이터 shape: {self.df.shape}")
        print(f"시점 범위: {self.df['시점'].min()} ~ {self.df['시점'].max()}")
        
        # price_columns 설정 (호환성을 위해)
        self.price_columns = ['가격']
        
    def get_merged_cell_mapping(self, file_path):
        """엑셀 파일에서 병합된 셀 정보를 읽어서 지역 매핑 생성"""
        region_mapping = {}
        
        try:
            print("엑셀 파일 읽는 중...")
            
            # openpyxl로 엑셀 파일 열기
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # 시트 선택
            sheet_name = None
            for name in workbook.sheetnames:
                if '47' in name and '㎡' in name and '아파트' in name:
                    sheet_name = name
                    break
            
            if sheet_name is None:
                sheet_name = workbook.sheetnames[0]
            
            worksheet = workbook[sheet_name]
            print(f"시트: {sheet_name}")
            
            # 모든 컬럼에 대해 지역명 매핑 생성
            sido_mapping = {}
            sigungu_mapping = {}
            
            # 1단계: 모든 셀 값 읽기
            for col in range(1, worksheet.max_column + 1):
                # 2행 (시/도)
                cell_2 = worksheet.cell(row=2, column=col)
                if cell_2.value and str(cell_2.value).strip():
                    value = str(cell_2.value).strip()
                    sido_mapping[col] = value
                
                # 3행 (시/군/구)
                cell_3 = worksheet.cell(row=3, column=col)
                if cell_3.value and str(cell_3.value).strip():
                    value = str(cell_3.value).strip()
                    sigungu_mapping[col] = value
            
            # 2단계: 병합된 셀 처리
            for merged_range in worksheet.merged_cells.ranges:
                min_row, max_row = merged_range.min_row, merged_range.max_row
                min_col, max_col = merged_range.min_col, merged_range.max_col
                
                # 2행 병합
                if min_row <= 2 <= max_row:
                    master_cell = worksheet.cell(row=2, column=min_col)
                    if master_cell.value and str(master_cell.value).strip():
                        merged_value = str(master_cell.value).strip()
                        for col in range(min_col, max_col + 1):
                            sido_mapping[col] = merged_value
                
                # 3행 병합
                if min_row <= 3 <= max_row:
                    master_cell = worksheet.cell(row=3, column=min_col)
                    if master_cell.value and str(master_cell.value).strip():
                        merged_value = str(master_cell.value).strip()
                        for col in range(min_col, max_col + 1):
                            sigungu_mapping[col] = merged_value
                
                # 2-3행 함께 병합
                if min_row <= 2 and max_row >= 3:
                    master_cell = worksheet.cell(row=min_row, column=min_col)
                    if master_cell.value and str(master_cell.value).strip():
                        region_value = str(master_cell.value).strip()
                        
                        if ' ' in region_value:
                            parts = region_value.split(' ', 1)
                            for col in range(min_col, max_col + 1):
                                sido_mapping[col] = parts[0]
                                sigungu_mapping[col] = parts[1]
                        else:
                            for col in range(min_col, max_col + 1):
                                sido_mapping[col] = region_value
                                sigungu_mapping[col] = region_value
            
            # 3단계: 빈 셀 채우기 (시/도만)
            current_sido = None
            for col in range(1, worksheet.max_column + 1):
                if col in sido_mapping and sido_mapping[col]:
                    current_sido = sido_mapping[col]
                elif current_sido and col not in sido_mapping:
                    sido_mapping[col] = current_sido
            
            # 4단계: 유효한 지역만 매핑 (B열부터)
            exclude_terms = ['구분', '지역', '시점', '년도']
            
            for col in range(2, worksheet.max_column + 1):  # B열부터
                pandas_idx = col - 1
                
                sido = sido_mapping.get(col, '').strip()
                sigungu = sigungu_mapping.get(col, '').strip()
                
                # 시/도만 있고 시/군/구가 없는 경우 처리
                if sido and not sigungu:
                    if sido in ['전국', '서울특별시'] or 'Total' in sido or 'Seoul' in sido:
                        sigungu = sido
                
                # 제외할 용어 체크
                if sido in exclude_terms or sigungu in exclude_terms:
                    continue
                
                # 유효한 지역명이 있으면 매핑
                if sido and sigungu:
                    region_mapping[pandas_idx] = (sido, sigungu)
            
            workbook.close()
            print(f"매핑 완료: {len(region_mapping)}개 지역")
            
        except Exception as e:
            print(f"openpyxl 처리 실패: {e}")
            return self.get_basic_region_mapping()
        
        return region_mapping
    
    def get_basic_region_mapping(self):
        """기본 방식으로 지역 매핑 생성"""
        region_mapping = {}
        
        # 2행과 3행에서 지역 정보 추출
        sido_row = self.df.iloc[1]
        sigungu_row = self.df.iloc[2]
        
        exclude_terms = ['구분', '지역', '시점', '년도']
        
        current_sido = None
        for col_idx in range(1, len(self.df.columns)):
            if col_idx == 0:
                continue
                
            # 시/도 찾기
            sido_val = sido_row.iloc[col_idx] if col_idx < len(sido_row) else None
            if pd.notna(sido_val) and str(sido_val).strip():
                potential_sido = str(sido_val).strip()
                if potential_sido not in exclude_terms:
                    current_sido = potential_sido
            
            # 시/군/구 찾기
            sigungu_val = sigungu_row.iloc[col_idx] if col_idx < len(sigungu_row) else None
            sigungu = None
            
            if pd.notna(sigungu_val) and str(sigungu_val).strip():
                sigungu = str(sigungu_val).strip()
            elif current_sido and current_sido in ['전국', '서울특별시']:
                sigungu = current_sido
            
            # 제외할 용어 체크
            if sigungu and sigungu in exclude_terms:
                continue
            
            if current_sido and sigungu and current_sido not in exclude_terms:
                region_mapping[col_idx] = (current_sido, sigungu)
        
        return region_mapping
        
    def update_region_combos(self):
        """지역 콤보박스와 리스트박스 업데이트"""
        if self.df is None or not hasattr(self, 'region_list'):
            return
            
        self.region_combo['values'] = self.region_list
        
        self.region_listbox.delete(0, tk.END)
        for region in self.region_list:
            self.region_listbox.insert(tk.END, region)
        
        self.region_var.set('')
        self.selected_regions_var.set("없음")
                    
    def plot_graph(self):
        """시계열 그래프 그리기"""
        if self.df is None:
            messagebox.showerror("오류", "먼저 데이터를 로드해주세요.")
            return
        
        mode = self.selection_mode.get()
        selected_regions = []
        
        if mode == "single":
            selected_region = self.region_var.get()
            if not selected_region:
                messagebox.showerror("오류", "지역을 선택해주세요.")
                return
            selected_regions = [selected_region]
        else:
            selected_indices = self.region_listbox.curselection()
            if not selected_indices:
                messagebox.showerror("오류", "하나 이상의 지역을 선택해주세요.")
                return
            selected_regions = [self.region_listbox.get(i) for i in selected_indices]
            
            if len(selected_regions) > 10:
                messagebox.showwarning("경고", "성능상 최대 10개 지역까지만 선택 가능합니다.")
                selected_regions = selected_regions[:10]
        
        try:
            # 기존 그래프 제거
            for widget in self.graph_frame.winfo_children():
                widget.destroy()
                
            # 새 그래프 생성
            fig, ax = plt.subplots(figsize=(14, 8))
            
            # 색상 팔레트 설정
            colors = plt.cm.tab10(np.linspace(0, 1, len(selected_regions)))
            
            price_col = '가격'
            chart_type = self.chart_type.get()
            base_point_type = self.base_point.get()
            base_year = self.base_year_var.get() if hasattr(self, 'base_year_var') else None
            all_plot_data = []
            base_dates = []  # 각 지역의 기준점 날짜 저장
            
            # 각 지역별로 그래프 그리기
            for i, region in enumerate(selected_regions):
                # 지역 데이터 필터링
                mask = self.df['지역'] == region
                filtered_data = self.df[mask].copy()
                
                if filtered_data.empty:
                    continue
                    
                # 데이터 정렬 및 정리
                filtered_data = filtered_data.sort_values('시점')
                plot_data = filtered_data.dropna(subset=['시점', price_col])
                
                if plot_data.empty:
                    continue
                
                # 시점을 인덱스로 설정 (calculate_change_rate에서 사용)
                plot_data_indexed = plot_data.set_index('시점')[price_col]
                
                # 차트 유형에 따른 데이터 처리
                if chart_type == "change_rate":
                    # 상승/하락률 계산
                    y_data, base_date = self.calculate_change_rate(plot_data_indexed, base_point_type, base_year)
                    y_label = "상승/하락률 (%)"
                    y_format = lambda x, p: f'{x:+.1f}%'
                    
                    # 기준점 날짜 저장
                    if base_date:
                        base_dates.append(base_date)
                    
                    # 특정 연도 기준인 경우 기준 연도 이후 데이터만 사용
                    if base_point_type == "year" and base_date:
                        # plot_data도 기준 연도 이후로 필터링
                        plot_data = plot_data[plot_data['시점'] >= base_date].copy()
                        
                else:
                    # 절대가격
                    y_data = plot_data[price_col]
                    y_label = "평균평단가 (만원/평)"
                    y_format = lambda x, p: f'{x:,.0f}'
                
                all_plot_data.append((region, plot_data, y_data))
                
                # 선 그래프 그리기
                color = colors[i]
                if chart_type == "change_rate":
                    # 상승/하락률 차트의 경우 인덱스(시점) 사용
                    line = ax.plot(y_data.index, y_data.values, 
                                 marker='o', linewidth=2, markersize=3, 
                                 label=region, color=color)
                else:
                    # 절대가격 차트
                    line = ax.plot(plot_data['시점'], y_data, 
                                 marker='o', linewidth=2, markersize=3, 
                                 label=region, color=color)
                
                # 마지막 데이터 포인트에 지역명과 값 표시
                if len(plot_data) > 0:
                    if chart_type == "change_rate":
                        last_date = y_data.index[-1]
                        last_y = y_data.iloc[-1]
                    else:
                        last_point = plot_data.iloc[-1]
                        last_date = last_point['시점']
                        last_y = y_data.iloc[-1]
                    
                    # 지역명을 짧게 표시
                    short_name = region.split()[-1] if ' ' in region else region
                    
                    # 항상 평균평단가를 표시
                    last_price = plot_data.iloc[-1][price_col]
                    price_text = f'{last_price:,.0f}'
                    
                    ax.annotate(f'{short_name} {price_text}', 
                               xy=(last_date, last_y),
                               xytext=(5, 0), 
                               textcoords='offset points',
                               bbox=dict(boxstyle='round,pad=0.2', facecolor=color, alpha=0.7),
                               fontsize=8,
                               fontweight='bold',
                               ha='left',
                               va='center',
                               clip_on=False)  # 클리핑 방지
            
            if not all_plot_data:
                messagebox.showerror("오류", "선택된 지역들에 유효한 데이터가 없습니다.")
                return
            
            # 상승/하락률 차트인 경우 0% 기준선 추가
            if chart_type == "change_rate":
                ax.axhline(y=0, color='gray', linestyle='--', alpha=0.7, linewidth=1.5, label='0% 기준선')
                
                # 특정 연도 기준인 경우 기준점에 수직선 추가
                if base_point_type == "year" and base_dates:
                    # 모든 지역의 기준점이 같다고 가정하고 첫 번째 사용
                    base_date = base_dates[0]
                    ax.axvline(x=base_date, color='red', linestyle=':', alpha=0.7, linewidth=1.5, 
                              label=f'기준점 ({base_date.strftime("%Y-%m")})')
            
            # 그래프 설정
            if len(selected_regions) == 1:
                if chart_type == "change_rate":
                    if base_point_type == "first":
                        base_text = "첫 데이터 대비"
                    elif base_point_type == "previous":
                        base_text = "이전 기간 대비"
                    else:
                        base_text = f"{base_year}년 기준"
                    title = f'{selected_regions[0]} 상승/하락률 ({base_text})'
                else:
                    title = f'{selected_regions[0]} 평균평단가'
            else:
                if chart_type == "change_rate":
                    if base_point_type == "first":
                        base_text = "첫 데이터 대비"
                    elif base_point_type == "previous":
                        base_text = "이전 기간 대비"
                    else:
                        base_text = f"{base_year}년 기준"
                    title = f'지역별 상승/하락률 비교 ({base_text}, {len(selected_regions)}개 지역)'
                else:
                    title = f'지역별 평균평단가 비교 ({len(selected_regions)}개 지역)'
            
            ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
            ax.set_xlabel('시점', fontsize=12)
            ax.set_ylabel(y_label, fontsize=12)
            ax.grid(True, alpha=0.3)
            
            # y축 포맷 설정
            ax.yaxis.set_major_formatter(plt.FuncFormatter(y_format))
            
            # x축 날짜 포맷 설정
            fig.autofmt_xdate()
            
            # 범례 설정
            if len(selected_regions) > 1 or chart_type == "change_rate":
                ax.legend(loc='upper left', bbox_to_anchor=(0, 1), ncol=min(3, len(selected_regions)))
            
            # 여백 조정 - 오른쪽 여백을 충분히 확보
            plt.subplots_adjust(left=0.1, right=0.85, top=0.9, bottom=0.15)
            
            # 그래프를 GUI에 추가
            canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            # 통계 정보 표시
            if all_plot_data:
                if chart_type == "change_rate":
                    self.show_change_rate_statistics(all_plot_data, base_point_type)
                else:
                    first_region, first_data, first_y = all_plot_data[0]
                    self.show_statistics(first_data[price_col], first_region)
            
            # 선택된 지역 저장
            self.save_settings()
            
        except Exception as e:
            messagebox.showerror("오류", f"그래프 생성 중 오류가 발생했습니다: {str(e)}")
    
    def show_change_rate_statistics(self, all_plot_data, base_point_type):
        """상승/하락률 통계 정보 표시"""
        if base_point_type == "first":
            base_text = "첫 데이터 대비"
        elif base_point_type == "previous":
            base_text = "이전 기간 대비"
        else:
            base_year = self.base_year_var.get() if hasattr(self, 'base_year_var') else '2020'
            base_text = f"{base_year}년 기준"
        
        if len(all_plot_data) == 1:
            # 단일 지역
            region, data, change_rates = all_plot_data[0]
            if hasattr(change_rates, 'dropna'):
                valid_rates = change_rates.dropna()
            else:
                valid_rates = change_rates
            
            if len(valid_rates) > 0:
                stats_text = f"""
📊 {region} 상승/하락률 통계 ({base_text}):
• 평균 변화율: {valid_rates.mean():+.2f}%
• 최대 상승률: {valid_rates.max():+.2f}%
• 최대 하락률: {valid_rates.min():+.2f}%
• 최종 변화율: {valid_rates.iloc[-1]:+.2f}%
• 표준편차: {valid_rates.std():.2f}%
• 데이터 개수: {len(valid_rates)}개
                """
                print(stats_text)
        else:
            # 다중 지역
            print(f"\n📊 선택된 {len(all_plot_data)}개 지역 최종 변화율 ({base_text}):")
            for region, data, change_rates in all_plot_data:
                if len(change_rates) > 0:
                    if hasattr(change_rates, 'iloc'):
                        final_rate = change_rates.iloc[-1]
                    else:
                        final_rate = change_rates[-1]
                    print(f"• {region}: {final_rate:+.2f}%")
            
    def show_statistics(self, price_data, region_name):
        """절대가격 통계 정보 표시"""
        stats_text = f"""
📊 {region_name} 평균평단가 통계 (1평 기준):
• 평균: {price_data.mean():.1f}만원/평
• 최고: {price_data.max():.1f}만원/평
• 최저: {price_data.min():.1f}만원/평
• 표준편차: {price_data.std():.1f}만원/평
• 데이터 개수: {len(price_data)}개
        """
        print(stats_text)

    def on_selection_mode_change(self):
        """선택 모드 변경 시 UI 업데이트"""
        mode = self.selection_mode.get()
        
        if mode == "single":
            self.region_combo.configure(state="readonly")
            self.region_listbox.configure(state="normal")
            self.selected_regions_var.set("위의 드롭박스에서 선택하세요")
        else:
            self.region_combo.configure(state="readonly")
            self.region_listbox.configure(state="normal")
            self.selected_regions_var.set("아래 리스트에서 Ctrl+클릭으로 다중 선택하세요")

    def on_listbox_select(self, event):
        """리스트박스 선택 시 선택된 지역 표시 업데이트"""
        if self.selection_mode.get() == "multiple":
            selected_indices = self.region_listbox.curselection()
            selected_regions = [self.region_listbox.get(i) for i in selected_indices]
            
            if selected_regions:
                if len(selected_regions) <= 3:
                    self.selected_regions_var.set(", ".join(selected_regions))
                else:
                    self.selected_regions_var.set(f"{', '.join(selected_regions[:3])} 등 {len(selected_regions)}개")
            else:
                self.selected_regions_var.set("없음")

    def save_settings(self):
        """현재 설정을 파일에 저장"""
        try:
            settings = {
                'last_file_path': self.file_path_var.get(),
                'last_selected_region': self.region_var.get() if hasattr(self, 'region_var') else '',
                'window_geometry': self.root.geometry(),
                'chart_type': self.chart_type.get() if hasattr(self, 'chart_type') else 'change_rate',
                'base_point': self.base_point.get() if hasattr(self, 'base_point') else 'first',
                'base_year': self.base_year_var.get() if hasattr(self, 'base_year_var') else '2020'
            }
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"설정 저장 중 오류: {e}")

    def load_settings(self):
        """저장된 설정을 불러와서 적용"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                
                # 파일 경로 복원
                last_file_path = settings.get('last_file_path', '')
                if last_file_path and os.path.exists(last_file_path):
                    self.file_path_var.set(last_file_path)
                    
                    # 자동으로 데이터 로드 시도
                    try:
                        self.load_data_silent()
                        
                        # 마지막 선택된 지역 복원
                        last_region = settings.get('last_selected_region', '')
                        if last_region and hasattr(self, 'region_list') and last_region in self.region_list:
                            self.region_var.set(last_region)
                            
                    except Exception as e:
                        print(f"자동 데이터 로드 실패: {e}")
                
                # 차트 설정 복원
                chart_type = settings.get('chart_type', 'change_rate')
                if hasattr(self, 'chart_type'):
                    self.chart_type.set(chart_type)
                    
                base_point = settings.get('base_point', 'first')
                if hasattr(self, 'base_point'):
                    self.base_point.set(base_point)
                
                base_year = settings.get('base_year', '2020')
                if hasattr(self, 'base_year_var'):
                    self.base_year_var.set(base_year)
                
                # 창 크기 복원
                window_geometry = settings.get('window_geometry', '')
                if window_geometry:
                    try:
                        self.root.geometry(window_geometry)
                    except:
                        pass
                        
        except Exception as e:
            print(f"설정 로드 중 오류: {e}")

    def load_data_silent(self):
        """자동 로드용 - 오류 메시지 없이 데이터 로드"""
        file_path = self.file_path_var.get()
        if not file_path or not os.path.exists(file_path):
            return False
            
        try:
            # 파일 경로 정규화 및 권한 확인
            if not os.access(file_path, os.R_OK):
                return False
            
            # 엑셀 파일의 시트 목록 확인
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
            
            # '47.㎡당아파트평균매매' 시트 확인
            target_sheet = None
            for sheet in sheet_names:
                if '47' in sheet and '㎡' in sheet and '아파트' in sheet:
                    target_sheet = sheet
                    break
                    
            if target_sheet is None:
                # 첫 번째 시트 사용
                target_sheet = sheet_names[0] if sheet_names else None
                
            if target_sheet is None:
                return False
            
            # 해당 시트 읽기
            self.df = pd.read_excel(file_path, sheet_name=target_sheet)
            
            # 데이터 전처리
            self.preprocess_data()
            
            # 지역 콤보박스 업데이트
            self.update_region_combos()
            
            print(f"자동 로드 완료: {target_sheet} ({len(self.df)}개 행)")
            return True
            
        except Exception as e:
            print(f"자동 로드 실패: {e}")
            return False

    def on_closing(self):
        """프로그램 종료 시 호출"""
        self.save_settings()
        self.root.destroy()

def main():
    root = tk.Tk()
    app = ApartmentPriceAnalyzer(root)
    root.mainloop()

if __name__ == "__main__":
    main()