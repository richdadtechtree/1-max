"""
네이버 부동산 매물 조회 프로그램
- 지역(법정동) 선택 후 아파트 매물 조회
- Selenium을 사용하여 페이지 스크롤 방식으로 데이터 수집
- Excel 파일로 저장
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
import os
import time
import threading
import re
from datetime import datetime
import logging
import warnings

warnings.filterwarnings('ignore')

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


class NaverRealEstateApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("네이버 부동산 매물 조회")
        self.root.geometry("1200x700")

        # 설정 파일 경로
        self.settings_file = os.path.join(os.getcwd(), 'naver_real_estate_settings.json')

        # 기본 설정
        self.lawdong_path = "C:/law-dong/law-dong.txt"
        self.download_path = os.path.join(os.getcwd(), "매물조회결과")

        # 법정동 데이터
        self.sido_list = []
        self.sigungu_dict = {}
        self.dong_dict = {}
        self.region_codes = {}

        # 매물 데이터
        self.properties_data = []

        # WebDriver
        self.driver = None

        # 조회 중단 플래그
        self.stop_search = False

        # 설정 로드
        self.load_settings()

        # 법정동 파일 로드
        if not self.load_lawdong_file():
            messagebox.showerror("오류", "법정동 코드 파일을 로드할 수 없습니다.\n설정에서 파일 경로를 확인해주세요.")

        # GUI 설정
        self.setup_gui()

        # 폴더 생성
        if not os.path.exists(self.download_path):
            os.makedirs(self.download_path)

    def load_settings(self):
        """설정 파일 로드"""
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    self.lawdong_path = settings.get('lawdong_path', self.lawdong_path)
                    self.download_path = settings.get('download_path', self.download_path)
                logging.info(f"설정 로드 완료: {self.settings_file}")
        except Exception as e:
            logging.error(f"설정 로드 오류: {e}")

    def save_settings(self):
        """설정 파일 저장"""
        try:
            settings = {
                'lawdong_path': self.lawdong_path,
                'download_path': self.download_path
            }
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            logging.info("설정 저장 완료")
        except Exception as e:
            logging.error(f"설정 저장 오류: {e}")

    def load_lawdong_file(self):
        """법정동 코드 파일 로드"""
        try:
            if not os.path.exists(self.lawdong_path):
                logging.error(f"법정동 파일이 존재하지 않습니다: {self.lawdong_path}")
                return False

            for encoding in ['cp949', 'euc-kr', 'utf-8']:
                try:
                    with open(self.lawdong_path, 'r', encoding=encoding) as file:
                        law_dong_data = []
                        for line in file:
                            parts = line.strip().split('\t')
                            if len(parts) < 2:
                                continue
                            code = parts[0].strip()
                            name = parts[1].strip()
                            if any('폐지' in part for part in parts):
                                continue
                            law_dong_data.append({
                                'code': code,
                                'name': name,
                                'sido_code': code[:2],
                                'sigungu_code': code[2:5],
                                'dong_code': code[5:]
                            })

                        self.sido_list = []
                        self.sigungu_dict = {}
                        self.dong_dict = {}
                        self.region_codes = {}

                        sido_data = [item for item in law_dong_data if item['code'].endswith('00000000')]
                        for sido in sido_data:
                            sido_name = sido['name']
                            self.sido_list.append(sido_name)
                            self.sigungu_dict[sido_name] = []

                        sigungu_data = [item for item in law_dong_data
                                       if item['dong_code'] == '00000' and not item['code'].endswith('00000000')]

                        for item in sigungu_data:
                            names = item['name'].split()
                            if len(names) >= 2:
                                sido_name = names[0]
                                sigungu_name = ' '.join(names[1:])

                                if sido_name in self.sigungu_dict:
                                    if sigungu_name not in self.sigungu_dict[sido_name]:
                                        self.sigungu_dict[sido_name].append(sigungu_name)
                                        key = f"{sido_name} {sigungu_name}"
                                        self.dong_dict[key] = []

                        dong_data = [item for item in law_dong_data if item['dong_code'] != '00000']

                        for item in dong_data:
                            names = item['name'].split()
                            if len(names) >= 3:
                                sido_name = names[0]
                                dong_name = names[-1]
                                sigungu_name = ' '.join(names[1:-1])

                                key = f"{sido_name} {sigungu_name}"
                                if key in self.dong_dict:
                                    if dong_name not in self.dong_dict[key]:
                                        self.dong_dict[key].append(dong_name)

                                full_name = f"{sido_name} {sigungu_name} {dong_name}"
                                self.region_codes[full_name] = item['code']

                        logging.info(f"법정동 파일 로드 완료: {len(self.sido_list)}개 시도, {len(self.dong_dict)}개 시군구")
                        return True

                except UnicodeDecodeError:
                    continue

            logging.error("법정동 파일 인코딩을 확인할 수 없습니다.")
            return False

        except Exception as e:
            logging.error(f"법정동 파일 로드 오류: {e}")
            return False

    def setup_driver(self):
        """Chrome WebDriver 설정 (undetected-chromedriver 사용)"""
        try:
            if self.driver is not None:
                return self.driver

            self.update_status("Chrome 브라우저 초기화 중 (undetected mode)...")

            # undetected_chromedriver 옵션 설정
            options = uc.ChromeOptions()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-infobars")
            options.add_argument("--disable-extensions")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")

            # undetected_chromedriver로 브라우저 생성
            self.driver = uc.Chrome(options=options, use_subprocess=True)

            logging.info("Chrome WebDriver 초기화 완료 (undetected mode)")
            return self.driver

        except Exception as e:
            logging.error(f"WebDriver 초기화 오류: {e}")
            self.update_status(f"브라우저 초기화 오류: {str(e)}")
            return None

    def close_driver(self):
        """WebDriver 종료"""
        try:
            if self.driver:
                self.driver.quit()
                self.driver = None
                logging.info("WebDriver 종료")
        except Exception as e:
            logging.error(f"WebDriver 종료 오류: {e}")

    def save_debug_html(self, driver, name):
        """디버깅용 HTML 저장"""
        try:
            debug_path = os.path.join(os.path.dirname(__file__), f"debug_{name}.html")
            with open(debug_path, 'w', encoding='utf-8') as f:
                f.write(driver.page_source)
            logging.info(f"디버그 HTML 저장: {debug_path}")
        except Exception as e:
            logging.error(f"디버그 HTML 저장 오류: {e}")

    def load_all_complexes(self, driver):
        """'더보기' 버튼을 클릭하여 전체 단지 목록 로드"""
        try:
            more_btn_selectors = [
                "button[class*='button-more']",
                "button[class*='ButtonMore']",
                "[class*='more'] button",
                "button[data-nlogs*='more']",
                "button:contains('더보기')"
            ]

            click_count = 0
            max_clicks = 20  # 최대 20번 클릭 (안전장치)

            while click_count < max_clicks:
                more_btn = None

                # 더보기 버튼 찾기
                for selector in more_btn_selectors:
                    try:
                        if 'contains' in selector:
                            # XPath로 텍스트 검색
                            buttons = driver.find_elements(By.XPATH, "//button[contains(text(), '더보기')]")
                        else:
                            buttons = driver.find_elements(By.CSS_SELECTOR, selector)

                        for btn in buttons:
                            if btn.is_displayed() and '더보기' in btn.text:
                                more_btn = btn
                                break
                        if more_btn:
                            break
                    except:
                        continue

                if not more_btn:
                    logging.info(f"더보기 버튼 클릭 완료 (총 {click_count}회)")
                    break

                try:
                    # 스크롤하여 버튼 보이게
                    driver.execute_script("arguments[0].scrollIntoView(true);", more_btn)
                    time.sleep(0.5)

                    # 버튼 클릭
                    driver.execute_script("arguments[0].click();", more_btn)
                    click_count += 1
                    logging.info(f"더보기 버튼 클릭 ({click_count}회)")
                    time.sleep(1)

                except Exception as e:
                    logging.debug(f"더보기 버튼 클릭 오류: {e}")
                    break

            # 스크롤을 맨 위로
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)

        except Exception as e:
            logging.error(f"전체 단지 로드 오류: {e}")

    def click_autocomplete_item(self, driver, keyword):
        """자동완성 목록에서 항목 클릭"""
        try:
            time.sleep(1.5)

            # 자동완성 목록 선택자들
            autocomplete_selectors = [
                "[class*='autocomplete'] li",
                "[class*='Autocomplete'] li",
                "[class*='suggest'] li",
                "[class*='Suggest'] li",
                "[class*='dropdown'] li",
                "[class*='search'] ul li",
                "[class*='result'] li a",
                "[role='listbox'] [role='option']",
                "[class*='layer'] li",
                "[class*='list'] li a[href*='complex']"
            ]

            for selector in autocomplete_selectors:
                try:
                    items = driver.find_elements(By.CSS_SELECTOR, selector)
                    for item in items:
                        if item.is_displayed():
                            item_text = item.text.strip()
                            if item_text:
                                logging.info(f"자동완성 항목 발견: {item_text[:50]}")
                                item.click()
                                time.sleep(2)
                                return True
                except:
                    continue

            # 방법 2: 아파트/단지 링크 직접 클릭
            try:
                complex_links = driver.find_elements(By.CSS_SELECTOR, "a[href*='complexes']")
                for link in complex_links:
                    if link.is_displayed():
                        logging.info(f"단지 링크 클릭: {link.text[:30] if link.text else 'no text'}")
                        link.click()
                        time.sleep(2)
                        return True
            except:
                pass

            return False
        except Exception as e:
            logging.error(f"자동완성 클릭 오류: {e}")
            return False

    def setup_gui(self):
        """GUI 설정"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 지역 선택
        region_frame = ttk.LabelFrame(main_frame, text="지역 선택", padding="10")
        region_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(region_frame, text="시도:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.sido_var = tk.StringVar()
        self.sido_combo = ttk.Combobox(region_frame, textvariable=self.sido_var,
                                        values=self.sido_list, state="readonly", width=15)
        self.sido_combo.grid(row=0, column=1, padx=5, pady=5)
        self.sido_combo.bind('<<ComboboxSelected>>', self.on_sido_selected)

        ttk.Label(region_frame, text="시군구:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.sigungu_var = tk.StringVar()
        self.sigungu_combo = ttk.Combobox(region_frame, textvariable=self.sigungu_var,
                                           state="readonly", width=15)
        self.sigungu_combo.grid(row=0, column=3, padx=5, pady=5)
        self.sigungu_combo.bind('<<ComboboxSelected>>', self.on_sigungu_selected)

        ttk.Label(region_frame, text="동:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.dong_var = tk.StringVar()
        self.dong_combo = ttk.Combobox(region_frame, textvariable=self.dong_var,
                                        state="readonly", width=15)
        self.dong_combo.grid(row=0, column=5, padx=5, pady=5)

        self.search_button = ttk.Button(region_frame, text="매물 조회", command=self.start_search)
        self.search_button.grid(row=0, column=6, padx=20, pady=5)

        self.stop_button = ttk.Button(region_frame, text="중단", command=self.stop_search_process, state="disabled")
        self.stop_button.grid(row=0, column=7, padx=5, pady=5)

        # 거래 유형 선택
        option_frame = ttk.LabelFrame(main_frame, text="거래 유형", padding="10")
        option_frame.pack(fill="x", pady=(0, 10))

        self.trade_sale = tk.BooleanVar(value=True)
        self.trade_jeonse = tk.BooleanVar(value=True)
        self.trade_monthly = tk.BooleanVar(value=True)

        ttk.Checkbutton(option_frame, text="매매", variable=self.trade_sale).pack(side="left", padx=10)
        ttk.Checkbutton(option_frame, text="전세", variable=self.trade_jeonse).pack(side="left", padx=10)
        ttk.Checkbutton(option_frame, text="월세", variable=self.trade_monthly).pack(side="left", padx=10)

        ttk.Button(option_frame, text="설정", command=self.show_settings).pack(side="right", padx=10)

        # 진행 상태
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill="x", pady=(0, 10))

        self.status_var = tk.StringVar(value="대기 중...")
        self.status_label = ttk.Label(progress_frame, textvariable=self.status_var)
        self.status_label.pack(fill="x")

        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.pack(fill="x", pady=5)

        # 결과 테이블
        result_frame = ttk.LabelFrame(main_frame, text="조회 결과", padding="10")
        result_frame.pack(fill="both", expand=True, pady=(0, 10))

        columns = ('단지명', '거래유형', '전용면적', '층', '매매가', '보증금', '월세', '중개사명')
        self.tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=15)

        col_widths = {'단지명': 180, '거래유형': 70, '전용면적': 80, '층': 80,
                      '매매가': 100, '보증금': 100, '월세': 80, '중개사명': 150}
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 100), anchor='center')

        scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar_y.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")

        # 하단 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x")

        self.result_count_var = tk.StringVar(value="총 0건")
        ttk.Label(button_frame, textvariable=self.result_count_var).pack(side="left")

        self.export_button = ttk.Button(button_frame, text="Excel 저장", command=self.export_to_excel, state="disabled")
        self.export_button.pack(side="right", padx=5)

        ttk.Button(button_frame, text="결과 초기화", command=self.clear_results).pack(side="right", padx=5)

        # 브라우저 종료 버튼
        ttk.Button(button_frame, text="브라우저 종료", command=self.close_driver).pack(side="right", padx=5)

    def on_sido_selected(self, event):
        sido = self.sido_var.get()
        self.sigungu_combo['values'] = self.sigungu_dict.get(sido, [])
        self.sigungu_var.set('')
        self.dong_var.set('')
        self.dong_combo['values'] = []

    def on_sigungu_selected(self, event):
        sido = self.sido_var.get()
        sigungu = self.sigungu_var.get()
        key = f"{sido} {sigungu}"
        self.dong_combo['values'] = self.dong_dict.get(key, [])
        self.dong_var.set('')

    def show_settings(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("설정")
        dialog.geometry("500x200")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="법정동 코드 파일:").grid(row=0, column=0, sticky="w", pady=5)
        lawdong_var = tk.StringVar(value=self.lawdong_path)
        ttk.Entry(frame, textvariable=lawdong_var, width=40).grid(row=0, column=1, padx=5, pady=5)

        def browse_lawdong():
            path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
            if path:
                lawdong_var.set(path)

        ttk.Button(frame, text="찾기", command=browse_lawdong).grid(row=0, column=2, pady=5)

        ttk.Label(frame, text="저장 경로:").grid(row=1, column=0, sticky="w", pady=5)
        download_var = tk.StringVar(value=self.download_path)
        ttk.Entry(frame, textvariable=download_var, width=40).grid(row=1, column=1, padx=5, pady=5)

        def browse_download():
            path = filedialog.askdirectory()
            if path:
                download_var.set(path)

        ttk.Button(frame, text="찾기", command=browse_download).grid(row=1, column=2, pady=5)

        def save():
            self.lawdong_path = lawdong_var.get()
            self.download_path = download_var.get()
            self.save_settings()

            if self.load_lawdong_file():
                self.sido_combo['values'] = self.sido_list
                messagebox.showinfo("설정", "설정이 저장되었습니다.")
            else:
                messagebox.showerror("오류", "법정동 파일을 로드할 수 없습니다.")
            dialog.destroy()

        ttk.Button(frame, text="저장", command=save).grid(row=2, column=1, pady=20)

    def start_search(self):
        dong = self.dong_var.get()
        if not dong:
            messagebox.showwarning("경고", "동을 선택해주세요.")
            return

        if not any([self.trade_sale.get(), self.trade_jeonse.get(), self.trade_monthly.get()]):
            messagebox.showwarning("경고", "최소 하나의 거래 유형을 선택해주세요.")
            return

        self.search_button.config(state="disabled")
        self.stop_button.config(state="normal")
        self.export_button.config(state="disabled")
        self.stop_search = False

        self.clear_results()

        thread = threading.Thread(target=self.search_properties, daemon=True)
        thread.start()

    def stop_search_process(self):
        self.stop_search = True
        self.update_status("검색 중단 중...")

    def search_properties(self):
        """fin.land.naver.com/home에서 매물 검색"""
        try:
            sido = self.sido_var.get()
            sigungu = self.sigungu_var.get()
            dong = self.dong_var.get()

            search_keyword = f"{sigungu} {dong}"
            self.update_status(f"'{search_keyword}' 검색 준비 중...")

            # WebDriver 설정
            driver = self.setup_driver()
            if not driver:
                self.update_status("브라우저 초기화 실패")
                self.root.after(0, self.search_completed)
                return

            all_properties = []

            try:
                # fin.land.naver.com/home 접속
                self.update_status("fin.land.naver.com 접속 중...")
                driver.get("https://fin.land.naver.com/home")
                time.sleep(4)

                # 디버깅: 페이지 HTML 저장
                self.save_debug_html(driver, "home_page")

                # 검색창 찾기 및 검색어 입력
                self.update_status(f"'{search_keyword}' 검색 중...")

                search_success = False

                # 방법 1: 검색 버튼/아이콘 클릭 후 입력
                try:
                    search_btn_selectors = [
                        "button[class*='search']",
                        "[class*='search'] button",
                        "[class*='Search']",
                        "a[href*='search']",
                        "[class*='gnb'] [class*='search']"
                    ]
                    for btn_sel in search_btn_selectors:
                        try:
                            search_btn = driver.find_element(By.CSS_SELECTOR, btn_sel)
                            if search_btn.is_displayed():
                                search_btn.click()
                                time.sleep(1)
                                logging.info(f"검색 버튼 클릭: {btn_sel}")
                                break
                        except:
                            continue
                except:
                    pass

                # 방법 2: 검색창 직접 찾기
                search_selectors = [
                    "input[placeholder*='아파트']",
                    "input[placeholder*='검색']",
                    "input[placeholder*='지역']",
                    "input[placeholder*='단지']",
                    "[class*='search'] input",
                    "[class*='Search'] input",
                    "input[type='search']",
                    "input[type='text']"
                ]

                for selector in search_selectors:
                    try:
                        search_input = WebDriverWait(driver, 3).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        if search_input.is_displayed():
                            search_input.click()
                            time.sleep(0.5)
                            search_input.clear()
                            search_input.send_keys(search_keyword)
                            time.sleep(2)  # 자동완성 대기

                            # 자동완성 목록에서 선택 시도
                            autocomplete_found = self.click_autocomplete_item(driver, search_keyword)

                            if not autocomplete_found:
                                # 자동완성이 없으면 Enter
                                search_input.send_keys(Keys.ENTER)

                            time.sleep(3)
                            search_success = True
                            logging.info(f"검색창 찾음: {selector}")
                            break
                    except Exception as e:
                        logging.debug(f"선택자 {selector} 실패: {e}")
                        continue

                if not search_success:
                    # URL로 직접 검색
                    self.update_status("URL로 직접 검색 시도...")
                    import urllib.parse
                    encoded_keyword = urllib.parse.quote(search_keyword)
                    driver.get(f"https://fin.land.naver.com/search?query={encoded_keyword}")
                    time.sleep(3)

                # 검색 결과 페이지에서 단지 목록 수집
                self.update_status("검색 결과 분석 중...")

                # 현재 URL 확인
                current_url = driver.current_url
                logging.info(f"현재 URL: {current_url}")

                # 디버깅: 검색 결과 HTML 저장
                self.save_debug_html(driver, "search_result")

                # 페이지 스크롤하여 동적 콘텐츠 로드
                for _ in range(3):
                    driver.execute_script("window.scrollBy(0, 500);")
                    time.sleep(1)
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(1)

                # 단지 링크 수집
                complex_list = self.collect_complex_from_search(driver)

                if not complex_list:
                    self.update_status("검색 결과가 없습니다.")
                    self.root.after(0, self.search_completed)
                    return

                self.update_status(f"{len(complex_list)}개 단지 발견. 매물 조회 시작...")

                # 각 단지의 매물 조회
                total = len(complex_list)
                for idx, (complex_id, complex_name) in enumerate(complex_list):
                    if self.stop_search:
                        break

                    self.update_status(f"'{complex_name}' 매물 조회 중... ({idx+1}/{total})")
                    self.update_progress((idx + 1) / total * 100)

                    properties = self.get_complex_articles_fin(driver, complex_id, complex_name)

                    # 거래 유형 필터링
                    for prop in properties:
                        trade_type = prop.get('거래유형', '')
                        if (trade_type == '매매' and self.trade_sale.get()) or \
                           (trade_type == '전세' and self.trade_jeonse.get()) or \
                           (trade_type == '월세' and self.trade_monthly.get()):
                            all_properties.append(prop)

                    time.sleep(1.5)

            except Exception as e:
                logging.error(f"검색 중 오류: {e}")
                import traceback
                traceback.print_exc()

            self.properties_data = all_properties
            self.root.after(0, lambda: self.display_results(all_properties))

            if self.stop_search:
                self.update_status(f"검색 중단됨. {len(all_properties)}건 조회됨.")
            else:
                self.update_status(f"검색 완료. 총 {len(all_properties)}건 조회됨.")

        except Exception as e:
            logging.error(f"검색 오류: {e}")
            self.update_status(f"오류 발생: {str(e)}")
        finally:
            self.root.after(0, self.search_completed)

    def collect_complex_from_search(self, driver):
        """검색 결과에서 단지 목록 수집 (버튼 클릭 방식)"""
        complex_list = []
        seen_ids = set()

        try:
            time.sleep(2)

            # "더보기" 버튼을 클릭하여 전체 단지 목록 로드
            self.load_all_complexes(driver)

            # 단지 섹션에서 버튼 항목 찾기
            item_selectors = [
                "[class*='SearchResultList'] button[class*='link']",
                "[class*='searchResult'] button",
                "li[class*='item'] button[class*='link']",
                "[class*='list'] li button"
            ]

            buttons = []
            for selector in item_selectors:
                try:
                    found = driver.find_elements(By.CSS_SELECTOR, selector)
                    if found:
                        # 아파트 단지만 필터링 (아파트 배지가 있는 항목)
                        for btn in found:
                            try:
                                text = btn.text
                                if '아파트' in text:
                                    buttons.append(btn)
                            except:
                                buttons.append(btn)
                        if buttons:
                            logging.info(f"단지 버튼 발견: {selector}, {len(buttons)}개")
                            break
                except:
                    continue

            if not buttons:
                logging.warning("단지 버튼을 찾지 못함")
                # 방법 2: 링크에서 단지 ID 추출 시도
                all_links = driver.find_elements(By.TAG_NAME, "a")
                for link in all_links:
                    try:
                        href = link.get_attribute("href")
                        if href and "/complexes/" in href:
                            match = re.search(r'/complexes/(\d+)', href)
                            if match:
                                complex_id = match.group(1)
                                if complex_id not in seen_ids:
                                    seen_ids.add(complex_id)
                                    text = link.text.strip()
                                    complex_name = text.split('\n')[0] if text else f"단지_{complex_id}"
                                    complex_list.append((complex_id, complex_name))
                    except:
                        continue

            # 버튼 클릭하여 단지 ID 수집 (전체 단지)
            total_buttons = len(buttons)
            logging.info(f"총 {total_buttons}개 단지 수집 시작...")
            self.update_status(f"총 {total_buttons}개 단지 수집 중...")

            for i in range(total_buttons):
                if self.stop_search:
                    break

                try:
                    # 매번 버튼 목록 새로 조회 (DOM 변경 대응)
                    current_buttons = []
                    for selector in item_selectors:
                        try:
                            found = driver.find_elements(By.CSS_SELECTOR, selector)
                            for btn in found:
                                if '아파트' in btn.text:
                                    current_buttons.append(btn)
                            if current_buttons:
                                break
                        except:
                            continue

                    if i >= len(current_buttons):
                        break

                    btn = current_buttons[i]
                    btn_text = btn.text.strip()

                    # 단지명 추출 (아파트\n단지명\n주소 형태)
                    lines = btn_text.split('\n')
                    complex_name = lines[1] if len(lines) > 1 else lines[0]
                    complex_name = complex_name.replace('아파트', '').strip()

                    # 버튼 클릭
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(1.5)

                    # URL에서 complexId 추출
                    current_url = driver.current_url
                    match = re.search(r'/complexes/(\d+)', current_url)
                    if match:
                        complex_id = match.group(1)
                        if complex_id not in seen_ids:
                            seen_ids.add(complex_id)
                            complex_list.append((complex_id, complex_name))
                            logging.info(f"단지 수집 ({i+1}/{total_buttons}): {complex_name} (ID: {complex_id})")

                    # 뒤로 가기
                    driver.back()
                    time.sleep(1)

                    # 진행률 업데이트
                    if i % 5 == 0:
                        self.update_status(f"단지 수집 중... ({i+1}/{total_buttons})")

                except Exception as e:
                    logging.debug(f"버튼 {i} 클릭 오류: {e}")
                    try:
                        driver.back()
                        time.sleep(1)
                    except:
                        pass
                    continue

            logging.info(f"수집된 단지 수: {len(complex_list)}")

        except Exception as e:
            logging.error(f"단지 목록 수집 오류: {e}")

        return complex_list[:30]  # 최대 30개

    def get_complex_articles_fin(self, driver, complex_id, complex_name):
        """fin.land.naver.com 단지 페이지에서 매물 수집"""
        articles = []

        try:
            # 단지 매물 페이지 접속
            url = f"https://fin.land.naver.com/complexes/{complex_id}?tab=article"
            driver.get(url)
            time.sleep(2)

            # 페이지 스크롤하며 매물 수집
            collected_texts = set()
            scroll_count = 0

            while scroll_count < 5:
                # 매물 항목 찾기 - 다양한 선택자 시도
                item_selectors = [
                    "[class*='ArticleItem']",
                    "[class*='articleItem']",
                    "[class*='article_item']",
                    "[class*='item_article']",
                    "li[class*='article']",
                    "[data-article-id]"
                ]

                items = []
                for selector in item_selectors:
                    try:
                        found = driver.find_elements(By.CSS_SELECTOR, selector)
                        if found:
                            items = found
                            break
                    except:
                        continue

                for item in items:
                    try:
                        text = item.text
                        if not text or len(text) < 5:
                            continue

                        # 중복 체크
                        text_hash = hash(text[:80])
                        if text_hash in collected_texts:
                            continue
                        collected_texts.add(text_hash)

                        # 매물 정보 파싱
                        article = self.parse_fin_article(text, complex_name)
                        if article and article.get('거래유형'):
                            articles.append(article)
                    except:
                        continue

                # 스크롤
                driver.execute_script("window.scrollBy(0, 400);")
                time.sleep(0.8)
                scroll_count += 1

        except Exception as e:
            logging.error(f"단지 매물 수집 오류 ({complex_name}): {e}")

        return articles

    def parse_fin_article(self, text, complex_name):
        """fin.land.naver.com 매물 텍스트 파싱"""
        try:
            article = {
                '단지명': complex_name,
                '거래유형': '',
                '전용면적': '',
                '층': '',
                '매매가': '',
                '보증금': '',
                '월세': '',
                '중개사명': ''
            }

            lines = text.split('\n')

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # 거래유형
                if '매매' in line and not article['거래유형']:
                    article['거래유형'] = '매매'
                elif '전세' in line and not article['거래유형']:
                    article['거래유형'] = '전세'
                elif '월세' in line and not article['거래유형']:
                    article['거래유형'] = '월세'

                # 면적
                area_match = re.search(r'(\d+\.?\d*)\s*㎡', line)
                if area_match and not article['전용면적']:
                    article['전용면적'] = f"{area_match.group(1)}㎡"

                # 층
                floor_match = re.search(r'(\d+)\s*층', line)
                if floor_match and not article['층']:
                    article['층'] = f"{floor_match.group(1)}층"

                # 가격 패턴
                # 억 단위
                price_match = re.search(r'(\d+억\s*\d*,?\d*만?|\d+억)', line)
                if price_match:
                    price = price_match.group(1)
                    if article['거래유형'] == '매매' and not article['매매가']:
                        article['매매가'] = price
                    elif article['거래유형'] == '전세' and not article['보증금']:
                        article['보증금'] = price
                    elif article['거래유형'] == '월세' and not article['보증금']:
                        article['보증금'] = price

                # 만원 단위
                if not price_match:
                    won_match = re.search(r'(\d+,?\d*만)', line)
                    if won_match:
                        price = won_match.group(1)
                        if article['거래유형'] == '월세' and article['보증금'] and not article['월세']:
                            article['월세'] = price

                # 중개사
                if ('부동산' in line or '공인' in line) and not article['중개사명']:
                    article['중개사명'] = line[:25]

            return article if article['거래유형'] else None

        except:
            return None

    def search_via_fin_land(self, driver, search_keyword):
        """fin.land.naver.com에서 검색"""
        all_properties = []

        try:
            import urllib.parse
            encoded_keyword = urllib.parse.quote(search_keyword)

            # fin.land.naver.com 검색 페이지
            url = f"https://fin.land.naver.com/search?query={encoded_keyword}"
            self.update_status(f"fin.land.naver.com 검색 중...")
            driver.get(url)
            time.sleep(3)

            # 아파트 단지 목록 찾기
            try:
                complex_items = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "[class*='complex'], [class*='item'], .search_item"))
                )

                complex_links = []
                for item in complex_items[:20]:  # 최대 20개 단지
                    try:
                        link = item.find_element(By.TAG_NAME, "a")
                        href = link.get_attribute("href")
                        if href and "complexes" in href:
                            complex_links.append(href)
                    except:
                        pass

                self.update_status(f"{len(complex_links)}개 단지 발견")

                for idx, link in enumerate(complex_links):
                    if self.stop_search:
                        break

                    self.update_status(f"단지 매물 조회 중... ({idx+1}/{len(complex_links)})")
                    self.update_progress((idx + 1) / len(complex_links) * 100)

                    driver.get(link)
                    time.sleep(2)

                    # 매물 탭 클릭
                    try:
                        article_tab = driver.find_element(By.XPATH, "//a[contains(text(), '매물') or contains(@href, 'article')]")
                        article_tab.click()
                        time.sleep(2)
                    except:
                        pass

                    # 매물 정보 추출
                    properties = self.extract_properties_from_complex_page(driver)
                    all_properties.extend(properties)

            except Exception as e:
                logging.error(f"단지 검색 오류: {e}")

        except Exception as e:
            logging.error(f"fin.land 검색 오류: {e}")

        return all_properties

    def extract_properties_from_page(self, driver):
        """현재 페이지에서 매물 정보 추출"""
        properties = []

        try:
            # 다양한 선택자로 매물 항목 찾기
            selectors = [
                ".item_inner", ".article_item", ".complex_item",
                "[class*='article']", "[class*='item']", ".lst_complex li"
            ]

            items = []
            for selector in selectors:
                try:
                    found = driver.find_elements(By.CSS_SELECTOR, selector)
                    if found:
                        items = found
                        break
                except:
                    continue

            for item in items:
                try:
                    prop = self.parse_property_element(item)
                    if prop and prop.get('단지명'):
                        properties.append(prop)
                except:
                    continue

        except Exception as e:
            logging.error(f"매물 추출 오류: {e}")

        return properties

    def extract_properties_from_complex_page(self, driver):
        """단지 상세 페이지에서 매물 정보 추출"""
        properties = []

        try:
            # 단지명 추출
            complex_name = ""
            try:
                name_elem = driver.find_element(By.CSS_SELECTOR, "h2, .complex_title, [class*='title']")
                complex_name = name_elem.text.strip()
            except:
                pass

            # 매물 목록 스크롤
            scroll_container = None
            try:
                scroll_container = driver.find_element(By.CSS_SELECTOR, "[class*='article_list'], [class*='scroll']")
            except:
                pass

            # 페이지 스크롤하며 매물 수집
            for _ in range(5):
                items = driver.find_elements(By.CSS_SELECTOR, "[class*='article_item'], [class*='item'], li[class*='article']")

                for item in items:
                    try:
                        prop = self.parse_property_element(item, complex_name)
                        if prop and prop not in properties:
                            properties.append(prop)
                    except:
                        continue

                # 스크롤
                if scroll_container:
                    driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scroll_container)
                else:
                    driver.execute_script("window.scrollBy(0, 500);")
                time.sleep(1)

        except Exception as e:
            logging.error(f"단지 매물 추출 오류: {e}")

        return properties

    def parse_property_element(self, element, default_complex_name=""):
        """매물 요소에서 정보 추출"""
        try:
            text = element.text
            if not text:
                return None

            lines = text.split('\n')

            prop = {
                '단지명': default_complex_name,
                '거래유형': '',
                '전용면적': '',
                '층': '',
                '매매가': '',
                '보증금': '',
                '월세': '',
                '중개사명': ''
            }

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # 단지명 (한글로 시작하고 길이가 있는 경우)
                if not prop['단지명'] and len(line) > 2 and re.match(r'^[가-힣]', line):
                    if '아파트' in line or '단지' in line or len(line) <= 30:
                        prop['단지명'] = line

                # 거래유형
                if '매매' in line:
                    prop['거래유형'] = '매매'
                elif '전세' in line:
                    prop['거래유형'] = '전세'
                elif '월세' in line:
                    prop['거래유형'] = '월세'

                # 면적
                area_match = re.search(r'(\d+\.?\d*)\s*㎡', line)
                if area_match and not prop['전용면적']:
                    prop['전용면적'] = f"{area_match.group(1)}㎡"

                # 층
                floor_match = re.search(r'(\d+층|\d+/\d+층)', line)
                if floor_match:
                    prop['층'] = floor_match.group(1)

                # 가격 (억, 만원)
                price_match = re.search(r'(\d+억\s*\d*만?|\d+,?\d*만원|\d+억)', line)
                if price_match:
                    price = price_match.group(1)
                    if prop['거래유형'] == '매매' and not prop['매매가']:
                        prop['매매가'] = price
                    elif prop['거래유형'] == '전세' and not prop['보증금']:
                        prop['보증금'] = price
                    elif prop['거래유형'] == '월세':
                        if not prop['보증금']:
                            prop['보증금'] = price
                        elif not prop['월세']:
                            prop['월세'] = price

                # 중개사
                if '공인' in line or '부동산' in line:
                    prop['중개사명'] = line[:20]

            return prop if prop['단지명'] or prop['거래유형'] else None

        except Exception as e:
            return None

    def display_results(self, properties):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for prop in properties:
            values = (
                prop.get('단지명', ''),
                prop.get('거래유형', ''),
                prop.get('전용면적', ''),
                prop.get('층', ''),
                prop.get('매매가', ''),
                prop.get('보증금', ''),
                prop.get('월세', ''),
                prop.get('중개사명', '')
            )
            self.tree.insert('', 'end', values=values)

        self.result_count_var.set(f"총 {len(properties)}건")

    def clear_results(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.properties_data = []
        self.result_count_var.set("총 0건")
        self.progress_bar['value'] = 0

    def search_completed(self):
        self.search_button.config(state="normal")
        self.stop_button.config(state="disabled")
        if self.properties_data:
            self.export_button.config(state="normal")

    def update_status(self, message):
        self.root.after(0, lambda: self.status_var.set(message))

    def update_progress(self, value):
        self.root.after(0, lambda: self.progress_bar.configure(value=value))

    def export_to_excel(self):
        if not self.properties_data:
            messagebox.showwarning("경고", "저장할 데이터가 없습니다.")
            return

        try:
            sigungu = self.sigungu_var.get()
            dong = self.dong_var.get()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"매물조회_{sigungu}_{dong}_{timestamp}.xlsx"
            filepath = os.path.join(self.download_path, filename)

            df = pd.DataFrame(self.properties_data)
            df['조회일시'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='전체매물')

            self.format_excel(filepath)

            messagebox.showinfo("저장 완료", f"파일이 저장되었습니다.\n{filepath}")
            logging.info(f"Excel 저장 완료: {filepath}")

        except Exception as e:
            logging.error(f"Excel 저장 오류: {e}")
            messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다.\n{str(e)}")

    def format_excel(self, filepath):
        try:
            wb = load_workbook(filepath)
            ws = wb.active

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(filepath)

        except Exception as e:
            logging.error(f"Excel 서식 적용 오류: {e}")

    def run(self):
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()

    def on_closing(self):
        self.close_driver()
        self.root.destroy()


if __name__ == "__main__":
    app = NaverRealEstateApp()
    app.run()
