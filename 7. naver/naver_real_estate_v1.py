"""
네이버 부동산 매물 조회 프로그램 v1
- 단지명 직접 입력 방식
- 검색 결과에서 단지 선택 후 매물 조회
- Excel 파일로 저장
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
import os
import time
import threading
import re
from datetime import datetime
import logging
import warnings

warnings.filterwarnings('ignore')

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


class NaverRealEstateAppV1:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("네이버 부동산 매물 조회 v1 (단지명 검색)")
        self.root.geometry("1200x700")

        # 설정 파일 경로
        self.settings_file = os.path.join(os.path.dirname(__file__), 'naver_real_estate_v1_settings.json')

        # 기본 설정
        self.download_path = os.path.join(os.path.dirname(__file__), "매물조회결과")

        # 매물 데이터
        self.properties_data = []

        # WebDriver
        self.driver = None

        # 조회 중단 플래그
        self.stop_search = False

        # 설정 로드
        self.load_settings()

        # GUI 설정
        self.setup_gui()

        # 폴더 생성
        if not os.path.exists(self.download_path):
            os.makedirs(self.download_path)

    def load_settings(self):
        """설정 파일 로드"""
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    self.download_path = settings.get('download_path', self.download_path)
                logging.info(f"설정 로드 완료: {self.settings_file}")
        except Exception as e:
            logging.error(f"설정 로드 오류: {e}")

    def save_settings(self):
        """설정 파일 저장"""
        try:
            settings = {
                'download_path': self.download_path
            }
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
            logging.info("설정 저장 완료")
        except Exception as e:
            logging.error(f"설정 저장 오류: {e}")

    def setup_driver(self):
        """Chrome WebDriver 설정 (undetected-chromedriver 사용)"""
        try:
            # 기존 드라이버가 있으면 유효한지 확인
            if self.driver is not None:
                try:
                    # 세션이 유효한지 테스트
                    _ = self.driver.current_url
                    return self.driver
                except Exception:
                    # 세션이 끊어졌으면 드라이버 정리
                    logging.info("기존 브라우저 세션 만료, 재시작...")
                    try:
                        self.driver.quit()
                    except:
                        pass
                    self.driver = None

            self.update_status("Chrome 브라우저 초기화 중 (undetected mode)...")

            # undetected_chromedriver 옵션 설정
            options = uc.ChromeOptions()
            options.add_argument("--start-maximized")
            options.add_argument("--disable-infobars")
            options.add_argument("--disable-extensions")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")

            # undetected_chromedriver로 브라우저 생성 (버전 자동 매칭)
            self.driver = uc.Chrome(
                options=options,
                use_subprocess=True,
                version_main=144  # 현재 설치된 Chrome 버전에 맞춤
            )

            logging.info("Chrome WebDriver 초기화 완료 (undetected mode)")
            return self.driver

        except Exception as e:
            logging.error(f"WebDriver 초기화 오류: {e}")
            self.update_status(f"브라우저 초기화 오류: {str(e)}")
            self.driver = None
            return None

    def close_driver(self):
        """WebDriver 종료"""
        try:
            if self.driver:
                self.driver.quit()
                self.driver = None
                logging.info("WebDriver 종료")
        except Exception as e:
            logging.error(f"WebDriver 종료 오류: {e}")

    def setup_gui(self):
        """GUI 설정"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill="both", expand=True)

        # 단지명 입력
        search_frame = ttk.LabelFrame(main_frame, text="단지명 검색", padding="10")
        search_frame.pack(fill="x", pady=(0, 10))

        ttk.Label(search_frame, text="단지명:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.complex_name_var = tk.StringVar()
        self.complex_entry = ttk.Entry(search_frame, textvariable=self.complex_name_var, width=40)
        self.complex_entry.grid(row=0, column=1, padx=5, pady=5)
        self.complex_entry.bind('<Return>', lambda e: self.start_search())

        self.search_button = ttk.Button(search_frame, text="검색", command=self.start_search)
        self.search_button.grid(row=0, column=2, padx=10, pady=5)

        self.stop_button = ttk.Button(search_frame, text="중단", command=self.stop_search_process, state="disabled")
        self.stop_button.grid(row=0, column=3, padx=5, pady=5)

        # 거래 유형 선택
        option_frame = ttk.LabelFrame(main_frame, text="거래 유형", padding="10")
        option_frame.pack(fill="x", pady=(0, 10))

        self.trade_sale = tk.BooleanVar(value=True)
        self.trade_jeonse = tk.BooleanVar(value=True)
        self.trade_monthly = tk.BooleanVar(value=True)

        ttk.Checkbutton(option_frame, text="매매", variable=self.trade_sale).pack(side="left", padx=10)
        ttk.Checkbutton(option_frame, text="전세", variable=self.trade_jeonse).pack(side="left", padx=10)
        ttk.Checkbutton(option_frame, text="월세", variable=self.trade_monthly).pack(side="left", padx=10)

        ttk.Button(option_frame, text="설정", command=self.show_settings).pack(side="right", padx=10)

        # 진행 상태
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill="x", pady=(0, 10))

        self.status_var = tk.StringVar(value="단지명을 입력하고 검색 버튼을 클릭하세요.")
        self.status_label = ttk.Label(progress_frame, textvariable=self.status_var)
        self.status_label.pack(fill="x")

        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.pack(fill="x", pady=5)

        # 결과 테이블
        result_frame = ttk.LabelFrame(main_frame, text="조회 결과", padding="10")
        result_frame.pack(fill="both", expand=True, pady=(0, 10))

        columns = ('단지명', '거래유형', '전용면적', '층', '매매가', '보증금', '월세', '중개사명')
        self.tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=15)

        col_widths = {'단지명': 180, '거래유형': 70, '전용면적': 80, '층': 80,
                      '매매가': 100, '보증금': 100, '월세': 80, '중개사명': 150}
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 100), anchor='center')

        scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar_y.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")

        # 하단 버튼
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x")

        self.result_count_var = tk.StringVar(value="총 0건")
        ttk.Label(button_frame, textvariable=self.result_count_var).pack(side="left")

        self.export_button = ttk.Button(button_frame, text="Excel 저장", command=self.export_to_excel, state="disabled")
        self.export_button.pack(side="right", padx=5)

        ttk.Button(button_frame, text="결과 초기화", command=self.clear_results).pack(side="right", padx=5)

        # 브라우저 종료 버튼
        ttk.Button(button_frame, text="브라우저 종료", command=self.close_driver).pack(side="right", padx=5)

    def show_settings(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("설정")
        dialog.geometry("500x150")
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="저장 경로:").grid(row=0, column=0, sticky="w", pady=5)
        download_var = tk.StringVar(value=self.download_path)
        ttk.Entry(frame, textvariable=download_var, width=40).grid(row=0, column=1, padx=5, pady=5)

        def browse_download():
            path = filedialog.askdirectory()
            if path:
                download_var.set(path)

        ttk.Button(frame, text="찾기", command=browse_download).grid(row=0, column=2, pady=5)

        def save():
            self.download_path = download_var.get()
            self.save_settings()
            messagebox.showinfo("설정", "설정이 저장되었습니다.")
            dialog.destroy()

        ttk.Button(frame, text="저장", command=save).grid(row=1, column=1, pady=20)

    def start_search(self):
        complex_name = self.complex_name_var.get().strip()
        if not complex_name:
            messagebox.showwarning("경고", "단지명을 입력해주세요.")
            return

        if not any([self.trade_sale.get(), self.trade_jeonse.get(), self.trade_monthly.get()]):
            messagebox.showwarning("경고", "최소 하나의 거래 유형을 선택해주세요.")
            return

        self.search_button.config(state="disabled")
        self.stop_button.config(state="normal")
        self.export_button.config(state="disabled")
        self.stop_search = False

        self.clear_results()

        thread = threading.Thread(target=self.search_complex, daemon=True)
        thread.start()

    def stop_search_process(self):
        self.stop_search = True
        self.update_status("검색 중단 중...")

    def search_complex(self):
        """단지명으로 검색하여 매물 조회"""
        try:
            complex_name = self.complex_name_var.get().strip()
            self.update_status(f"'{complex_name}' 검색 준비 중...")

            # WebDriver 설정
            driver = self.setup_driver()
            if not driver:
                self.update_status("브라우저 초기화 실패")
                self.root.after(0, self.search_completed)
                return

            all_properties = []

            try:
                # fin.land.naver.com/home 접속
                self.update_status("fin.land.naver.com 접속 중...")
                driver.get("https://fin.land.naver.com/home")
                time.sleep(4)

                # 검색창 찾기 및 검색어 입력
                self.update_status(f"'{complex_name}' 검색 중...")

                search_success = False

                # 검색 버튼/아이콘 클릭 후 입력
                try:
                    search_btn_selectors = [
                        "button[class*='search']",
                        "[class*='search'] button",
                        "[class*='Search']",
                        "a[href*='search']",
                        "[class*='gnb'] [class*='search']"
                    ]
                    for btn_sel in search_btn_selectors:
                        try:
                            search_btn = driver.find_element(By.CSS_SELECTOR, btn_sel)
                            if search_btn.is_displayed():
                                search_btn.click()
                                time.sleep(1)
                                logging.info(f"검색 버튼 클릭: {btn_sel}")
                                break
                        except:
                            continue
                except:
                    pass

                # 검색창 직접 찾기
                search_selectors = [
                    "input[placeholder*='아파트']",
                    "input[placeholder*='검색']",
                    "input[placeholder*='지역']",
                    "input[placeholder*='단지']",
                    "[class*='search'] input",
                    "[class*='Search'] input",
                    "input[type='search']",
                    "input[type='text']"
                ]

                for selector in search_selectors:
                    try:
                        search_input = WebDriverWait(driver, 3).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        if search_input.is_displayed():
                            search_input.click()
                            time.sleep(0.5)
                            search_input.clear()
                            search_input.send_keys(complex_name)
                            time.sleep(2)  # 자동완성 대기
                            search_input.send_keys(Keys.ENTER)
                            time.sleep(3)
                            search_success = True
                            logging.info(f"검색창 찾음: {selector}")
                            break
                    except Exception as e:
                        logging.debug(f"선택자 {selector} 실패: {e}")
                        continue

                if not search_success:
                    # URL로 직접 검색
                    self.update_status("URL로 직접 검색 시도...")
                    import urllib.parse
                    encoded_keyword = urllib.parse.quote(complex_name)
                    driver.get(f"https://fin.land.naver.com/search?query={encoded_keyword}")
                    time.sleep(3)

                # 검색 결과 페이지에서 단지 선택
                self.update_status("검색 결과에서 단지 찾는 중...")

                current_url = driver.current_url
                logging.info(f"현재 URL: {current_url}")

                # 단지 클릭하여 상세 페이지로 이동
                complex_id, actual_complex_name = self.select_complex_from_search(driver, complex_name)

                if not complex_id:
                    self.update_status("검색 결과가 없습니다.")
                    self.root.after(0, self.search_completed)
                    return

                self.update_status(f"'{actual_complex_name}' 매물 조회 중...")

                # 매물 수집 - API 방식 먼저 시도
                self.update_status(f"'{actual_complex_name}' API로 매물 조회 중...")
                all_properties = self.get_articles_via_api(driver, complex_id, actual_complex_name)

                # API 실패 또는 결과 없으면 스크롤 방식으로 fallback
                if not all_properties:
                    self.update_status(f"'{actual_complex_name}' 스크롤 방식으로 매물 조회 중...")
                    all_properties = self.get_complex_articles(driver, complex_id, actual_complex_name)

                # 거래 유형 필터링
                filtered_properties = []
                for prop in all_properties:
                    trade_type = prop.get('거래유형', '')
                    if (trade_type == '매매' and self.trade_sale.get()) or \
                       (trade_type == '전세' and self.trade_jeonse.get()) or \
                       (trade_type == '월세' and self.trade_monthly.get()):
                        filtered_properties.append(prop)

                all_properties = filtered_properties

            except Exception as e:
                logging.error(f"검색 중 오류: {e}")
                import traceback
                traceback.print_exc()

            self.properties_data = all_properties
            self.root.after(0, lambda: self.display_results(all_properties))

            if self.stop_search:
                self.update_status(f"검색 중단됨. {len(all_properties)}건 조회됨.")
            else:
                self.update_status(f"검색 완료. 총 {len(all_properties)}건 조회됨.")

        except Exception as e:
            logging.error(f"검색 오류: {e}")
            self.update_status(f"오류 발생: {str(e)}")
        finally:
            self.root.after(0, self.search_completed)

    def select_complex_from_search(self, driver, search_keyword):
        """검색 결과에서 첫 번째 아파트 단지 선택"""
        try:
            time.sleep(2)

            # 단지 버튼 찾기
            item_selectors = [
                "[class*='SearchResultList'] button[class*='link']",
                "[class*='searchResult'] button",
                "li[class*='item'] button[class*='link']",
                "[class*='list'] li button"
            ]

            for selector in item_selectors:
                try:
                    buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                    for btn in buttons:
                        try:
                            text = btn.text
                            if '아파트' in text:
                                # 단지명 추출
                                lines = text.split('\n')
                                complex_name = lines[1] if len(lines) > 1 else lines[0]
                                complex_name = complex_name.replace('아파트', '').strip()

                                logging.info(f"단지 발견: {complex_name}")

                                # 버튼 클릭
                                driver.execute_script("arguments[0].click();", btn)
                                time.sleep(2)

                                # URL에서 complexId 추출
                                current_url = driver.current_url
                                match = re.search(r'/complexes/(\d+)', current_url)
                                if match:
                                    complex_id = match.group(1)
                                    logging.info(f"단지 선택: {complex_name} (ID: {complex_id})")
                                    return complex_id, complex_name

                        except Exception as e:
                            logging.debug(f"버튼 처리 오류: {e}")
                            continue
                except:
                    continue

            # 링크 방식으로 시도
            all_links = driver.find_elements(By.TAG_NAME, "a")
            for link in all_links:
                try:
                    href = link.get_attribute("href")
                    if href and "/complexes/" in href:
                        match = re.search(r'/complexes/(\d+)', href)
                        if match:
                            complex_id = match.group(1)
                            text = link.text.strip()
                            complex_name = text.split('\n')[0] if text else search_keyword
                            link.click()
                            time.sleep(2)
                            logging.info(f"단지 선택 (링크): {complex_name} (ID: {complex_id})")
                            return complex_id, complex_name
                except:
                    continue

            return None, None

        except Exception as e:
            logging.error(f"단지 선택 오류: {e}")
            return None, None

    def save_debug_html(self, driver, name):
        """디버깅용 HTML 저장"""
        try:
            debug_path = os.path.join(os.path.dirname(__file__), f"debug_{name}.html")
            with open(debug_path, 'w', encoding='utf-8') as f:
                f.write(driver.page_source)
            logging.info(f"디버그 HTML 저장: {debug_path}")
        except Exception as e:
            logging.error(f"디버그 HTML 저장 오류: {e}")

    def get_browser_cookies(self, driver):
        """브라우저에서 쿠키 가져오기"""
        cookies = {}
        try:
            for cookie in driver.get_cookies():
                cookies[cookie['name']] = cookie['value']
        except:
            pass
        return cookies

    def get_articles_via_api(self, driver, complex_id, complex_name):
        """API를 통해 매물 목록 가져오기 - 스크롤 방식으로 우회"""
        # API 방식이 불안정하므로 스크롤 방식을 우선 사용
        # 빈 리스트 반환하여 스크롤 방식으로 fallback
        logging.info("API 방식 스킵, 스크롤 방식 사용")
        return []

    def get_articles_via_requests(self, driver, complex_id, complex_name):
        """requests 라이브러리로 API 호출 (fallback) - 사용 안함"""
        return []

    def _old_get_articles_via_requests(self, driver, complex_id, complex_name):
        """requests 라이브러리로 API 호출 (fallback)"""
        import requests

        articles = []

        try:
            # 쿠키 가져오기
            cookies = self.get_browser_cookies(driver)

            # 헤더 설정
            headers = {
                'User-Agent': driver.execute_script("return navigator.userAgent"),
                'Accept': 'application/json, text/plain, */*',
                'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                'Referer': f'https://fin.land.naver.com/complexes/{complex_id}?tab=article',
                'Origin': 'https://fin.land.naver.com'
            }

            # 세션 생성
            session = requests.Session()
            session.headers.update(headers)

            # 쿠키 설정
            for name, value in cookies.items():
                session.cookies.set(name, value, domain='.naver.com')

            page = 1
            max_pages = 50

            while page <= max_pages:
                if self.stop_search:
                    break

                api_url = f"https://fin.land.naver.com/front-api/v1/complex/article/list?complexNumber={complex_id}&page={page}&userChannelType=PC"

                try:
                    response = session.get(api_url, timeout=10)

                    if response.status_code == 200:
                        data = response.json()

                        if data.get('isSuccess') and data.get('result'):
                            result = data['result']
                            article_list = result.get('list', [])

                            if not article_list:
                                break

                            for item in article_list:
                                article = self.parse_api_article(item, complex_name)
                                if article:
                                    articles.append(article)

                            logging.info(f"Requests API 페이지 {page}: {len(article_list)}개 수집")

                            if result.get('hasMore', False) or len(article_list) >= 20:
                                page += 1
                                time.sleep(0.5)
                            else:
                                break
                        else:
                            break
                    elif response.status_code == 429:
                        logging.warning("API 요청 제한 (429)")
                        time.sleep(3)
                        continue
                    else:
                        break

                except Exception as e:
                    logging.error(f"Requests API 오류: {e}")
                    break

        except Exception as e:
            logging.error(f"Requests fallback 오류: {e}")

        return articles

    def parse_api_article(self, item, complex_name):
        """API 응답에서 매물 정보 파싱"""
        try:
            article = {
                '단지명': complex_name,
                '거래유형': '',
                '전용면적': '',
                '층': '',
                '매매가': '',
                '보증금': '',
                '월세': '',
                '중개사명': ''
            }

            # 거래 유형
            trade_type = item.get('tradeTypeName', '')
            if '매매' in trade_type:
                article['거래유형'] = '매매'
            elif '전세' in trade_type:
                article['거래유형'] = '전세'
            elif '월세' in trade_type:
                article['거래유형'] = '월세'
            else:
                article['거래유형'] = trade_type

            # 면적
            area = item.get('exclusiveArea') or item.get('area') or item.get('supplyArea', '')
            if area:
                article['전용면적'] = f"{area}㎡"

            # 층
            floor = item.get('floor') or item.get('floorInfo', '')
            if floor:
                article['층'] = f"{floor}층" if not str(floor).endswith('층') else str(floor)

            # 가격
            deal_price = item.get('dealPrice') or item.get('price', 0)
            warranty_price = item.get('warrantyPrice') or item.get('deposit', 0)
            rent_price = item.get('rentPrice') or item.get('monthlyRent', 0)

            def format_price(price):
                if not price:
                    return ''
                try:
                    price = int(price)
                    if price >= 10000:
                        억 = price // 10000
                        만 = price % 10000
                        if 만 > 0:
                            return f"{억}억 {만:,}만"
                        return f"{억}억"
                    return f"{price:,}만"
                except:
                    return str(price)

            if article['거래유형'] == '매매':
                article['매매가'] = format_price(deal_price)
            elif article['거래유형'] == '전세':
                article['보증금'] = format_price(warranty_price or deal_price)
            elif article['거래유형'] == '월세':
                article['보증금'] = format_price(warranty_price)
                article['월세'] = format_price(rent_price)

            # 중개사
            broker = item.get('realtorName') or item.get('brokerageName', '')
            article['중개사명'] = broker[:25] if broker else ''

            return article if article['거래유형'] else None

        except Exception as e:
            logging.debug(f"API 매물 파싱 오류: {e}")
            return None

    def click_more_button(self, driver):
        """더보기 버튼 클릭 (개선된 버전)"""
        try:
            # 다양한 더보기 버튼 셀렉터
            more_selectors = [
                "button[class*='more']",
                "button[class*='More']",
                "[class*='more'] button",
                "[class*='More'] button",
                "button[class*='load']",
                "button[class*='Load']",
                "[class*='pagination'] button",
                "a[class*='more']",
                "[class*='btn'][class*='more']",
                "button[data-nlogs*='more']"
            ]

            for selector in more_selectors:
                try:
                    buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                    for btn in buttons:
                        try:
                            if btn.is_displayed():
                                btn_text = btn.text.lower()
                                if '더보기' in btn_text or '더 보기' in btn_text or 'more' in btn_text or '다음' in btn_text:
                                    driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                                    time.sleep(0.3)
                                    driver.execute_script("arguments[0].click();", btn)
                                    logging.info(f"더보기 버튼 클릭: {btn_text[:20]}")
                                    time.sleep(1)
                                    return True
                        except:
                            continue
                except:
                    continue

            # 텍스트로 버튼 찾기
            try:
                all_buttons = driver.find_elements(By.TAG_NAME, "button")
                for btn in all_buttons:
                    try:
                        if btn.is_displayed():
                            btn_text = btn.text
                            if '더보기' in btn_text or '더 보기' in btn_text or '매물 더보기' in btn_text:
                                driver.execute_script("arguments[0].scrollIntoView(true);", btn)
                                time.sleep(0.3)
                                driver.execute_script("arguments[0].click();", btn)
                                logging.info(f"더보기 버튼 클릭 (텍스트): {btn_text[:20]}")
                                time.sleep(1)
                                return True
                    except:
                        continue
            except:
                pass

            return False
        except:
            return False

    def get_complex_articles(self, driver, complex_id, complex_name):
        """단지 페이지에서 매물 수집 (속도 최적화)"""
        articles = []

        try:
            # 단지 매물 페이지 접속
            url = f"https://fin.land.naver.com/complexes/{complex_id}?tab=article"
            driver.get(url)
            time.sleep(2)  # 3초 → 2초

            # 페이지 스크롤하며 매물 수집
            collected_texts = set()
            scroll_count = 0
            max_scroll = 50  # 100 → 50으로 줄이고 효율적으로
            no_new_count = 0

            self.update_status(f"'{complex_name}' 매물 수집 중...")

            # 더보기 버튼 먼저 빠르게 클릭
            for _ in range(10):
                if self.stop_search:
                    break
                if not self.click_more_button(driver):
                    break
                time.sleep(0.5)  # 1초 → 0.5초

            while scroll_count < max_scroll:
                if self.stop_search:
                    break

                # 매물 항목 찾기 - 가장 효과적인 셀렉터만 사용
                items = driver.find_elements(By.CSS_SELECTOR, "li[class*='ArticleCard'][class*='__item']")
                if not items:
                    items = driver.find_elements(By.CSS_SELECTOR, "[class*='ArticleCard'][class*='item']")

                new_items_count = 0
                for item in items:
                    try:
                        text = item.text
                        if not text or len(text) < 20:
                            continue

                        text_key = text[:50]
                        if text_key in collected_texts:
                            continue
                        collected_texts.add(text_key)
                        new_items_count += 1

                        article = self.parse_article(text, complex_name)
                        if article and article.get('거래유형'):
                            articles.append(article)
                    except:
                        continue

                # 5회마다 상태 업데이트
                if scroll_count % 5 == 0:
                    self.update_status(f"'{complex_name}' 수집 중... ({len(articles)}건)")

                if new_items_count == 0:
                    no_new_count += 1
                else:
                    no_new_count = 0

                # 연속 5회 새 항목 없으면 종료 (10 → 5)
                if no_new_count >= 5:
                    break

                # 빠른 스크롤
                driver.execute_script("window.scrollBy(0, 800);")
                time.sleep(0.3)  # 0.5초 → 0.3초
                scroll_count += 1

                self.update_progress(min((scroll_count / max_scroll) * 100, 99))

            logging.info(f"수집된 매물 수: {len(articles)}")

        except Exception as e:
            logging.error(f"매물 수집 오류 ({complex_name}): {e}")

        return articles

    def parse_article(self, text, complex_name):
        """매물 텍스트 파싱"""
        try:
            article = {
                '단지명': complex_name,
                '거래유형': '',
                '전용면적': '',
                '층': '',
                '매매가': '',
                '보증금': '',
                '월세': '',
                '중개사명': ''
            }

            lines = text.split('\n')

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # 거래유형
                if '매매' in line and not article['거래유형']:
                    article['거래유형'] = '매매'
                elif '전세' in line and not article['거래유형']:
                    article['거래유형'] = '전세'
                elif '월세' in line and not article['거래유형']:
                    article['거래유형'] = '월세'

                # 면적 - "132.69㎡/97.78㎡" 형식에서 뒤의 숫자(전용면적) 추출
                if not article['전용면적']:
                    # 두 개의 면적이 있는 경우 (공급/전용)
                    dual_area_match = re.search(r'[\d.]+\s*㎡\s*/\s*([\d.]+)\s*㎡', line)
                    if dual_area_match:
                        article['전용면적'] = f"{dual_area_match.group(1)}㎡"
                    else:
                        # 단일 면적
                        single_area_match = re.search(r'([\d.]+)\s*㎡', line)
                        if single_area_match:
                            article['전용면적'] = f"{single_area_match.group(1)}㎡"

                # 층 - "15/20층" 형식에서 앞의 숫자(실제 층) 추출
                if not article['층']:
                    # 두 개의 층수가 있는 경우 (현재층/총층)
                    dual_floor_match = re.search(r'(\d+)\s*/\s*\d+\s*층', line)
                    if dual_floor_match:
                        article['층'] = f"{dual_floor_match.group(1)}층"
                    else:
                        # 단일 층수
                        single_floor_match = re.search(r'(\d+)\s*층', line)
                        if single_floor_match:
                            article['층'] = f"{single_floor_match.group(1)}층"

                # 가격 패턴 - 억 단위
                price_match = re.search(r'(\d+억\s*\d*,?\d*만?|\d+억)', line)
                if price_match:
                    price = price_match.group(1)
                    if article['거래유형'] == '매매' and not article['매매가']:
                        article['매매가'] = price
                    elif article['거래유형'] == '전세' and not article['보증금']:
                        article['보증금'] = price
                    elif article['거래유형'] == '월세' and not article['보증금']:
                        article['보증금'] = price

                # 만원 단위
                if not price_match:
                    won_match = re.search(r'(\d+,?\d*만)', line)
                    if won_match:
                        price = won_match.group(1)
                        if article['거래유형'] == '월세' and article['보증금'] and not article['월세']:
                            article['월세'] = price

                # 중개사
                if ('부동산' in line or '공인' in line) and not article['중개사명']:
                    article['중개사명'] = line[:25]

            return article if article['거래유형'] else None

        except:
            return None

    def display_results(self, properties):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for prop in properties:
            values = (
                prop.get('단지명', ''),
                prop.get('거래유형', ''),
                prop.get('전용면적', ''),
                prop.get('층', ''),
                prop.get('매매가', ''),
                prop.get('보증금', ''),
                prop.get('월세', ''),
                prop.get('중개사명', '')
            )
            self.tree.insert('', 'end', values=values)

        self.result_count_var.set(f"총 {len(properties)}건")

    def clear_results(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.properties_data = []
        self.result_count_var.set("총 0건")
        self.progress_bar['value'] = 0

    def search_completed(self):
        self.search_button.config(state="normal")
        self.stop_button.config(state="disabled")
        if self.properties_data:
            self.export_button.config(state="normal")

    def update_status(self, message):
        self.root.after(0, lambda: self.status_var.set(message))

    def update_progress(self, value):
        self.root.after(0, lambda: self.progress_bar.configure(value=value))

    def export_to_excel(self):
        if not self.properties_data:
            messagebox.showwarning("경고", "저장할 데이터가 없습니다.")
            return

        try:
            complex_name = self.complex_name_var.get().strip()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # 파일명에 사용 불가한 문자 제거
            safe_name = re.sub(r'[\\/*?:"<>|]', '', complex_name)
            filename = f"매물조회_{safe_name}_{timestamp}.xlsx"
            filepath = os.path.join(self.download_path, filename)

            df = pd.DataFrame(self.properties_data)
            df['조회일시'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='매물목록')

            self.format_excel(filepath)

            messagebox.showinfo("저장 완료", f"파일이 저장되었습니다.\n{filepath}")
            logging.info(f"Excel 저장 완료: {filepath}")

        except Exception as e:
            logging.error(f"Excel 저장 오류: {e}")
            messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다.\n{str(e)}")

    def format_excel(self, filepath):
        try:
            wb = load_workbook(filepath)
            ws = wb.active

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(filepath)

        except Exception as e:
            logging.error(f"Excel 서식 적용 오류: {e}")

    def run(self):
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()

    def on_closing(self):
        self.close_driver()
        self.root.destroy()


if __name__ == "__main__":
    app = NaverRealEstateAppV1()
    app.run()
